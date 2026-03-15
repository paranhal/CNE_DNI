# -*- coding: utf-8 -*-
"""전부하 최종데이터 빠른 분석: 소형 파일 1개만, W열='평균값 데이터' 행만."""
from pathlib import Path
from collections import defaultdict
import openpyxl

DATA_DIR = Path(r"G:\내 드라이브\CNE\최종데이터")
COL_W = 23

# 계룡 1차 (약 11MB) 등 상대적으로 작은 파일 먼저
SMALL = "전부하_원데이터_1차_20260315_133425_계룡.xlsx"


def main():
    path = DATA_DIR / SMALL
    if not path.exists():
        print("파일 없음:", path)
        return
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    header = [ws.cell(1, c).value for c in range(1, 25)]
    print("헤더(1~24열):", header)
    print()
    by_school_device = defaultdict(lambda: defaultdict(int))
    total_w = 0
    for r in range(2, min(ws.max_row + 1, 50000)):
        if ws.cell(r, COL_W).value != "평균값 데이터":
            continue
        total_w += 1
        school = ws.cell(r, 3).value
        code = ws.cell(r, 22).value
        memo1 = ws.cell(r, 6).value
        key_school = code or school or "(빈값)"
        key_dev = memo1 if memo1 is not None else "(빈값)"
        by_school_device[key_school][key_dev] += 1
    wb.close()
    print(f"파일: {path.name}")
    print(f"W열 '평균값 데이터' 행 수: {total_w}")
    print(f"학교 수: {len(by_school_device)}")
    for school, devs in sorted(by_school_device.items())[:15]:
        total_dev = sum(devs.values())
        counts = list(devs.values())
        ok600 = sum(1 for c in counts if c == 600)
        print(f"  {school}: 장비 {len(devs)}개, 총 행 {total_dev}, 600개인 장비 {ok600}/{len(devs)}")
        for dev, cnt in list(devs.items())[:3]:
            print(f"      {dev} -> {cnt}행")
    print("\n(빠른 분석 완료)")


if __name__ == "__main__":
    main()
