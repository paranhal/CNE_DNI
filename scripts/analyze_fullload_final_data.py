# -*- coding: utf-8 -*-
"""
G:\\내 드라이브\\CNE\\최종데이터 의 전부하 원데이터 지역별 파일 분석.
W열(23열) = "평균값 데이터" 인 행만 사용. 학교별·장비(메모1)별 600개 데이터 확인.
"""
import os
from pathlib import Path
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("openpyxl 필요: pip install openpyxl")
    raise

DATA_DIR = Path(r"G:\내 드라이브\CNE\최종데이터")
COL_W = 23  # 평균값 데이터
HEADER_ROW = 1

# 원데이터 열 인덱스 (1-based, OUTPUT_HEADERS 순서)
COL_SCHOOL = 3      # 학교
COL_SCHOOL_CODE = 22  # 학교코드
COL_MEMO1 = 6       # 메모1 (장비 식별)
COL_MEASURE_TYPE = 9  # 측정유형 (전부하 등)


def analyze_file(path: Path) -> dict:
    """한 파일에서 W열='평균값 데이터' 행만 로드해 학교·장비별 개수 집계."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    # 시트 이름 확인
    sheet_name = ws.title
    # 헤더 확인 (1행)
    header = [ws.cell(HEADER_ROW, c).value for c in range(1, 25)]
    # W열 값으로 필터: 2행부터 순회
    by_school_device = defaultdict(lambda: defaultdict(int))  # school -> memo1 -> count
    total_avg_rows = 0
    for r in range(2, ws.max_row + 1):
        val_w = ws.cell(r, COL_W).value
        if val_w != "평균값 데이터":
            continue
        total_avg_rows += 1
        school = ws.cell(r, COL_SCHOOL).value
        school_code = ws.cell(r, COL_SCHOOL_CODE).value
        memo1 = ws.cell(r, COL_MEMO1).value
        key_school = school_code if school_code else school
        if key_school is None:
            key_school = "(빈값)"
        key_device = memo1 if memo1 is not None else "(빈값)"
        by_school_device[key_school][key_device] += 1
    wb.close()
    return {
        "path": str(path.name),
        "sheet": sheet_name,
        "header_preview": header[:5],
        "total_avg_rows": total_avg_rows,
        "by_school_device": dict(by_school_device),
        "schools": list(by_school_device.keys()),
    }


def main():
    if not DATA_DIR.exists():
        print(f"[오류] 폴더 없음: {DATA_DIR}")
        return
    files = sorted(DATA_DIR.glob("전부하_원데이터_*.xlsx"))
    print(f"전부하 원데이터 파일 수: {len(files)}\n")
    all_results = []
    for path in files:
        try:
            res = analyze_file(path)
            all_results.append(res)
        except Exception as e:
            print(f"[예외] {path.name}: {e}")
            all_results.append({"path": path.name, "error": str(e)})
    # 요약 출력
    out_lines = []
    out_lines.append("=" * 60)
    out_lines.append("전부하 최종데이터 분석 (W열 = '평균값 데이터' 행만)")
    out_lines.append("=" * 60)
    total_rows_all = 0
    for res in all_results:
        if "error" in res:
            out_lines.append(f"\n[파일] {res['path']} -> 오류: {res['error']}")
            continue
        out_lines.append(f"\n[파일] {res['path']}")
        out_lines.append(f"  시트: {res['sheet']}, '평균값 데이터' 행 수: {res['total_avg_rows']}")
        total_rows_all += res["total_avg_rows"]
        by_sd = res["by_school_device"]
        out_lines.append(f"  학교 수: {len(by_sd)}")
        not_600 = []
        for school, devices in by_sd.items():
            for device, count in devices.items():
                if count != 600:
                    not_600.append((school, device, count))
        if not_600:
            out_lines.append(f"  600개 아님: {len(not_600)}건 -> {not_600[:10]}")
        else:
            out_lines.append("  학교·장비별 모두 600개 OK")
        # 학교별 장비 수
        for school in list(by_sd.keys())[:5]:
            devs = by_sd[school]
            out_lines.append(f"    {school}: 장비 {len(devs)}개, 행수 {sum(devs.values())}")
        if len(by_sd) > 5:
            out_lines.append(f"    ... 외 {len(by_sd) - 5}개 학교")
    out_lines.append("\n" + "=" * 60)
    out_lines.append(f"전체 '평균값 데이터' 행 합계: {total_rows_all}")
    out_lines.append("=" * 60)
    text = "\n".join(out_lines)
    print(text)
    out_path = Path(__file__).resolve().parent.parent / "docs" / "전부하_최종데이터_분석결과.txt"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(text)
    print(f"\n결과 저장: {out_path}")


if __name__ == "__main__":
    main()
