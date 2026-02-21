"""
원본 가상자산 엑셀: 오류 행 수정 후 수정본 저장, 기존 파일은 이름뒤에 OLD 붙여 백업.

- 원본 파일 → 동일 폴더에 {원본이름}_OLD.xlsx 로 복사
- 수정 적용한 워크북 → {원본이름}_수정본.xlsx 로 저장
- 수정 가능 케이스: 관리코드 8자리 학교코드 → 같은 학교명의 9자리 코드로 치환, 학교정보 8자리 → 장비 시트에서 9자리 추출해 치환
"""

from __future__ import annotations

import re
import shutil
from pathlib import Path
from typing import Any

import pandas as pd
import openpyxl

from .sheet_defs import VA_DATA_SHEETS_ALL, get_va_header_row
from .data_quality import (
    SCHOOL_CODE_PATTERN,
    is_valid_school_code,
)

# 관리코드 오기재 보정 규칙 (학교코드 부분): (패턴, 치환) → 치환 후 9자리 형식이면 적용
# 10C → 108 (C 오타), VH → HS (고등학교)
TYPO_RULES_PREFIX: list[tuple[str, str]] = [
    ("10C", "108"),
    ("VH", "HS"),
]

# 8자리 학교코드 패턴 (오류 케이스)
SCHOOL_CODE_8_DIGIT = re.compile(r"^N\d{8}(ES|MS|HS|SS)$")


def _header_row(ws: openpyxl.worksheet.worksheet.Worksheet, sheet_name: str) -> int:
    return get_va_header_row(sheet_name)


def _col_index_by_header(ws: openpyxl.worksheet.worksheet.Worksheet, header_row: int, name: str) -> int | None:
    """1-based column index for header name, or None."""
    for col, cell in enumerate(ws[header_row], start=1):
        if cell.value and str(cell.value).strip() == name:
            return col
    return None


def _build_school_name_to_9digit_code(ws_school: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, str]:
    """학교정보 시트에서 학교명 -> 9자리 학교코드 매핑 (9자리만)."""
    out = {}
    header_row = _header_row(ws_school, "학교정보")
    col_code = _col_index_by_header(ws_school, header_row, "학교코드")
    col_name = _col_index_by_header(ws_school, header_row, "학교명")
    if col_code is None or col_name is None:
        return out
    for row in range(header_row + 1, ws_school.max_row + 1):
        code_val = ws_school.cell(row=row, column=col_code).value
        name_val = ws_school.cell(row=row, column=col_name).value
        if pd.isna(name_val) or not str(name_val).strip():
            continue
        name = str(name_val).strip()
        if pd.isna(code_val):
            continue
        s = str(code_val).strip()
        if is_valid_school_code(s):
            out[name] = s
    return out


def _build_school_name_to_9digit_from_equipment(
    wb: openpyxl.Workbook,
    equipment_sheets: tuple[str, ...],
) -> dict[str, set[str]]:
    """장비 시트들에서 학교명 -> 9자리 학교코드 집합 (관리번호에서 추출)."""
    out: dict[str, set[str]] = {}
    for sheet_name in equipment_sheets:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        header_row = _header_row(ws, sheet_name)
        col_mgmt = _col_index_by_header(ws, header_row, "관리번호")
        col_name = _col_index_by_header(ws, header_row, "학교명")
        if col_mgmt is None or col_name is None:
            continue
        for row in range(header_row + 1, ws.max_row + 1):
            mgmt = ws.cell(row=row, column=col_mgmt).value
            name_val = ws.cell(row=row, column=col_name).value
            if pd.isna(name_val) or not str(name_val).strip():
                continue
            name = str(name_val).strip()
            if pd.isna(mgmt) or "-" not in str(mgmt):
                continue
            prefix = str(mgmt).strip().split("-", 1)[0]
            if is_valid_school_code(prefix):
                out.setdefault(name, set()).add(prefix)
    return out


def _fix_management_code_value(
    value: str,
    school_name_to_code: dict[str, str],
    school_name: str | None,
) -> str | None:
    """관리번호 값이 8자리 학교코드-접미사 형태면 9자리로 치환한 값 반환, 불가면 None."""
    value = value.strip()
    if "-" not in value:
        return None
    prefix, suffix = value.split("-", 1)
    prefix, suffix = prefix.strip(), suffix.strip()
    if not SCHOOL_CODE_8_DIGIT.match(prefix):
        return None
    if not school_name:
        return None
    code_9 = school_name_to_code.get(school_name)
    if not code_9:
        return None
    return f"{code_9}-{suffix}"


def _fix_management_code_typos(value: str) -> str | None:
    """관리번호 학교코드 부분에 오타 보정 규칙 적용 후 유효하면 반환."""
    value = value.strip()
    if "-" not in value:
        return None
    prefix, suffix = value.split("-", 1)
    prefix, suffix = prefix.strip(), suffix.strip()
    for old, new in TYPO_RULES_PREFIX:
        new_prefix = prefix.replace(old, new)
        if new_prefix != prefix and is_valid_school_code(new_prefix):
            return f"{new_prefix}-{suffix}"
    return None


def _fix_school_code_value(
    school_name: str | None,
    name_to_codes: dict[str, set[str]],
) -> str | None:
    """학교정보의 학교코드: 해당 학교명의 9자리 코드가 유일하면 반환."""
    if not school_name:
        return None
    codes = name_to_codes.get(school_name)
    if not codes or len(codes) != 1:
        return None
    return next(iter(codes))


def _write_modification_log(
    log_path: Path,
    records: list[dict[str, Any]],
    source_file_name: str,
) -> None:
    """수정 이력 CSV 저장. 검증용."""
    log_path = Path(log_path)
    log_path.parent.mkdir(parents=True, exist_ok=True)
    if not records:
        df = pd.DataFrame(columns=["원본파일", "시트", "행", "컬럼", "수정전_값", "수정후_값"])
    else:
        df = pd.DataFrame(records)
    df.to_csv(log_path, index=False, encoding="utf-8-sig")


def run_fix_and_save(
    va_path: Path | str,
    issue_csv_path: Path | str | None = None,
    output_dir: Path | str | None = None,
    modification_log_path: Path | str | None = None,
) -> tuple[Path, Path, int, Path]:
    """
    원본 복사(OLD) → 수정 적용 → 수정본 저장 → 수정 이력 CSV 생성.
    Returns: (path_old, path_revised, num_fixed, path_modification_log).
    """
    va_path = Path(va_path)
    if not va_path.exists():
        raise FileNotFoundError(f"원본 파일 없음: {va_path}")

    parent = va_path.parent
    stem = va_path.stem
    ext = va_path.suffix
    out_dir_for_log = parent
    path_old = parent / f"{stem}_OLD{ext}"
    path_revised = parent / f"{stem}_수정본{ext}"
    if output_dir is not None:
        out = Path(output_dir)
        out.mkdir(parents=True, exist_ok=True)
        path_old = out / f"{stem}_OLD{ext}"
        path_revised = out / f"{stem}_수정본{ext}"
        out_dir_for_log = out
    if modification_log_path is None:
        modification_log_path = out_dir_for_log / f"수정이력_{stem}.csv"
    else:
        modification_log_path = Path(modification_log_path)

    # 1) 원본을 _OLD.xlsx 로 복사
    shutil.copy2(va_path, path_old)

    # 2) 이슈 목록 로드
    if issue_csv_path is None:
        # output/품질검사_이상목록.csv 기본
        proj = va_path.resolve().parent
        for _ in range(5):
            cand = proj / "output" / "품질검사_이상목록.csv"
            if cand.exists():
                issue_csv_path = cand
                break
            proj = proj.parent
        if issue_csv_path is None:
            issue_csv_path = Path(__file__).resolve().parent.parent / "output" / "품질검사_이상목록.csv"
    issue_csv_path = Path(issue_csv_path)
    if not issue_csv_path.exists():
        raise FileNotFoundError(f"이상 목록 CSV 없음: {issue_csv_path}")

    issues_df = pd.read_csv(issue_csv_path, encoding="utf-8-sig")
    modification_log: list[dict[str, Any]] = []

    if issues_df.empty:
        wb = openpyxl.load_workbook(va_path)
        wb.save(path_revised)
        wb.close()
        _write_modification_log(modification_log_path, modification_log, str(va_path.name))
        return path_old, path_revised, 0, modification_log_path

    # 3) 워크북 로드 (수정용)
    wb = openpyxl.load_workbook(va_path)

    # 4) 학교명 -> 9자리 학교코드 (학교정보 시트)
    if "학교정보" in wb.sheetnames:
        school_name_to_code = _build_school_name_to_9digit_code(wb["학교정보"])
    else:
        school_name_to_code = {}

    # 5) 학교명 -> 9자리 코드 집합 (장비 시트, 학교정보 학교코드 수정용)
    from .sheet_defs import VA_DATA_SHEETS_EQUIPMENT
    name_to_codes = _build_school_name_to_9digit_from_equipment(wb, VA_DATA_SHEETS_EQUIPMENT)
    # 학교정보에서 이미 9자리인 것도 보강
    for name, code in school_name_to_code.items():
        name_to_codes.setdefault(name, set()).add(code)

    num_fixed = 0
    for _, row in issues_df.iterrows():
        sheet_name = row.get("sheet")
        excel_row = row.get("excel_row")
        column = row.get("column")
        value = row.get("value")
        issue_type = row.get("issue_type")
        if pd.isna(sheet_name) or sheet_name not in wb.sheetnames:
            continue
        if pd.isna(excel_row) or not isinstance(excel_row, (int, float)):
            continue
        excel_row = int(excel_row)
        ws = wb[sheet_name]
        header_row = _header_row(ws, sheet_name)
        col_idx = _col_index_by_header(ws, header_row, column)
        if col_idx is None:
            continue

        new_value: str | None = None
        if issue_type == "관리코드_형식_오류" and column == "관리번호":
            col_school = _col_index_by_header(ws, header_row, "학교명")
            school_name = None
            if col_school is not None:
                school_name = ws.cell(row=excel_row, column=col_school).value
                if not pd.isna(school_name):
                    school_name = str(school_name).strip()
            new_value = _fix_management_code_value(
                str(value) if not pd.isna(value) else "",
                school_name_to_code,
                school_name,
            )
            if new_value is None:
                new_value = _fix_management_code_typos(str(value) if not pd.isna(value) else "")
        elif issue_type == "학교코드_형식_오류" and column == "학교코드" and sheet_name == "학교정보":
            col_school = _col_index_by_header(ws, header_row, "학교명")
            school_name = None
            if col_school is not None:
                school_name = ws.cell(row=excel_row, column=col_school).value
                if not pd.isna(school_name):
                    school_name = str(school_name).strip()
            new_value = _fix_school_code_value(school_name, name_to_codes)

        if new_value is not None:
            old_val_str = "" if pd.isna(value) else str(value).strip()
            ws.cell(row=excel_row, column=col_idx).value = new_value
            modification_log.append({
                "원본파일": va_path.name,
                "시트": sheet_name,
                "행": excel_row,
                "컬럼": column,
                "수정전_값": old_val_str,
                "수정후_값": new_value,
            })
            num_fixed += 1

    wb.save(path_revised)
    wb.close()
    _write_modification_log(modification_log_path, modification_log, va_path.name)
    return path_old, path_revised, num_fixed, modification_log_path


# 품질검사에서 사용하는 기본 원본 경로 (config 없을 때)
DEFAULT_VA_PATH = Path(
    "/Users/paranhal/Library/CloudStorage/GoogleDrive-paranhanl66@gmail.com"
    "/내 드라이브/20260215_D드라이브/Lee_20260202/충남_통합_가상자산DB_Lee.xlsx"
)


def main() -> None:
    import sys
    try:
        from .config_loader import get_path
        base = get_path("raw_data_root")
        va_path = base / "충남_통합_가상자산DB_Lee.xlsx"
    except Exception:
        va_path = DEFAULT_VA_PATH
    if not va_path.exists():
        va_path = DEFAULT_VA_PATH
    if len(sys.argv) > 1:
        va_path = Path(sys.argv[1])
    out_dir = None
    if len(sys.argv) > 2:
        out_dir = Path(sys.argv[2])

    if not va_path.exists():
        print("파일 없음:", va_path)
        return
    path_old, path_revised, num_fixed, path_log = run_fix_and_save(va_path, output_dir=out_dir)
    print("OLD 백업:", path_old)
    print("수정본 저장:", path_revised)
    print("수정 반영 건수:", num_fixed)
    print("수정 이력(검증용):", path_log)


if __name__ == "__main__":
    main()
