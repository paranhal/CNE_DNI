# -*- coding: utf-8 -*-
"""
DNI_ISP_MEANSURE.XLSX → DNI_ISP_MEASURE.XLSX 변환 스크립트 (대전 ISP 1·2차 측정 통합)

[입력]
- D:\CNE_DNI\src\measure\DNI\DNI_ISP_MEANSURE.XLSX
  - 시트명: "대전" (가정)
  - 열 구성(1-based 열 번호 기준):
    - E(5)~I(9): 최종 측정값(다운로드, 업로드, RTT, RSSI, CH) → 이 스크립트에서 채움
    - J(10)~N(14): 1차 측정값
    - O(15): 1차 측정값 판정
    - P(16)~T(20): 2차 측정값
    - U(21): 2차 측정값 판정

[출력]
- D:\CNE_DNI\src\measure\DNI\DNI_ISP_MEASURE.XLSX
  - 시트명: "ISP측정"
  - E~I 열에 최종 측정값이 들어 있는 형태 (CNE_ISP_MEASURE.XLSX와 동일 구조 가정)

[최종 측정값 산정 규칙]
- 1차(O열)가 "양호"이면: J~N(1차) 값을 최종(E~I)에 사용
- 그렇지 않고 2차(U열)가 "양호"이면: P~T(2차) 값을 최종(E~I)에 사용
- 둘 다 "양호"가 아니면: 최종(E~I)은 비워 둠(기존 값 유지 또는 빈칸)
"""
from __future__ import annotations

import os
import sys
from typing import Optional, Tuple, List

from openpyxl import load_workbook

if getattr(sys, "frozen", False):
    _BASE_DIR = os.path.dirname(sys.executable)
else:
    _BASE_DIR = os.path.dirname(os.path.abspath(__file__))
_RUN_DIR = os.getcwd()

INPUT_FILENAME = "DNI_ISP_MEANSURE.XLSX"
OUTPUT_FILENAME = "DNI_ISP_MEASURE.XLSX"

def _pick_existing(candidates):
    for p in candidates:
        if os.path.isfile(p):
            return p
    return candidates[0]


def _resolve_input_path():
    return _pick_existing(
        [
            os.path.join(_RUN_DIR, INPUT_FILENAME),
            os.path.join(_RUN_DIR, "DNI", INPUT_FILENAME),
            os.path.join(_BASE_DIR, INPUT_FILENAME),
            os.path.join(_BASE_DIR, "DNI", INPUT_FILENAME),
        ]
    )


def _resolve_output_path(input_path):
    in_dir = os.path.dirname(input_path) if input_path else _RUN_DIR
    if in_dir and os.path.isdir(in_dir):
        return os.path.join(in_dir, OUTPUT_FILENAME)
    return os.path.join(_RUN_DIR, OUTPUT_FILENAME)


INPUT_PATH = _resolve_input_path()
OUTPUT_PATH = _resolve_output_path(INPUT_PATH)


def _log(msg: str) -> None:
    print(msg, flush=True)


def find_source_sheet_name(wb) -> str:
    """
    원본 시트명 추론.
    - 우선 "대전" 시트가 있으면 사용
    - 없으면 "ISP측정" 또는 첫 번째 시트를 사용
    """
    if "대전" in wb.sheetnames:
        return "대전"
    if "ISP측정" in wb.sheetnames:
        return "ISP측정"
    return wb.sheetnames[0]


def find_header_row(ws, max_scan: int = 10) -> int:
    """
    헤더 행 추정:
    - 앞쪽 몇 행에서 '다운로드', '업로드' 등의 텍스트가 포함된 행을 찾음
    - 없으면 1행을 헤더로 간주
    """
    keywords = ("다운로드", "업로드", "RTT", "RSSI", "CH")
    for r in range(1, min(ws.max_row, max_scan) + 1):
        for c in range(1, min(ws.max_column, 30) + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and any(kw in v for kw in keywords):
                return r
    return 1


def select_final_values(
    first_vals: List[object],
    second_vals: List[object],
    judge_first: Optional[object],
    judge_second: Optional[object],
) -> Optional[List[object]]:
    """
    최종 측정값 선택 로직.

    - judge_first == "양호" → first_vals 반환
    - 그렇지 않고 judge_second == "양호" → second_vals 반환
    - 둘 다 "양호"가 아니면:
        - 1차 값이 하나라도 있으면 first_vals
        - 1차 값이 전부 비어 있으면 second_vals
        - 그래도 없으면 None
    """
    j1 = str(judge_first).strip() if judge_first is not None else ""
    j2 = str(judge_second).strip() if judge_second is not None else ""

    if j1 == "양호":
        return first_vals
    if j2 == "양호":
        return second_vals
    # 둘 다 "양호"가 아닌 경우: 1차 우선, 1차가 없으면 2차 사용
    def _has_value(values: List[object]) -> bool:
        return any(v not in (None, "") for v in values)

    if _has_value(first_vals):
        return first_vals
    if _has_value(second_vals):
        return second_vals
    return None


def process_dni_isp_meansure(
    input_path: str = INPUT_PATH,
    output_path: str = OUTPUT_PATH,
) -> bool:
    if not os.path.isfile(input_path):
        _log(f"[오류] 입력 파일을 찾을 수 없습니다: {input_path}")
        return False

    _log("=" * 60)
    _log("[DNI ISP] DNI_ISP_MEANSURE.XLSX → DNI_ISP_MEASURE.XLSX 변환 시작")
    _log(f"입력:  {input_path}")
    _log(f"출력:  {output_path}")
    _log("=" * 60)

    wb = load_workbook(input_path, data_only=True)

    src_sheet_name = find_source_sheet_name(wb)
    ws = wb[src_sheet_name]

    header_row = find_header_row(ws)
    data_start_row = header_row + 1

    _log(f"원본 시트: {src_sheet_name}")
    _log(f"헤더 행: {header_row} → 데이터 시작 행: {data_start_row}")

    # 열 번호 (1-based)
    COL_E = 5   # 최종 다운로드
    COL_F = 6   # 최종 업로드
    COL_G = 7   # 최종 RTT
    COL_H = 8   # 최종 RSSI
    COL_I = 9   # 최종 CH

    COL_J = 10  # 1차 다운로드
    COL_K = 11  # 1차 업로드
    COL_L = 12  # 1차 RTT
    COL_M = 13  # 1차 RSSI
    COL_N = 14  # 1차 CH
    COL_O = 15  # 1차 판정

    COL_P = 16  # 2차 다운로드
    COL_Q = 17  # 2차 업로드
    COL_R = 18  # 2차 RTT
    COL_S = 19  # 2차 RSSI
    COL_T = 20  # 2차 CH
    COL_U = 21  # 2차 판정

    updated_rows = 0
    total_rows = 0

    for r in range(data_start_row, ws.max_row + 1):
        total_rows += 1

        first_vals = [
            ws.cell(row=r, column=COL_J).value,
            ws.cell(row=r, column=COL_K).value,
            ws.cell(row=r, column=COL_L).value,
            ws.cell(row=r, column=COL_M).value,
            ws.cell(row=r, column=COL_N).value,
        ]
        second_vals = [
            ws.cell(row=r, column=COL_P).value,
            ws.cell(row=r, column=COL_Q).value,
            ws.cell(row=r, column=COL_R).value,
            ws.cell(row=r, column=COL_S).value,
            ws.cell(row=r, column=COL_T).value,
        ]
        judge_first = ws.cell(row=r, column=COL_O).value
        judge_second = ws.cell(row=r, column=COL_U).value

        final_vals = select_final_values(first_vals, second_vals, judge_first, judge_second)
        if final_vals is None:
            # 둘 다 "양호"가 아니면 최종값을 채우지 않음 (기존 값 유지 또는 빈칸)
            continue

        ws.cell(row=r, column=COL_E, value=final_vals[0])
        ws.cell(row=r, column=COL_F, value=final_vals[1])
        ws.cell(row=r, column=COL_G, value=final_vals[2])
        ws.cell(row=r, column=COL_H, value=final_vals[3])
        ws.cell(row=r, column=COL_I, value=final_vals[4])

        updated_rows += 1

    _log(f"총 데이터 행 수: {total_rows}행")
    _log(f"최종 측정값이 채워진 행 수: {updated_rows}행")

    # CNE_ISP_MEASURE 구조와 맞추기 위해 시트명을 'ISP측정'으로 변경
    ws.title = "ISP측정"

    # 새 파일로 저장 (원본은 그대로 유지)
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    wb.close()

    _log(f"[완료] 변환 완료 → {output_path}")
    return True


def main() -> bool:
    return process_dni_isp_meansure()


if __name__ == "__main__":
    main()

