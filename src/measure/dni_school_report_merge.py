# -*- coding: utf-8 -*-
"""
대전(DNI) 학교별 리포트 통합 - 가로형 Excel 생성

학교별_리포트 폴더의 개별 xlsx 파일들을 읽어
학교 1행 × 항목 N열 가로형으로 통합 저장
"""
from __future__ import print_function
import sys
import io
import os
from datetime import datetime

if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from tqdm import tqdm

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DNI_DIR = os.path.join(BASE_DIR, "DNI")
RUN_DIR = os.getcwd()

SHEET_NAME = "문제점분석"
DATA_ROWS = list(range(2, 35))  # 2~34행
F_COL = 6
H_COL = 8

ITEM_NAMES = {
    2: "100M급 장비 수",
    3: "모델명, 제조사",
    4: "PoE 사용포트 수",
    5: "장비간 연결구성",
    6: "케이블 타입",
    7: "포트 속도 설정",
    8: "2.4GHz 사용 여부",
    9: "보안 설정",
    10: "불량 AP 수",
    11: "신호강도 (4지점)",
    12: "채널 대역폭",
    13: "채널 설정",
    14: "채널 중첩 및 간섭",
    15: "AP 모델, 위치",
    16: "SSID, BSSID",
    17: "WiFi 규격",
    18: "계위 단계 수",
    19: "Cat5 사용 수",
    20: "망분리 현황",
    21: "전부하 평균속도",
    22: "교실별 개별속도",
    23: "무선 다운로드",
    24: "무선 업로드",
    25: "신호강도",
    26: "채널 중첩수",
    27: "채널설정 확인",
    28: "지연시간(RTT)",
    29: "집선↔외부망 다운로드",
    30: "집선↔외부망 업로드",
    31: "패킷손실률",
    32: "Jitter",
    33: "L2↔집선 다운로드",
    34: "L2↔집선 업로드",
}


def _input(prompt):
    try:
        return input(prompt).strip()
    except (EOFError, KeyboardInterrupt):
        print("\n[중단] 사용자 입력으로 종료합니다.")
        sys.exit(1)


def _pick_file(prompt, candidates):
    existing = [p for p in candidates if os.path.isfile(p)]
    if existing:
        print(f"\n{prompt}")
        for i, p in enumerate(existing, 1):
            print(f"  {i}. {p}")
        print("  0. 직접 경로 입력")
        s = _input("번호 선택 (Enter: 1번): ")
        if not s:
            return existing[0]
        if s == "0":
            return _input("전체 경로 입력: ")
        try:
            idx = int(s)
            if 1 <= idx <= len(existing):
                return existing[idx - 1]
        except ValueError:
            pass
        print("[경고] 잘못된 입력. 1번 사용")
        return existing[0]
    print(f"\n{prompt}")
    print("  후보 파일이 없어 직접 경로를 입력하세요.")
    return _input("전체 경로 입력: ")


def _pick_report_dir():
    candidates = [
        os.path.join(RUN_DIR, "학교별_리포트"),
        os.path.join(RUN_DIR, "학교별_리포트_V1.1"),
        os.path.join(BASE_DIR, "DNI", "학교별_리포트"),
        os.path.join(BASE_DIR, "CNE", "학교별_리포트"),
        os.path.join(BASE_DIR, "DNI", "학교별_리포트_V1.1"),
        os.path.join(BASE_DIR, "CNE", "학교별_리포트_V1.1"),
    ]
    existing = [p for p in candidates if os.path.isdir(p)]
    if existing:
        print("\n학교별 리포트 폴더를 선택하세요.")
        for i, p in enumerate(existing, 1):
            print(f"  {i}. {p}")
        print("  0. 직접 경로 입력")
        s = _input("번호 선택 (Enter: 1번): ")
        if not s:
            return existing[0]
        if s == "0":
            return _input("폴더 전체 경로 입력: ")
        try:
            idx = int(s)
            if 1 <= idx <= len(existing):
                return existing[idx - 1]
        except ValueError:
            pass
        print("[경고] 잘못된 입력. 1번 사용")
        return existing[0]
    print("\n학교별 리포트 폴더를 찾지 못했습니다. 직접 경로를 입력하세요.")
    return _input("폴더 전체 경로 입력: ")


def load_region_map(stats_file):
    """Sheet1에서 학교코드 → 지역 매핑"""
    code_to_region = {}
    if not os.path.isfile(stats_file):
        return code_to_region
    wb = load_workbook(stats_file, data_only=True, read_only=True)
    if "Sheet1" in wb.sheetnames:
        ws = wb["Sheet1"]
        for r in range(2, ws.max_row + 1):
            code = str(ws.cell(r, 1).value or "").strip()
            region = str(ws.cell(r, 2).value or "").strip()
            if code and region:
                code_to_region[code] = region
    wb.close()
    return code_to_region


def extract_school_data(filepath):
    """학교 리포트 파일에서 F열·H열 데이터 + 정상/개선필요 카운트 추출"""
    wb = load_workbook(filepath, data_only=True, read_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    f_vals = {}
    h_vals = {}
    for r in DATA_ROWS:
        f_vals[r] = ws.cell(r, F_COL).value
        h_vals[r] = ws.cell(r, H_COL).value
    wb.close()
    return f_vals, h_vals


def build_headers():
    """가로형 헤더 구성"""
    headers = ["학교코드", "학교명", "지역"]
    for r in DATA_ROWS:
        name = ITEM_NAMES.get(r, f"항목{r}")
        headers.append(f"{name}_측정값")
        headers.append(f"{name}_판정")
    headers.append("정상_수")
    headers.append("개선필요_수")
    return headers


def main():
    print("=" * 50)
    print("[대전(DNI) 학교별 리포트 통합] 가로형 생성")
    print("=" * 50)

    report_dir = _pick_report_dir()
    stats_file = _pick_file(
        "통계 파일(TOTAL_MEASURE_LIST)을 선택하세요.",
        [
            os.path.join(RUN_DIR, "DNI_TOTAL_MEASURE_LIST_V1.xlsx"),
            os.path.join(RUN_DIR, "TOTAL_MEASURE_LIST_V1.xlsx"),
            os.path.join(RUN_DIR, "CNE", "TOTAL_MEASURE_LIST_V1.xlsx"),
            os.path.join(BASE_DIR, "DNI", "DNI_TOTAL_MEASURE_LIST_V1.xlsx"),
            os.path.join(BASE_DIR, "DNI", "TOTAL_MEASURE_LIST_V1.xlsx"),
            os.path.join(BASE_DIR, "CNE", "TOTAL_MEASURE_LIST_V1.xlsx"),
        ],
    )
    default_out_name = "DNI_학교별_문제점분석_통합.xlsx"
    out_name = _input(f"\n출력 파일명 입력 (Enter: {default_out_name}): ")
    out_name = out_name if out_name else default_out_name
    if not out_name.lower().endswith(".xlsx"):
        out_name += ".xlsx"
    output_file = os.path.join(RUN_DIR, out_name)

    print(f"\n선택된 리포트 폴더: {report_dir}")
    print(f"선택된 통계 파일: {stats_file}")
    print(f"출력 파일: {output_file}")

    if not os.path.isdir(report_dir):
        print(f"[오류] 리포트 폴더 없음: {report_dir}")
        sys.exit(1)
    if not os.path.isfile(stats_file):
        print(f"[오류] 통계 파일 없음: {stats_file}")
        sys.exit(1)

    files = []
    for dp, _, fs in os.walk(report_dir):
        for f in fs:
            if f.endswith(".xlsx") and not f.startswith("~"):
                files.append(os.path.join(dp, f))
    files = sorted(files)
    print(f"대상 파일: {len(files)}개")

    code_to_region = load_region_map(stats_file)

    headers = build_headers()
    rows_data = []

    for filepath in tqdm(files, desc="파일 읽기", unit="교"):
        fname = os.path.basename(filepath)
        parts = fname.rsplit("_", 1)
        if len(parts) < 2:
            continue
        school_code = parts[-1].replace(".xlsx", "").replace("_백업", "")
        school_name = parts[0] if parts[0] != school_code else ""
        region = code_to_region.get(school_code, "")

        try:
            f_vals, h_vals = extract_school_data(filepath)
        except Exception as e:
            print(f"\n[경고] {fname} 읽기 실패: {e}")
            continue

        row = [school_code, school_name, region]
        count_ok = 0
        count_bad = 0
        for r in DATA_ROWS:
            fv = f_vals.get(r)
            hv = h_vals.get(r)
            row.append(fv if fv is not None else "")
            row.append(hv if hv is not None else "")
            s = str(hv or "").strip()
            if s == "정상":
                count_ok += 1
            elif s == "개선필요":
                count_bad += 1
        row.append(count_ok)
        row.append(count_bad)
        rows_data.append(row)

    print(f"\n읽은 학교: {len(rows_data)}개")

    # Excel 작성
    wb = Workbook()
    ws = wb.active
    ws.title = "문제점분석_통합"

    # 스타일
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    red_font = Font(color="FF0000")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 헤더
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    # 데이터
    for ri, row in enumerate(rows_data, 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if isinstance(val, str) and val == "개선필요":
                cell.font = red_font

    # 열 너비 조정
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 8
    for c in range(4, len(headers) + 1):
        ws.column_dimensions[ws.cell(1, c).column_letter].width = 15

    # 필터
    ws.auto_filter.ref = f"A1:{ws.cell(1, len(headers)).column_letter}{len(rows_data) + 1}"

    # 행 고정
    ws.freeze_panes = "D2"

    try:
        wb.save(output_file)
        print(f"[완료] 저장: {output_file}")
    except PermissionError:
        alt = output_file.replace(".xlsx", f"_{datetime.now().strftime('%H%M%S')}.xlsx")
        wb.save(alt)
        print(f"[완료] 저장: {alt}")

    print(f"  행: {len(rows_data)}개 학교, 열: {len(headers)}개 항목")


if __name__ == "__main__":
    main()
