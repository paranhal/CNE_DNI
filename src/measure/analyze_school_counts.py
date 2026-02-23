# -*- coding: utf-8 -*-
"""
원본 자료 통계: 장비 데이터가 있는 학교 수, 학교코드 수 분석
"""
from __future__ import print_function
import sys
import io
import os

if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

from openpyxl import load_workbook

_MEASURE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, _MEASURE_DIR)
from measure_utils import extract_school_code_from_mgmt_num


def find_header_row(ws):
    for row in range(1, min(6, ws.max_row + 1)):
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None and "장비관리번호" in str(val).strip():
                return row
    for row in range(1, min(6, ws.max_row + 1)):
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                s = str(val).strip().lower()
                if any(kw in s for kw in ["다운로드", "down", "업로드", "up", "mbps"]):
                    return row
    return 2


def is_numeric(val):
    if val is None:
        return False
    if isinstance(val, (int, float)):
        return True
    try:
        float(str(val).replace(",", "").replace("%", ""))
        return True
    except ValueError:
        return False


def analyze_fullload():
    """전부하측정 시트 분석"""
    path = os.path.join(_MEASURE_DIR, "CNE", "CNE_FULLLOAD_MEASURE.xlsx")
    if not os.path.isfile(path):
        print(f"[전부하] 파일 없음: {path}")
        return
    wb = load_workbook(path, data_only=True)
    ws = wb["전부하측정"]
    header_row = find_header_row(ws)
    col_mgmt, col_dl, col_ul, col_m3, col_m4 = 3, 4, 5, 6, 7

    schools_any = set()       # 장비 1개라도 있는 학교코드
    schools_full = set()      # 4개 측정값 모두 있는 장비가 있는 학교코드
    school_equip_count = {}   # 학교별 장비 수 (전체)
    school_full_count = {}    # 학교별 완전 데이터 장비 수
    total_rows = ws.max_row - header_row
    skipped = 0

    for r in range(header_row + 1, ws.max_row + 1):
        mgmt = ws.cell(r, col_mgmt).value
        school_code = extract_school_code_from_mgmt_num(mgmt)
        if not school_code or school_code == "0":
            skipped += 1
            continue

        def _get(col):
            v = ws.cell(r, col).value
            return float(str(v).replace(",", "")) if is_numeric(v) else None

        dl, ul, m3, m4 = _get(col_dl), _get(col_ul), _get(col_m3), _get(col_m4)
        has_any = dl is not None or ul is not None or m3 is not None or m4 is not None
        has_full = dl is not None and ul is not None and m3 is not None and m4 is not None

        schools_any.add(school_code)
        school_equip_count[school_code] = school_equip_count.get(school_code, 0) + 1

        if has_full:
            schools_full.add(school_code)
            school_full_count[school_code] = school_full_count.get(school_code, 0) + 1

    print("=" * 60)
    print("[전부하측정] CNE_FULLLOAD_MEASURE.xlsx")
    print("=" * 60)
    print(f"  헤더 행: {header_row}, 데이터 행 범위: {header_row+1} ~ {ws.max_row} (총 {total_rows}행)")
    print(f"  관리번호 없음/무효로 스킵한 행: {skipped}행")
    print(f"  장비가 1개라도 있는 학교코드 수: {len(schools_any)}개")
    print(f"  4개 측정값 모두 있는 장비가 있는 학교코드 수: {len(schools_full)}개")
    print(f"  전체 데이터 행 수: {sum(school_equip_count.values())}")
    print(f"  완전 데이터 장비 수: {sum(school_full_count.values())}")
    if schools_any - schools_full:
        print(f"  [참고] 완전 데이터 없는 학교 수: {len(schools_any - schools_full)}개")
    wb.close()
    print()


def analyze_isp():
    """ISP측정 시트 분석"""
    path = os.path.join(_MEASURE_DIR, "CNE", "CNE_ISP_MEASURE.XLSX")
    if not os.path.isfile(path):
        print(f"[ISP측정] 파일 없음: {path}")
        return
    wb = load_workbook(path, data_only=True)
    ws = wb["ISP측정"]
    header_row = find_header_row(ws)
    col_mgmt, col_dl, col_ul, col_rtt, col_rssi, col_ch = 3, 5, 6, 7, 8, 9

    schools_any = set()
    schools_full = set()
    school_equip_count = {}
    school_full_count = {}
    total_rows = ws.max_row - header_row
    skipped = 0

    for r in range(header_row + 1, ws.max_row + 1):
        mgmt = ws.cell(r, col_mgmt).value
        school_code = extract_school_code_from_mgmt_num(mgmt)
        if not school_code or school_code == "0":
            skipped += 1
            continue

        def _get(col):
            v = ws.cell(r, col).value
            return float(str(v).replace(",", "")) if is_numeric(v) else None

        dl, ul, rtt, rssi, ch = _get(col_dl), _get(col_ul), _get(col_rtt), _get(col_rssi), _get(col_ch)
        has_any = dl is not None or ul is not None or rtt is not None or rssi is not None or ch is not None
        has_full = all(x is not None for x in (dl, ul, rtt, rssi, ch))

        schools_any.add(school_code)
        school_equip_count[school_code] = school_equip_count.get(school_code, 0) + 1

        if has_full:
            schools_full.add(school_code)
            school_full_count[school_code] = school_full_count.get(school_code, 0) + 1

    print("=" * 60)
    print("[ISP측정] CNE_ISP_MEASURE.XLSX")
    print("=" * 60)
    print(f"  헤더 행: {header_row}, 데이터 행 범위: {header_row+1} ~ {ws.max_row} (총 {total_rows}행)")
    print(f"  관리번호 없음/무효로 스킵한 행: {skipped}행")
    print(f"  장비가 1개라도 있는 학교코드 수: {len(schools_any)}개")
    print(f"  5개 측정값 모두 있는 장비가 있는 학교코드 수: {len(schools_full)}개")
    print(f"  전체 데이터 행 수: {sum(school_equip_count.values())}")
    print(f"  완전 데이터 장비 수: {sum(school_full_count.values())}")
    if schools_any - schools_full:
        print(f"  [참고] 완전 데이터 없는 학교 수: {len(schools_any - schools_full)}개")
    wb.close()
    print()


def main():
    print("\n[원본 자료 통계] 장비 데이터가 있는 학교 수, 학교코드 수\n")
    analyze_fullload()
    analyze_isp()
    print("=" * 60)
    print("종료")


if __name__ == "__main__":
    main()
