# -*- coding: utf-8 -*-
"""
전부하 측정 데이터 전처리

1. CNE_FULLLOAD_MEASURE.xlsx의 "전부하측정" 시트에서 학교별 평균 계산
2. 새 시트 "전부하측정_학교별평균" 추가
3. 열: 학교코드(A), 학교명, 다운로드, 업로드, (F열), (G열), 다운로드 진단, 업로드 진단
4. 진단 기준: 375 Mbps 이상 → 양호, 미만 → 미흡 (ISP와 동일)
5. 실 데이터: 소숫점 1자리, 단위 표시
6. 마지막에 "전체" 행 추가
"""
from __future__ import print_function
import sys
import io

if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

print("[전부하 측정 전처리] 스크립트 시작", flush=True)

import os
from openpyxl import load_workbook
from tqdm import tqdm

_MEASURE_DIR = os.path.dirname(os.path.abspath(__file__))
if _MEASURE_DIR not in sys.path:
    sys.path.insert(0, _MEASURE_DIR)

try:
    from measure_utils import extract_school_code_from_mgmt_num
    from fullload_preprocess_config import (
        COL_MGMT_NUM_DEFAULT,
        FULLLOAD_MEASURE_CANDIDATES,
        SHEET_FULLLOAD_SOURCE,
        SHEET_FULLLOAD_AVG,
        COL_DOWNLOAD,
        COL_UPLOAD,
        COL_MEASURE3,
        COL_MEASURE4,
        FULLLOAD_THRESHOLD_MBPS,
        SCHOOL_LIST_SEARCH_DIRS,
        NUMBER_FORMAT_DOWNLOAD,
        NUMBER_FORMAT_UPLOAD,
        NUMBER_FORMAT_MS,
        NUMBER_FORMAT_DEFAULT,
    )
except Exception as e:
    print(f"[오류] config 로드 실패: {e}", flush=True)
    sys.exit(1)


def _log(msg):
    print(msg, flush=True)


def load_school_name_map(region):
    """학교 리스트에서 학교코드 → 학교명 매핑 로드"""
    import csv
    candidates = [
        (f"school_reg_list_{region}.csv", "csv"),
        (f"SCHOOL_REG_LIST_{region}.csv", "csv"),
        (f"school_reg_list_{region}.xlsx", "xlsx"),
        (f"SCHOOL_REG_LIST_{region}.xlsx", "xlsx"),
        ("school_reg_list.csv", "csv"),
        ("SCHOOL_REG_LIST.csv", "csv"),
    ]
    for base_dir in SCHOOL_LIST_SEARCH_DIRS:
        for fname, ext in candidates:
            path = os.path.join(base_dir, fname)
            if not os.path.isfile(path):
                continue
            code_to_name = {}
            try:
                if ext == "csv":
                    with open(path, "r", encoding="utf-8-sig") as f:
                        reader = csv.reader(f)
                        header = next(reader, None)
                        rows = list(reader)
                else:
                    wb = load_workbook(path, read_only=True, data_only=True)
                    ws = wb.active
                    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
                    rows = [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)] for r in range(2, ws.max_row + 1)]
                    wb.close()
                if not header:
                    continue
                code_col = name_col = None
                for i, h in enumerate(header):
                    s = str(h or "").strip().lower()
                    if "학교코드" in s or "code" in s:
                        code_col = i
                    if "학교명" in s or "name" in s:
                        name_col = i
                if code_col is None:
                    code_col = 0
                if name_col is None:
                    name_col = 2
                for row in rows:
                    if len(row) > max(code_col, name_col):
                        code = str(row[code_col] or "").strip()
                        name = str(row[name_col] or "").strip()
                        if code:
                            code_to_name[code] = name
                            code_to_name[code[:12]] = name
                if code_to_name:
                    _log(f"  학교 리스트 로드: {path} ({len(code_to_name)}개)")
                return code_to_name
            except Exception:
                pass
    _log("  [참고] 학교 리스트 없음 → 학교명 빈칸")
    return {}


def find_mgmt_column(ws, header_row):
    """헤더에서 '장비관리번호' 또는 '관리번호' 열 인덱스 반환 (1-based)"""
    keywords = ["장비관리번호", "관리번호"]
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val is not None:
            s = str(val).strip()
            if any(kw in s for kw in keywords):
                return col
    return None


def find_header_row(ws):
    """헤더 행 찾기 - '장비관리번호' 또는 '다운로드'/'Down' 포함된 행 (실제 컬럼 헤더)"""
    # 1순위: 장비관리번호가 있는 행 (실제 데이터 헤더)
    for row in range(1, min(6, ws.max_row + 1)):
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None and "장비관리번호" in str(val).strip():
                return row
    # 2순위: 다운로드/Down, 업로드/Up 등 측정 컬럼 헤더가 있는 행
    for row in range(1, min(6, ws.max_row + 1)):
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                s = str(val).strip().lower()
                if any(kw in s for kw in ["다운로드", "down", "업로드", "up", "mbps"]):
                    return row
    return 2


def is_numeric(val):
    if val is None:
        return False
    if isinstance(val, (int, float)):
        return True
    s = str(val).strip()
    if not s:
        return False
    try:
        float(s.replace(",", "").replace("%", ""))
        return True
    except ValueError:
        return False


def _find_measure_columns(ws, header_row):
    """헤더에서 다운로드, 업로드, F열, G열 인덱스 반환 (1-based). 없으면 config 기본값."""
    col_dl = col_ul = col_m3 = col_m4 = None
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val is None:
            continue
        s = str(val).strip().lower().replace("\n", " ")
        if "down" in s or "다운로드" in s:
            if col_dl is None:
                col_dl = col
            elif col_ul is None:
                col_ul = col  # 두 번째 Down은 업로드 근처
        elif "up" in s or "업로드" in s:
            if col_ul is None:
                col_ul = col
        elif "rssi" in s and col_m3 is None:
            col_m3 = col
        elif s == "ch" and col_m4 is None:
            col_m4 = col
    # D,E,F,G 순서로 첫 4개 측정열 사용 (헤더 검색 실패 시)
    if col_dl is None:
        col_dl = COL_DOWNLOAD
    if col_ul is None:
        col_ul = COL_UPLOAD
    if col_m3 is None:
        col_m3 = COL_MEASURE3
    if col_m4 is None:
        col_m4 = COL_MEASURE4
    return col_dl, col_ul, col_m3, col_m4


def process_fullload_measure(region="CNE"):
    """전부하 측정 데이터 전처리: 학교별 평균 계산 후 새 시트 추가"""
    candidates = FULLLOAD_MEASURE_CANDIDATES.get(region, [])
    input_path = None
    for p in candidates:
        if os.path.isfile(p):
            input_path = p
            break
    if not input_path:
        _log(f"[오류] 파일을 찾을 수 없습니다. 다음 경로를 확인하세요:")
        for p in candidates:
            _log(f"       {p}")
        return
    _log(f"[진행] 지역: {region} | 파일: {input_path}")

    code_to_name = load_school_name_map(region)

    wb = load_workbook(input_path, data_only=True)
    if SHEET_FULLLOAD_SOURCE not in wb.sheetnames:
        _log(f"[오류] 시트 '{SHEET_FULLLOAD_SOURCE}'가 없습니다.")
        wb.close()
        return

    ws_src = wb[SHEET_FULLLOAD_SOURCE]
    header_row = find_header_row(ws_src)
    # 헤더 기반 열 인덱스 동적 검색 (컬럼 순서 변경에 대응)
    col_mgmt = find_mgmt_column(ws_src, header_row)
    if col_mgmt is None:
        col_mgmt = COL_MGMT_NUM_DEFAULT
        _log(f"  [참고] 헤더에서 '장비관리번호' 미검색 → C열({col_mgmt}) 사용")

    col_dl, col_ul, col_m3, col_m4 = _find_measure_columns(ws_src, header_row)
    _log(f"  열 매핑: 장비관리번호={col_mgmt}, 다운로드={col_dl}, 업로드={col_ul}, 측정3={col_m3}, 측정4={col_m4}")

    hdr_f = ws_src.cell(row=header_row, column=col_m3).value or "측정값3"
    hdr_g = ws_src.cell(row=header_row, column=col_m4).value or "측정값4"

    max_row = ws_src.max_row

    # 학교별 데이터 수집
    school_data = {}
    for r in tqdm(range(header_row + 1, max_row + 1), desc=f"[{region}] 데이터 읽기", unit="행"):
        mgmt = ws_src.cell(row=r, column=col_mgmt).value
        school_code = extract_school_code_from_mgmt_num(mgmt)
        if not school_code or school_code == "0":
            continue

        def _get_num(col):
            if col is None:
                return None
            v = ws_src.cell(row=r, column=col).value
            if is_numeric(v):
                try:
                    return float(str(v).replace(",", "").replace("%", ""))
                except ValueError:
                    return None
            return None

        dl = _get_num(col_dl)
        ul = _get_num(col_ul)
        m3 = _get_num(col_m3)
        m4 = _get_num(col_m4)

        # 데이터가 있는 장비만 포함 (4개 측정값 모두 있는 행만)
        if dl is None or ul is None or m3 is None or m4 is None:
            continue

        if school_code not in school_data:
            school_data[school_code] = {"rows": []}
        school_data[school_code]["rows"].append((dl, ul, m3, m4))

    # 학교별 평균 계산 (데이터 있는 장비들만, 소숫점 1자리)
    def _avg_from_rows(rows, idx):
        if not rows:
            return ""
        return round(sum(r[idx] for r in rows) / len(rows), 1)

    avg_rows = []
    for school_code in tqdm(sorted(school_data.keys()), desc=f"[{region}] 평균 계산", unit="학교"):
        rows = school_data[school_code]["rows"]
        if not rows:
            continue
        school_name = code_to_name.get(school_code) or code_to_name.get(school_code[:12]) or ""

        avg_dl = _avg_from_rows(rows, 0)
        avg_ul = _avg_from_rows(rows, 1)
        avg_m3 = _avg_from_rows(rows, 2)
        avg_m4 = _avg_from_rows(rows, 3)

        diag_dl = "양호" if (isinstance(avg_dl, (int, float)) and avg_dl >= FULLLOAD_THRESHOLD_MBPS) else "미흡"
        diag_ul = "양호" if (isinstance(avg_ul, (int, float)) and avg_ul >= FULLLOAD_THRESHOLD_MBPS) else "미흡"

        avg_rows.append({
            "school_code": school_code,
            "school_name": school_name,
            "download": avg_dl,
            "upload": avg_ul,
            "measure3": avg_m3,
            "measure4": avg_m4,
            "diag_dl": diag_dl,
            "diag_ul": diag_ul,
        })

    # 전체 평균 행
    if avg_rows:
        all_dl = [r["download"] for r in avg_rows if isinstance(r["download"], (int, float))]
        all_ul = [r["upload"] for r in avg_rows if isinstance(r["upload"], (int, float))]
        all_m3 = [r["measure3"] for r in avg_rows if isinstance(r["measure3"], (int, float))]
        all_m4 = [r["measure4"] for r in avg_rows if isinstance(r["measure4"], (int, float))]
        total_dl = round(sum(all_dl) / len(all_dl), 1) if all_dl else ""
        total_ul = round(sum(all_ul) / len(all_ul), 1) if all_ul else ""
        total_m3 = round(sum(all_m3) / len(all_m3), 1) if all_m3 else ""
        total_m4 = round(sum(all_m4) / len(all_m4), 1) if all_m4 else ""
        diag_dl = "양호" if (isinstance(total_dl, (int, float)) and total_dl >= FULLLOAD_THRESHOLD_MBPS) else "미흡"
        diag_ul = "양호" if (isinstance(total_ul, (int, float)) and total_ul >= FULLLOAD_THRESHOLD_MBPS) else "미흡"
        avg_rows.append({
            "school_code": "",
            "school_name": "전체",
            "download": total_dl,
            "upload": total_ul,
            "measure3": total_m3,
            "measure4": total_m4,
            "diag_dl": diag_dl,
            "diag_ul": diag_ul,
        })

    # 기존 AVG 시트 제거 후 새로 생성
    if SHEET_FULLLOAD_AVG in wb.sheetnames:
        del wb[SHEET_FULLLOAD_AVG]
    ws_avg = wb.create_sheet(SHEET_FULLLOAD_AVG)

    # 헤더: 학교코드(A), 학교명, 다운로드, 업로드, (F열명), (G열명), 다운로드 진단, 업로드 진단
    headers = ["학교코드", "학교명", "다운로드", "업로드", str(hdr_f), str(hdr_g), "다운로드 진단", "업로드 진단"]
    for c, h in enumerate(headers, 1):
        ws_avg.cell(row=1, column=c, value=h)

    for ri, row in enumerate(avg_rows, 2):
        ws_avg.cell(row=ri, column=1, value=row["school_code"])
        ws_avg.cell(row=ri, column=2, value=row["school_name"])
        ws_avg.cell(row=ri, column=3, value=row["download"])
        ws_avg.cell(row=ri, column=4, value=row["upload"])
        ws_avg.cell(row=ri, column=5, value=row["measure3"])
        ws_avg.cell(row=ri, column=6, value=row["measure4"])
        ws_avg.cell(row=ri, column=7, value=row["diag_dl"])
        ws_avg.cell(row=ri, column=8, value=row["diag_ul"])

        # 숫자 셀에 단위 서식 적용 (표시용)
        for col_idx, fmt in [(3, NUMBER_FORMAT_DOWNLOAD), (4, NUMBER_FORMAT_UPLOAD),
                             (5, NUMBER_FORMAT_DEFAULT), (6, NUMBER_FORMAT_DEFAULT)]:
            cell = ws_avg.cell(row=ri, column=col_idx)
            if cell.value != "" and isinstance(cell.value, (int, float)):
                cell.number_format = fmt

    wb.save(input_path)
    wb.close()
    _log(f"[완료] {input_path}")
    _log(f"       {SHEET_FULLLOAD_AVG}: {len(avg_rows)}행")


def main():
    _log("=" * 50)
    _log("[학교별 측정 값 현황] 전부하 측정 전처리 시작")
    _log("=" * 50)
    import argparse
    parser = argparse.ArgumentParser(description="전부하 측정 전처리")
    parser.add_argument("--region", "-r", choices=["CNE", "DNI", "ALL"], default="CNE",
                        help="지역 (CNE=충남, DNI=대전, ALL=둘 다)")
    args = parser.parse_args()
    region = args.region
    _log(f"실행 지역: {region}")
    try:
        if region == "ALL":
            for r in ["CNE", "DNI"]:
                process_fullload_measure(r)
        else:
            process_fullload_measure(region)
    except Exception as e:
        _log(f"[오류] {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    _log("=" * 50)
    _log("종료")


if __name__ == "__main__":
    main()
