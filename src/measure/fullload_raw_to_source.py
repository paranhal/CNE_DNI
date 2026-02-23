# -*- coding: utf-8 -*-
"""
전부하 원시 데이터 → 통계용 원본 → 학교별 평균

1. FULLLOAD_RAWA_1.xlsx에서 1차/2차 선택 로직 적용 → 통계용 원본 생성
2. 학교코드 수정 (학교명 기준)
3. 학교별 평균 및 판정
4. 전체 로그 기록
"""
from __future__ import print_function
import sys
import io
import os
import csv
from datetime import datetime
from openpyxl import load_workbook, Workbook
from tqdm import tqdm

if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

_MEASURE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, _MEASURE_DIR)

try:
    from measure_utils import extract_school_code_from_mgmt_num
    from fullload_raw_config import (
        FULLLOAD_RAW_CANDIDATES,
        FULLLOAD_OUTPUT,
        FULLLOAD_STATS_OUTPUT,
        SHEET_1ST_CANDIDATES,
        SHEET_2ND_CANDIDATES,
        SHEET_SOURCE,
        SHEET_AVG,
        THRESHOLD_MBPS,
        LOG_DIR,
        LOG_PREFIX,
        CNE_DIR,
    )
    from fullload_preprocess_config import SCHOOL_LIST_SEARCH_DIRS
except Exception as e:
    print(f"[오류] config 로드 실패: {e}", flush=True)
    sys.exit(1)

# 로거
_log_lines = []


def log(msg, also_print=True):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {msg}"
    _log_lines.append(line)
    if also_print:
        print(msg, flush=True)


def save_log():
    os.makedirs(LOG_DIR, exist_ok=True)
    fname = f"{LOG_PREFIX}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    path = os.path.join(LOG_DIR, fname)
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(_log_lines))
    log(f"로그 저장: {path}", also_print=True)


def load_school_list():
    """학교 리스트: 학교명 → 학교코드, 학교코드 → 학교명"""
    name_to_code = {}
    code_to_name = {}
    for base_dir in SCHOOL_LIST_SEARCH_DIRS:
        for fname in ["school_reg_list_CNE.xlsx", "school_reg_list_CNE.csv", "SCHOOL_REG_LIST_CNE.xlsx"]:
            path = os.path.join(base_dir, fname)
            if not os.path.isfile(path):
                continue
            try:
                if path.endswith(".csv"):
                    with open(path, "r", encoding="utf-8-sig") as f:
                        reader = csv.reader(f)
                        header = next(reader, None)
                        rows = list(reader)
                else:
                    wb = load_workbook(path, read_only=True, data_only=True)
                    ws = wb.active
                    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
                    rows = [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)] for r in range(2, ws.max_row + 1)]
                    wb.close()
                code_col = name_col = 0
                for i, h in enumerate(header or []):
                    s = str(h or "").lower()
                    if "학교코드" in s or "code" in s:
                        code_col = i
                    if "학교명" in s or "name" in s:
                        name_col = i
                for row in rows:
                    if len(row) > max(code_col, name_col):
                        code = str(row[code_col] or "").strip()
                        name = str(row[name_col] or "").strip()
                        if code and name:
                            name_to_code[name] = code
                            code_to_name[code] = name
                if name_to_code:
                    log(f"학교 리스트 로드: {path} ({len(name_to_code)}개)")
                    return name_to_code, code_to_name
            except Exception as e:
                log(f"  학교 리스트 읽기 실패: {path} - {e}")
    log("  [경고] 학교 리스트 없음")
    return name_to_code, code_to_name


def find_sheet(wb, candidates):
    for c in candidates:
        if c in wb.sheetnames:
            return c
    return None


def find_header_row(ws):
    """측정 컬럼(Down/Up) 있는 행을 헤더로. 2차는 2행 헤더라 Down이 row2에 있음."""
    found_down = None
    found_mgmt = None
    for row in range(1, min(8, ws.max_row + 1)):
        for col in range(1, min(ws.max_column + 1, 20)):
            val = ws.cell(row=row, column=col).value
            if val is None:
                continue
            s = str(val).strip().lower()
            if "down" in s or "다운로드" in s:
                found_down = row
            if "장비관리번호" in str(val) or "관리번호" in str(val):
                found_mgmt = row
    return found_down or found_mgmt or 1


def find_columns(ws, header_row, search_rows=3):
    """장비관리번호, 학교명, Down, Up, RSSI, CH 열 인덱스 (1-based). 2행 헤더 지원."""
    cols = {}
    start = max(1, header_row - 1)
    for hr in range(start, min(header_row + search_rows, ws.max_row + 1)):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=hr, column=c).value
            if val is None:
                continue
            s = str(val).strip().lower().replace("\n", " ")
            if ("장비관리번호" in str(val) or "관리번호" in str(val)) and "mgmt" not in cols:
                cols["mgmt"] = c
            elif "학교명" in str(val) and "school_name" not in cols:
                cols["school_name"] = c
            elif ("down" in s or "다운로드" in s) and "dl" not in cols:
                cols["dl"] = c
            elif ("up" in s or "업로드" in s) and "ul" not in cols:
                cols["ul"] = c
            elif "rssi" in s and "rssi" not in cols:
                cols["rssi"] = c
            elif s.strip() == "ch" and "ch" not in cols:
                cols["ch"] = c
    return cols


def is_numeric(v):
    if v is None:
        return False
    try:
        float(str(v).replace(",", "").replace("%", ""))
        return True
    except ValueError:
        return False


def get_num(v):
    if not is_numeric(v):
        return None
    try:
        return float(str(v).replace(",", "").replace("%", ""))
    except ValueError:
        return None


def diagnose(dl, ul):
    """375 Mbps 기준: 둘 다 >= 375 → 양호, 아니면 미흡"""
    if dl is None or ul is None:
        return "미흡"
    return "양호" if (dl >= THRESHOLD_MBPS and ul >= THRESHOLD_MBPS) else "미흡"


def read_sheet_data(wb, sheet_name, name_to_code):
    """시트에서 데이터 읽기, 학교코드 수정"""
    ws = wb[sheet_name]
    hr = find_header_row(ws)
    cols = find_columns(ws, hr)
    if "mgmt" not in cols or "dl" not in cols or "ul" not in cols:
        log(f"  [경고] {sheet_name}: 필수 열 없음 (mgmt, dl, ul) - cols={cols}")
        return [], []

    data = []
    fixes = []
    for r in range(hr + 1, ws.max_row + 1):
        mgmt = ws.cell(r, cols["mgmt"]).value
        school_name = ws.cell(r, cols.get("school_name", 0) or 2).value if cols.get("school_name") else None
        school_name = str(school_name or "").strip()

        dl = get_num(ws.cell(r, cols["dl"]).value)
        ul = get_num(ws.cell(r, cols["ul"]).value)
        rssi = get_num(ws.cell(r, cols["rssi"]).value) if cols.get("rssi") else None
        ch = get_num(ws.cell(r, cols["ch"]).value) if cols.get("ch") else None

        sc_from_mgmt = extract_school_code_from_mgmt_num(mgmt)
        sc_from_name = name_to_code.get(school_name, "") if school_name else ""

        # 학교코드: 학교명이 있으면 학교명으로 조회한 코드 우선 사용
        if sc_from_name and sc_from_mgmt != sc_from_name:
            fixes.append((mgmt, sc_from_mgmt, sc_from_name, school_name))
        sc_use = sc_from_name if (school_name and sc_from_name) else (sc_from_mgmt or sc_from_name)

        data.append({
            "mgmt": mgmt,
            "school_name": school_name,
            "school_code": sc_use,
            "dl": dl,
            "ul": ul,
            "rssi": rssi,
            "ch": ch,
            "diag": diagnose(dl, ul),
            "src_row": r,  # 원본 시트 행 번호 (1-based)
            "src_sheet": sheet_name,
            "sc_from_mgmt": sc_from_mgmt or "",  # 장비관리번호에서 추출한 코드
            "sc_from_name": sc_from_name or "",  # 학교명으로 조회한 코드
        })
    return data, fixes


def build_stat_source(wb, name_to_code):
    """1차/2차 선택 로직으로 통계용 원본 데이터 구성"""
    sheet_1st = find_sheet(wb, SHEET_1ST_CANDIDATES)
    sheet_2nd = find_sheet(wb, SHEET_2ND_CANDIDATES)
    if not sheet_1st and not sheet_2nd:
        log("[오류] 1차/2차 시트를 찾을 수 없습니다.")
        return [], {}

    data_1st, fixes_1st = read_sheet_data(wb, sheet_1st, name_to_code) if sheet_1st else ([], [])
    data_2nd, fixes_2nd = read_sheet_data(wb, sheet_2nd, name_to_code) if sheet_2nd else ([], [])
    if not data_1st and not data_2nd:
        log("[오류] 1차/2차 시트에서 데이터를 읽을 수 없습니다.")
        return [], {}

    # 학교코드 수정 로그 (전체는 로그파일에, 요약만 콘솔)
    all_fixes = fixes_1st + fixes_2nd
    for mgmt, wrong, correct, name in all_fixes:
        log(f"  [학교코드 수정] 장비={mgmt} | 잘못된코드={wrong} → 올바른코드={correct} (학교명={name})", also_print=False)
    if all_fixes:
        log(f"  [학교코드 수정] 총 {len(all_fixes)}건 (상세는 로그파일 참조)")

    # 장비별로 1차/2차 매칭 (장비관리번호 기준)
    by_mgmt_1st = {str(r["mgmt"]).strip(): r for r in data_1st if r["mgmt"]}
    by_mgmt_2nd = {str(r["mgmt"]).strip(): r for r in data_2nd if r["mgmt"]} if data_2nd else {}

    # 통계용 수집
    stats = {
        "diff_1st_only": [],      # 1차만 있는 장비 (2차 측정 장비와 다름): mgmt, 1차시트행, 학교명
        "diff_2nd_only": [],      # 2차만 있는 장비 (1차 측정 장비와 다름): mgmt, 2차시트행, 학교명
        "fail_1st_no_2nd": [],   # 1차 미흡 + 2차 미측정: mgmt, 1차시트행, 학교명
        "ok_1st_has_2nd": [],    # 1차 양호 + 2차 측정함: mgmt, 1차시트행, 2차시트행, 학교명
        "no_school_code": [],    # 학교코드 없음: (src_sheet, src_row, school_name, school_name_있었는지, mgmt, dl, ul, sc_from_mgmt, sc_from_name)
        "n_1st_raw": len(data_1st),
        "n_2nd_raw": len(data_2nd),
        "sheet_1st": sheet_1st or "",
        "sheet_2nd": sheet_2nd or "",
    }

    selected = []
    sel_1st_only = sel_1st_ok = sel_2nd_ok = sel_both_fail = sel_2nd_only = 0
    for mgmt, r1 in by_mgmt_1st.items():
        r2 = by_mgmt_2nd.get(mgmt)
        if r2 is None:
            selected.append(r1)
            sel_1st_only += 1
            stats["diff_1st_only"].append((mgmt, r1.get("src_row"), r1.get("school_name", "")))
            if r1["diag"] == "미흡":
                stats["fail_1st_no_2nd"].append((mgmt, r1.get("src_row"), r1.get("school_name", "")))
            if not r1.get("school_code"):
                _add_no_school_code(stats["no_school_code"], r1)
        else:
            if r1["diag"] == "양호":
                selected.append(r1)
                sel_1st_ok += 1
                stats["ok_1st_has_2nd"].append((mgmt, r1.get("src_row"), r2.get("src_row"), r1.get("school_name", "")))
                if not r1.get("school_code"):
                    _add_no_school_code(stats["no_school_code"], r1)
            elif r2["diag"] == "양호":
                selected.append(r2)
                sel_2nd_ok += 1
                if not r2.get("school_code"):
                    _add_no_school_code(stats["no_school_code"], r2)
            else:
                selected.append(r1)
                sel_both_fail += 1
                if not r1.get("school_code"):
                    _add_no_school_code(stats["no_school_code"], r1)

    for mgmt, r2 in by_mgmt_2nd.items():
        if mgmt not in by_mgmt_1st:
            selected.append(r2)
            sel_2nd_only += 1
            stats["diff_2nd_only"].append((mgmt, r2.get("src_row"), r2.get("school_name", "")))
            if not r2.get("school_code"):
                _add_no_school_code(stats["no_school_code"], r2)

    log(f"  [선택 요약] 1차만:{sel_1st_only}, 1차양호:{sel_1st_ok}, 2차양호사용:{sel_2nd_ok}, 둘다미흡:{sel_both_fail}, 2차만:{sel_2nd_only}")
    if stats.get("no_school_code"):
        log(f"  [학교코드 없음] {len(stats['no_school_code'])}건 (상세는 통계 파일 '학교코드_없음' 시트 참조)")

    return selected, stats


def _add_no_school_code(lst, r):
    """학교코드 없음 목록에 추가 (중복 방지: 동일 src_sheet+src_row는 제외)"""
    key = (r.get("src_sheet"), r.get("src_row"))
    for item in lst:
        if (item[0], item[1]) == key:
            return
    school_name = r.get("school_name") or ""
    lst.append((
        r.get("src_sheet", ""),
        r.get("src_row"),
        school_name,
        "있음" if school_name else "없음",
        r.get("mgmt"),
        r.get("dl"),
        r.get("ul"),
        r.get("sc_from_mgmt", ""),
        r.get("sc_from_name", ""),
    ))


def _write_stats_file(stats, n_selected, n_avg_schools):
    """통계 파일(Excel) 생성: 케이스별 건수, 장비관리번호·행번호, 원본 vs 생성 차이"""
    wb = Workbook()
    ws = wb.active
    ws.title = "요약"

    n_1st = stats.get("n_1st_raw", 0)
    n_2nd = stats.get("n_2nd_raw", 0)
    n_orig_total = n_1st + n_2nd
    diff_1st_only = stats.get("diff_1st_only", [])
    diff_2nd_only = stats.get("diff_2nd_only", [])
    fail_1st_no_2nd = stats.get("fail_1st_no_2nd", [])
    ok_1st_has_2nd = stats.get("ok_1st_has_2nd", [])
    no_school_code = stats.get("no_school_code", [])

    sheet_1st = stats.get("sheet_1st", "1차측정")
    sheet_2nd = stats.get("sheet_2nd", "2차측정")

    # 요약 시트: 원본 vs 생성 차이
    ws.cell(1, 1, value="[원본 vs 생성 자료 차이]")
    ws.cell(2, 1, value="구분")
    ws.cell(2, 2, value="행 수")
    ws.cell(3, 1, value="원본 1차 시트")
    ws.cell(3, 2, value=n_1st)
    ws.cell(4, 1, value="원본 2차 시트")
    ws.cell(4, 2, value=n_2nd)
    ws.cell(5, 1, value="원본 합계(1차+2차)")
    ws.cell(5, 2, value=n_orig_total)
    ws.cell(6, 1, value="생성(통계용 원본)")
    ws.cell(6, 2, value=n_selected)
    ws.cell(7, 1, value="차이(원본합계 - 생성)")
    ws.cell(7, 2, value=n_orig_total - n_selected)
    ws.cell(8, 1, value="(차이 원인: 1차·2차 같은 장비는 1행만 유지)")
    ws.cell(9, 1, value=f"행번호 기준: 1차시트={sheet_1st}, 2차시트={sheet_2nd}")
    ws.cell(10, 1, value="")
    ws.cell(11, 1, value="[케이스별 건수]")
    ws.cell(12, 1, value="구분")
    ws.cell(12, 2, value="건수")
    n_diff_total = len(diff_1st_only) + len(diff_2nd_only)
    ws.cell(13, 1, value="1차·2차 측정 장비가 다른 경우 (합계)")
    ws.cell(13, 2, value=n_diff_total)
    ws.cell(14, 1, value="  - 1차만 있는 장비")
    ws.cell(14, 2, value=len(diff_1st_only))
    ws.cell(15, 1, value="  - 2차만 있는 장비")
    ws.cell(15, 2, value=len(diff_2nd_only))
    ws.cell(16, 1, value="1차 미흡 + 2차 미측정")
    ws.cell(16, 2, value=len(fail_1st_no_2nd))
    ws.cell(17, 1, value="1차 양호 + 2차 측정함")
    ws.cell(17, 2, value=len(ok_1st_has_2nd))
    ws.cell(18, 1, value="학교코드 없음")
    ws.cell(18, 2, value=len(no_school_code))

    # 시트2: 학교코드 없음 (상세)
    ws_nosc = wb.create_sheet("학교코드_없음", 1)
    ws_nosc.cell(1, 1, value="순번")
    ws_nosc.cell(1, 2, value="원본시트")
    ws_nosc.cell(1, 3, value="원본_행번호")
    ws_nosc.cell(1, 4, value="학교명")
    ws_nosc.cell(1, 5, value="학교명_있었는지")
    ws_nosc.cell(1, 6, value="장비관리번호")
    ws_nosc.cell(1, 7, value="Down(Mbps)")
    ws_nosc.cell(1, 8, value="Up(Mbps)")
    ws_nosc.cell(1, 9, value="장비에서_추출한코드")
    ws_nosc.cell(1, 10, value="학교명으로_조회한코드")
    for i, row in enumerate(no_school_code, 2):
        src_sheet, src_row, school_name, name_yn, mgmt, dl, ul, sc_mgmt, sc_name = row
        ws_nosc.cell(i, 1, value=i - 1)
        ws_nosc.cell(i, 2, value=src_sheet)
        ws_nosc.cell(i, 3, value=src_row)
        ws_nosc.cell(i, 4, value=school_name or "")
        ws_nosc.cell(i, 5, value=name_yn)
        ws_nosc.cell(i, 6, value=mgmt)
        ws_nosc.cell(i, 7, value=dl if dl is not None else "")
        ws_nosc.cell(i, 8, value=ul if ul is not None else "")
        ws_nosc.cell(i, 9, value=sc_mgmt or "")
        ws_nosc.cell(i, 10, value=sc_name or "")

    # 시트3: 1차만 (장비 다름)
    ws1 = wb.create_sheet("1차만_장비다름")
    ws1.cell(1, 1, value="순번")
    ws1.cell(1, 2, value="장비관리번호")
    ws1.cell(1, 3, value="1차시트_행번호")
    ws1.cell(1, 4, value="학교명")
    for i, (mgmt, row, name) in enumerate(diff_1st_only, 2):
        ws1.cell(i, 1, value=i - 1)
        ws1.cell(i, 2, value=mgmt)
        ws1.cell(i, 3, value=row)
        ws1.cell(i, 4, value=name or "")

    # 시트4: 2차만 (장비 다름)
    ws2 = wb.create_sheet("2차만_장비다름")
    ws2.cell(1, 1, value="순번")
    ws2.cell(1, 2, value="장비관리번호")
    ws2.cell(1, 3, value="2차시트_행번호")
    ws2.cell(1, 4, value="학교명")
    for i, (mgmt, row, name) in enumerate(diff_2nd_only, 2):
        ws2.cell(i, 1, value=i - 1)
        ws2.cell(i, 2, value=mgmt)
        ws2.cell(i, 3, value=row)
        ws2.cell(i, 4, value=name or "")

    # 시트5: 1차 미흡 + 2차 미측정
    ws3 = wb.create_sheet("1차미흡_2차미측정")
    ws3.cell(1, 1, value="순번")
    ws3.cell(1, 2, value="장비관리번호")
    ws3.cell(1, 3, value="1차시트_행번호")
    ws3.cell(1, 4, value="학교명")
    for i, (mgmt, row, name) in enumerate(fail_1st_no_2nd, 2):
        ws3.cell(i, 1, value=i - 1)
        ws3.cell(i, 2, value=mgmt)
        ws3.cell(i, 3, value=row)
        ws3.cell(i, 4, value=name or "")

    # 시트6: 1차 양호 + 2차 측정함
    ws4 = wb.create_sheet("1차양호_2차측정함")
    ws4.cell(1, 1, value="순번")
    ws4.cell(1, 2, value="장비관리번호")
    ws4.cell(1, 3, value="1차시트_행번호")
    ws4.cell(1, 4, value="2차시트_행번호")
    ws4.cell(1, 5, value="학교명")
    for i, (mgmt, r1, r2, name) in enumerate(ok_1st_has_2nd, 2):
        ws4.cell(i, 1, value=i - 1)
        ws4.cell(i, 2, value=mgmt)
        ws4.cell(i, 3, value=r1)
        ws4.cell(i, 4, value=r2)
        ws4.cell(i, 5, value=name or "")

    try:
        wb.save(FULLLOAD_STATS_OUTPUT)
        log(f"[완료] 통계 파일: {FULLLOAD_STATS_OUTPUT}")
    except PermissionError:
        alt = FULLLOAD_STATS_OUTPUT.replace(".xlsx", f"_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        wb.save(alt)
        log(f"[경고] 통계 파일이 사용 중이어서 대체 저장: {alt}")


def main():
    log("=" * 60)
    log("[전부하] FULLLOAD_RAWA_1 → 통계용 원본 → 학교별 평균")
    log("=" * 60)

    raw_path = None
    for p in FULLLOAD_RAW_CANDIDATES:
        if os.path.isfile(p):
            raw_path = p
            break
    if not raw_path:
        log(f"[오류] 파일 없음. 다음 경로 확인:")
        for p in FULLLOAD_RAW_CANDIDATES:
            log(f"       {p}")
        save_log()
        sys.exit(1)
    log(f"입력 파일: {raw_path}")

    name_to_code, code_to_name = load_school_list()

    wb_raw = load_workbook(raw_path, data_only=True)
    selected, stats = build_stat_source(wb_raw, name_to_code)
    wb_raw.close()

    if not selected:
        log("[오류] 선택된 데이터 없음")
        save_log()
        sys.exit(1)

    log(f"통계용 원본 행 수: {len(selected)}")

    # 통계용 원본 시트 생성 (CNE_FULLLOAD_MEASURE.xlsx)
    os.makedirs(CNE_DIR, exist_ok=True)
    wb_out = Workbook()
    ws_src = wb_out.active
    ws_src.title = SHEET_SOURCE

    headers = ["구분", "학교명", "장비관리번호", "Down(Mbps)", "Up(Mbps)", "*RSSI", "CH", "비고"]
    for c, h in enumerate(headers, 1):
        ws_src.cell(row=1, column=c, value=h)

    for ri, row in enumerate(selected, 2):
        ws_src.cell(ri, 1, value="")
        ws_src.cell(ri, 2, value=row["school_name"])
        ws_src.cell(ri, 3, value=row["mgmt"])
        ws_src.cell(ri, 4, value=row["dl"])
        ws_src.cell(ri, 5, value=row["ul"])
        ws_src.cell(ri, 6, value=row["rssi"])
        ws_src.cell(ri, 7, value=row["ch"])
        ws_src.cell(ri, 8, value=row["diag"])

    # 학교별 평균 (다운로드/업로드 있는 장비만, rssi/ch는 있으면 포함)
    school_rows = {}
    for r in selected:
        if r["dl"] is None or r["ul"] is None:
            continue
        sc = r["school_code"]
        if not sc:
            continue
        if sc not in school_rows:
            school_rows[sc] = []
        school_rows[sc].append((r["dl"], r["ul"], r.get("rssi"), r.get("ch")))

    avg_rows = []
    for sc in sorted(school_rows.keys()):
        rows = school_rows[sc]
        avg_dl = round(sum(r[0] for r in rows) / len(rows), 1)
        avg_ul = round(sum(r[1] for r in rows) / len(rows), 1)
        rssi_vals = [r[2] for r in rows if r[2] is not None]
        ch_vals = [r[3] for r in rows if r[3] is not None]
        avg_rssi = round(sum(rssi_vals) / len(rssi_vals), 1) if rssi_vals else ""
        avg_ch = round(sum(ch_vals) / len(ch_vals), 1) if ch_vals else ""
        diag_dl = "양호" if avg_dl >= THRESHOLD_MBPS else "미흡"
        diag_ul = "양호" if avg_ul >= THRESHOLD_MBPS else "미흡"
        school_name = code_to_name.get(sc) or (selected[0]["school_name"] if selected else "")
        for r in selected:
            if r["school_code"] == sc:
                school_name = r["school_name"] or school_name
                break
        avg_rows.append({
            "school_code": sc,
            "school_name": school_name,
            "download": avg_dl,
            "upload": avg_ul,
            "rssi": avg_rssi,
            "ch": avg_ch,
            "diag_dl": diag_dl,
            "diag_ul": diag_ul,
        })

    ws_avg = wb_out.create_sheet(SHEET_AVG)
    hdrs = ["학교코드", "학교명", "다운로드", "업로드", "*RSSI", "CH", "다운로드 진단", "업로드 진단"]
    for c, h in enumerate(hdrs, 1):
        ws_avg.cell(row=1, column=c, value=h)
    for ri, row in enumerate(avg_rows, 2):
        ws_avg.cell(ri, 1, value=row["school_code"])
        ws_avg.cell(ri, 2, value=row["school_name"])
        ws_avg.cell(ri, 3, value=row["download"])
        ws_avg.cell(ri, 4, value=row["upload"])
        ws_avg.cell(ri, 5, value=row["rssi"])
        ws_avg.cell(ri, 6, value=row["ch"])
        ws_avg.cell(ri, 7, value=row["diag_dl"])
        ws_avg.cell(ri, 8, value=row["diag_ul"])

    try:
        wb_out.save(FULLLOAD_OUTPUT)
        log(f"[완료] {FULLLOAD_OUTPUT}")
    except PermissionError:
        alt_path = FULLLOAD_OUTPUT.replace(".xlsx", f"_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        wb_out.save(alt_path)
        log(f"[경고] 원본 파일이 사용 중이어서 대체 경로에 저장: {alt_path}")
        log(f"       Excel에서 {os.path.basename(FULLLOAD_OUTPUT)} 파일을 닫은 후 다시 실행하세요.")
    log(f"       전부하측정: {len(selected)}행")
    log(f"       전부하측정_학교별평균: {len(avg_rows)}개 학교")

    # 통계 파일 생성
    _write_stats_file(stats, len(selected), len(avg_rows))

    log("=" * 60)
    save_log()


if __name__ == "__main__":
    main()
