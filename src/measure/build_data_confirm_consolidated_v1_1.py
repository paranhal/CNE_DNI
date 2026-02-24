# -*- coding: utf-8 -*-
"""
개별 학교 리포트 데이터 → 통합 파일 (행/열 전환) V1.1

[입력]
- CNE/학교별_리포트_V1.1/*.xlsx (개별 학교 리포트)
- TOTAL_MEASURE_LIST Sheet1 (지역 정보)
- 최종_데이터확인_템플릿.xlsx (템플릿)

[출력]
- CNE/데이터확인_통합_V1.1.xlsx
- A열: 지역, B열: 학교코드, C열: 학교명, D열~: 측정 데이터

[매핑] 검증파일(계룡고 행/열 전환) 기준으로 개별리포트 F열(행) → 통합 열
- 템플릿 C4~C22: 수집데이터 R2~R20 (1:1)
- 템플릿 C23~C36: 품질측정 R21~R34 (C23=R21, C24=R22, ...)
"""
from __future__ import print_function
import sys
import io
import os
import re
import copy

if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

from openpyxl import load_workbook, Workbook
from tqdm import tqdm

_MEASURE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, _MEASURE_DIR)

from school_report_config_v1_1 import TOTAL_MEASURE_LIST, CNE_DIR, J_OUTPUT_MAP, J_COL

# 출력 파일 (V1.1 전용 경로)
REPORT_DIR = os.path.join(CNE_DIR, "학교별_리포트_V1.1")
OUTPUT_FILE = os.path.join(CNE_DIR, "데이터확인_통합_V1.1.xlsx")
TEMPLATE_PATH = os.path.join(_MEASURE_DIR, "최종_데이터확인_템플릿.xlsx")

# J_OUTPUT_MAP 행 순서
DATA_ROWS = [rd[0] for rd in J_OUTPUT_MAP]

# 개별리포트 행 → 통합파일 열 매핑 (검증파일 계룡고 행/열전환 기준)
# 검증 C4=R2, C5=R4, C6=R6, C8=R14, C9=R18, C10=R19, C11=R20, C12=R21, C13=R11, C14=R12, C15=R13,
# C17=R15, C19=R17, C23=R22, C24=R23, C25=R24, C26=R25, C28=R26, C29=R27, C30=R28, C31=R29,
# C32=R30, C33=R31, C34=R32, C35=R33, C36=R34
# C7(장비연결), C16, C18, C20, C21, C22, C27: 검증에서 비어있음 → 빈칸 유지
COLS_TO_CLEAR = [7, 16, 18, 20, 21, 22, 27]  # 데이터 없이 빈칸으로 둘 열

# 템플릿 구조: 723행=제목(헤더), 724행=실데이터 시작
HEADER_ROW = 723
DATA_START_ROW = 724
TEMPLATE_DATA_COLS = {
    2: 4,   4: 5,   6: 6,   11: 13,  12: 14,  13: 15,  14: 8,   15: 17,  17: 19,
    18: 9,  19: 10, 20: 11, 21: 12,  22: 23,  23: 24,  24: 25,  25: 26,  26: 28,
    27: 29, 28: 30, 29: 31, 30: 32,  31: 33,  32: 34,  33: 35,  34: 36,
}


def load_region_code_name():
    """TOTAL_MEASURE_LIST Sheet1에서 지역, 학교코드, 학교명"""
    code_to_region = {}
    code_to_name = {}
    if not os.path.isfile(TOTAL_MEASURE_LIST):
        return code_to_region, code_to_name
    wb = load_workbook(TOTAL_MEASURE_LIST, data_only=True)
    if "Sheet1" not in wb.sheetnames:
        wb.close()
        return code_to_region, code_to_name
    ws = wb["Sheet1"]
    for r in range(2, ws.max_row + 1):
        code = str(ws.cell(r, 1).value or "").strip()
        region = str(ws.cell(r, 2).value or "").strip()
        name = str(ws.cell(r, 3).value or "").strip()
        if code:
            code_to_region[code] = region
            code_to_name[code] = name
    wb.close()
    return code_to_region, code_to_name


def extract_school_code_from_filename(fname):
    """파일명에서 학교코드 추출 (학교명_학교코드.xlsx)"""
    base = os.path.splitext(fname)[0]
    m = re.search(r"_([A-Z0-9]{12,})$", base)
    if m:
        return m.group(1)
    return None


def read_report_data(report_path):
    """개별 리포트에서 F열(측정값) 데이터 추출 - R2~R34 전체"""
    data = {}
    try:
        wb = load_workbook(report_path, data_only=True)
        ws = wb.active
        if "문제점분석" in wb.sheetnames:
            ws = wb["문제점분석"]
        for row in range(2, 35):
            val = ws.cell(row=row, column=J_COL).value
            data[row] = val if val is not None else ""
        wb.close()
    except Exception:
        pass
    return data


def save_template_row(ws, row, max_col):
    """템플릿 행의 값·스타일 저장 (덮어쓰기 전에 호출)"""
    cache = []
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        try:
            font = copy.copy(cell.font) if cell.font else None
        except Exception:
            font = None
        try:
            border = copy.copy(cell.border) if cell.border else None
        except Exception:
            border = None
        try:
            fill = copy.copy(cell.fill) if cell.fill else None
        except Exception:
            fill = None
        try:
            alignment = copy.copy(cell.alignment) if cell.alignment else None
        except Exception:
            alignment = None
        cache.append({
            "value": cell.value,
            "font": font,
            "border": border,
            "fill": fill,
            "number_format": cell.number_format,
            "alignment": alignment,
        })
    return cache


def apply_template_row(ws, row, template_cache):
    """저장된 템플릿 행을 대상 행에 적용"""
    for c, data in enumerate(template_cache, start=1):
        dst = ws.cell(row=row, column=c)
        dst.value = data["value"]
        if data["font"]:
            dst.font = data["font"]
        if data["border"]:
            dst.border = data["border"]
        if data["fill"]:
            dst.fill = data["fill"]
        if data["number_format"]:
            dst.number_format = data["number_format"]
        if data["alignment"]:
            dst.alignment = data["alignment"]


def main():
    print("=" * 50)
    print("[데이터확인 통합 V1.1] 개별 리포트 → 통합 (행/열 전환)")
    print("=" * 50)
    code_to_region, code_to_name = load_region_code_name()
    if not os.path.isdir(REPORT_DIR):
        print(f"[오류] 리포트 폴더 없음: {REPORT_DIR}")
        sys.exit(1)
    files = [f for f in os.listdir(REPORT_DIR) if f.endswith(".xlsx")]
    school_files = []
    for f in files:
        code = extract_school_code_from_filename(f)
        if code:
            school_files.append((code, os.path.join(REPORT_DIR, f)))
    school_files.sort(key=lambda x: x[0])
    print(f"대상 학교: {len(school_files)}개")
    if not os.path.isfile(TEMPLATE_PATH):
        print(f"[오류] 템플릿 없음: {TEMPLATE_PATH}")
        sys.exit(1)
    wb_out = load_workbook(TEMPLATE_PATH)
    ws_out = wb_out.active
    if "데이터확인" in wb_out.sheetnames:
        ws_out = wb_out["데이터확인"]
    max_col = max(36, ws_out.max_column)
    template_row_data = save_template_row(ws_out, DATA_START_ROW, max_col)
    for ri, (school_code, report_path) in enumerate(tqdm(school_files, desc="통합", unit="교"), DATA_START_ROW):
        if ri > DATA_START_ROW:
            apply_template_row(ws_out, ri, template_row_data)
        ws_out.cell(ri, 1, value=code_to_region.get(school_code, ""))
        ws_out.cell(ri, 2, value=school_code)
        ws_out.cell(ri, 3, value=code_to_name.get(school_code, ""))
        row_data = read_report_data(report_path)
        for data_row, template_col in TEMPLATE_DATA_COLS.items():
            val = row_data.get(data_row, "")
            ws_out.cell(ri, template_col, value=val)
        for col in COLS_TO_CLEAR:
            ws_out.cell(ri, col, value="")
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    wb_out.save(OUTPUT_FILE)
    print(f"[완료] {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
