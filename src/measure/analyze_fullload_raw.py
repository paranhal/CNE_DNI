# -*- coding: utf-8 -*-
"""FULLLOAD_RAWA_1.xlsx 구조 분석 (파일 있을 때 실행)"""
import sys
import io
import os

if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from openpyxl import load_workbook

path = os.path.join(os.path.dirname(__file__), "CNE", "FULLLOAD_RAWA_1.xlsx")
if not os.path.isfile(path):
    print(f"파일 없음: {path}")
    sys.exit(1)

wb = load_workbook(path, data_only=True)
print("시트 목록:", wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print(f"\n=== {name} ===")
    print(f"  행: {ws.max_row}, 열: {ws.max_column}")
    for r in range(1, min(5, ws.max_row + 1)):
        row = [ws.cell(r, c).value for c in range(1, min(15, ws.max_column + 1))]
        print(f"  Row {r}: {row}")
wb.close()
