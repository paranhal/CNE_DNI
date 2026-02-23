# -*- coding: utf-8 -*-
"""
무선망(ISP) 측정 데이터 전처리

1. CNE_ISP_MEASURE.XLSX의 "ISP측정" 시트에서 학교별 평균 계산
2. 새 시트 "ISP측정_학교별평균" 추가
3. 열: 학교명, 학교코드, 다운로드, 업로드, RTT, RSSI, CH, 다운로드 진단, 업로드 진단
4. 진단 기준: 375 Mbps 이상 → 양호, 미만 → 미흡
5. 마지막에 "전체" 행 추가 (전체 평균)
"""
from __future__ import print_function
import sys
import io

# 한글 출력 깨짐 방지 (PowerShell/CMD UTF-8)
if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

print("[무선망 ISP 전처리] 스크립트 시작", flush=True)

import os
from openpyxl import load_workbook
from tqdm import tqdm

# config 로드
_MEASURE_DIR = os.path.dirname(os.path.abspath(__file__))
if _MEASURE_DIR not in sys.path:
    sys.path.insert(0, _MEASURE_DIR)

try:
    from measure_utils import extract_school_code_from_mgmt_num
    from wireless_preprocess_config import (
        ISP_MEASURE_FILE,
        ISP_MEASURE_CANDIDATES,
        SHEET_ISP_SOURCE,
        SHEET_ISP_AVG,
        COL_MGMT_NUM,
        COL_DOWNLOAD,
        COL_UPLOAD,
        COL_RTT,
        COL_RSSI,
        COL_CH,
        ISP_THRESHOLD_MBPS,
        SCHOOL_LIST_SEARCH_DIRS,
    )
except Exception as e:
    print(f"[오류] config 로드 실패: {e}", flush=True)
    sys.exit(1)


def _log(msg):
    print(msg, flush=True)


def load_school_name_map(region):
    """학교 리스트에서 학교코드 → 학교명 매핑 로드"""
    import csv
    candidates = [
        (f"school_reg_list_{region}.csv", "csv"),
        (f"SCHOOL_REG_LIST_{region}.csv", "csv"),
        (f"school_reg_list_{region}.xlsx", "xlsx"),
        (f"SCHOOL_REG_LIST_{region}.xlsx", "xlsx"),
        ("school_reg_list.csv", "csv"),
        ("SCHOOL_REG_LIST.csv", "csv"),
    ]
    for base_dir in SCHOOL_LIST_SEARCH_DIRS:
        for fname, ext in candidates:
            path = os.path.join(base_dir, fname)
            if not os.path.isfile(path):
                continue
            code_to_name = {}
            try:
                if ext == "csv":
                    with open(path, "r", encoding="utf-8-sig") as f:
                        reader = csv.reader(f)
                        header = next(reader, None)
                        rows = list(reader)
                else:
                    wb = load_workbook(path, read_only=True, data_only=True)
                    ws = wb.active
                    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
                    rows = [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)] for r in range(2, ws.max_row + 1)]
                    wb.close()
                if not header:
                    continue
                code_col = name_col = None
                for i, h in enumerate(header):
                    s = str(h or "").strip().lower()
                    if "학교코드" in s or "code" in s:
                        code_col = i
                    if "학교명" in s or "name" in s:
                        name_col = i
                if code_col is None:
                    code_col = 0
                if name_col is None:
                    name_col = 2
                for row in rows:
                    if len(row) > max(code_col, name_col):
                        code = str(row[code_col] or "").strip()
                        name = str(row[name_col] or "").strip()
                        if code:
                            code_to_name[code] = name
                            code_to_name[code[:12]] = name
                if code_to_name:
                    _log(f"  학교 리스트 로드: {path} ({len(code_to_name)}개)")
                return code_to_name
            except Exception:
                pass
    _log("  [참고] 학교 리스트 없음 → 학교명 빈칸")
    return {}


def is_numeric(val):
    if val is None:
        return False
    if isinstance(val, (int, float)):
        return True
    s = str(val).strip()
    if not s:
        return False
    try:
        float(s.replace(",", "").replace("%", ""))
        return True
    except ValueError:
        return False


def find_header_row(ws):
    """헤더 행 찾기 (1~5행 중 '장비관리번호','다운로드','업로드' 등 포함된 행)"""
    for row in range(1, min(6, ws.max_row + 1)):
        for col in range(1, min(ws.max_column + 1, 30)):
            val = ws.cell(row=row, column=col).value
            if val and isinstance(val, str):
                s = val.strip()
                if any(kw in s for kw in ["장비관리번호", "다운로드", "업로드", "RTT", "RSSI", "CH"]):
                    return row
    return 1


def process_isp_measure(region="CNE"):
    """ISP 측정 데이터 전처리: 학교별 평균 계산 후 새 시트 추가"""
    candidates = ISP_MEASURE_CANDIDATES.get(region, [ISP_MEASURE_FILE[region]])
    input_path = None
    for p in candidates:
        if os.path.isfile(p):
            input_path = p
            break
    if not input_path:
        _log(f"[오류] 파일을 찾을 수 없습니다. 다음 경로를 확인하세요:")
        for p in candidates:
            _log(f"       {p}")
        return
    _log(f"[진행] 지역: {region} | 파일: {input_path}")

    code_to_name = load_school_name_map(region)

    wb = load_workbook(input_path, data_only=True)
    if SHEET_ISP_SOURCE not in wb.sheetnames:
        _log(f"[오류] 시트 '{SHEET_ISP_SOURCE}'가 없습니다.")
        wb.close()
        return

    ws_src = wb[SHEET_ISP_SOURCE]
    header_row = find_header_row(ws_src)
    max_row = ws_src.max_row
    max_col = ws_src.max_column

    # 학교별 데이터 수집
    school_data = {}  # school_code -> {downloads, uploads, rtts, rssis, chs}
    for r in tqdm(range(header_row + 1, max_row + 1), desc=f"[{region}] 데이터 읽기", unit="행"):
        mgmt = ws_src.cell(row=r, column=COL_MGMT_NUM).value
        school_code = extract_school_code_from_mgmt_num(mgmt)
        if not school_code:
            continue

        def _get_num(col):
            v = ws_src.cell(row=r, column=col).value
            if is_numeric(v):
                try:
                    return float(str(v).replace(",", "").replace("%", ""))
                except ValueError:
                    return None
            return None

        dl = _get_num(COL_DOWNLOAD)
        ul = _get_num(COL_UPLOAD)
        rtt = _get_num(COL_RTT)
        rssi = _get_num(COL_RSSI)
        ch = _get_num(COL_CH)

        # 데이터가 있는 장비만 포함 (5개 측정값 모두 있는 행만)
        if dl is None or ul is None or rtt is None or rssi is None or ch is None:
            continue

        if school_code not in school_data:
            school_data[school_code] = {"rows": []}
        school_data[school_code]["rows"].append((dl, ul, rtt, rssi, ch))

    # 학교별 평균 계산 (데이터 있는 장비들만)
    def _avg_from_rows(rows, idx):
        if not rows:
            return ""
        return round(sum(r[idx] for r in rows) / len(rows), 2)

    avg_rows = []
    for school_code in tqdm(sorted(school_data.keys()), desc=f"[{region}] 평균 계산", unit="학교"):
        rows = school_data[school_code]["rows"]
        if not rows:
            continue
        school_name = code_to_name.get(school_code) or code_to_name.get(school_code[:12]) or ""

        avg_dl = _avg_from_rows(rows, 0)
        avg_ul = _avg_from_rows(rows, 1)
        avg_rtt = _avg_from_rows(rows, 2)
        avg_rssi = _avg_from_rows(rows, 3)
        avg_ch = _avg_from_rows(rows, 4)

        if isinstance(avg_dl, str):
            avg_dl = ""
        if isinstance(avg_ul, str):
            avg_ul = ""

        diag_dl = "양호" if (isinstance(avg_dl, (int, float)) and avg_dl >= ISP_THRESHOLD_MBPS) else "미흡"
        diag_ul = "양호" if (isinstance(avg_ul, (int, float)) and avg_ul >= ISP_THRESHOLD_MBPS) else "미흡"

        avg_rows.append({
            "school_name": school_name,
            "school_code": school_code,
            "download": avg_dl,
            "upload": avg_ul,
            "rtt": avg_rtt,
            "rssi": avg_rssi,
            "ch": avg_ch,
            "diag_dl": diag_dl,
            "diag_ul": diag_ul,
        })

    # 전체 평균 행 (전부하)
    if avg_rows:
        all_dl = [r["download"] for r in avg_rows if isinstance(r["download"], (int, float))]
        all_ul = [r["upload"] for r in avg_rows if isinstance(r["upload"], (int, float))]
        all_rtt = [r["rtt"] for r in avg_rows if isinstance(r["rtt"], (int, float))]
        all_rssi = [r["rssi"] for r in avg_rows if isinstance(r["rssi"], (int, float))]
        all_ch = [r["ch"] for r in avg_rows if isinstance(r["ch"], (int, float))]
        total_dl = round(sum(all_dl) / len(all_dl), 2) if all_dl else ""
        total_ul = round(sum(all_ul) / len(all_ul), 2) if all_ul else ""
        total_rtt = round(sum(all_rtt) / len(all_rtt), 2) if all_rtt else ""
        total_rssi = round(sum(all_rssi) / len(all_rssi), 2) if all_rssi else ""
        total_ch = round(sum(all_ch) / len(all_ch), 2) if all_ch else ""
        diag_dl = "양호" if (isinstance(total_dl, (int, float)) and total_dl >= ISP_THRESHOLD_MBPS) else "미흡"
        diag_ul = "양호" if (isinstance(total_ul, (int, float)) and total_ul >= ISP_THRESHOLD_MBPS) else "미흡"
        avg_rows.append({
            "school_name": "전체",
            "school_code": "",
            "download": total_dl,
            "upload": total_ul,
            "rtt": total_rtt,
            "rssi": total_rssi,
            "ch": total_ch,
            "diag_dl": diag_dl,
            "diag_ul": diag_ul,
        })

    # 기존 AVG 시트가 있으면 제거 후 새로 생성
    if SHEET_ISP_AVG in wb.sheetnames:
        del wb[SHEET_ISP_AVG]
    ws_avg = wb.create_sheet(SHEET_ISP_AVG)

    # 헤더: 학교명, 학교코드, 다운로드, 업로드, RTT, RSSI, CH, 다운로드 진단, 업로드 진단
    headers = ["학교명", "학교코드", "다운로드", "업로드", "RTT", "RSSI", "CH", "다운로드 진단", "업로드 진단"]
    for c, h in enumerate(headers, 1):
        ws_avg.cell(row=1, column=c, value=h)

    for ri, row in enumerate(avg_rows, 2):
        ws_avg.cell(row=ri, column=1, value=row["school_name"])
        ws_avg.cell(row=ri, column=2, value=row["school_code"])
        ws_avg.cell(row=ri, column=3, value=row["download"])
        ws_avg.cell(row=ri, column=4, value=row["upload"])
        ws_avg.cell(row=ri, column=5, value=row["rtt"])
        ws_avg.cell(row=ri, column=6, value=row["rssi"])
        ws_avg.cell(row=ri, column=7, value=row["ch"])
        ws_avg.cell(row=ri, column=8, value=row["diag_dl"])
        ws_avg.cell(row=ri, column=9, value=row["diag_ul"])

    wb.save(input_path)
    wb.close()


def main():
    _log("=" * 50)
    _log("[학교별 측정 값 현황] 무선망 ISP 전처리 시작")
    _log("=" * 50)
    import argparse
    parser = argparse.ArgumentParser(description="무선망 ISP 측정 전처리")
    parser.add_argument("--region", "-r", choices=["CNE", "DNI", "ALL"], default="CNE",
                        help="지역 (CNE=충남, DNI=대전, ALL=둘 다)")
    args = parser.parse_args()
    region = args.region
    _log(f"실행 지역: {region}")
    try:
        if region == "ALL":
            for r in ["CNE", "DNI"]:
                process_isp_measure(r)
        else:
            process_isp_measure(region)
    except Exception as e:
        _log(f"[오류] {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    _log("=" * 50)
    _log("종료")


if __name__ == "__main__":
    main()
