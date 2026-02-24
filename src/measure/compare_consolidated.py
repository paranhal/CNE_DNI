# -*- coding: utf-8 -*-
"""검증파일 vs 개별리포트 vs 통합출력 비교"""
import os
from openpyxl import load_workbook

BASE = os.path.dirname(os.path.abspath(__file__))
CNE = os.path.join(BASE, "CNE")
REPORT_PATH = os.path.join(CNE, "학교별_리포트_V1.1", "계룡고등학교_N108140237HS.xlsx")
VERIFY_PATH = os.path.join(CNE, "데이터확인_통합_V1.1_검증.xlsx")
OUTPUT_PATH = os.path.join(CNE, "데이터확인_통합_V1.1.xlsx")

J_COL = 6  # F열
G_COL = 7   # G열
L_COL = 8   # H열

def main():
    # 1. 개별 리포트 (문제점분석): 행2~34, F열(6), G열(7), H열(8)
    print("=== 개별 리포트 (계룡고등학교) 문제점분석 ===")
    wb_r = load_workbook(REPORT_PATH, data_only=True)
    ws_r = wb_r["문제점분석"] if "문제점분석" in wb_r.sheetnames else wb_r.active
    for r in range(1, 35):
        a = ws_r.cell(r, 1).value  # A열 요구사항/항목
        f = ws_r.cell(r, J_COL).value
        g = ws_r.cell(r, G_COL).value
        h = ws_r.cell(r, L_COL).value
        a_str = str(a)[:40] if a else ""
        f_str = str(f)[:50] if f else ""
        print("R%d A=%s | F=%s | G=%s | H=%s" % (r, a_str, repr(f_str)[:40], repr(str(g)[:20]) if g else "None", str(h)[:15] if h else ""))
    wb_r.close()

    # 2. 검증파일 (행/열 전환된 구조): 계룡고등학교 행 찾기
    print("\n=== 검증파일 (행/열 전환) 구조 ===")
    wb_v = load_workbook(VERIFY_PATH, data_only=True)
    ws_v = wb_v.active
    if "데이터확인" in wb_v.sheetnames:
        ws_v = wb_v["데이터확인"]
    print("Max row=%d, max col=%d" % (ws_v.max_row, ws_v.max_column))
    print("Row1 (헤더):")
    for c in range(1, min(45, ws_v.max_column + 1)):
        v = ws_v.cell(1, c).value
        print("  C%d: %s" % (c, repr(str(v)[:50]) if v else "None"))
    # 계룡고등학교 행 찾기 (N108140237HS)
    gye_row = None
    for r in range(2, min(800, ws_v.max_row + 1)):
        code = ws_v.cell(r, 2).value
        name = ws_v.cell(r, 3).value
        if code == "N108140237HS" or (name and "계룡" in str(name) and "고등" in str(name)):
            gye_row = r
            print("\n계룡고등학교 Row%d:" % r)
            for c in range(1, min(45, ws_v.max_column + 1)):
                v = ws_v.cell(r, c).value
                if v is not None and str(v).strip():
                    print("  C%d: %s" % (c, repr(str(v)[:70])))
            break
    if gye_row is None:
        print("Row2 (첫 데이터행):")
        for c in range(1, min(45, ws_v.max_column + 1)):
            v = ws_v.cell(2, c).value
            print("  C%d: %s" % (c, repr(str(v)[:60]) if v else "None"))
    wb_v.close()

    # 3. 통합출력에서 계룡고 행 찾기 + 검증과 비교
    print("\n=== 통합출력 (계룡고등학교 행) ===")
    wb_o = load_workbook(OUTPUT_PATH, data_only=True)
    ws_o = wb_o["데이터확인"] if "데이터확인" in wb_o.sheetnames else wb_o.active
    out_row = None
    for r in range(2, min(800, ws_o.max_row + 1)):
        name = ws_o.cell(r, 3).value
        if name and "계룡" in str(name) and "고" in str(name):
            out_row = r
            print("Row %d: %s" % (r, name))
            for c in range(1, min(45, ws_o.max_column + 1)):
                v = ws_o.cell(r, c).value
                if v is not None and str(v).strip():
                    print("  C%d: %s" % (c, repr(str(v)[:60])))
            break
    wb_o.close()

    # 4. 검증 vs 통합출력 비교 (계룡고)
    if gye_row and out_row:
        print("\n=== 검증 vs 통합출력 비교 (C4~C36) ===")
        wb_v2 = load_workbook(VERIFY_PATH, data_only=True)
        ws_v2 = wb_v2["데이터확인"] if "데이터확인" in wb_v2.sheetnames else wb_v2.active
        wb_o2 = load_workbook(OUTPUT_PATH, data_only=True)
        ws_o2 = wb_o2["데이터확인"] if "데이터확인" in wb_o2.sheetnames else wb_o2.active
        diff = []
        for c in range(4, 37):
            v_verify = ws_v2.cell(gye_row, c).value
            v_out = ws_o2.cell(out_row, c).value
            vv = str(v_verify or "").strip()
            vo = str(v_out or "").strip()
            if vv != vo:
                diff.append("C%d: 검증=%s | 출력=%s" % (c, repr(vv[:50]), repr(vo[:50])))
        if diff:
            print("차이 %d건:" % len(diff))
            for d in diff[:20]:
                print("  ", d)
        else:
            print("일치!")
        wb_v2.close()
        wb_o2.close()

if __name__ == "__main__":
    main()
