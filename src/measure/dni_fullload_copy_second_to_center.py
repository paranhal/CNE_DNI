# -*- coding: utf-8 -*-
"""
DNO_FULLLOAD_MEANSURE_수정.xlsx 전부하_통합 시트 2차 결과 복사

[동작]
- V열(22) 학교명과 B열(2) 학교명이 같은 학교끼리 매칭.
- 같은 학교 내 행 순서대로, 2차 데이터(X~AC)를 1차 행의 N~S에 복사.
  N←X, O←Y, P←Z, Q←AA, R←AB, S←AC
"""

from __future__ import annotations

import os
from collections import defaultdict

from openpyxl import load_workbook


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DNI_DIR = os.path.join(BASE_DIR, "DNI")
INPUT_PATH = os.path.join(DNI_DIR, "DNO_FULLLOAD_MEANSURE_수정.xlsx")


def _log(msg: str) -> None:
    print(msg, flush=True)


def main() -> None:
    if not os.path.isfile(INPUT_PATH):
        _log(f"[오류] 파일이 없습니다: {INPUT_PATH}")
        return

    _log("=" * 60)
    _log("[DNI 전부하] V열·B열 학교명 매칭 → X~AC → N~S 복사")
    _log(f"대상: {INPUT_PATH}")
    _log("=" * 60)

    wb = load_workbook(INPUT_PATH, data_only=False)
    ws = wb["전부하_통합"] if "전부하_통합" in wb.sheetnames else wb[wb.sheetnames[0]]
    max_row = ws.max_row
    _log(f"시트: {ws.title} | 행 수: {max_row}")

    COL_B = 2       # B열: 1차 학교명
    COL_V = 22      # V열: 2차 학교명
    SRC_X  = 24;  SRC_Y  = 25;  SRC_Z  = 26
    SRC_AA = 27;  SRC_AB = 28;  SRC_AC = 29
    DST_N = 14;  DST_O = 15;  DST_P = 16
    DST_Q = 17;  DST_R = 18;  DST_S = 19
    DST_COLS = (DST_N, DST_O, DST_P, DST_Q, DST_R, DST_S)
    SRC_COLS = (SRC_X, SRC_Y, SRC_Z, SRC_AA, SRC_AB, SRC_AC)

    def _s(v):
        return (str(v).strip() if v is not None else "") or ""

    # ── 1단계: N~S 비우기 ──
    for r in range(2, max_row + 1):
        for c in DST_COLS:
            ws.cell(row=r, column=c).value = None
    _log("[1단계] N~S 비움")

    # ── 2단계: 1차 행(B열 학교명) 수집 ──
    rows_1st = defaultdict(list)   # 학교명 → [row, ...]
    for r in range(2, max_row + 1):
        name = _s(ws.cell(row=r, column=COL_B).value)
        if name:
            rows_1st[name].append(r)

    # ── 3단계: 2차 행(V열 학교명 + X~AC 데이터) 수집 ──
    data_2nd = defaultdict(list)   # 학교명 → [(x,y,z,aa,ab,ac), ...]
    for r in range(2, max_row + 1):
        name = _s(ws.cell(row=r, column=COL_V).value)
        if not name:
            continue
        vals = tuple(ws.cell(row=r, column=c).value for c in SRC_COLS)
        if all(v is None for v in vals):
            continue
        data_2nd[name].append(vals)

    _log(f"[2단계] 1차(B열): {len(rows_1st)}개 학교, {sum(len(v) for v in rows_1st.values())}행")
    _log(f"[3단계] 2차(V열): {len(data_2nd)}개 학교, {sum(len(v) for v in data_2nd.values())}행")

    common = set(rows_1st) & set(data_2nd)
    _log(f"  공통 학교: {len(common)}개")
    if not common:
        _log(f"  1차 샘플: {list(rows_1st.keys())[:3]}")
        _log(f"  2차 샘플: {list(data_2nd.keys())[:3]}")

    # ── 4단계: 같은 학교명끼리 행 순서대로 N~S에 복사 ──
    updated = 0
    for school in common:
        first_rows = sorted(rows_1st[school])
        second_vals = data_2nd[school]
        for i, row in enumerate(first_rows):
            if i >= len(second_vals):
                break
            vals = second_vals[i]
            for ci, dst_c in enumerate(DST_COLS):
                ws.cell(row=row, column=dst_c).value = vals[ci]
            updated += 1

    _log(f"[4단계] N~S 복사 완료: {updated}행")

    # 샘플 확인
    for school in list(common)[:3]:
        n_filled = sum(
            1 for r in rows_1st[school]
            if ws.cell(row=r, column=DST_N).value is not None
        )
        _log(f"  {school}: 1차={len(rows_1st[school])}행, 2차={len(data_2nd[school])}행, N채움={n_filled}")

    from datetime import datetime
    out_path = INPUT_PATH
    try:
        wb.save(out_path)
    except PermissionError:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = os.path.join(DNI_DIR, f"DNO_FULLLOAD_MEANSURE_수정_merged_{stamp}.xlsx")
        _log(f"[경고] 원본 열림 → 대신 저장: {out_path}")
        wb.save(out_path)
    wb.close()
    _log(f"[완료] 저장: {out_path}")


if __name__ == "__main__":
    main()
