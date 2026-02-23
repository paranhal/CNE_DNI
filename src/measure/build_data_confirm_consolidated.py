# -*- coding: utf-8 -*-
"""
개별 학교 리포트 데이터 → 통합 파일 (행/열 전환)

[입력]
- CNE/학교별_리포트/*.xlsx (개별 학교 리포트)
- TOTAL_MEASURE_LIST Sheet1 (지역 정보)
- 최종_데이터확인_템플릿.xlsx (템플릿, 있으면 사용)

[출력]
- CNE/데이터확인_통합.xlsx (별도 파일)
- A열: 지역, B열: 학교코드, C열: 학교명, D열~: 측정 데이터 (개별 F11,F22,... 값을 열로)
"""
from __future__ import print_function
import sys
import io
import os
import re
import copy

if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

from openpyxl import load_workbook, Workbook
from tqdm import tqdm

_MEASURE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, _MEASURE_DIR)

from school_report_config import TOTAL_MEASURE_LIST, CNE_DIR, J_OUTPUT_MAP, J_COL

# 출력 파일 (별도 저장)
REPORT_DIR = os.path.join(CNE_DIR, "학교별_리포트")
OUTPUT_FILE = os.path.join(CNE_DIR, "데이터확인_통합.xlsx")
TEMPLATE_PATH = os.path.join(_MEASURE_DIR, "최종_데이터확인_템플릿.xlsx")

# J_OUTPUT_MAP 행 순서
DATA_ROWS = [rd[0] for rd in J_OUTPUT_MAP]

# 최종_데이터확인_템플릿 열 매핑: {개별리포트 행: 템플릿 열}
# C1~3: 지역,학교코드,학교명 / C4~22: 전수조사(미입력) / C23~36: 측정데이터
TEMPLATE_DATA_COLS = {
    11: 13,   # F11 ISP RSSI → C13 신호강도(4지점)
    22: 23,   # F22 전부하 → C23 10개교실 평균속도
    23: 24,   # F23 ISP C → C24 교실별 개별속도
    24: 25,   # F24 ISP D → C25 무선 다운로드
    25: 26,   # F25 ISP RSSI → C26 무선 업로드
    28: 30,   # F28 → C30 지연시간(RTT)
    29: 31,   # F29 → C31 집선↔외부망 다운로드
    30: 32,   # F30 → C32 집선↔외부망 업로드
    31: 33,   # F31 → C33 패킷손실률
    32: 34,   # F32 → C34 Jitter
    33: 35,   # F33 → C35 L2↔집선 다운로드
    34: 36,   # F34 → C36 L2↔집선 업로드
}


def load_region_code_name():
    """TOTAL_MEASURE_LIST Sheet1에서 지역, 학교코드, 학교명"""
    code_to_region = {}
    code_to_name = {}
    if not os.path.isfile(TOTAL_MEASURE_LIST):
        return code_to_region, code_to_name
    wb = load_workbook(TOTAL_MEASURE_LIST, data_only=True)
    if "Sheet1" not in wb.sheetnames:
        wb.close()
        return code_to_region, code_to_name
    ws = wb["Sheet1"]
    for r in range(2, ws.max_row + 1):
        code = str(ws.cell(r, 1).value or "").strip()
        region = str(ws.cell(r, 2).value or "").strip()
        name = str(ws.cell(r, 3).value or "").strip()
        if code:
            code_to_region[code] = region
            code_to_name[code] = name
    wb.close()
    return code_to_region, code_to_name


def extract_school_code_from_filename(fname):
    """파일명에서 학교코드 추출 (학교명_학교코드.xlsx)"""
    base = os.path.splitext(fname)[0]
    m = re.search(r"_([A-Z0-9]{12,})$", base)
    if m:
        return m.group(1)
    return None


def read_report_data(report_path):
    """개별 리포트에서 F열(측정값) 데이터 추출"""
    data = {}
    try:
        wb = load_workbook(report_path, data_only=True)
        ws = wb.active
        if "문제점분석" in wb.sheetnames:
            ws = wb["문제점분석"]
        for i, row in enumerate(DATA_ROWS):
            val = ws.cell(row=row, column=J_COL).value
            data[row] = val if val is not None else ""
        wb.close()
    except Exception:
        pass
    return data


def save_template_row(ws, row, max_col):
    """템플릿 행의 값·스타일 저장 (덮어쓰기 전에 호출)"""
    cache = []
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        try:
            font = copy.copy(cell.font) if cell.font else None
        except Exception:
            font = None
        try:
            border = copy.copy(cell.border) if cell.border else None
        except Exception:
            border = None
        try:
            fill = copy.copy(cell.fill) if cell.fill else None
        except Exception:
            fill = None
        try:
            alignment = copy.copy(cell.alignment) if cell.alignment else None
        except Exception:
            alignment = None
        cache.append({
            "value": cell.value,
            "font": font,
            "border": border,
            "fill": fill,
            "number_format": cell.number_format,
            "alignment": alignment,
        })
    return cache


def apply_template_row(ws, row, template_cache):
    """저장된 템플릿 행을 대상 행에 적용"""
    for c, data in enumerate(template_cache, start=1):
        dst = ws.cell(row=row, column=c)
        dst.value = data["value"]
        if data["font"]:
            dst.font = data["font"]
        if data["border"]:
            dst.border = data["border"]
        if data["fill"]:
            dst.fill = data["fill"]
        if data["number_format"]:
            dst.number_format = data["number_format"]
        if data["alignment"]:
            dst.alignment = data["alignment"]


def main():
    print("=" * 50)
    print("[데이터확인 통합] 개별 리포트 → 통합 (행/열 전환)")
    print("=" * 50)
    code_to_region, code_to_name = load_region_code_name()
    if not os.path.isdir(REPORT_DIR):
        print(f"[오류] 리포트 폴더 없음: {REPORT_DIR}")
        sys.exit(1)
    files = [f for f in os.listdir(REPORT_DIR) if f.endswith(".xlsx")]
    school_files = []
    for f in files:
        code = extract_school_code_from_filename(f)
        if code:
            school_files.append((code, os.path.join(REPORT_DIR, f)))
    school_files.sort(key=lambda x: x[0])
    print(f"대상 학교: {len(school_files)}개")
    if not os.path.isfile(TEMPLATE_PATH):
        print(f"[오류] 템플릿 없음: {TEMPLATE_PATH}")
        sys.exit(1)
    wb_out = load_workbook(TEMPLATE_PATH)
    ws_out = wb_out.active
    if "데이터확인" in wb_out.sheetnames:
        ws_out = wb_out["데이터확인"]
    # 템플릿 R1=헤더, R2=고정값(상수). R2의 고정값을 3행 이후에도 복사
    max_col = max(36, ws_out.max_column)
    template_row2 = save_template_row(ws_out, 2, max_col)
    data_start_row = 2
    for ri, (school_code, report_path) in enumerate(tqdm(school_files, desc="통합", unit="교"), data_start_row):
        if ri > 2:
            apply_template_row(ws_out, ri, template_row2)
        ws_out.cell(ri, 1, value=code_to_region.get(school_code, ""))
        ws_out.cell(ri, 2, value=school_code)
        ws_out.cell(ri, 3, value=code_to_name.get(school_code, ""))
        row_data = read_report_data(report_path)
        for data_row, template_col in TEMPLATE_DATA_COLS.items():
            val = row_data.get(data_row, "")
            ws_out.cell(ri, template_col, value=val)
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    wb_out.save(OUTPUT_FILE)
    print(f"[완료] {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
