# -*- coding: utf-8 -*-
"""템플릿, 결과, 원본 데이터 구조 분석"""
import sys
import io
import os

if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from openpyxl import load_workbook

BASE = os.path.dirname(os.path.abspath(__file__))
CNE = os.path.join(BASE, "CNE")

def analyze(path, name, rows_to_show=40):
    if not os.path.isfile(path):
        print(f"[{name}] 파일 없음: {path}")
        return
    print(f"\n{'='*60}")
    print(f"[{name}] {path}")
    print("="*60)
    wb = load_workbook(path, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- 시트: {sheet_name} ({ws.max_row}행 x {ws.max_column}열) ---")
        for r in range(1, min(rows_to_show, ws.max_row + 1)):
            row_vals = []
            for c in range(1, min(15, ws.max_column + 1)):
                v = ws.cell(r, c).value
                row_vals.append(str(v)[:15] if v is not None else "")
            print(f"  R{r}: {row_vals}")
    wb.close()

# 측정값_템플릿.xlsx
analyze(os.path.join(BASE, "측정값_템플릿.xlsx"), "템플릿")
# 학교별_리포트 내 파일
report_dir = os.path.join(CNE, "학교별_리포트")
if os.path.isdir(report_dir):
    files = [f for f in os.listdir(report_dir) if f.endswith(".xlsx")]
    if files:
        # N10C035903HS가 포함된 파일 또는 첫 파일
        target = next((f for f in files if "N10C035903HS" in f), files[0])
        analyze(os.path.join(report_dir, target), f"결과파일({target})", 45)
# TOTAL_MEASURE_LIST
for fname in ["TOTAL_MEASURE_LIST_V1.xlsx", "TOTAL_MEASURE_LIST_V1.XLSX"]:
    p = os.path.join(CNE, fname)
    if os.path.isfile(p):
        analyze(p, "원본데이터", 5)
        break
