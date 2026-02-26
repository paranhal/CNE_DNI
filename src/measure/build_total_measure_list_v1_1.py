# -*- coding: utf-8 -*-
"""
통계 데이터 통합 → TOTAL_MEASURE_LIST_V1.XLSX (V1.1)

전부하측정_학교별평균, ISP측정_학교별평균, 집선ISP, CNE_WIRED_MEANSURE_AVG 시트 통합
+ 신규: 학교별통신장비현황, POE, 케이블통계, AP_장비통계
"""
from __future__ import print_function
import sys
import io
import os

if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

from openpyxl import load_workbook, Workbook

_MEASURE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, _MEASURE_DIR)

from school_report_config_v1_1 import TOTAL_MEASURE_LIST, CNE_DIR

TRACE_SCHOOL_CODE = (os.environ.get("TRACE_SCHOOL_CODE") or "").strip()

# 소스 파일 및 시트 (있으면 복사, 없으면 스킵)
SOURCES = [
    (os.path.join(CNE_DIR, "CNE_FULLLOAD_MEASURE.xlsx"), "전부하측정_학교별평균"),
    (os.path.join(CNE_DIR, "CNE_ISP_MEASURE.XLSX"), "ISP측정_학교별평균"),
    (os.path.join(CNE_DIR, "집선ISP.xlsx"), "집선ISP"),
    (os.path.join(CNE_DIR, "집선ISP_측정.xlsx"), "집선ISP"),
    (os.path.join(CNE_DIR, "CNE_WIRED_MEANSURE_V1.XLSX"), "CNE_WIRED_MEANSURE_AVG"),
    # 신규 시트 (파일·시트 있으면 통합)
    (os.path.join(CNE_DIR, "학교별통신장비현황.xlsx"), "학교별통신장비현황"),
    (os.path.join(CNE_DIR, "CNE_POE_LIST.xlsx"), "POE"),
    (os.path.join(CNE_DIR, "케이블통계.xlsx"), "케이블통계"),
    (os.path.join(CNE_DIR, "CNE_AP_LIST.xlsx"), "AP_장비통계"),
    (os.path.join(CNE_DIR, "충남AP.xlsx"), "충남AP"),
]


def main():
    print("=" * 50)
    print("[통합 V1.1] TOTAL_MEASURE_LIST_V1.1.XLSX 생성")
    print("=" * 50)
    wb = Workbook()
    wb.remove(wb.active)
    copied = set()
    trace_summary = []
    for path, sheet_name in SOURCES:
        if sheet_name in copied:
            continue
        if not os.path.isfile(path):
            continue
        try:
            wb_src = load_workbook(path, data_only=True)
            if sheet_name not in wb_src.sheetnames:
                wb_src.close()
                continue
            ws_src = wb_src[sheet_name]
            ws_dst = wb.create_sheet(sheet_name)
            for r in range(1, ws_src.max_row + 1):
                for c in range(1, ws_src.max_column + 1):
                    v = ws_src.cell(r, c).value
                    ws_dst.cell(r, c, value=v)
            wb_src.close()
            print(f"  [복사] {sheet_name} <- {os.path.basename(path)}")
            copied.add(sheet_name)

            if TRACE_SCHOOL_CODE:
                # 학교코드 위치 탐색 후 추적 대상 행 샘플 출력
                code_col = 1
                for c in range(1, min(50, ws_dst.max_column + 1)):
                    h = str(ws_dst.cell(1, c).value or "").strip().lower()
                    if "학교코드" in h or h == "code" or h.endswith("code"):
                        code_col = c
                        break
                hit_row = None
                for r in range(2, ws_dst.max_row + 1):
                    v = str(ws_dst.cell(r, code_col).value or "").strip()
                    if v == TRACE_SCHOOL_CODE:
                        hit_row = r
                        break
                if hit_row:
                    vals = [ws_dst.cell(hit_row, c).value for c in range(1, min(11, ws_dst.max_column + 1))]
                    trace_summary.append((sheet_name, hit_row, vals))
                else:
                    trace_summary.append((sheet_name, None, []))
        except Exception as e:
            print(f"  [오류] {path}: {e}")
    if len(copied) == 0:
        print("[오류] 복사된 시트 없음")
        sys.exit(1)
    os.makedirs(os.path.dirname(TOTAL_MEASURE_LIST), exist_ok=True)
    wb.save(TOTAL_MEASURE_LIST)
    print(f"[완료] {TOTAL_MEASURE_LIST}")
    if TRACE_SCHOOL_CODE:
        print(f"[TRACE][TOTAL 병합] 학교코드={TRACE_SCHOOL_CODE}")
        for sn, rr, vals in trace_summary:
            if rr is None:
                print(f"  - {sn}: 없음")
            else:
                print(f"  - {sn}: row={rr} A~J={vals}")


if __name__ == "__main__":
    main()
