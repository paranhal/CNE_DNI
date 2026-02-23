# -*- coding: utf-8 -*-
"""
719개 학교 기준 자료 유무 확인 (실제 데이터 값)

school_reg_list_CNE 719개 학교를 키로,
TOTAL_MEASURE_LIST 각 시트에서 실제 측정값을 추출하여 통합
"""
from __future__ import print_function
import sys
import io
import os
import csv
import re

if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

from openpyxl import load_workbook, Workbook

_MEASURE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, _MEASURE_DIR)

from school_report_config import TOTAL_MEASURE_LIST, CNE_DIR, J_OUTPUT_MAP

SCHOOL_LIST_SEARCH_DIRS = [_MEASURE_DIR, os.path.join(os.path.dirname(_MEASURE_DIR), "split")]
OUTPUT_FILE = os.path.join(CNE_DIR, "학교별_자료유무_확인.xlsx")


def load_full_school_list():
    """719개 대상 학교 리스트 로드 (학교코드 순서 유지)"""
    for base_dir in SCHOOL_LIST_SEARCH_DIRS:
        for fname in ["school_reg_list_CNE.xlsx", "school_reg_list_CNE.csv", "SCHOOL_REG_LIST_CNE.xlsx"]:
            path = os.path.join(base_dir, fname)
            if not os.path.isfile(path):
                continue
            try:
                codes = []
                code_to_name = {}
                if path.endswith(".csv"):
                    with open(path, "r", encoding="utf-8-sig") as f:
                        reader = csv.reader(f)
                        header = next(reader, None)
                        rows = list(reader)
                    code_col = name_col = 0
                    for i, h in enumerate(header or []):
                        s = str(h or "").lower()
                        if "학교코드" in s or "code" in s:
                            code_col = i
                        if "학교명" in s or "name" in s:
                            name_col = i
                    for row in rows:
                        if len(row) > max(code_col, name_col):
                            code = str(row[code_col] or "").strip()
                            name = str(row[name_col] or "").strip()
                            if code:
                                codes.append(code)
                                code_to_name[code] = name
                else:
                    wb = load_workbook(path, read_only=True, data_only=True)
                    ws = wb.active
                    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
                    code_col = name_col = 0
                    for i, h in enumerate(header or []):
                        s = str(h or "").lower()
                        if "학교코드" in s or "code" in s:
                            code_col = i
                        if "학교명" in s or "name" in s:
                            name_col = i
                    for r in range(2, ws.max_row + 1):
                        code = str(ws.cell(r, code_col + 1).value or "").strip()
                        name = str(ws.cell(r, name_col + 1).value or "").strip()
                        if code:
                            codes.append(code)
                            code_to_name[code] = name
                    wb.close()
                if codes:
                    return codes, code_to_name
            except Exception:
                pass
    return [], {}


def load_sheet_by_school(wb, sheet_name):
    """시트에서 학교코드별 행 데이터 로드 {school_code: row_index}"""
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    code_col = 1
    for c in range(1, min(6, ws.max_column + 1)):
        v = ws.cell(1, c).value
        if v and ("학교코드" in str(v) or "code" in str(v).lower()):
            code_col = c
            break
    result = {}
    for r in range(2, ws.max_row + 1):
        sc = str(ws.cell(r, code_col).value or "").strip()
        if sc:
            result[sc] = r
    return result


def get_cell_value(ws, row, col):
    v = ws.cell(row, col).value
    if v is None or (isinstance(v, str) and not str(v).strip()):
        return ""
    return v


def main():
    print("=" * 50)
    print("[학교별 자료 유무 확인] 719개 학교, 실제 데이터 값")
    print("=" * 50)
    all_schools, code_to_name = load_full_school_list()
    if not all_schools:
        print("[오류] school_reg_list_CNE에서 학교 리스트를 로드할 수 없습니다.")
        sys.exit(1)
    print(f"대상 학교: {len(all_schools)}개")
    if not os.path.isfile(TOTAL_MEASURE_LIST):
        print(f"[오류] 통계 파일 없음: {TOTAL_MEASURE_LIST}")
        sys.exit(1)
    wb = load_workbook(TOTAL_MEASURE_LIST, data_only=True)
    # 시트별 학교코드→행 매핑
    sheet_rows = {}
    for sheet_name in wb.sheetnames:
        sheet_rows[sheet_name] = load_sheet_by_school(wb, sheet_name)
    # J_OUTPUT_MAP 기반 컬럼 정의: (시트, 열1, 열2, 제목)
    cols = [
        ("학교코드", None, None, None),
        ("학교명", None, None, None),
        ("전부하_다운로드", "전부하측정_학교별평균", 3, None),
        ("전부하_업로드", "전부하측정_학교별평균", 4, None),
        ("ISP_다운로드", "ISP측정_학교별평균", 3, None),
        ("ISP_업로드", "ISP측정_학교별평균", 4, None),
        ("집선ISP_E", "집선ISP", 5, None),
        ("집선ISP_C", "집선ISP", 3, None),
        ("집선ISP_D", "집선ISP", 4, None),
        ("CNE_WIRED_E", "CNE_WIRED_MEANSURE_AVG", 5, None),
        ("CNE_WIRED_F", "CNE_WIRED_MEANSURE_AVG", 6, None),
        ("CNE_WIRED_D1", "CNE_WIRED_MEANSURE_AVG", 4, None),
        ("CNE_WIRED_D2", "CNE_WIRED_MEANSURE_AVG", 4, None),
    ]
    headers = [c[0] for c in cols]
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "학교별_자료유무"
    for c, h in enumerate(headers, 1):
        ws_out.cell(row=1, column=c, value=h)
    for ri, school_code in enumerate(all_schools, 2):
        ws_out.cell(ri, 1, value=school_code)
        ws_out.cell(ri, 2, value=code_to_name.get(school_code, ""))
        for ci, (_, sheet_name, col1, _) in enumerate(cols[2:], 3):
            if not sheet_name or not col1:
                continue
            val = ""
            if school_code in sheet_rows.get(sheet_name, {}):
                row = sheet_rows[sheet_name][school_code]
                ws = wb[sheet_name]
                val = get_cell_value(ws, row, col1)
            ws_out.cell(ri, ci, value=val)
    wb.close()
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    wb_out.save(OUTPUT_FILE)
    print(f"[완료] {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
