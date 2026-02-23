# -*- coding: utf-8 -*-
"""N10C035903HS 원본 데이터 및 결과 파일 J/L열 확인"""
import os
from openpyxl import load_workbook

BASE = os.path.dirname(os.path.abspath(__file__))
CNE = os.path.join(BASE, "CNE")
code = "N10C035903HS"

# 1. TOTAL_MEASURE_LIST에서 N10C035903HS 데이터
path = os.path.join(CNE, "TOTAL_MEASURE_LIST_V1.xlsx")
if not os.path.isfile(path):
    path = os.path.join(CNE, "TOTAL_MEASURE_LIST_V1.XLSX")
print("=" * 60)
print("[원본] TOTAL_MEASURE_LIST - N10C035903HS")
print("=" * 60)
wb = load_workbook(path, data_only=True)
for sn in wb.sheetnames:
    ws = wb[sn]
    code_col = 1
    for c in range(1, min(6, ws.max_column + 1)):
        v = ws.cell(1, c).value
        if v and ("학교코드" in str(v) or "학교명" in str(v)):
            code_col = 1 if "학교코드" in str(v) else 2
            if "학교명" in str(v) and "학교코드" not in str(v):
                code_col = 2
            break
    for c in range(1, min(6, ws.max_column + 1)):
        v = ws.cell(1, c).value
        if v and "학교코드" in str(v):
            code_col = c
            break
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, code_col).value
        if v and str(v).strip() == code:
            row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            print(f"\n{sn}: {row}")
            break
wb.close()

# 2. 결과 파일 J/L열 (행 20,22,23,27,28,29,30,31,32,33)
report_dir = os.path.join(CNE, "학교별_리포트")
target = None
for f in os.listdir(report_dir or []):
    if f.endswith(".xlsx") and code in f:
        target = f
        break
if not target and os.path.isdir(report_dir):
    target = next((f for f in os.listdir(report_dir) if f.endswith(".xlsx")), None)
if target:
    path = os.path.join(report_dir, target)
    print("\n" + "=" * 60)
    print(f"[결과] {target} - J열(측정값), L열(평가결과)")
    print("=" * 60)
    wb = load_workbook(path, data_only=True)
    ws = wb["문제점분석기본"] if "문제점분석기본" in wb.sheetnames else wb.active
    # 분류코드 A20,A22,A23,A27,A28,A29,A30,A31,A32,A33 해당 행
    code_rows = {"A20": 30, "A22": 32, "A23": 33, "A27": 37, "A28": 38, "A29": 39, "A30": 40, "A31": 41, "A32": 42, "A33": 43}
    for cls_code, row in sorted(code_rows.items(), key=lambda x: x[1]):
        j_val = ws.cell(row, 10).value
        k_val = ws.cell(row, 11).value
        l_val = ws.cell(row, 12).value
        print("  %s R%d: J=%r  K=%r  L=%r" % (cls_code, row, j_val, k_val, l_val))
    wb.close()
