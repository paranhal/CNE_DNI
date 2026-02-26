# -*- coding: utf-8 -*-
"""
대전(DNI) ISP 측정값 + 학교 리스트(315개) 통합

[입력]
- 학교 리스트: src/split/school_reg_list_DNI.csv
  - 컬럼: 학교코드, 지역, 학교명
- ISP 평균: src/measure/DNI/DNI_ISP_MEASURE.XLSX
  - 시트: 두 개 존재 (이름은 깨져있을 수 있음)
  - 평균 시트(두 번째 시트)의 1행 헤더:
    - 학교명, 학교코드, 다운로드, 업로드, RTT, RSSI, CH, 다운로드 진단, 업로드 진단

[출력]
- 같은 통합 파일 내 새 시트 추가:
  - 시트명: "ISP_315학교"
  - 행: 315개 (school_reg_list_DNI 기준 전 학교)
  - 컬럼:
    - 학교코드, 지역, 학교명, 다운로드, 업로드, RTT, RSSI, CH, 다운로드 진단, 업로드 진단
  - ISP 데이터가 없는 학교는 ISP 관련 컬럼을 빈칸으로 둠
"""
from __future__ import annotations

import csv
import os
import sys
from typing import Dict, Any

from openpyxl import load_workbook

if getattr(sys, "frozen", False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
RUN_DIR = os.getcwd()

def _pick_existing(candidates):
    for p in candidates:
        if os.path.isfile(p):
            return p
    return candidates[0]


SCHOOL_LIST_PATH = _pick_existing(
    [
        os.path.join(RUN_DIR, "school_reg_list_DNI.csv"),
        os.path.join(RUN_DIR, "split", "school_reg_list_DNI.csv"),
        os.path.join(BASE_DIR, "split", "school_reg_list_DNI.csv"),
        os.path.join(os.path.dirname(BASE_DIR), "split", "school_reg_list_DNI.csv"),
    ]
)
DNI_ISP_MEASURE_PATH = _pick_existing(
    [
        os.path.join(RUN_DIR, "DNI_ISP_MEASURE.XLSX"),
        os.path.join(RUN_DIR, "DNI", "DNI_ISP_MEASURE.XLSX"),
        os.path.join(BASE_DIR, "DNI_ISP_MEASURE.XLSX"),
        os.path.join(BASE_DIR, "DNI", "DNI_ISP_MEASURE.XLSX"),
    ]
)
TRACE_SCHOOL_CODE = (os.environ.get("TRACE_SCHOOL_CODE") or "").strip()


def _log(msg: str) -> None:
    print(msg, flush=True)


def load_school_list(path: str) -> list[dict[str, Any]]:
    """school_reg_list_DNI.csv 로드 → [{code, region, name}, ...]"""
    schools: list[dict[str, Any]] = []
    if not os.path.isfile(path):
        _log(f"[오류] 학교 리스트 파일을 찾을 수 없습니다: {path}")
        return schools

    with open(path, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        header = next(reader, None)
        if not header:
            _log("[오류] 학교 리스트 헤더가 비어 있습니다.")
            return schools

        # 헤더 인덱스: 학교코드, 지역, 학교명
        code_idx = region_idx = name_idx = 0
        for i, h in enumerate(header):
            s = (h or "").strip()
            if "학교코드" in s or "code" in s:
                code_idx = i
            elif "지역" in s:
                region_idx = i
            elif "학교명" in s or "name" in s:
                name_idx = i

        for row in reader:
            if len(row) <= code_idx:
                continue
            code = (row[code_idx] or "").strip()
            if not code:
                continue
            region = (row[region_idx] or "").strip() if len(row) > region_idx else ""
            name = (row[name_idx] or "").strip() if len(row) > name_idx else ""
            schools.append({"code": code, "region": region, "name": name})

    _log(f"[학교 리스트] {len(schools)}개 로드 (파일: {path})")
    return schools


def load_isp_avg_by_school(path: str) -> Dict[str, dict[str, Any]]:
    """
    DNI_ISP_MEASURE.XLSX의 평균 시트에서 학교코드별 ISP 데이터 로드
    - 시트 구조 헤더(1행):
      학교명, 학교코드, 다운로드, 업로드, RTT, RSSI, CH, 다운로드 진단, 업로드 진단
    """
    if not os.path.isfile(path):
        _log(f"[오류] ISP 통계 파일을 찾을 수 없습니다: {path}")
        return {}

    wb = load_workbook(path, data_only=True)
    sheetnames = wb.sheetnames
    if len(sheetnames) < 2:
        _log(f"[오류] 평균 시트(두 번째 시트)가 없습니다. 시트 목록: {sheetnames}")
        wb.close()
        return {}

    ws = wb[sheetnames[1]]  # 두 번째 시트: ISP측정_학교별평균
    max_row, max_col = ws.max_row, ws.max_column
    headers = [ws.cell(1, c).value for c in range(1, max_col + 1)]

    # 컬럼 인덱스 찾기
    idx_name = idx_code = idx_dl = idx_ul = idx_rtt = idx_rssi = idx_ch = idx_diag_dl = idx_diag_ul = None
    for i, h in enumerate(headers):
        s = str(h or "").strip()
        if s == "학교명":
            idx_name = i
        elif s == "학교코드":
            idx_code = i
        elif s == "다운로드":
            idx_dl = i
        elif s == "업로드":
            idx_ul = i
        elif s == "RTT":
            idx_rtt = i
        elif s == "RSSI":
            idx_rssi = i
        elif s == "CH":
            idx_ch = i
        elif s == "다운로드 진단":
            idx_diag_dl = i
        elif s == "업로드 진단":
            idx_diag_ul = i

    if idx_code is None:
        _log("[오류] 평균 시트에서 '학교코드' 열을 찾을 수 없습니다.")
        wb.close()
        return {}

    by_school: Dict[str, dict[str, Any]] = {}
    for r in range(2, max_row + 1):
        code = (ws.cell(r, idx_code + 1).value or "").strip()
        if not code:
            continue
        rec: dict[str, Any] = {
            "school_name": ws.cell(r, idx_name + 1).value if idx_name is not None else "",
            "download": ws.cell(r, idx_dl + 1).value if idx_dl is not None else "",
            "upload": ws.cell(r, idx_ul + 1).value if idx_ul is not None else "",
            "rtt": ws.cell(r, idx_rtt + 1).value if idx_rtt is not None else "",
            "rssi": ws.cell(r, idx_rssi + 1).value if idx_rssi is not None else "",
            "ch": ws.cell(r, idx_ch + 1).value if idx_ch is not None else "",
            "diag_dl": ws.cell(r, idx_diag_dl + 1).value if idx_diag_dl is not None else "",
            "diag_ul": ws.cell(r, idx_diag_ul + 1).value if idx_diag_ul is not None else "",
        }
        by_school[code] = rec

    wb.close()
    _log(f"[ISP 평균] {len(by_school)}개 학교 로드 (파일: {path})")
    return by_school


def merge_to_new_sheet(school_list_path=None, isp_measure_path=None):
    school_list_path = school_list_path or SCHOOL_LIST_PATH
    isp_measure_path = isp_measure_path or DNI_ISP_MEASURE_PATH
    schools = load_school_list(school_list_path)
    isp_by_school = load_isp_avg_by_school(isp_measure_path)
    if not schools:
        return False

    if not os.path.isfile(isp_measure_path):
        _log(f"[오류] ISP 파일이 없습니다: {isp_measure_path}")
        return False

    wb = load_workbook(isp_measure_path, data_only=True)

    sheet_name = "ISP_315학교"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws_out = wb.create_sheet(sheet_name)

    headers = [
        "학교코드",
        "지역",
        "학교명",
        "다운로드",
        "업로드",
        "RTT",
        "RSSI",
        "CH",
        "다운로드 진단",
        "업로드 진단",
    ]
    for c, h in enumerate(headers, 1):
        ws_out.cell(row=1, column=c, value=h)

    missing_count = 0
    for i, s in enumerate(schools, start=2):
        code = s["code"]
        region = s.get("region", "")
        name = s.get("name", "")
        ws_out.cell(row=i, column=1, value=code)
        ws_out.cell(row=i, column=2, value=region)
        ws_out.cell(row=i, column=3, value=name)

        rec = isp_by_school.get(code)
        if not rec:
            missing_count += 1
            if TRACE_SCHOOL_CODE and code == TRACE_SCHOOL_CODE:
                _log("[TRACE][ISP_315학교] 매칭 데이터 없음 -> 빈칸 유지")
            continue

        ws_out.cell(row=i, column=4, value=rec.get("download", ""))
        ws_out.cell(row=i, column=5, value=rec.get("upload", ""))
        ws_out.cell(row=i, column=6, value=rec.get("rtt", ""))
        ws_out.cell(row=i, column=7, value=rec.get("rssi", ""))
        ws_out.cell(row=i, column=8, value=rec.get("ch", ""))
        ws_out.cell(row=i, column=9, value=rec.get("diag_dl", ""))
        ws_out.cell(row=i, column=10, value=rec.get("diag_ul", ""))
        if TRACE_SCHOOL_CODE and code == TRACE_SCHOOL_CODE:
            _log(
                "[TRACE][ISP_315학교] 저장 "
                f"DL={rec.get('download','')} UL={rec.get('upload','')} "
                f"RTT={rec.get('rtt','')} RSSI={rec.get('rssi','')} CH={rec.get('ch','')}"
            )

    wb.save(isp_measure_path)
    wb.close()

    _log(f"[완료] {sheet_name} 시트 생성 완료 (총 {len(schools)}개 학교, ISP 데이터 없는 학교 {missing_count}개)")
    return True


def main():
    _log("=" * 60)
    _log("[DNI ISP] school_reg_list_DNI + ISP측정_학교별평균 통합 (315개 기준)")
    _log("=" * 60)
    return merge_to_new_sheet()


if __name__ == "__main__":
    main()

