# -*- coding: utf-8 -*-
"""
유선망 1차 측정 결과 전처리 - 충남/대전 공통

1. 학교별 파일 통합 → {지역}_TOTAL 시트
2. 학교별 평균 계산 → {지역}_WIRED_MEANSURE_AVG 시트
3. {지역}_WIRED_MEANSURE_V1.XLSX 저장
"""
from __future__ import print_function
import sys
import io

# 한글 출력 깨짐 방지 (PowerShell/CMD UTF-8)
if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

print("[유선망 전처리] 스크립트 시작", flush=True)

# ========== 실행 지역 선택 ==========
RUN_REGION = "CNE"   # "CNE" = 충남  |  "DNI" = 대전  |  "ALL" = 둘 다

import os

# 실행 시작 즉시 출력 (버퍼링 방지)
def _log(msg):
    print(msg, flush=True)
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from tqdm import tqdm

# config 로드 (스크립트와 동일 폴더)
_MEASURE_DIR = os.path.dirname(os.path.abspath(__file__))
if _MEASURE_DIR not in sys.path:
    sys.path.insert(0, _MEASURE_DIR)

try:
    from wired_preprocess_config import (
        WIRED_1ST_SOURCE,
        WIRED_OUTPUT_FILE,
        SHEET_NAMES,
        SCHOOL_CODE_PATTERN,
        COLUMN_DISPLAY_FORMATS,
        THROUGHPUT_THRESHOLD_MBPS,
        SCHOOL_LIST_SEARCH_DIRS,
        AVG_COLUMN_SELECT,
    )
except Exception as e:
    print(f"[오류] config 로드 실패: {e}", flush=True)
    sys.exit(1)


def load_school_name_map(region):
    """학교 리스트에서 학교코드 → 학교명 매핑 로드 (CSV/Excel)"""
    import csv
    candidates = [
        (f"school_reg_list_{region}.csv", "csv"),
        (f"SCHOOL_REG_LIST_{region}.csv", "csv"),
        (f"school_reg_list_{region}.xlsx", "xlsx"),
        (f"SCHOOL_REG_LIST_{region}.xlsx", "xlsx"),
        ("school_reg_list.csv", "csv"),
        ("SCHOOL_REG_LIST.csv", "csv"),
    ]
    for base_dir in SCHOOL_LIST_SEARCH_DIRS:
        for fname, ext in candidates:
            path = os.path.join(base_dir, fname)
            if not os.path.isfile(path):
                continue
            code_to_name = {}
            try:
                if ext == "csv":
                    with open(path, "r", encoding="utf-8-sig") as f:
                        reader = csv.reader(f)
                        header = next(reader, None)
                        rows = list(reader)
                else:
                    wb = load_workbook(path, read_only=True, data_only=True)
                    ws = wb.active
                    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
                    rows = [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)] for r in range(2, ws.max_row + 1)]
                    wb.close()
                if not header:
                    continue
                code_col = name_col = None
                for i, h in enumerate(header):
                    s = str(h or "").strip().lower()
                    if "학교코드" in s or "code" in s:
                        code_col = i
                    if "학교명" in s or "name" in s:
                        name_col = i
                if code_col is None:
                    code_col = 0
                if name_col is None:
                    name_col = 2
                for row in rows:
                    if len(row) > max(code_col, name_col):
                        code = str(row[code_col] or "").strip()
                        name = str(row[name_col] or "").strip()
                        if code:
                            code_to_name[code] = name
                            code_to_name[code[:12]] = name
                if code_to_name:
                    _log(f"  학교 리스트 로드: {path} ({len(code_to_name)}개)")
                return code_to_name
            except Exception:
                pass
    _log("  [참고] 학교 리스트 없음 → 학교명 빈칸")
    return {}


def find_throughput_column_index(headers):
    """Avg Throughput (Mbps) 열 인덱스 반환 (0-based, all_headers 기준)"""
    for i, h in enumerate(headers):
        if h and ("throughput" in str(h).lower() or "mbps" in str(h).lower()):
            return i
    return None


def get_column_format(header_text):
    """헤더 텍스트에 맞는 열 표시 형식 반환 (숫자+단위)"""
    if not header_text:
        return None
    s = str(header_text).lower()
    for keywords, fmt in COLUMN_DISPLAY_FORMATS:
        if any(kw.lower() in s for kw in keywords):
            return fmt
    return None


def extract_school_code_from_filename(filename):
    """파일명에서 학교코드 추출 (G107441266MS 형식 또는 12자리 숫자)"""
    base = os.path.splitext(filename)[0]
    m = SCHOOL_CODE_PATTERN.search(base)
    return m.group(1) if m else base[:20]  # 못 찾으면 파일명 앞 20자


def find_header_row(ws):
    """헤더 행 찾기 (1~5행 중 '학교코드','관리번호','측정' 등 포함된 행)"""
    for row in range(1, min(6, ws.max_row + 1)):
        for col in range(1, min(ws.max_column + 1, 30)):
            val = ws.cell(row=row, column=col).value
            if val and isinstance(val, str):
                s = val.strip()
                if any(kw in s for kw in ["학교코드", "관리번호", "측정", "대역폭", "지연", "패킷"]):
                    return row
    return 1


def is_numeric(val):
    """숫자형 값 여부"""
    if val is None:
        return False
    if isinstance(val, (int, float)):
        return True
    s = str(val).strip()
    if not s:
        return False
    try:
        float(s.replace(",", "").replace("%", ""))
        return True
    except ValueError:
        return False


def get_numeric_columns(ws, header_row):
    """헤더 행 기준으로 숫자형 데이터 열 인덱스 목록 (1-based)"""
    numeric_cols = []
    for col in range(1, ws.max_column + 1):
        # 데이터 행 샘플로 확인 (헤더 다음 3행)
        for r in range(header_row + 1, min(header_row + 4, ws.max_row + 1)):
            v = ws.cell(row=r, column=col).value
            if is_numeric(v):
                numeric_cols.append(col)
                break
    return numeric_cols


def read_school_file(path, school_code):
    """학교별 엑셀 파일 읽기 → (header_row, headers, data_rows, numeric_cols)"""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    if ws.max_row < 2:
        wb.close()
        return None, [], [], []

    header_row = find_header_row(ws)
    headers = [ws.cell(row=header_row, column=c).value for c in range(1, ws.max_column + 1)]
    numeric_cols = get_numeric_columns(ws, header_row)

    data_rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        row = []
        for c in range(1, ws.max_column + 1):
            row.append(ws.cell(row=r, column=c).value)
        data_rows.append(row)

    wb.close()
    return header_row, headers, data_rows, numeric_cols


def integrate_and_average(region="CNE"):
    """통합 + 학교별 평균 계산"""
    source_dir = WIRED_1ST_SOURCE[region]
    output_path = WIRED_OUTPUT_FILE[region]
    sheet_names = SHEET_NAMES[region]

    _log(f"[진행] 지역: {region} | 소스: {source_dir}")

    if not os.path.isdir(source_dir):
        _log(f"[오류] 소스 폴더가 없습니다: {source_dir}")
        _log("       폴더를 생성하고 학교별 엑셀 파일을 넣어주세요.")
        os.makedirs(os.path.dirname(source_dir), exist_ok=True)
        os.makedirs(source_dir, exist_ok=True)
        return

    # 학교별 xlsx 파일 수집
    files = [f for f in os.listdir(source_dir) if f.lower().endswith((".xlsx", ".xls"))]
    if not files:
        _log(f"[오류] 엑셀 파일이 없습니다: {source_dir}")
        return

    all_headers = None
    all_rows = []
    school_data = {}  # school_code -> {headers, rows, numeric_cols}
    for fname in tqdm(sorted(files), desc=f"[{region}] 파일 읽기", unit="파일"):
        fpath = os.path.join(source_dir, fname)
        if not os.path.isfile(fpath):
            continue
        school_code = extract_school_code_from_filename(fname)
        result = read_school_file(fpath, school_code)
        if result[0] is None:
            continue
        _, headers, data_rows, numeric_cols = result

        if all_headers is None:
            all_headers = ["학교코드"] + [str(h) if h is not None else "" for h in headers]

        for row in data_rows:
            all_rows.append([school_code] + row)

        if school_code not in school_data:
            school_data[school_code] = {"headers": headers, "rows": [], "numeric_cols": numeric_cols}
        school_data[school_code]["rows"].extend(data_rows)
        school_data[school_code]["numeric_cols"] = numeric_cols

    # 학교코드/학교명: TOTAL B열(학교코드), C열(학교명) = 소스 파일 1열, 2열
    throughput_col_idx = find_throughput_column_index(all_headers)

    # 학교별 평균 계산
    avg_rows = []
    n_cols = len(all_headers) - 1  # 학교코드 제외
    for school_code, data in tqdm(sorted(school_data.items()), desc=f"[{region}] 평균 계산", unit="학교"):
        rows = data["rows"]
        numeric_cols = data["numeric_cols"]
        equip_count = len(rows)
        # TOTAL B열=소스1열(학교코드), C열=소스2열(학교명)
        school_code_val = rows[0][0] if rows and len(rows[0]) > 0 else school_code
        school_name_val = rows[0][1] if rows and len(rows[0]) > 1 else ""

        if not rows:
            avg_rows.append([school_code_val, school_name_val, 0] + [""] * n_cols + [""])
            continue

        avg_line = []
        for c in range(1, n_cols + 1):
            if c in numeric_cols:
                vals = []
                for r in rows:
                    if c <= len(r):
                        v = r[c - 1]
                        if is_numeric(v):
                            try:
                                x = float(str(v).replace(",", "").replace("%", ""))
                                vals.append(x)
                            except ValueError:
                                pass
                avg_line.append(round(sum(vals) / len(vals), 4) if vals else "")
            else:
                avg_line.append(rows[0][c - 1] if c <= len(rows[0]) else "")

        # 진단결과: Avg Throughput >= 700 → 양호, 이하면 미흡
        diagnosis = ""
        tp_idx = (throughput_col_idx - 1) if throughput_col_idx and throughput_col_idx > 0 else None
        if tp_idx is not None and tp_idx < len(avg_line):
            try:
                tp_val = avg_line[tp_idx]
                if is_numeric(tp_val):
                    tp = float(str(tp_val).replace(",", ""))
                    diagnosis = "양호" if tp >= THROUGHPUT_THRESHOLD_MBPS else "미흡"
            except (ValueError, TypeError):
                pass

        avg_rows.append([school_code_val, school_name_val, equip_count] + avg_line + [diagnosis])

    # 엑셀 저장
    wb = Workbook()
    ws_total = wb.active
    ws_total.title = sheet_names["total"]

    # 열별 표시 형식 (헤더 → 단위 서식)
    col_formats = [get_column_format(h) for h in all_headers]

    for c, h in enumerate(all_headers, 1):
        ws_total.cell(row=1, column=c, value=h)
    for ri, row in tqdm(enumerate(all_rows, 2), total=len(all_rows), desc=f"[{region}] TOTAL 시트 쓰기", unit="행"):
        for ci, val in enumerate(row, 1):
            cell = ws_total.cell(row=ri, column=ci, value=val)
            if ci <= len(col_formats) and col_formats[ci - 1] and is_numeric(val):
                cell.number_format = col_formats[ci - 1]

    # AVG 시트 열: 학교코드(D), 학교명(E), 장비개수, K, L, M, N, 진단결과(K열 700Mbps 기준)
    keep_cols = AVG_COLUMN_SELECT.get("keep_cols", [11, 12, 13, 14])  # K,L,M,N

    def _filter_avg_row(row):
        """row: [school_code, school_name, equip_count, ...avg_line..., diagnosis]"""
        selected = [row[c - 1] for c in keep_cols if c <= len(row)]
        diagnosis = row[-1] if row else ""
        return [row[0], row[1], row[2]] + selected + [diagnosis]

    full_avg_headers = ["학교코드", "학교명", "장비개수"] + list(all_headers[1:]) + ["진단결과"]

    def _filter_avg_headers():
        selected_h = [full_avg_headers[c - 1] for c in keep_cols if c <= len(full_avg_headers)]
        return ["학교코드", "학교명", "장비개수"] + selected_h + ["진단결과"]

    avg_headers = _filter_avg_headers()
    def _fmt(col_1based):
        if col_1based <= 3:
            return None
        idx = col_1based - 3
        return col_formats[idx] if idx < len(col_formats) else None
    avg_col_formats = [None, None, None] + [_fmt(c) for c in keep_cols] + [None]

    filtered_avg_rows = [_filter_avg_row(r) for r in avg_rows]

    ws_avg = wb.create_sheet(sheet_names["avg"])
    for c, h in enumerate(avg_headers, 1):
        ws_avg.cell(row=1, column=c, value=h)
    for ri, row in tqdm(enumerate(filtered_avg_rows, 2), total=len(filtered_avg_rows), desc=f"[{region}] AVG 시트 쓰기", unit="행"):
        for ci, val in enumerate(row, 1):
            cell = ws_avg.cell(row=ri, column=ci, value=val)
            if ci <= len(avg_col_formats) and avg_col_formats[ci - 1] and is_numeric(val):
                cell.number_format = avg_col_formats[ci - 1]

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    sheet_total = sheet_names["total"]
    sheet_avg = sheet_names["avg"]
    _log(f"[완료] {output_path}")
    _log(f"       {sheet_total}: {len(all_rows)}행")
    _log(f"       {sheet_avg}: {len(avg_rows)}개 학교")


def main():
    _log("=" * 50)
    _log("[학교별 측정 값 현황] 유선망 전처리 시작")
    _log("=" * 50)
    import argparse
    parser = argparse.ArgumentParser(description="유선망 1차 측정 결과 전처리")
    parser.add_argument("--region", "-r", choices=["CNE", "DNI", "ALL"], default=RUN_REGION,
                        help=f"지역 (CNE=충남, DNI=대전, ALL=둘 다, 기본값: {RUN_REGION})")
    args = parser.parse_args()
    region = args.region
    _log(f"실행 지역: {region}")
    try:
        if region == "ALL":
            for r in ["CNE", "DNI"]:
                integrate_and_average(r)
        else:
            integrate_and_average(region)
    except Exception as e:
        _log(f"[오류] {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    _log("=" * 50)
    _log("종료")


if __name__ == "__main__":
    main()
