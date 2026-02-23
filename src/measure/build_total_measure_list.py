# -*- coding: utf-8 -*-
"""
통계 데이터 통합 → TOTAL_MEASURE_LIST_V1.XLSX

전부하측정_학교별평균, ISP측정_학교별평균, 집선ISP, CNE_WIRED_MEANSURE_AVG 시트 통합
"""
from __future__ import print_function
import sys
import io
import os

if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

_MEASURE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, _MEASURE_DIR)

from school_report_config import TOTAL_MEASURE_LIST, CNE_DIR

# 소스 파일 및 시트
SOURCES = [
    (os.path.join(CNE_DIR, "CNE_FULLLOAD_MEASURE.xlsx"), "전부하측정_학교별평균"),
    (os.path.join(CNE_DIR, "CNE_ISP_MEASURE.XLSX"), "ISP측정_학교별평균"),
    (os.path.join(CNE_DIR, "집선ISP.xlsx"), "집선ISP"),
    (os.path.join(CNE_DIR, "집선ISP_측정.xlsx"), "집선ISP"),
    (os.path.join(CNE_DIR, "CNE_WIRED_MEANSURE_V1.XLSX"), "CNE_WIRED_MEANSURE_AVG"),
]


def copy_sheet(ws_src, wb_dst, sheet_name):
    """시트 복사 (데이터만, 서식 유지 위해 값 복사)"""
    ws_dst = wb_dst.create_sheet(sheet_name)
    for row in ws_src.iter_rows():
        for cell in row:
            dst_cell = ws_dst.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                dst_cell.font = cell.font.copy()
                dst_cell.border = cell.border.copy()
                dst_cell.fill = cell.fill.copy()
                dst_cell.number_format = cell.number_format
                dst_cell.alignment = cell.alignment.copy()
    return ws_dst


def main():
    print("=" * 50)
    print("[통합] TOTAL_MEASURE_LIST_V1.XLSX 생성")
    print("=" * 50)
    wb = Workbook()
    wb.remove(wb.active)
    copied = set()
    for path, sheet_name in SOURCES:
        if sheet_name in copied:
            continue
        if not os.path.isfile(path):
            continue
        try:
            wb_src = load_workbook(path, data_only=True)
            if sheet_name not in wb_src.sheetnames:
                wb_src.close()
                continue
            ws_src = wb_src[sheet_name]
            ws_dst = wb.create_sheet(sheet_name)
            for r in range(1, ws_src.max_row + 1):
                for c in range(1, ws_src.max_column + 1):
                    v = ws_src.cell(r, c).value
                    ws_dst.cell(r, c, value=v)
            wb_src.close()
            print(f"  [복사] {sheet_name} <- {os.path.basename(path)}")
            copied.add(sheet_name)
        except Exception as e:
            print(f"  [오류] {path}: {e}")
    if len(copied) == 0:
        print("[오류] 복사된 시트 없음")
        sys.exit(1)
    os.makedirs(os.path.dirname(TOTAL_MEASURE_LIST), exist_ok=True)
    wb.save(TOTAL_MEASURE_LIST)
    print(f"[완료] {TOTAL_MEASURE_LIST}")


if __name__ == "__main__":
    main()
