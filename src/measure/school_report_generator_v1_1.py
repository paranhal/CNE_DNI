# -*- coding: utf-8 -*-
"""
학교별 측정 리포트 생성 (템플릿 사용) V1.1

측정값_템플릿.xlsx 복사 → J열 측정값, L열 판정 입력 → 학교명_학교코드.xlsx 저장
대상 719개 학교 중 통계 데이터 없는 학교는 로그에 기록

[함수 구조]
- 섹션 1: 시트/셀 유틸 (시트명 보정, 행 선택, 헤더 열 탐색)
- 섹션 2: 파일/경로 (파일명 정리, 템플릿/통계/출력 경로 해석)
- 섹션 3: 사용자 입력 (선택 프롬프트)
- 섹션 4: 학교 목록 로딩 (전체 목록, 통계 내 메타, 학교별 행 인덱스)
- 섹션 5: 측정값 추출·포맷 (시트→값, 숫자/문자 포맷, 출력 문자열)
- 섹션 6: 판정 (기준값 연산 → 정상/개선필요)
- 섹션 7: 리포트 생성 (템플릿 채우기: J열, L열, 요약 셀)
- 섹션 8: 오케스트레이션 (경로 설정, 데이터 로드, 로그, 일괄 생성)

TODO(2차): load_full_school_list 내 CSV/Excel 헤더 파싱 공통화(_parse_school_table).
TODO(2차): format_output_value 를 fmt별 _format_output_* 로 분리 검토.
"""
from __future__ import print_function
import sys
import io
import os
import csv
import re
from datetime import datetime
from collections import Counter

if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

from openpyxl import load_workbook
from openpyxl.styles import Font
from tqdm import tqdm

if getattr(sys, "frozen", False):
    _MEASURE_DIR = os.path.dirname(sys.executable)
else:
    _MEASURE_DIR = os.path.dirname(os.path.abspath(__file__))
_RUN_DIR = os.getcwd()
sys.path.insert(0, _MEASURE_DIR)

from school_report_config_v1_1 import (
    TEMPLATE_CANDIDATES,
    TOTAL_MEASURE_LIST,
    OUTPUT_DIR,
    J_OUTPUT_MAP,
    L_JUDGMENT_MAP,
    J_COL,
    L_COL,
    G_COL,
    FONT_BLACK_ROWS,
    JUDGE_BY_V2_ROWS,
    JUDGE_BOTH_ROWS,
    JUDGMENT_ROW_START,
    JUDGMENT_ROW_END,
    LOG_DIR,
    LOG_PREFIX,
)

SCHOOL_LIST_SEARCH_DIRS = [_RUN_DIR, _MEASURE_DIR, os.path.join(os.path.dirname(_MEASURE_DIR), "split")]
TEMPLATE_SHEET = "문제점분석"


# ---------------------------------------------------------------------------
# 섹션 1: 시트/셀 유틸 (시트명 보정, 행 선택, 헤더 열 탐색)
# ---------------------------------------------------------------------------

def _norm_sheet_name(name):
    """시트명 비교용 정규화 (공백/구분자 제거, 소문자)."""
    s = str(name or "").strip().lower()
    return re.sub(r"[^0-9a-z가-힣]", "", s)


def _resolve_sheet_name(wb_stats, target_name):
    """설정 시트명을 실제 통계 파일 시트명으로 보정."""
    if target_name in wb_stats.sheetnames:
        return target_name

    target_norm = _norm_sheet_name(target_name)
    if not target_norm:
        return None

    norm_to_real = {_norm_sheet_name(s): s for s in wb_stats.sheetnames}
    if target_norm in norm_to_real:
        return norm_to_real[target_norm]

    alias_map = {
        _norm_sheet_name("CNE_WIRED_MEANSURE_AVG"): [
            "DNI_WIRED_MEANSURE_AVG",
            "WIRED_MEANSURE_AVG",
        ],
        _norm_sheet_name("AP_장비통계"): [
            "무선AP",
            "AP장비통계",
            "충남AP",
        ],
        _norm_sheet_name("충남AP"): [
            "무선AP",
            "AP장비통계",
            "AP_장비통계",
        ],
        _norm_sheet_name("집선ISP"): [
            "유선ISP",
            "백본ISP",
        ],
    }
    for alias in alias_map.get(target_norm, []):
        alias_norm = _norm_sheet_name(alias)
        if alias_norm in norm_to_real:
            return norm_to_real[alias_norm]

    if target_norm == _norm_sheet_name("CNE_WIRED_MEANSURE_AVG"):
        for s in wb_stats.sheetnames:
            n = _norm_sheet_name(s)
            if "wired" in n and "meansure" in n and "avg" in n:
                return s
    if target_norm in (_norm_sheet_name("AP_장비통계"), _norm_sheet_name("충남AP")):
        ap_candidates = [s for s in wb_stats.sheetnames
                         if "ap" in _norm_sheet_name(s) and "isp" not in _norm_sheet_name(s)]
        if ap_candidates:
            return ap_candidates[0]

    return None


def _pick_best_row(ws, rows, row_def):
    """동일 학교 다중행 중 매핑 열 기준으로 가장 값이 많은 행 선택."""
    if not rows:
        return None
    if len(rows) == 1:
        return rows[0]

    col1, col2 = row_def[2], row_def[3]
    cols = []
    if isinstance(col1, list):
        cols.extend(c for c in col1 if isinstance(c, int))
    elif isinstance(col1, int):
        cols.append(col1)
    if isinstance(col2, int):
        cols.append(col2)
    cols = sorted(set(cols))
    if not cols:
        return rows[0]

    best_row = rows[0]
    best_score = -1
    for r in rows:
        score = sum(
            1
            for c in cols
            if ws.cell(r, c).value is not None and str(ws.cell(r, c).value).strip() != ""
        )
        if score > best_score:
            best_row = r
            best_score = score
    return best_row


def _find_header_col(ws, include_keywords, exclude_keywords=None):
    """헤더 키워드로 열 번호 탐색."""
    exclude_keywords = exclude_keywords or []
    for c in range(1, min(80, ws.max_column + 1)):
        h = str(ws.cell(1, c).value or "").strip().lower()
        if not h:
            continue
        if any(k in h for k in include_keywords) and not any(k in h for k in exclude_keywords):
            return c
    return None


def _resolve_isp_cols_by_header(ws):
    """ISP 학교별평균 시트의 DL/UL/RSSI 열을 헤더 기반으로 보정."""
    dl_col = _find_header_col(ws, ["다운로드", "dl", "download"], ["진단", "평가"])
    ul_col = _find_header_col(ws, ["업로드", "up", "upload"], ["진단", "평가"])
    rssi_col = _find_header_col(ws, ["rssi", "*rssi"], ["진단", "평가"])
    return {
        11: rssi_col,
        23: dl_col,
        24: ul_col,
        25: rssi_col,
    }


# ---------------------------------------------------------------------------
# 섹션 2: 파일/경로 (파일명 정리, 템플릿·통계·출력 경로)
# ---------------------------------------------------------------------------

def sanitize_filename(s):
    """파일명에 사용 불가 문자 제거."""
    if not s:
        return ""
    s = str(s).strip()
    for c in r'\/:*?"<>|':
        s = s.replace(c, "_")
    return s[:50]


def find_template():
    """측정값 템플릿 xlsx 경로 탐색 (실행 디렉터리 → 설정 후보)."""
    run_candidates = [
        os.path.join(_RUN_DIR, "최종_측정값_템플릿.xlsx"),
        os.path.join(_RUN_DIR, "측정값_템플릿.xlsx"),
    ]
    for p in run_candidates:
        if os.path.isfile(p):
            return p
    for template_dir in [os.path.join(_RUN_DIR, "측정값_템플릿"), os.path.join(_RUN_DIR, "측정밗_템플릿")]:
        if os.path.isdir(template_dir):
            for f in os.listdir(template_dir):
                if f.endswith((".xlsx", ".xls")):
                    return os.path.join(template_dir, f)
    for p in TEMPLATE_CANDIDATES:
        if os.path.isfile(p):
            return p
    for template_dir in [os.path.join(_MEASURE_DIR, "측정값_템플릿"), os.path.join(_MEASURE_DIR, "측정밗_템플릿")]:
        if os.path.isdir(template_dir):
            for f in os.listdir(template_dir):
                if f.endswith((".xlsx", ".xls")):
                    return os.path.join(template_dir, f)
    return None


def _get_total_measure_candidate_list():
    """통합 통계 파일 후보 경로 리스트 (단일 정의)."""
    return [
        os.path.join(_RUN_DIR, "DNI_TOTAL_MEASURE_LIST_V1.xlsx"),
        os.path.join(_RUN_DIR, "TOTAL_MEASURE_LIST_V1.xlsx"),
        os.path.join(_RUN_DIR, "CNE", "TOTAL_MEASURE_LIST_V1.xlsx"),
        os.path.join(_MEASURE_DIR, "CNE", "TOTAL_MEASURE_LIST_V1.xlsx"),
        TOTAL_MEASURE_LIST,
    ]


def resolve_total_measure_path():
    """통합 통계 파일 기본 경로 (존재하면 해당 경로, 없으면 첫 후보 반환)."""
    candidates = _get_total_measure_candidate_list()
    for p in candidates:
        if p and os.path.isfile(p):
            return p
    return candidates[0]


def resolve_output_dir():
    """출력 디렉터리: 현재 실행 폴더 하위 학교별_리포트."""
    return os.path.join(_RUN_DIR, "학교별_리포트")


def resolve_log_dir():
    """로그 디렉터리: 현재 실행 폴더 하위 logs."""
    return os.path.join(_RUN_DIR, "logs")


# ---------------------------------------------------------------------------
# 섹션 3: 사용자 입력 (선택 프롬프트)
# ---------------------------------------------------------------------------

def _input(prompt):
    """안전한 input; EOF/키보드 중단 시 종료."""
    try:
        return input(prompt).strip()
    except (EOFError, KeyboardInterrupt):
        print("\n[중단] 사용자 입력으로 종료합니다.")
        sys.exit(1)


def _discover_total_measure_candidates():
    """실행 디렉터리 및 고정 후보에서 통계 파일 목록 수집."""
    candidates = list(_get_total_measure_candidate_list())
    try:
        for f in os.listdir(_RUN_DIR):
            if not f.lower().endswith(".xlsx"):
                continue
            if "total_measure_list_v1" in f.lower():
                candidates.append(os.path.join(_RUN_DIR, f))
    except Exception:
        pass

    unique = []
    seen = set()
    for p in candidates:
        if not p:
            continue
        np = os.path.normcase(os.path.abspath(p))
        if np in seen:
            continue
        seen.add(np)
        if os.path.isfile(p):
            unique.append(p)
    return unique


def select_total_measure_path():
    """통합 통계 파일 경로를 사용자 선택 또는 입력으로 결정."""
    found = _discover_total_measure_candidates()
    if not found:
        default_path = resolve_total_measure_path()
        print("[안내] 자동으로 찾은 통계 파일이 없습니다.")
        s = _input(
            f"통계 파일 경로를 입력하세요 (Enter: 기본값 사용)\n기본값: {default_path}\n> "
        )
        return s if s else default_path

    print("\n통계 파일을 선택하세요.")
    for i, p in enumerate(found, start=1):
        print(f"  {i}. {p}")
    print("  0. 직접 경로 입력")
    s = _input("번호 선택 (Enter: 1번): ")
    if not s:
        return found[0]
    if s == "0":
        direct = _input("통계 파일 전체 경로 입력: ")
        return direct
    try:
        idx = int(s)
        if 1 <= idx <= len(found):
            return found[idx - 1]
    except ValueError:
        pass
    print("[경고] 잘못된 입력입니다. 1번 파일을 사용합니다.")
    return found[0]


def select_output_dir():
    """출력 폴더 사용자 지정 (기본: 학교별_리포트)."""
    default_name = "학교별_리포트"
    default_path = os.path.join(_RUN_DIR, default_name)
    s = _input(f"\n출력 폴더명을 입력하세요 (Enter: {default_name})\n> ")
    if not s:
        return default_path
    s = s.strip().strip("\"'")
    if os.path.isabs(s):
        return s
    return os.path.join(_RUN_DIR, s)


def select_output_layout():
    """출력 구조: 지역 폴더(region) 또는 단일 폴더(flat)."""
    print("\n출력 폴더 구조를 선택하세요.")
    print("  1. 지역명(개수) 하위 폴더로 저장")
    print("  2. 한 폴더에 모두 저장")
    s = _input("번호 선택 (Enter: 1번): ")
    if not s:
        return "region"
    if s == "2":
        return "flat"
    return "region"


# ---------------------------------------------------------------------------
# 섹션 4: 학교 목록 로딩 (전체 목록, 통계 메타, 학교별 행 인덱스)
# ---------------------------------------------------------------------------

def load_full_school_list():
    """719개 대상 학교 리스트 로드 (school_reg_list_CNE 등 검색)."""
    for base_dir in SCHOOL_LIST_SEARCH_DIRS:
        for fname in ["school_reg_list_CNE.xlsx", "school_reg_list_CNE.csv", "SCHOOL_REG_LIST_CNE.xlsx"]:
            path = os.path.join(base_dir, fname)
            if not os.path.isfile(path):
                continue
            try:
                codes = []
                code_to_name = {}
                code_to_region = {}
                if path.endswith(".csv"):
                    with open(path, "r", encoding="utf-8-sig") as f:
                        reader = csv.reader(f)
                        header = next(reader, None)
                        rows = list(reader)
                    code_col = name_col = region_col = 0
                    for i, h in enumerate(header or []):
                        s = str(h or "").lower()
                        if "학교코드" in s or "code" in s:
                            code_col = i
                        if "학교명" in s or "name" in s:
                            name_col = i
                        if "지역" in s or "region" in s:
                            region_col = i
                    for row in rows:
                        if len(row) > max(code_col, name_col, region_col):
                            code = str(row[code_col] or "").strip()
                            name = str(row[name_col] or "").strip()
                            region = str(row[region_col] or "").strip()
                            if code:
                                codes.append(code)
                                code_to_name[code] = name
                                code_to_region[code] = region
                else:
                    wb = load_workbook(path, read_only=True, data_only=True)
                    ws = wb.active
                    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
                    code_col = name_col = region_col = 0
                    for i, h in enumerate(header or []):
                        s = str(h or "").lower()
                        if "학교코드" in s or "code" in s:
                            code_col = i
                        if "학교명" in s or "name" in s:
                            name_col = i
                        if "지역" in s or "region" in s:
                            region_col = i
                    for r in range(2, ws.max_row + 1):
                        code = str(ws.cell(r, code_col + 1).value or "").strip()
                        name = str(ws.cell(r, name_col + 1).value or "").strip()
                        region = str(ws.cell(r, region_col + 1).value or "").strip()
                        if code:
                            codes.append(code)
                            code_to_name[code] = name
                            code_to_region[code] = region
                    wb.close()
                if codes:
                    return codes, code_to_name, code_to_region
            except Exception:
                pass
    return [], {}, {}


def load_school_meta_from_sheet1(wb_stats):
    """통계 파일의 Sheet1에서 학교코드/지역/학교명 메타 로드."""
    target_sheet = None
    for n in wb_stats.sheetnames:
        if str(n).strip().lower() == "sheet1":
            target_sheet = n
            break
    if not target_sheet:
        return [], {}, {}

    ws = wb_stats[target_sheet]
    headers = [str(ws.cell(1, c).value or "").strip().lower() for c in range(1, ws.max_column + 1)]

    code_col = name_col = region_col = None
    for i, h in enumerate(headers, 1):
        if code_col is None and ("학교코드" in h or "code" == h or h.endswith("code")):
            code_col = i
        if name_col is None and ("학교명" in h or "name" == h or h.endswith("name")):
            name_col = i
        if region_col is None and ("지역" in h or "region" == h or h.endswith("region")):
            region_col = i

    if not code_col or not name_col:
        return [], {}, {}

    codes = []
    code_to_name = {}
    code_to_region = {}
    for r in range(2, ws.max_row + 1):
        code = str(ws.cell(r, code_col).value or "").strip()
        if not code:
            continue
        name = str(ws.cell(r, name_col).value or "").strip()
        region = str(ws.cell(r, region_col).value or "").strip() if region_col else ""
        codes.append(code)
        code_to_name[code] = name
        code_to_region[code] = region
    return codes, code_to_name, code_to_region


def load_stats_by_school(wb_stats):
    """통계 워크북에서 학교코드별 (시트명 → 데이터 행 리스트) 로드."""
    by_school = {}
    for real_sheet_name in wb_stats.sheetnames:
        ws = wb_stats[real_sheet_name]
        code_col = 1
        for c in range(1, min(50, ws.max_column + 1)):
            v = ws.cell(1, c).value
            if v and ("학교코드" in str(v) or "code" in str(v).lower()):
                code_col = c
                break
        for r in range(2, ws.max_row + 1):
            sc = str(ws.cell(r, code_col).value or "").strip()
            if not sc:
                continue
            if sc not in by_school:
                by_school[sc] = {}
            by_school[sc].setdefault(real_sheet_name, []).append(r)
    return by_school


def get_school_name_from_stats(wb_stats, school_data):
    """학교 리스트에 이름이 없을 때 통계 시트에서 학교명 보강."""
    for sheet_name, rows in school_data.items():
        resolved = _resolve_sheet_name(wb_stats, sheet_name)
        if not resolved:
            continue
        ws = wb_stats[resolved]
        row_list = rows if isinstance(rows, list) else [rows]
        row = row_list[0] if row_list else None
        if row is None:
            continue
        name_col = None
        for c in range(1, min(50, ws.max_column + 1)):
            v = ws.cell(1, c).value
            s = str(v or "").strip().lower()
            if "학교명" in s or s == "name":
                name_col = c
                break
        if name_col:
            n = str(ws.cell(row, name_col).value or "").strip()
            if n:
                return n
    return ""


# ---------------------------------------------------------------------------
# 섹션 5: 측정값 추출·포맷 (시트→값, 숫자/문자 포맷, 출력 문자열)
# ---------------------------------------------------------------------------

def get_school_values(wb_stats, school_code, school_data, row_def):
    """한 학교의 특정 출력행에 대한 원시값 조회.
    row_def: (row, sheet, col1, col2, format_type)
    format_type: None=단순, "cable"=케이블통계, "fullload"=전부하 Mbps, "fixed"=고정값
    반환: (v1, v2) 또는 (None, None)
    """
    sheet_name = row_def[1]
    if sheet_name == "fixed":
        return (row_def[2], None)
    if sheet_name == "h_only":
        return (None, None)
    resolved_sheet = _resolve_sheet_name(wb_stats, sheet_name)
    if not resolved_sheet or resolved_sheet not in school_data:
        return None, None

    ws = wb_stats[resolved_sheet]
    rows = school_data[resolved_sheet]
    row_list = rows if isinstance(rows, list) else [rows]
    data_row = _pick_best_row(ws, row_list, row_def)
    if data_row is None:
        return None, None

    sheet_norm = _norm_sheet_name(row_def[1])
    if "isp" in sheet_norm and "학교별평균" in row_def[1]:
        dynamic_cols = _resolve_isp_cols_by_header(ws)
        report_row = row_def[0]
        if report_row in dynamic_cols and dynamic_cols[report_row]:
            v = ws.cell(data_row, dynamic_cols[report_row]).value
            return v, None

    col1, col2 = row_def[2], row_def[3]
    if isinstance(col1, list):
        vals = [ws.cell(data_row, c).value for c in col1]
        return vals, None
    v1 = ws.cell(data_row, col1).value if col1 else None
    v2 = ws.cell(data_row, col2).value if col2 else None
    return v1, v2


def get_numeric(val):
    """값을 숫자로 변환 (없으면 None)."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "").replace("%", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def format_value(v):
    """단일 값을 출력용 문자열/숫자로 정리."""
    if v is None or (isinstance(v, str) and not str(v).strip()):
        return ""
    if isinstance(v, (int, float)):
        return v
    return str(v).strip()


def format_output_value(v1, v2, row_def):
    """row_def 포맷 타입에 따라 J열 출력 문자열 생성."""
    fmt = row_def[4] if len(row_def) >= 5 else None
    if fmt == "h_only":
        return ""
    if fmt == "fixed":
        return str(row_def[2]) if row_def[2] is not None else ""
    if fmt == "location5":
        g_val = format_value(v1) if v1 is not None else ""
        h_val = format_value(v2) if v2 is not None else ""
        parts = [
            f"위치 5 - {g_val}" if g_val else "위치 5 - ",
            f"위치 5 외 - {h_val}" if h_val else "위치 5 외 - ",
            "보관 및 확인불가 - 없음",
        ]
        return "\n".join(parts)
    if fmt == "n_ac_ax" and isinstance(v1, (list, tuple)):
        d_val = format_value(v1[0]) if len(v1) > 0 else ""
        e_val = format_value(v1[1]) if len(v1) > 1 else ""
        f_val = format_value(v1[2]) if len(v1) > 2 else ""
        parts = [
            f"N - {d_val}식",
            f"AC - {e_val}식",
            f"AX - {f_val}식",
        ]
        return "\n".join(parts)
    if fmt == "cable" and isinstance(v1, (list, tuple)):
        labels = ("SM", "MM", "CAT6", "CAT5e", "CAT5")
        parts = []
        for i, lbl in enumerate(labels):
            val = format_value(v1[i]) if i < len(v1) and v1[i] is not None else ""
            parts.append(f"{lbl}({val})" if val else f"{lbl}()")
        return "\n".join(parts)
    if fmt == "fullload":
        c_val = format_value(v1) if v1 is not None else ""
        d_val = format_value(v2) if v2 is not None else ""
        parts = []
        if c_val:
            parts.append(f"{c_val}(Mbps)")
        if d_val:
            parts.append(f"{d_val}(Mbps)")
        return "\n".join(parts) if parts else ""
    if row_def[3] is not None and not isinstance(row_def[2], list):
        return f"{format_value(v1)} / {format_value(v2)}" if (v1 or v2) else (format_value(v1) or format_value(v2))
    return format_value(v1)


# ---------------------------------------------------------------------------
# 섹션 6: 판정 (기준값 연산 → 정상/개선필요)
# ---------------------------------------------------------------------------

def judge(val, op, threshold, val2=None):
    """연산(op)과 기준값(threshold)으로 정상/개선필요 판정. 미해당 시 빈 문자열."""
    if op == "always":
        return str(threshold) if threshold else "정상"
    if op == "has_value":
        s = str(val or "").strip()
        return "개선필요" if s else "정상"
    if op == "zero_or_empty_ok":
        s = str(val or "").strip()
        if not s:
            return "정상"
        n = get_numeric(val)
        return "정상" if n is not None and n <= threshold else "개선필요"
    if op == "split_exact":
        s = str(val or "").strip()
        if s == "분리":
            return "정상"
        if s == "미분리":
            return "개선필요"
        return "개선필요"
    if op == "both_ge":
        n1 = get_numeric(val)
        n2 = get_numeric(val2) if val2 is not None else None
        if n1 is not None and n1 <= threshold:
            return "개선필요"
        if n2 is not None and n2 <= threshold:
            return "개선필요"
        return "정상"
    if op == "split":
        s = str(val or "").strip()
        return "정상" if "분리" in s else "개선필요"
    if op == "ge_before_keyword":
        s = str(val or "").strip()
        m = re.search(r"(\d+)\s*계위", s)
        if not m:
            return ""
        n = get_numeric(m.group(1))
        if n is None:
            return ""
        return "개선필요" if n >= threshold else "정상"
    n = get_numeric(val)
    if n is None:
        return ""
    if op == "ge":
        return "정상" if n >= threshold else "개선필요"
    if op == "le":
        return "정상" if n <= threshold else "개선필요"
    return ""


# ---------------------------------------------------------------------------
# 섹션 7: 리포트 생성 (J열 측정값, L열 판정, 요약 셀)
# ---------------------------------------------------------------------------

def _compute_row_output(wb_stats, school_data, row_def):
    """한 출력행에 대한 J열 문자열과 판정용 원시값 (v1, v2) 계산.
    반환: (j_str, v1, v2) — J열에 쓸 문자열과 L열 판정에 쓸 값."""
    v1, v2 = get_school_values(wb_stats, None, school_data, row_def)
    j_str = format_output_value(v1, v2, row_def)
    return j_str, v1, v2


def _compute_judgment_for_row(row, v1, v2, l_map):
    """L열 판정 문자열 계산. l_map: {row -> (op, threshold)}."""
    if row not in l_map:
        return None
    op, threshold = l_map[row]
    val_for_judge = v2 if row in JUDGE_BY_V2_ROWS else v1
    if isinstance(val_for_judge, (list, tuple)):
        val_for_judge = val_for_judge[0] if val_for_judge else None
    val2_for_judge = v2 if row in JUDGE_BOTH_ROWS and not isinstance(v1, (list, tuple)) else None
    result = judge(val_for_judge, op, threshold, val2_for_judge)
    if not result and v1 is not None and v2 is not None and not isinstance(v1, (list, tuple)):
        result = judge(v2, op, threshold)
    return result


def _fill_measurement_and_judgment(ws, wb_stats, school_data):
    """J열 측정값과 L열 판정을 채운다 (계산은 _compute_* 에 위임)."""
    l_map = {row: (op, th) for row, op, th in L_JUDGMENT_MAP}
    for row_def in J_OUTPUT_MAP:
        row = row_def[0]
        j_str, v1, v2 = _compute_row_output(wb_stats, school_data, row_def)
        fmt = row_def[4] if len(row_def) >= 5 else None
        if fmt != "h_only":
            cell = ws.cell(row=row, column=J_COL)
            cell.value = j_str
            if row in FONT_BLACK_ROWS:
                cell.font = Font(color="000000")
        result = _compute_judgment_for_row(row, v1, v2, l_map)
        if result is not None:
            l_cell = ws.cell(row=row, column=L_COL, value=result)
            if result == "개선필요":
                l_cell.font = Font(color="FF0000")


def _fill_summary_cells(ws):
    """G21 고정값, G36/L36·G37/L37 정상·개선필요 개수 채우기."""
    ws.cell(row=21, column=G_COL, value="375 Mhz 이상")
    count_정상 = 0
    count_개선필요 = 0
    for r in range(JUDGMENT_ROW_START, JUDGMENT_ROW_END + 1):
        val = ws.cell(row=r, column=L_COL).value
        s = str(val or "").strip()
        if s == "정상":
            count_정상 += 1
        elif s == "개선필요":
            count_개선필요 += 1
    ws.cell(row=36, column=G_COL, value="정상")
    ws.cell(row=36, column=L_COL, value=count_정상)
    ws.cell(row=37, column=G_COL, value="개선필요")
    ws.cell(row=37, column=L_COL, value=count_개선필요)


def generate_school_report(template_path, wb_stats, school_code, school_data):
    """템플릿을 열어 J열(측정값), L열(판정), 요약 셀을 채운 워크북 반환."""
    wb = load_workbook(template_path)
    ws = wb[TEMPLATE_SHEET] if TEMPLATE_SHEET in wb.sheetnames else wb.active
    _fill_measurement_and_judgment(ws, wb_stats, school_data)
    _fill_summary_cells(ws)
    return wb


# ---------------------------------------------------------------------------
# 섹션 8: 오케스트레이션 (경로 설정, 데이터 로드, 로그, 일괄 생성)
# ---------------------------------------------------------------------------

def _ensure_template_path():
    """템플릿 경로 확인, 없으면 메시지 출력 후 sys.exit(1). 반환: template_path."""
    template_path = find_template()
    if not template_path:
        check_paths = [
            os.path.join(_RUN_DIR, "최종_측정값_템플릿.xlsx"),
            os.path.join(_RUN_DIR, "측정값_템플릿.xlsx"),
        ] + TEMPLATE_CANDIDATES
        print(f"[오류] 템플릿 없음. 확인: {check_paths}")
        sys.exit(1)
    return template_path


def _normalize_region(region):
    """지역 문자열 정규화: 빈 값 → '미분류'."""
    if not region or not str(region).strip():
        return "미분류"
    return str(region).strip()


def _setup_paths():
    """템플릿·통계·출력·로그 경로 및 출력 레이아웃을 사용자와 함께 결정.
    반환: (template_path, total_measure_path, output_dir, log_dir, output_layout)
    """
    template_path = _ensure_template_path()
    total_measure_path = select_total_measure_path()
    output_dir = select_output_dir()
    output_layout = select_output_layout()
    log_dir = resolve_log_dir()
    return template_path, total_measure_path, output_dir, log_dir, output_layout


def _load_workbook_and_school_lists(total_measure_path):
    """통계 워크북 로드 및 학교 목록·메타·학교별 데이터 준비.
    반환: (wb_stats, all_schools, code_to_name, code_to_region, by_school)
    """
    if not os.path.isfile(total_measure_path):
        print(f"[오류] 통계 파일 없음: {total_measure_path}")
        sys.exit(1)
    wb_stats = load_workbook(total_measure_path, data_only=True)
    sheet_codes, sheet_code_to_name, sheet_code_to_region = load_school_meta_from_sheet1(wb_stats)
    if sheet_codes:
        all_schools = sheet_codes
        code_to_name = sheet_code_to_name
        code_to_region = sheet_code_to_region
        print(f"[안내] sheet1 기준 학교 메타 로드: {len(all_schools)}개")
    else:
        all_schools, code_to_name, code_to_region = load_full_school_list()
        if not all_schools:
            print("[경고] 학교 리스트 없음. 통계에 있는 학교만 처리합니다.")

    by_school = load_stats_by_school(wb_stats)
    if not by_school:
        print("[오류] 학교별 데이터 없음")
        sys.exit(1)
    return wb_stats, all_schools, code_to_name, code_to_region, by_school


def _build_missing_schools_log_lines(all_schools, schools_with_data, code_to_name):
    """통계 데이터 없는 학교 로그용 문장 리스트 생성. (로그 출력 지점 정리용)"""
    missing = [sc for sc in all_schools if sc not in schools_with_data]
    lines = [
        f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 시작",
        f"대상 학교: {len(all_schools)}개",
        f"통계 데이터 있음: {len(schools_with_data)}개",
        f"통계 데이터 없음: {len(missing)}개",
    ]
    if missing:
        lines.append("")
        lines.append("[통계 데이터 없는 학교 (리포트 미생성)]")
        for sc in sorted(missing):
            lines.append(f"  {sc}  {code_to_name.get(sc, '')}")
    return lines, missing


def _write_missing_schools_log(log_path, all_schools, schools_with_data, code_to_name):
    """통계 데이터 없는 학교 목록을 로그 파일에 기록."""
    log_lines, missing = _build_missing_schools_log_lines(
        all_schools, schools_with_data, code_to_name
    )
    with open(log_path, "w", encoding="utf-8") as f:
        f.write("\n".join(log_lines))
    return missing


def _generate_and_save_all(
    template_path,
    wb_stats,
    output_dir,
    output_layout,
    by_school,
    code_to_name,
    code_to_region,
):
    """리포트 생성 및 저장. 저장 실패/이름 없는 학교는 제외하고 로그용 목록 반환.
    반환: (generated_count, skipped_codes)
    """
    os.makedirs(output_dir, exist_ok=True)
    save_codes = sorted(by_school.keys())
    region_totals = Counter()
    for sc in save_codes:
        region_totals[_normalize_region(code_to_region.get(sc, ""))] += 1

    skipped_codes = []
    generated_count = 0
    for school_code in tqdm(save_codes, desc="학교별 생성", unit="교"):
        school_name = code_to_name.get(school_code, "")
        if not school_name:
            school_name = get_school_name_from_stats(wb_stats, by_school[school_code])
        safe_name = sanitize_filename(school_name)
        if not safe_name:
            skipped_codes.append(school_code)
            continue
        school_data = by_school[school_code]
        wb = generate_school_report(template_path, wb_stats, school_code, school_data)
        out_name = f"{school_code}_{safe_name}.xlsx"
        region = _normalize_region(code_to_region.get(school_code, ""))
        if output_layout == "region":
            region_dir = os.path.join(output_dir, f"{region}({region_totals[region]})")
            os.makedirs(region_dir, exist_ok=True)
            out_path = os.path.join(region_dir, out_name)
        else:
            out_path = os.path.join(output_dir, out_name)
        try:
            wb.save(out_path)
            generated_count += 1
        except PermissionError:
            skipped_codes.append(school_code)
        finally:
            try:
                wb.close()
            except Exception:
                pass
    return generated_count, skipped_codes


def _append_skipped_log(log_path, skipped_codes):
    """학교명 없음/저장 실패로 제외된 학교 목록을 로그에 추가."""
    if not skipped_codes:
        return
    with open(log_path, "a", encoding="utf-8") as f:
        f.write("\n\n[학교명 없음/저장 실패로 제외]\n")
        for sc in sorted(set(skipped_codes)):
            f.write(f"  {sc}\n")


def main():
    """진입점: 경로 설정 → 데이터 로드 → 로그 초기화 → 학교별 리포트 생성 → 로그 보완."""
    print("=" * 50)
    print("[학교별 측정 리포트 V1.1] 생성 (템플릿 사용)")
    print("=" * 50)

    template_path, total_measure_path, output_dir, log_dir, output_layout = _setup_paths()
    print(f"템플릿: {template_path}")
    print(f"통계: {total_measure_path}")

    wb_stats, all_schools, code_to_name, code_to_region, by_school = _load_workbook_and_school_lists(total_measure_path)
    schools_with_data = set(by_school.keys())

    os.makedirs(log_dir, exist_ok=True)
    log_path = os.path.join(log_dir, f"{LOG_PREFIX}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")
    # 로그 출력 지점 1: 실행 시작 + 통계 없는 학교 목록
    missing = _write_missing_schools_log(log_path, all_schools, schools_with_data, code_to_name)

    print(f"출력: {output_dir}")
    print(f"출력 구조: {'지역 폴더' if output_layout == 'region' else '단일 폴더'}")
    if missing:
        print(f"[로그] 통계 데이터 없는 학교 {len(missing)}개 → {log_path}")

    generated_count, skipped_codes = _generate_and_save_all(
        template_path, wb_stats, output_dir, output_layout,
        by_school, code_to_name, code_to_region,
    )
    wb_stats.close()

    # 로그 출력 지점 2: 학교명 없음/저장 실패로 제외된 학교
    _append_skipped_log(log_path, skipped_codes)
    if skipped_codes:
        print(f"[주의] 학교명/저장문제로 제외된 학교: {len(skipped_codes)}개")
    print(f"[완료] {generated_count}개 학교 리포트 생성")


if __name__ == "__main__":
    main()
