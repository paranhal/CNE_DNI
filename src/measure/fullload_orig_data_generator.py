# -*- coding: utf-8 -*-
"""
전부하(EDU) 1차/2차 원데이터 생성 프로그램

기획: docs/07_전부하_EDU_원데이터_생성_기획.md
- 실행 폴더(및 raw_data)의 파일 목록 → 사용자 파일·시트·열 매핑 선택 (ISP와 동일)
- 학교당 최대 10대, 20초 간격 시작 → 10분 구간 장비당 600개 (평균=1차/2차), RTT 없음
- "시작 전" 구간: 1번 180개 … 9번 20개, 속도 변동·장비 추가 시 ~1Mbps 감소, RTT 0.3·RSSI -0.5 차이
- 시간: 14:30:00 ~ 17:00 이전
- 요구사항 1: 차수별 평균 기준 600개 생성(평균=Down,Up,RSSI), 5개 학교 CSV 기반 분산 추후 반영 가능.
- 요구사항 2: 10번 전 1~9번 데이터, 차수 평균보다 나쁘지 않게(속도>=target), 장비↑속도↓.
- 출력: 측정서버=EDU, Standard=802.11ax, CH→Freq, RTT=빈 값(전 구간).
"""

from __future__ import annotations

import csv
import json
import os
import sys
import random
from collections import defaultdict
from datetime import datetime, time, timedelta
from statistics import mean, stdev
from typing import Any

if getattr(sys, "frozen", False):
    _BASE_DIR = os.path.dirname(os.path.abspath(sys.executable))
    _MEIPASS = getattr(sys, "_MEIPASS", _BASE_DIR)
    _MEASURE_BUNDLE = os.path.join(_MEIPASS, "measure")
    if os.path.isdir(_MEASURE_BUNDLE) and _MEASURE_BUNDLE not in sys.path:
        sys.path.insert(0, _MEASURE_BUNDLE)
    if _MEIPASS not in sys.path:
        sys.path.insert(0, _MEIPASS)
    _DATA_DIRS = [_BASE_DIR]
else:
    _BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    _RUN_DIR = os.getcwd()
    _DATA_DIRS = [
        _RUN_DIR,
        _BASE_DIR,
        os.path.join(_RUN_DIR, "row_data"),
        os.path.join(_RUN_DIR, "raw_data"),
        os.path.join(_BASE_DIR, "row_data"),
        os.path.join(_BASE_DIR, "raw_data"),
        os.path.join(os.path.dirname(_BASE_DIR), "row_data"),
        os.path.join(os.path.dirname(_BASE_DIR), "raw_data"),
    ]
if _BASE_DIR not in sys.path:
    sys.path.insert(0, _BASE_DIR)

# 출력 열 순서 (ISP_샘플과 동일 + W열 평균값 데이터)
OUTPUT_HEADERS = [
    "Date", "StartTime", "학교", "학년", "반", "메모1", "메모2",
    "측정순번", "측정유형", "측정서버", "DL", "UL", "RTT", "LOSS",
    "SSID", "BSSID", "Standard", "CH", "Freq.", "RSSI", "BW", "학교코드",
    "평균값 데이터",  # W열: 2번 구간(평균용 측정) 행에만 "평균값 데이터", 1번 구간은 빈칸
]

START_TIME = time(14, 30, 0)
END_BEFORE = time(17, 0, 0)
SEC_BETWEEN_DEVICES = 20
SEC_MAIN_RECORD = 600
MAX_DEVICES_PER_SCHOOL = 10
LOSS_FIXED = 0
# 요구사항 2: ramp-up은 차수 평균보다 나쁘면 안 됨(속도 >= target). 장비 늘어날수록 속도 감소.
RAMP_UP_DROP_MBPS_LO, RAMP_UP_DROP_MBPS_HI = 0.1, 1.0
RAMP_UP_FLUCTUATE_MBPS = 0.5
RSSI_DIFF_ON_DROP = -0.5
BASE_RAMP_RSSI = -50.0
# 5개 학교 분석 CSV 미사용 시 기본 분산
DL_UL_RATIO_LO, DL_UL_RATIO_HI = 0.85, 1.15
RSSI_MARGIN_DEFAULT = 5  # 평균 ±5
# 측정값 현실 반영: 혼자 측정이 최고, 장비 늘면 감소. 평균에서 너무 동떨어진 값은 최대 5개, 띄엄띄엄.
MAX_DEVIATION_RATIO = 0.20   # 평균 대비 ±20% 초과하면 극단값으로 간주 (상한 자체는 1.2배)
MAX_OUTLIER_COUNT = 5        # 평균과 동떨어진 자료 최대 개수
OUTLIER_SPREAD_STEP = 120    # 이상치 위치 간격(600/5 정도로 띄엄띄엄)

# 분석 결과 저장 (main에서 설정). None이면 기본값 사용.
_variance_params: dict[str, float] | None = None  # dl_cv, ul_cv, rssi_std

VARIANCE_JSON_NAME = "fullload_variance_params.json"


def _load_variance_params_from_json() -> dict[str, float] | None:
    """5개 학교 분산 분석 JSON에서 dl_cv, ul_cv, rssi_std 로드. 실행/실행파일/measure 폴더 순으로 탐색."""
    search_dirs = list(_DATA_DIRS)
    if getattr(sys, "frozen", False):
        _m = getattr(sys, "_MEIPASS", _BASE_DIR)
        measure_bundle = os.path.join(_m, "measure")
        if measure_bundle not in search_dirs:
            search_dirs.insert(0, measure_bundle)
    for d in search_dirs:
        path = os.path.join(d, VARIANCE_JSON_NAME)
        if os.path.isfile(path):
            try:
                with open(path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                if isinstance(data, dict) and ("dl_cv" in data or "ul_cv" in data or "rssi_std" in data):
                    return {
                        "dl_cv": float(data.get("dl_cv", 0.15)),
                        "ul_cv": float(data.get("ul_cv", 0.15)),
                        "rssi_std": float(data.get("rssi_std", 2.0)),
                    }
            except (json.JSONDecodeError, TypeError, ValueError):
                pass
    return None


# CH → Freq. (5GHz): CH 36 → 5180, CH +1 → Freq. +5 (CH 100 → 5500)
def _ch_to_freq(ch: int) -> int:
    if ch >= 36:
        return 5180 + (ch - 36) * 5
    return 2437  # 2.4GHz

STANDARD_WIFI = "802.11ax"  # 전부하 출력용 (802.11ac 아님)

HEADER_ROW = 3
SUBHEADER_ROW = 2


def _analyze_fullload_csv(csv_paths: list[str]) -> dict[str, float] | None:
    """
    5개 학교 전부하 CSV에서 장비(memo1)별 DL, UL, RSSI 평균·표준편차 산출 후
    분산 특성(계수 of variation, RSSI std) 반환. 요구사항 1 분산 로직에 사용.
    """
    if not csv_paths:
        return None
    by_device: dict[str, list[tuple[float, float, float]]] = defaultdict(list)
    for path in csv_paths:
        if not os.path.isfile(path):
            continue
        try:
            with open(path, "r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                if not reader.fieldnames:
                    continue
                for row in reader:
                    try:
                        dl = float(row.get("DL_Mbps", row.get("dl_mbps", "")) or 0)
                        ul = float(row.get("UL_Mbps", row.get("ul_mbps", "")) or 0)
                        rssi_s = row.get("RSSI", row.get("rssi", ""))
                        rssi = float(rssi_s) if rssi_s else -50.0
                    except (ValueError, TypeError):
                        continue
                    memo = (row.get("memo1") or "").strip()
                    if not memo:
                        continue
                    by_device[memo].append((dl, ul, rssi))
        except Exception:
            continue
    # 장비별 CV(std/mean), RSSI std
    dl_cvs, ul_cvs, rssi_stds = [], [], []
    for _memo, vals in by_device.items():
        if len(vals) < 10:
            continue
        dls = [v[0] for v in vals]
        uls = [v[1] for v in vals]
        rssis = [v[2] for v in vals]
        m_dl, m_ul = mean(dls), mean(uls)
        m_rssi = mean(rssis)
        if m_dl >= 0.1 and len(dls) >= 2:
            dl_cvs.append(stdev(dls) / m_dl)
        if m_ul >= 0.1 and len(uls) >= 2:
            ul_cvs.append(stdev(uls) / m_ul)
        if len(rssis) >= 2:
            rssi_stds.append(stdev(rssis))
    if not dl_cvs and not ul_cvs and not rssi_stds:
        return None
    dl_cv = mean(dl_cvs) if dl_cvs else 0.15
    ul_cv = mean(ul_cvs) if ul_cvs else 0.15
    rssi_std = mean(rssi_stds) if rssi_stds else 2.0
    return {"dl_cv": dl_cv, "ul_cv": ul_cv, "rssi_std": rssi_std}


def _log(msg: str) -> None:
    print(msg, flush=True)


# 취소(종료)로 인식할 입력 (대소문자 무시)
_CANCEL_KEYS = ("q", "quit", "취소", "exit", "종료")


def _check_cancel(s: str) -> None:
    """취소 키 입력이면 프로그램 종료."""
    if not s:
        return
    low = s.strip().lower()
    if low in _CANCEL_KEYS:
        _log("\n[취소] 프로그램을 종료합니다.")
        sys.exit(0)


def _input(prompt: str) -> str:
    try:
        out = input(prompt).strip()
        _check_cancel(out)
        return out
    except (EOFError, KeyboardInterrupt):
        _log("\n[중단] 사용자 입력으로 종료합니다.")
        sys.exit(1)


def _collect_files(exts=(".xlsx", ".xlsm")) -> list[str]:
    seen = set()
    out = []
    for d in _DATA_DIRS:
        if not d or not os.path.isdir(d):
            continue
        try:
            for f in os.listdir(d):
                if f.startswith(".") or f.startswith("~$"):
                    continue
                low = f.lower()
                if not any(low.endswith(e) for e in exts):
                    continue
                p = os.path.join(d, f)
                if not os.path.isfile(p):
                    continue
                key = os.path.normcase(os.path.abspath(p))
                if key in seen:
                    continue
                seen.add(key)
                out.append(p)
        except Exception:
            pass
    return sorted(out)


def _pick_file(prompt: str, candidates: list[str]) -> str:
    existing = [p for p in candidates if os.path.isfile(p)]
    if getattr(sys, "frozen", False) and _DATA_DIRS:
        _log(f"\n[데이터 위치] 실행 파일이 있는 폴더: {_DATA_DIRS[0]}")
    _log(f"\n{prompt}")
    if existing:
        for i, p in enumerate(existing, 1):
            _log(f"  {i}. {os.path.basename(p)}")
        _log("  0. 직접 경로 입력")
        _log("  Q. 취소(종료)")
        s = _input("번호 선택 (Enter: 1번): ")
        if not s:
            return existing[0]
        _check_cancel(s)
        if s == "0":
            return _input("전체 경로 입력: ")
        try:
            idx = int(s)
            if 1 <= idx <= len(existing):
                return existing[idx - 1]
        except ValueError:
            pass
        _log("[경고] 잘못된 입력. 1번 사용")
        return existing[0]
    _log("  후보 없음. 직접 경로 입력 (Q: 취소)")
    return _input("전체 경로 입력: ")


def _pick_sheet(wb, prompt: str) -> str:
    names = wb.sheetnames
    _log(f"\n{prompt}")
    for i, n in enumerate(names, 1):
        _log(f"  {i}. {n}")
    _log("  Q. 취소(종료)")
    s = _input("번호 선택 (Enter: 1번): ")
    if not s:
        return names[0]
    _check_cancel(s)
    try:
        idx = int(s)
        if 1 <= idx <= len(names):
            return names[idx - 1]
    except ValueError:
        pass
    return names[0]


def _col_letter_to_index(col: str) -> int | None:
    try:
        from openpyxl.utils import column_index_from_string
        return column_index_from_string(col.strip().upper())
    except Exception:
        return None


def _col_index_to_letter(c: int) -> str:
    try:
        from openpyxl.utils import get_column_letter
        return get_column_letter(c)
    except Exception:
        return str(c)


def _guess_col(ws, header_row: int, keywords: list[str], exclude: list[str] | None = None) -> int | None:
    exclude = exclude or []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is None:
            continue
        s = str(v).strip().lower()
        if any(k in s for k in exclude):
            continue
        if any(k in s for k in keywords):
            return c
    return None


# 충남 시트만 BK(63)까지 있을 수 있음. 나머지 시트는 데이터 있는 열까지만 표시.
MAX_DISPLAY_COL = 63  # 상한(이 열 수를 넘지 않음)
DISPLAY_BLOCK = 31   # 한 줄에 표시할 열 수 (가독성)


def _last_data_column(ws, header_row: int) -> int:
    """제목행(header_row)에서 값이 있는 마지막 열 인덱스(1-based)."""
    last = 0
    for c in range(1, min(ws.max_column, MAX_DISPLAY_COL) + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is not None and str(v).strip():
            last = c
    return last or 1


def _show_header_and_sample(ws, header_row: int, n_sample: int = 5) -> None:
    # 시트별로 데이터가 있는 열까지만 표시 (충남만 BK까지, 나머지는 실제 사용 열만)
    max_col = _last_data_column(ws, header_row)
    _log(f"\n[2행=제목 보조, 3행=제목행 / 샘플 데이터] (열 A ~ {_col_index_to_letter(max_col)}, 최대 {max_col}열)")
    start = SUBHEADER_ROW
    end = min(header_row + 1 + n_sample, ws.max_row + 1)
    cell_abbrev = 10  # 한 셀 표시 길이 (열 많을 때 가독성)
    for r in range(start, end):
        label = "보조" if r == SUBHEADER_ROW else ("제목" if r == header_row else f"행{r}")
        for block_start in range(1, max_col + 1, DISPLAY_BLOCK):
            block_end = min(block_start + DISPLAY_BLOCK, max_col + 1)
            parts = [f"({label})"]
            for c in range(block_start, block_end):
                v = ws.cell(row=r, column=c).value
                letter = _col_index_to_letter(c)
                parts.append(f"{letter}:{str(v)[:cell_abbrev]}" if v is not None else f"{letter}:")
            _log("  " + " | ".join(parts))
    if ws.max_column > max_col:
        _log(f"  (이 시트는 {ws.max_column}열까지 있음. 열 매핑 시 직접 열 문자 입력 가능)")


def _ask_col(ws, header_row: int, field_name: str, guess_keywords: list[str], default_guess: int | None) -> int | None:
    default_letter = _col_index_to_letter(default_guess) if default_guess is not None else ""
    if default_guess is not None:
        hint = ws.cell(row=header_row, column=default_guess).value
        s = _input(f"  {field_name} 열 (Enter: {default_letter} '{hint}'): ")
    else:
        s = _input(f"  {field_name} 열 (비우려면 0, 예: A, B, C): ")
    if not s and default_guess is not None:
        return default_guess
    if s == "0" or (s and s.strip().upper() == "0"):
        return None
    s = (s or "").strip().upper()
    try:
        c = int(s)
        if 1 <= c <= ws.max_column:
            return c
    except ValueError:
        pass
    c = _col_letter_to_index(s)
    if c is not None and 1 <= c <= ws.max_column:
        return c
    return default_guess


def _read_mapping(ws, header_row: int) -> dict[str, int]:
    g_school_code = _guess_col(ws, header_row, ["학교코드", "code"])
    g_school_name = _guess_col(ws, header_row, ["학교명", "학교"])
    g_region = _guess_col(ws, header_row, ["시군구", "지역", "region"])
    g_mgmt = _guess_col(ws, header_row, ["장비관리번호", "메모1", "관리번호"])
    # 제목행에 "다운로드", "Down (Mbps)" 등으로 표기된 열 검색
    g_dl = _guess_col(ws, header_row, ["다운로드", "dl", "download", "down"])
    g_ul = _guess_col(ws, header_row, ["업로드", "ul", "upload", "up"])
    g_rssi = _guess_col(ws, header_row, ["rssi", "신호세기"])
    g_ch = _guess_col(ws, header_row, ["ch", "채널"])

    _log("\n[열 매핑] 제목행(3행) 기준 열을 입력하세요. A, B, C, ... (Enter=자동검색값)")
    mapping = {}
    mapping["학교코드"] = _ask_col(ws, header_row, "학교코드", ["학교코드"], g_school_code)
    mapping["학교명"] = _ask_col(ws, header_row, "학교명", ["학교명"], g_school_name)
    mapping["시군구"] = _ask_col(ws, header_row, "시군구(지역, 지역별 파일 분산용)", ["시군구", "지역"], g_region)
    mapping["장비관리번호"] = _ask_col(ws, header_row, "장비관리번호(메모1)", ["장비"], g_mgmt)
    mapping["다운로드"] = _ask_col(ws, header_row, "다운로드", ["다운로드"], g_dl)
    mapping["업로드"] = _ask_col(ws, header_row, "업로드", ["업로드"], g_ul)
    mapping["RTT"] = None  # 전부하는 RTT 없음 → 매핑 질문 생략
    mapping["RSSI"] = _ask_col(ws, header_row, "RSSI", ["rssi"], g_rssi)
    mapping["CH"] = _ask_col(ws, header_row, "CH", ["ch"], g_ch)
    return mapping


def _to_float(v: Any) -> float | None:
    if v is None:
        return None
    try:
        s = str(v).strip().replace(",", "").replace("%", "")
        if not s:
            return None
        return float(s)
    except ValueError:
        return None


def _read_data_rows(ws, header_row: int, mapping: dict[str, int]) -> list[dict]:
    rows = []
    need_dl = mapping.get("다운로드") or mapping.get("업로드")
    if not mapping.get("학교코드") and not mapping.get("장비관리번호"):
        return rows
    for r in range(header_row + 1, ws.max_row + 1):
        def cv(key):
            col = mapping.get(key)
            if col is None:
                return None
            return ws.cell(row=r, column=col).value

        raw_code = cv("학교코드")
        raw_mgmt = cv("장비관리번호")
        if raw_code:
            school_code = str(raw_code).strip()[:12]
        elif raw_mgmt:
            try:
                from measure_utils import extract_school_code_from_mgmt_num
                school_code = extract_school_code_from_mgmt_num(raw_mgmt) or str(raw_mgmt).strip()[:12]
            except ImportError:
                school_code = str(raw_mgmt).strip().split("-")[0][:12] if "-" in str(raw_mgmt) else str(raw_mgmt).strip()[:12]
        else:
            school_code = ""
        if not school_code:
            continue
        school_name = (cv("학교명") or "")
        if school_name is not None:
            school_name = str(school_name).strip()
        region = (cv("시군구") or "")
        if region is not None:
            region = str(region).strip()
        else:
            region = ""
        mgmt = (cv("장비관리번호") or school_code or "")
        if mgmt is not None:
            mgmt = str(mgmt).strip()
        dl = _to_float(cv("다운로드"))
        ul = _to_float(cv("업로드"))
        rtt = _to_float(cv("RTT"))
        rssi = _to_float(cv("RSSI"))
        ch = _to_float(cv("CH"))
        if need_dl and dl is None and ul is None:
            continue
        rows.append({
            "school_code": school_code or "",
            "school_name": school_name or "",
            "region": region or "",
            "mgmt": mgmt or school_code,
            "dl": dl,
            "ul": ul,
            "rtt": rtt,
            "rssi": rssi,
            "ch": ch,
        })
    return rows


def _group_by_school_max_10(data: list[dict]) -> list[list[dict]]:
    """
    학교코드별로 묶고, 학교당 최대 10대만 사용 (순서 유지).
    측정 대수 = 1차/2차 시트의 해당 학교 행(장비) 수. 시트에 N행 있으면 N대.
    """
    from collections import OrderedDict
    by_school: dict[str, list[dict]] = OrderedDict()
    for row in data:
        sc = row.get("school_code") or ""
        if sc not in by_school:
            by_school[sc] = []
        if len(by_school[sc]) < MAX_DEVICES_PER_SCHOOL:
            by_school[sc].append(row)
    return list(by_school.values())


def _generate_ramp_up_for_device(
    target_dl: float, target_ul: float, target_rssi: float,
    num_points: int, seed: int,
) -> list[tuple[float, float, float]]:
    """
    요구사항 2: 혼자 측정이 최고값, 장비 늘어날수록 조금씩 감소. 차수 평균보다 나쁘면 안 됨(속도 >= target).
    상한 캡으로 비현실적 고속 방지 (평균의 1.2배 이하).
    """
    if num_points <= 0:
        return []
    random.seed(seed)
    out = []
    cap_hi_dl = target_dl * (1 + MAX_DEVIATION_RATIO)  # 1.2배
    cap_hi_ul = target_ul * (1 + MAX_DEVIATION_RATIO)
    # 시작은 평균보다 좋게(높게), 단 상한 이하
    dl = min(cap_hi_dl, target_dl * random.uniform(1.02, 1.15))
    ul = min(cap_hi_ul, target_ul * random.uniform(1.02, 1.15))
    rssi = target_rssi + random.uniform(0, 2)
    drop_every = 20  # 20초마다 장비 추가 → 감소
    for i in range(num_points):
        if i > 0 and i % drop_every == 0:
            drop_dl = random.uniform(RAMP_UP_DROP_MBPS_LO, RAMP_UP_DROP_MBPS_HI)
            drop_ul = random.uniform(RAMP_UP_DROP_MBPS_LO * 0.5, RAMP_UP_DROP_MBPS_HI * 0.5)
            dl = max(target_dl, min(cap_hi_dl, dl - drop_dl))
            ul = max(target_ul, min(cap_hi_ul, ul - drop_ul))
            rssi = max(-70, rssi + RSSI_DIFF_ON_DROP + random.uniform(-0.2, 0.2))
        else:
            fluc_dl = random.uniform(-RAMP_UP_FLUCTUATE_MBPS, RAMP_UP_FLUCTUATE_MBPS)
            fluc_ul = random.uniform(-RAMP_UP_FLUCTUATE_MBPS * 0.5, RAMP_UP_FLUCTUATE_MBPS * 0.5)
            dl = max(target_dl, min(cap_hi_dl, dl + fluc_dl))
            ul = max(target_ul, min(cap_hi_ul, ul + fluc_ul))
        out.append((round(dl, 2), round(ul, 2), round(rssi, 2)))
    return out


def _generate_600_rssi(mean_rssi: float, seed: int) -> list[float]:
    """600개 RSSI 생성. 평균 = mean_rssi. 5개 학교 CSV 분석 시 분산 반영."""
    random.seed(seed)
    total = 600 * mean_rssi
    params = _variance_params
    margin = params.get("rssi_std", RSSI_MARGIN_DEFAULT) * 2 if params else RSSI_MARGIN_DEFAULT
    lo = max(-70, mean_rssi - margin)
    hi = min(0, mean_rssi + margin)
    vals = []
    rem = total
    for i in range(599):
        need = rem / (600 - i)
        v_lo = max(lo, need - (600 - i - 1) * hi)
        v_hi = min(hi, need - (600 - i - 1) * lo)
        v_lo, v_hi = max(lo, min(v_lo, v_hi)), min(hi, max(v_lo, v_hi))
        v = random.uniform(v_lo, v_hi) if v_lo <= v_hi else need
        v = max(lo, min(hi, v))
        vals.append(v)
        rem -= v
    vals.append(max(lo, min(hi, rem)))
    # 합 보정으로 평균 정확히 맞춤
    s = sum(vals)
    vals[-1] = max(lo, min(hi, vals[-1] + (total - s)))
    return [round(x, 2) for x in vals]


def _redistribute_to_match_sum(vals: list[float], target_sum: float, cap_lo: float, cap_hi: float) -> None:
    """요구사항 1: sum(vals)이 정확히 target_sum이 되도록 재분배. 각 값은 [cap_lo, cap_hi] 유지."""
    n = len(vals)
    if n == 0:
        return
    eps = 1e-9
    current = sum(vals)
    diff = target_sum - current
    if abs(diff) < eps:
        return
    # 먼저 마지막 요소에 diff 반영 시도
    vals[-1] += diff
    if cap_lo <= vals[-1] <= cap_hi:
        return
    # 캡 밖이면 잘라내고 나머지를 다른 요소에 분산
    if vals[-1] > cap_hi:
        excess = vals[-1] - cap_hi
        vals[-1] = cap_hi
    else:
        excess = vals[-1] - cap_lo
        vals[-1] = cap_lo
    indices = list(range(n - 1))
    random.shuffle(indices)
    for i in indices:
        if abs(excess) < eps:
            break
        if excess > 0:
            room = cap_hi - vals[i]
            add = min(excess, room)
            vals[i] += add
            excess -= add
        else:
            room = vals[i] - cap_lo
            sub = min(-excess, room)
            vals[i] -= sub
            excess += sub
    # 수치 오차 잔여분은 마지막 요소에 반영(캡 초과 시에도 평균 우선)
    if abs(excess) >= eps:
        vals[-1] = max(cap_lo, min(cap_hi, vals[-1] + excess))


def _generate_600_with_mean(target_dl: float, target_ul: float, seed: int) -> list[tuple[float, float]]:
    """요구사항 1: 600개 (dl, ul) 생성. 평균 = target 정확 일치. 현실 반영: 상한 1.2배·하한 0.8배, 평균과 동떨어진 값 최대 5개 띄엄띄엄."""
    random.seed(seed)
    # 상·하한 캡: 혼자 측정이 최고이므로 극단값 방지 (평균의 0.8~1.2배)
    cap_lo_dl = max(0.1, target_dl * (1 - MAX_DEVIATION_RATIO))
    cap_hi_dl = target_dl * (1 + MAX_DEVIATION_RATIO)
    cap_lo_ul = max(0.1, target_ul * (1 - MAX_DEVIATION_RATIO))
    cap_hi_ul = target_ul * (1 + MAX_DEVIATION_RATIO)
    params = _variance_params
    if params:
        k = 0.5  # 분산은 좁게 (±0.5 sigma 수준)
        dl_cv = min(0.25, params.get("dl_cv", 0.15))
        ul_cv = min(0.25, params.get("ul_cv", 0.15))
        lo_dl = max(cap_lo_dl, target_dl * (1 - k * dl_cv))
        hi_dl = min(cap_hi_dl, target_dl * (1 + k * dl_cv))
        lo_ul = max(cap_lo_ul, target_ul * (1 - k * ul_cv))
        hi_ul = min(cap_hi_ul, target_ul * (1 + k * ul_cv))
    else:
        lo_dl = max(cap_lo_dl, target_dl * DL_UL_RATIO_LO)
        hi_dl = min(cap_hi_dl, target_dl * DL_UL_RATIO_HI)
        lo_ul = max(cap_lo_ul, target_ul * DL_UL_RATIO_LO)
        hi_ul = min(cap_hi_ul, target_ul * DL_UL_RATIO_HI)
    total_dl = 600 * target_dl
    total_ul = 600 * target_ul
    dls = []
    uls = []
    rem_dl, rem_ul = total_dl, total_ul
    n = 600
    for i in range(n - 1):
        need_dl = rem_dl / (n - i)
        need_ul = rem_ul / (n - i)
        dl_lo = max(lo_dl, need_dl - (n - i - 1) * hi_dl)
        dl_hi = min(hi_dl, need_dl - (n - i - 1) * lo_dl)
        dl_lo, dl_hi = max(lo_dl, min(dl_lo, dl_hi)), min(hi_dl, max(dl_lo, dl_hi))
        ul_lo = max(lo_ul, need_ul - (n - i - 1) * hi_ul)
        ul_hi = min(hi_ul, need_ul - (n - i - 1) * lo_ul)
        ul_lo, ul_hi = max(lo_ul, min(ul_lo, ul_hi)), min(hi_ul, max(ul_lo, ul_hi))
        dl = random.uniform(dl_lo, dl_hi) if dl_lo <= dl_hi else need_dl
        ul = random.uniform(ul_lo, ul_hi) if ul_lo <= ul_hi else need_ul
        dl = max(lo_dl, min(hi_dl, dl))
        ul = max(lo_ul, min(hi_ul, ul))
        dls.append(dl)
        uls.append(ul)
        rem_dl -= dl
        rem_ul -= ul
    dls.append(max(lo_dl, min(hi_dl, rem_dl)))
    uls.append(max(lo_ul, min(hi_ul, rem_ul)))
    _redistribute_to_match_sum(dls, total_dl, cap_lo_dl, cap_hi_dl)
    _redistribute_to_match_sum(uls, total_ul, cap_lo_ul, cap_hi_ul)
    # 평균과 동떨어진 값(0.9~1.1 밖) 최대 5개만 허용, 띄엄띄엄 위치에만
    def _limit_outliers(vals: list[float], target: float, cap_lo: float, cap_hi: float) -> None:
        inner_lo, inner_hi = target * 0.9, target * 1.1  # 이 안이면 "보통"
        outlier_indices = [i for i in range(n) if vals[i] < inner_lo or vals[i] > inner_hi]
        if len(outlier_indices) <= MAX_OUTLIER_COUNT:
            return
        allowed_set = set(min(i * OUTLIER_SPREAD_STEP, n - 1) for i in range(MAX_OUTLIER_COUNT))
        for i in outlier_indices:
            if i in allowed_set:
                continue
            vals[i] = target * random.uniform(0.92, 1.08)
        _redistribute_to_match_sum(vals, n * target, cap_lo, cap_hi)
    _limit_outliers(dls, target_dl, cap_lo_dl, cap_hi_dl)
    _limit_outliers(uls, target_ul, cap_lo_ul, cap_hi_ul)
    # 최종 평균 정확 일치 보정 (반올림 전에 합 일치)
    _redistribute_to_match_sum(dls, total_dl, cap_lo_dl, cap_hi_dl)
    _redistribute_to_match_sum(uls, total_ul, cap_lo_ul, cap_hi_ul)
    # 반올림 후에도 600개 평균 = target 되도록 마지막 값 보정
    rounded_dls = [round(x, 2) for x in dls]
    rounded_uls = [round(x, 2) for x in uls]
    rounded_dls[-1] = round(rounded_dls[-1] + (total_dl - sum(rounded_dls)), 2)
    rounded_uls[-1] = round(rounded_uls[-1] + (total_ul - sum(rounded_uls)), 2)
    return list(zip(rounded_dls, rounded_uls))


def _build_fullload_rows(
    school_groups: list[list[dict]],
    phase_label: str,
    seed_offset: int,
) -> list[list[Any]]:
    """전부하 원데이터 행: 학교별 ramp-up + 10분 구간. 학교당 장비 수 N에 따라 ramp (N-1)*20초, 10분 시작 시점 동적 적용."""
    out = []
    current = datetime.combine(datetime.today(), START_TIME)
    measure_date = datetime.today().date()

    for group_idx, devices in enumerate(school_groups):
        if not devices:
            continue
        # 17:00 넘으면 당일 측정 종료 → 다음 날 14:30로 넘겨 전교(712/192) 모두 출력
        while current.time() >= END_BEFORE:
            current = datetime.combine(current.date() + timedelta(days=1), START_TIME)
        measure_date = current.date()
        N = len(devices)
        base_time = current
        ramp_seconds = (N - 1) * SEC_BETWEEN_DEVICES if N >= 2 else 0

        # 20초 룰: 1번 장비 base_time+0, 2번 base_time+20, 3번 base_time+40, ... 동일 초에 여러 장비 측정 가능.
        # DL 전부 출력 후, 2번 마지막 DL 1초 뒤부터 UL 시작.
        seq_global = 0
        ramp_items: list[tuple[dict, list[tuple[float, float, float]]]] = []
        if ramp_seconds > 0:
            for dev_idx, dev in enumerate(devices[:-1]):
                start_t = dev_idx * SEC_BETWEEN_DEVICES
                num_points = ramp_seconds - start_t
                target_dl = dev.get("dl") if dev.get("dl") is not None else 50.0
                target_ul = dev.get("ul") if dev.get("ul") is not None else 20.0
                target_rssi = dev.get("rssi") if dev.get("rssi") is not None else BASE_RAMP_RSSI
                ramp_data = _generate_ramp_up_for_device(
                    target_dl, target_ul, target_rssi,
                    num_points, seed_offset + group_idx * 10000 + dev_idx * 1000,
                )
                ch_val = int(dev.get("ch")) if dev.get("ch") is not None else 100
                base = {
                    "학교": dev.get("school_name") or "",
                    "메모1": dev.get("mgmt") or dev.get("school_code") or "",
                    "CH": ch_val,
                    "학교코드": dev.get("school_code") or "",
                }
                ramp_items.append((base, ramp_data))
            # 1번 구간 ramp DL: 장비 d는 base_time + (d*20) + i 에 측정 (동일 초에 여러 장비 가능)
            for dev_idx, (base, ramp_data) in enumerate(ramp_items):
                start_t = dev_idx * SEC_BETWEEN_DEVICES
                for i, (dl, ul, rssi) in enumerate(ramp_data):
                    seq_global += 1
                    row_dt = base_time + timedelta(seconds=start_t + i)
                    row = [
                        row_dt.date(), row_dt.time(), base["학교"], "", "", base["메모1"], "AP#01", seq_global,
                        "전부하", "EDU", dl, None, None, LOSS_FIXED, "wi_cne_class_S", "",
                        STANDARD_WIFI, base["CH"], _ch_to_freq(base["CH"]), rssi, 80, base["학교코드"],
                        "",  # W열: 1번 구간(ramp)은 빈칸
                    ]
                    out.append(row)

        dev_10min: list[tuple[dict, list[tuple[float, float]], list[float], float]] = []
        for dev_idx, dev in enumerate(devices):
            target_dl = dev.get("dl") if dev.get("dl") is not None else 50.0
            target_ul = dev.get("ul") if dev.get("ul") is not None else 20.0
            target_rssi = dev.get("rssi") if dev.get("rssi") is not None else -52
            pairs = _generate_600_with_mean(
                target_dl, target_ul,
                seed_offset + group_idx * 1000 + dev_idx * 100,
            )
            rssi_list = _generate_600_rssi(
                target_rssi,
                seed_offset + group_idx * 1000 + dev_idx * 100 + 5000,
            )
            ch_val = int(dev.get("ch")) if dev.get("ch") is not None else 100
            base = {
                "학교": dev.get("school_name") or "",
                "메모1": dev.get("mgmt") or dev.get("school_code") or "",
                "CH": ch_val,
                "학교코드": dev.get("school_code") or "",
            }
            dev_10min.append((base, pairs, rssi_list, target_rssi))

        # 2번 구간 10분 DL: 장비 d는 base_time + ramp_seconds + (d*20) + i (동일 초에 여러 장비)
        for dev_idx, (base, pairs, rssi_list, target_rssi) in enumerate(dev_10min):
            start_t = dev_idx * SEC_BETWEEN_DEVICES
            for i, (dl, ul) in enumerate(pairs):
                seq_global += 1
                rssi_val = rssi_list[i] if i < len(rssi_list) else target_rssi
                row_dt = base_time + timedelta(seconds=ramp_seconds + start_t + i)
                row = [
                    row_dt.date(), row_dt.time(),
                    base["학교"], "", "", base["메모1"], "AP#01", seq_global,
                    "전부하", "EDU", dl, None, None, LOSS_FIXED, "wi_cne_class_S", "",
                    STANDARD_WIFI, base["CH"], _ch_to_freq(base["CH"]), rssi_val, 80, base["학교코드"],
                    "평균값 데이터",  # W열: 2번 구간(평균용 측정)
                ]
                out.append(row)

        # UL: 2번 마지막 DL 끝 1초 후 시작. 1번 UL → 2번 UL 동일하게 20초 간격 유지.
        last_dl_sec = ramp_seconds + (N - 1) * SEC_BETWEEN_DEVICES + SEC_MAIN_RECORD - 1
        start_ul = base_time + timedelta(seconds=last_dl_sec + 1)
        if ramp_seconds > 0:
            for dev_idx, (base, ramp_data) in enumerate(ramp_items):
                start_t = dev_idx * SEC_BETWEEN_DEVICES
                for i, (dl, ul, rssi) in enumerate(ramp_data):
                    seq_global += 1
                    row_dt = start_ul + timedelta(seconds=start_t + i)
                    row = [
                        row_dt.date(), row_dt.time(), base["학교"], "", "", base["메모1"], "AP#01", seq_global,
                        "전부하", "EDU", None, ul, None, LOSS_FIXED, "wi_cne_class_S", "",
                        STANDARD_WIFI, base["CH"], _ch_to_freq(base["CH"]), rssi, 80, base["학교코드"],
                        "",  # W열: 1번 구간(ramp)은 빈칸
                    ]
                    out.append(row)
        for dev_idx, (base, pairs, rssi_list, target_rssi) in enumerate(dev_10min):
            start_t = dev_idx * SEC_BETWEEN_DEVICES
            for i, (dl, ul) in enumerate(pairs):
                seq_global += 1
                rssi_val = rssi_list[i] if i < len(rssi_list) else target_rssi
                row_dt = start_ul + timedelta(seconds=ramp_seconds + start_t + i)
                row = [
                    row_dt.date(), row_dt.time(),
                    base["학교"], "", "", base["메모1"], "AP#01", seq_global,
                    "전부하", "EDU", None, ul, None, LOSS_FIXED, "wi_cne_class_S", "",
                    STANDARD_WIFI, base["CH"], _ch_to_freq(base["CH"]), rssi_val, 80, base["학교코드"],
                    "평균값 데이터",  # W열: 2번 구간(평균용 측정)
                ]
                out.append(row)

        # 다음 학교: 이 학교 DL 끝(last_dl_sec+1) + UL 전체 구간 후
        ul_total_sec = (ramp_seconds if ramp_seconds > 0 else 0) + (N - 1) * SEC_BETWEEN_DEVICES + SEC_MAIN_RECORD
        current = base_time + timedelta(seconds=last_dl_sec + 1 + ul_total_sec)

    return out


def _count_fullload_rows(school_groups: list[list[dict]]) -> int:
    """학교 그룹 리스트에 대해 생성될 전부하 행 수를 계산 (실제 생성 없이)."""
    total = 0
    for devices in school_groups:
        if not devices:
            continue
        N = len(devices)
        ramp_seconds = (N - 1) * SEC_BETWEEN_DEVICES if N >= 2 else 0
        ramp_rows = 2 * sum(ramp_seconds - d * SEC_BETWEEN_DEVICES for d in range(N - 1)) if ramp_seconds > 0 else 0
        main_rows = N * SEC_MAIN_RECORD * 2  # 10분 DL + UL
        total += ramp_rows + main_rows
    return total


# Excel 시트 최대 행 수(1,048,576). 지역별 분산 시 한 파일이 이 한도를 넘지 않도록 함.
EXCEL_MAX_DATA_ROWS = 1_048_575
# 한 지역 내에서 행이 한도 초과 시, 학교 단위로 잘라 쓸 때 한 파일당 최대 학교 수.
SCHOOLS_PER_FILE = 80


def _partition_groups_by_row_limit(
    groups: list[list[dict]], max_rows: int
) -> list[list[list[dict]]]:
    """
    학교 그룹 리스트를 행 한도 이하로 나누되, 파일별 행 수를 반반 정도로 균등 분할.
    학교 경계에서만 잘라서 한 학교 데이터가 여러 파일에 쪼개지지 않도록 함.
    """
    if not groups:
        return []
    counts = [_count_fullload_rows([g]) for g in groups]
    total = sum(counts)
    if total <= max_rows:
        return [groups]
    K = max(2, (total + max_rows - 1) // max_rows)  # 필요 파일 수(최소 2로 반반 분할)
    target_per_file = total / K
    buckets: list[list[list[dict]]] = []
    idx = 0
    for k in range(K):
        if idx >= len(groups):
            break
        bucket: list[list[dict]] = []
        bucket_rows = 0
        while idx < len(groups):
            c = counts[idx]
            if bucket_rows + c > max_rows and bucket:
                break
            if k < K - 1 and bucket_rows >= target_per_file and bucket:
                break
            bucket.append(groups[idx])
            bucket_rows += c
            idx += 1
        if bucket:
            buckets.append(bucket)
    return buckets


def _sanitize_region_for_filename(region: str) -> str:
    """파일명에 쓸 수 있도록 시군구 문자열 정리."""
    if not (region or "").strip():
        return "미분류"
    s = (region or "").strip()
    for c in r'\/:*?"<>|':
        s = s.replace(c, "_")
    return s[:50]  # 길이 제한


def _write_excel(rows: list[list[Any]], out_path: str) -> list[str]:
    """행을 하나의 Excel 파일로 저장. (분할은 run_phase에서 학교 단위로 수행)"""
    from openpyxl import Workbook
    if not rows:
        return []
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for c, h in enumerate(OUTPUT_HEADERS, 1):
        ws.cell(row=1, column=c, value=h)
    for r_idx, row in enumerate(rows, 2):
        for c_idx, val in enumerate(row, 1):
            if c_idx <= len(OUTPUT_HEADERS):
                ws.cell(row=r_idx, column=c_idx, value=val)
    wb.save(out_path)
    wb.close()
    _log(f"[저장] {out_path} ({len(rows)}행)")
    return [out_path]


def run_phase(wb, sheet_name: str, mapping: dict[str, int], phase_label: str, out_dir: str, seed_offset: int) -> str | None:
    ws = wb[sheet_name]
    data = _read_data_rows(ws, HEADER_ROW, mapping)
    if not data:
        _log(f"[경고] '{sheet_name}'에서 유효한 데이터 행이 없습니다.")
        return None
    data.sort(key=lambda x: (x.get("school_code") or "", x.get("mgmt") or ""))
    # 측정 대수 = 시트의 학교별 행 수 (1행=1대, 최대 10대)
    school_groups = _group_by_school_max_10(data)
    if not school_groups:
        return None
    # 지역별로 묶기 (각 학교 그룹의 첫 행에서 region 사용)
    by_region: dict[str, list[list[dict]]] = defaultdict(list)
    for group in school_groups:
        if not group:
            continue
        region = (group[0].get("region") or "").strip() or "미분류"
        by_region[region].append(group)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_no_region = os.path.join(out_dir, f"전부하_원데이터_{phase_label}_{stamp}")
    written: list[str] = []
    for region in sorted(by_region.keys()):
        groups_in_region = by_region[region]
        region_safe = _sanitize_region_for_filename(region)
        region_base = f"{base_no_region}_{region_safe}"
        # 한 지역 내 행이 너무 많으면 학교 수로 다시 분할
        start = 0
        part = 1
        while start < len(groups_in_region):
            chunk = groups_in_region[start : start + SCHOOLS_PER_FILE]
            total_rows = _count_fullload_rows(chunk)
            if total_rows > EXCEL_MAX_DATA_ROWS:
                # 행 한도 초과 시 학교 경계로 균등 분할(반반 정도), 한 파일에 최대치 채우지 않음
                sub_chunks = _partition_groups_by_row_limit(chunk, EXCEL_MAX_DATA_ROWS)
                for sub_idx, sub_chunk in enumerate(sub_chunks, 1):
                    out_rows = _build_fullload_rows(sub_chunk, phase_label, seed_offset)
                    if not out_rows:
                        continue
                    path = f"{region_base}_part{part}_{sub_idx}.xlsx"
                    written.extend(_write_excel(out_rows, path))
                start += len(chunk)
                part += 1
                continue
            out_rows = _build_fullload_rows(chunk, phase_label, seed_offset)
            if not out_rows:
                start += len(chunk)
                part += 1
                continue
            n_parts_region = (len(groups_in_region) + SCHOOLS_PER_FILE - 1) // SCHOOLS_PER_FILE
            if n_parts_region == 1 and start + len(chunk) >= len(groups_in_region):
                path = f"{region_base}.xlsx"
            else:
                path = f"{region_base}_part{part}.xlsx"
            written.extend(_write_excel(out_rows, path))
            start += len(chunk)
            part += 1
    if len(written) > 1:
        _log(f"  → 지역별로 총 {len(written)}개 파일로 분산 저장")
    return written[0] if written else None


def main() -> None:
    _log("=" * 60)
    _log("전부하(EDU) 1차/2차 원데이터 생성")
    _log("=" * 60)

    try:
        from openpyxl import load_workbook
    except ImportError:
        _log("[오류] openpyxl이 필요합니다. pip install openpyxl")
        sys.exit(1)

    files = _collect_files()
    if not files:
        path = _input("엑셀 파일 경로를 입력하세요 (Q: 취소): ")
    else:
        path = _pick_file("기초 데이터 파일(1차/2차 평균 데이터)을 선택하세요.", files)
    if not path or not os.path.isfile(path):
        _log("[오류] 파일을 찾을 수 없습니다.")
        sys.exit(1)

    wb = load_workbook(path, data_only=True)
    _log(f"\n파일: {path}")

    sheet_1 = _pick_sheet(wb, "1차 데이터 시트를 선택하세요.")
    _show_header_and_sample(wb[sheet_1], HEADER_ROW)
    mapping = _read_mapping(wb[sheet_1], HEADER_ROW)
    if not mapping.get("학교코드") and not mapping.get("장비관리번호"):
        _log("[오류] 학교코드 또는 장비관리번호 열이 필요합니다.")
        wb.close()
        sys.exit(1)

    out_dir = os.path.dirname(path)
    out_dir_custom = _input(f"\n출력 폴더 (Enter: {out_dir}, Q: 취소): ")
    if out_dir_custom.strip():
        out_dir = out_dir_custom.strip()

    # 5개 학교 전부하 분산: CSV 폴더 우선, 없으면 JSON(fullload_variance_params.json) 사용
    global _variance_params
    _variance_params = None
    csv_dir = _input("5개 학교 전부하 측정 CSV가 있는 폴더 경로 (Enter: JSON/기본 분산, Q: 취소): ").strip()
    if csv_dir and os.path.isdir(csv_dir):
        csv_paths = sorted([
            os.path.join(csv_dir, f) for f in os.listdir(csv_dir)
            if f.lower().endswith(".csv") and not f.startswith(".") and not f.startswith("~$")
        ])
        if csv_paths:
            _variance_params = _analyze_fullload_csv(csv_paths)
            if _variance_params:
                _log(f"[분산 분석] CSV {len(csv_paths)}개 적용 (dl_cv={_variance_params.get('dl_cv', 0):.3f}, ul_cv={_variance_params.get('ul_cv', 0):.3f}, rssi_std={_variance_params.get('rssi_std', 0):.2f})")
            else:
                _log("[분산 분석] 유효한 장비별 데이터 없음. JSON/기본 분산 시도.")
        else:
            _log("[분산 분석] 해당 폴더에 CSV 없음. JSON/기본 분산 시도.")
    if _variance_params is None:
        _variance_params = _load_variance_params_from_json()
        if _variance_params:
            _log(f"[분산 분석] JSON 적용 (dl_cv={_variance_params.get('dl_cv', 0):.3f}, ul_cv={_variance_params.get('ul_cv', 0):.3f}, rssi_std={_variance_params.get('rssi_std', 0):.2f})")
        else:
            _log("[분산 분석] JSON 없음. 기본 분산 사용.")

    seed_offset = random.randint(0, 0x7FFFFFFF)
    path_1 = run_phase(wb, sheet_1, mapping, "1차", out_dir, seed_offset)
    if path_1:
        _log(f"1차 원데이터: {path_1}")

    _log("\n2차 데이터도 생성합니다.")
    sheet_2 = _pick_sheet(wb, "2차 데이터 시트를 선택하세요 (같은 시트면 1차와 동일 번호).")
    use_same_mapping = _input("열 매핑을 1차와 동일하게 사용할까요? (Enter: 예, Q: 취소): ")
    if use_same_mapping.strip().lower() in ("n", "no", "0"):
        _show_header_and_sample(wb[sheet_2], HEADER_ROW)
        mapping = _read_mapping(wb[sheet_2], HEADER_ROW)
    path_2 = run_phase(wb, sheet_2, mapping, "2차", out_dir, seed_offset + 1)
    if path_2:
        _log(f"2차 원데이터: {path_2}")

    wb.close()
    _log("\n[완료]")


if __name__ == "__main__":
    main()
