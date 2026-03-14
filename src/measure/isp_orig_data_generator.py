# -*- coding: utf-8 -*-
"""
ISP 1차/2차 원데이터 생성 프로그램

기획: docs/06_원데이터_생성_프로그램_기획.md
- 실행 폴더(및 row_data/raw_data)의 파일 목록 → 사용자 파일 선택
- 선택 파일의 시트 목록 → 사용자 시트 선택 (1차, 2차 각각)
- 제목행 + 일부 데이터 표시 → 열 매핑(학교코드, 다운로드, 업로드, RTT, RSSI, CH, 장비 등) 선택
- 평균값이 나오도록 6개 원데이터 생성 (DL/UL 10~60 Mbps, RTT 0.7, RSSI -1)
- 시간: 14:00 시작, 학교 간 30분, 장비 간 3~4분, 회차 간 29초, DL-UL 15초, 17:00 전 종료
- 출력: ISP_샘플.xlsx와 동일 열 구조
"""

from __future__ import annotations

import os
import sys
import random
from datetime import datetime, time, timedelta
from typing import Any

if getattr(sys, "frozen", False):
    # 실행 파일(.exe / Mac 앱)이 있는 폴더 = 데이터/출력 기본 위치 (Windows·Mac 공통)
    _BASE_DIR = os.path.dirname(os.path.abspath(sys.executable))
    _MEIPASS = getattr(sys, "_MEIPASS", _BASE_DIR)
    _MEASURE_BUNDLE = os.path.join(_MEIPASS, "measure")
    if os.path.isdir(_MEASURE_BUNDLE) and _MEASURE_BUNDLE not in sys.path:
        sys.path.insert(0, _MEASURE_BUNDLE)
    if _MEIPASS not in sys.path:
        sys.path.insert(0, _MEIPASS)
    # 배포 시: 데이터 파일은 실행 파일과 같은 폴더에 둠 (다른 사람에게 exe 전달 시 동일 폴더 사용)
    _DATA_DIRS = [_BASE_DIR]
else:
    _BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    _RUN_DIR = os.getcwd()
    # 개발 시: 실행 위치 + 프로젝트 내 row_data/raw_data 등
    _DATA_DIRS = [
        _RUN_DIR,
        _BASE_DIR,
        os.path.join(_RUN_DIR, "row_data"),
        os.path.join(_RUN_DIR, "raw_data"),
        os.path.join(_BASE_DIR, "row_data"),
        os.path.join(_BASE_DIR, "raw_data"),
        os.path.join(os.path.dirname(_BASE_DIR), "row_data"),
        os.path.join(os.path.dirname(_BASE_DIR), "raw_data"),
    ]
if _BASE_DIR not in sys.path:
    sys.path.insert(0, _BASE_DIR)

_RUN_DIR = os.getcwd()

# 출력 열 순서 (ISP_샘플.xlsx 기준)
OUTPUT_HEADERS = [
    "Date", "StartTime", "학교", "학년", "반", "메모1", "메모2",
    "측정순번", "측정유형", "측정서버", "DL", "UL", "RTT", "LOSS",
    "SSID", "BSSID", "Standard", "CH", "Freq.", "RSSI", "BW", "학교코드",
]

# 값 규칙
DL_UL_MIN, DL_UL_MAX = 10.0, 60.0
RTT_FIXED = 0.7
RSSI_FIXED = -1
LOSS_FIXED = 0
START_TIME = time(14, 0, 0)
END_BEFORE = time(17, 0, 0)
SEC_DL_UL = 15
SEC_BETWEEN_SESSIONS = 29
SEC_MIN_DEVICE, SEC_MAX_DEVICE = 3 * 60, 4 * 60  # 3~4분
SEC_SCHOOL_TRAVEL = 30 * 60  # 30분


def _log(msg: str) -> None:
    print(msg, flush=True)


def _input(prompt: str) -> str:
    try:
        return input(prompt).strip()
    except (EOFError, KeyboardInterrupt):
        _log("\n[중단] 사용자 입력으로 종료합니다.")
        sys.exit(1)


def _collect_files(exts=(".xlsx", ".xlsm")) -> list[str]:
    seen = set()
    out = []
    for d in _DATA_DIRS:
        if not d or not os.path.isdir(d):
            continue
        try:
            for f in os.listdir(d):
                # 숨김 파일(.), 임시/잠금 파일(~$) 제외 → 일반적으로 보이는 파일만
                if f.startswith(".") or f.startswith("~$"):
                    continue
                low = f.lower()
                if not any(low.endswith(e) for e in exts):
                    continue
                p = os.path.join(d, f)
                if not os.path.isfile(p):
                    continue
                key = os.path.normcase(os.path.abspath(p))
                if key in seen:
                    continue
                seen.add(key)
                out.append(p)
        except Exception:
            pass
    return sorted(out)


def _pick_file(prompt: str, candidates: list[str]) -> str:
    existing = [p for p in candidates if os.path.isfile(p)]
    if getattr(sys, "frozen", False) and _DATA_DIRS:
        _log(f"\n[데이터 위치] 실행 파일이 있는 폴더: {_DATA_DIRS[0]}")
    _log(f"\n{prompt}")
    if existing:
        for i, p in enumerate(existing, 1):
            _log(f"  {i}. {os.path.basename(p)}")
        _log("  0. 직접 경로 입력")
        s = _input("번호 선택 (Enter: 1번): ")
        if not s:
            return existing[0]
        if s == "0":
            return _input("전체 경로 입력: ")
        try:
            idx = int(s)
            if 1 <= idx <= len(existing):
                return existing[idx - 1]
        except ValueError:
            pass
        _log("[경고] 잘못된 입력. 1번 사용")
        return existing[0]
    _log("  후보 없음. 직접 경로 입력")
    return _input("전체 경로 입력: ")


def _pick_sheet(wb, prompt: str) -> str:
    names = wb.sheetnames
    _log(f"\n{prompt}")
    for i, n in enumerate(names, 1):
        _log(f"  {i}. {n}")
    s = _input("번호 선택 (Enter: 1번): ")
    if not s:
        return names[0]
    try:
        idx = int(s)
        if 1 <= idx <= len(names):
            return names[idx - 1]
    except ValueError:
        pass
    return names[0]


# 제목 구조: 2행=제목 보조 행, 3행=제목행(헤더) → 4행부터 데이터
HEADER_ROW = 3
SUBHEADER_ROW = 2


def _col_letter_to_index(col: str) -> int | None:
    """엑셀 열 문자 → 1-based 열 번호. A=1, B=2, ..., Z=26, AA=27."""
    try:
        from openpyxl.utils import column_index_from_string
        return column_index_from_string(col.strip().upper())
    except Exception:
        return None


def _col_index_to_letter(c: int) -> str:
    """1-based 열 번호 → 엑셀 열 문자."""
    try:
        from openpyxl.utils import get_column_letter
        return get_column_letter(c)
    except Exception:
        return str(c)


def _guess_col(ws, header_row: int, keywords: list[str], exclude: list[str] | None = None) -> int | None:
    exclude = exclude or []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is None:
            continue
        s = str(v).strip().lower()
        if any(k in s for k in exclude):
            continue
        if any(k in s for k in keywords):
            return c
    return None


def _show_header_and_sample(ws, header_row: int, n_sample: int = 5) -> None:
    _log("\n[2행=제목 보조, 3행=제목행 / 샘플 데이터] (열: A, B, C, ...)")
    # 2행(보조), 3행(제목), 그 다음 샘플
    start = SUBHEADER_ROW
    end = min(header_row + 1 + n_sample, ws.max_row + 1)
    for r in range(start, end):
        label = "보조" if r == SUBHEADER_ROW else ("제목" if r == header_row else f"행{r}")
        parts = [f"({label})"]
        for c in range(1, min(ws.max_column + 1, 25)):
            v = ws.cell(row=r, column=c).value
            letter = _col_index_to_letter(c)
            parts.append(f"{letter}:{str(v)[:18]}" if v is not None else f"{letter}:")
        _log("  " + " | ".join(parts))


def _ask_col(ws, header_row: int, field_name: str, guess_keywords: list[str], default_guess: int | None) -> int | None:
    default_letter = _col_index_to_letter(default_guess) if default_guess is not None else ""
    if default_guess is not None:
        hint = ws.cell(row=header_row, column=default_guess).value
        s = _input(f"  {field_name} 열 (Enter: {default_letter} '{hint}'): ")
    else:
        s = _input(f"  {field_name} 열 (비우려면 0, 예: A, B, C): ")
    if not s and default_guess is not None:
        return default_guess
    if s == "0" or (s and s.strip().upper() == "0"):
        return None
    s = (s or "").strip().upper()
    # 숫자 입력도 허용 (기존 호환)
    try:
        c = int(s)
        if 1 <= c <= ws.max_column:
            return c
    except ValueError:
        pass
    # A, B, C, ... 형식
    c = _col_letter_to_index(s)
    if c is not None and 1 <= c <= ws.max_column:
        return c
    return default_guess


def _read_mapping(ws, header_row: int) -> dict[str, int]:
    """사용자 열 매핑. 반환: { '학교코드': 1, '학교명': 2, ... } 1-based."""
    g_school_code = _guess_col(ws, header_row, ["학교코드", "code"])
    g_school_name = _guess_col(ws, header_row, ["학교명", "학교"])
    g_mgmt = _guess_col(ws, header_row, ["장비관리번호", "메모1", "관리번호"])
    g_dl = _guess_col(ws, header_row, ["다운로드", "dl", "download"])
    g_ul = _guess_col(ws, header_row, ["업로드", "ul", "upload"])
    g_rtt = _guess_col(ws, header_row, ["rtt", "지연"])
    g_rssi = _guess_col(ws, header_row, ["rssi", "신호세기"])
    g_ch = _guess_col(ws, header_row, ["ch", "채널"])

    _log("\n[열 매핑] 제목행(3행) 기준 열을 입력하세요. A, B, C, ... (Enter=자동검색값)")
    mapping = {}
    mapping["학교코드"] = _ask_col(ws, header_row, "학교코드", ["학교코드"], g_school_code)
    mapping["학교명"] = _ask_col(ws, header_row, "학교명", ["학교명"], g_school_name)
    mapping["장비관리번호"] = _ask_col(ws, header_row, "장비관리번호(메모1)", ["장비"], g_mgmt)
    mapping["다운로드"] = _ask_col(ws, header_row, "다운로드", ["다운로드"], g_dl)
    mapping["업로드"] = _ask_col(ws, header_row, "업로드", ["업로드"], g_ul)
    mapping["RTT"] = _ask_col(ws, header_row, "RTT", ["rtt"], g_rtt)
    mapping["RSSI"] = _ask_col(ws, header_row, "RSSI", ["rssi"], g_rssi)
    mapping["CH"] = _ask_col(ws, header_row, "CH", ["ch"], g_ch)
    return mapping


def _to_float(v: Any) -> float | None:
    if v is None:
        return None
    try:
        s = str(v).strip().replace(",", "").replace("%", "")
        if not s:
            return None
        return float(s)
    except ValueError:
        return None


def _read_data_rows(ws, header_row: int, mapping: dict[str, int]) -> list[dict]:
    """데이터 행 읽기. 반환: [{학교코드, 학교명, 장비, dl, ul, rtt, rssi, ch}, ...]"""
    rows = []
    need_dl = mapping.get("다운로드") or mapping.get("업로드")
    if not mapping.get("학교코드") and not mapping.get("장비관리번호"):
        return rows
    for r in range(header_row + 1, ws.max_row + 1):
        def cv(key):
            col = mapping.get(key)
            if col is None:
                return None
            return ws.cell(row=r, column=col).value

        raw_code = cv("학교코드")
        raw_mgmt = cv("장비관리번호")
        if raw_code:
            school_code = str(raw_code).strip()[:12]
        elif raw_mgmt:
            try:
                from measure_utils import extract_school_code_from_mgmt_num
                school_code = extract_school_code_from_mgmt_num(raw_mgmt) or str(raw_mgmt).strip()[:12]
            except ImportError:
                school_code = str(raw_mgmt).strip().split("-")[0][:12] if "-" in str(raw_mgmt) else str(raw_mgmt).strip()[:12]
        else:
            school_code = ""
        if not school_code:
            continue
        school_name = (cv("학교명") or "")
        if school_name is not None:
            school_name = str(school_name).strip()
        mgmt = (cv("장비관리번호") or school_code or "")
        if mgmt is not None:
            mgmt = str(mgmt).strip()
        dl = _to_float(cv("다운로드"))
        ul = _to_float(cv("업로드"))
        rtt = _to_float(cv("RTT"))
        rssi = _to_float(cv("RSSI"))
        ch = _to_float(cv("CH"))
        if need_dl and dl is None and ul is None:
            continue
        rows.append({
            "school_code": school_code or "",
            "school_name": school_name or "",
            "mgmt": mgmt or school_code,
            "dl": dl,
            "ul": ul,
            "rtt": rtt,
            "rssi": rssi,
            "ch": ch,
        })
    return rows


def _generate_three_in_range(total: float, lo: float, hi: float, seed_offset: int = 0) -> list[float]:
    """
    3개 값을 생성하여 합 = total, 각 값 in [lo, hi].
    자연스러운 분산: 첫 두 값을 [lo, hi]에서 넓게 뽑고, 세 번째로 합 맞춤.
    """
    random.seed(seed_offset + int(total * 1000))
    # d1 가능 범위: 나머지 두 개가 [lo,hi]에 있으려면 total - d1 이 [2*lo, 2*hi]여야 함
    d1_lo = max(lo, total - 2 * hi)
    d1_hi = min(hi, total - 2 * lo)
    if d1_lo > d1_hi:
        d1 = (d1_lo + d1_hi) / 2
    else:
        d1 = random.uniform(d1_lo, d1_hi)
    # d2 가능 범위: d3 = total - d1 - d2 가 [lo, hi]여야 함
    d2_lo = max(lo, total - d1 - hi)
    d2_hi = min(hi, total - d1 - lo)
    if d2_lo > d2_hi:
        d2 = (d2_lo + d2_hi) / 2
    else:
        d2 = random.uniform(d2_lo, d2_hi)
    d3 = total - d1 - d2
    d3 = max(lo, min(hi, d3))
    # 소수점 2자리, 합 = total 유지: 두 개 반올림 후 세 번째로 보정
    o1 = round(d1, 2)
    o2 = round(d2, 2)
    o3 = round(total - o1 - o2, 2)
    o3 = max(lo, min(hi, o3))
    return [o1, o2, o3]


def _speed_range(avg: float) -> tuple[float, float]:
    """원본 평균 avg에 맞는 개별값 범위 [lo, hi]. 평균이 10~60이면 10~60, 그 밖이면 avg 주변."""
    if avg is None or (DL_UL_MIN <= avg <= DL_UL_MAX):
        return (DL_UL_MIN, DL_UL_MAX)
    if avg < DL_UL_MIN:
        return (max(0.1, avg - 5), DL_UL_MAX)
    return (max(DL_UL_MIN, avg * 0.85), avg * 1.15)


def _generate_six_values(avg_dl: float | None, avg_ul: float | None) -> list[tuple[float, float]]:
    """
    6회 측정 = 다운로드 3번 + 업로드 3번 (다운로드 6번·업로드 6번 아님).
    원본 평균이 나오도록 3개 DL값, 3개 UL값만 생성. 반환은 6행: (dl1,0),(0,ul1),(dl2,0),(0,ul2),(dl3,0),(0,ul3).
    - 3개 DL 합 = 3*avg_dl, 3개 UL 합 = 3*avg_ul (평균 보존). 원본 데이터는 절대 수정하지 않음.
    """
    if avg_dl is None:
        avg_dl = (DL_UL_MIN + DL_UL_MAX) / 2
    if avg_ul is None:
        avg_ul = (DL_UL_MIN + DL_UL_MAX) / 2
    total_dl = 3 * avg_dl
    total_ul = 3 * avg_ul
    lo_dl, hi_dl = _speed_range(avg_dl)
    lo_ul, hi_ul = _speed_range(avg_ul)
    seed = int(avg_dl * 100 + avg_ul * 10) & 0x7FFFFFFF
    dls = _generate_three_in_range(total_dl, lo_dl, hi_dl, seed)
    uls = _generate_three_in_range(total_ul, lo_ul, hi_ul, seed + 10000)
    return [
        (round(dls[0], 2), 0.0), (0.0, round(uls[0], 2)),
        (round(dls[1], 2), 0.0), (0.0, round(uls[1], 2)),
        (round(dls[2], 2), 0.0), (0.0, round(uls[2], 2)),
    ]


def _generate_six_in_range(total: float, lo: float, hi: float, seed_off: int) -> list[float]:
    """6개 값이 합 = total, 각 [lo, hi], 소수점 2자리. 평균 = total/6."""
    random.seed(seed_off + int(total * 100))
    v1 = random.uniform(lo, hi)
    v2 = random.uniform(lo, hi)
    v3 = random.uniform(lo, hi)
    v4 = random.uniform(lo, hi)
    s4 = v1 + v2 + v3 + v4
    v5_lo = max(lo, total - hi - s4)
    v5_hi = min(hi, total - lo - s4)
    v5 = random.uniform(v5_lo, v5_hi) if v5_lo <= v5_hi else (v5_lo + v5_hi) / 2
    v6 = total - s4 - v5
    v6 = max(lo, min(hi, v6))
    out = [round(v1, 2), round(v2, 2), round(v3, 2), round(v4, 2), round(v5, 2), round(v6, 2)]
    out[5] = round(total - out[0] - out[1] - out[2] - out[3] - out[4], 2)
    out[5] = max(lo, min(hi, out[5]))
    return out


def _generate_six_rtt_rssi(avg_rtt: float | None, avg_rssi: float | None, seed: int) -> tuple[list[float], list[float]]:
    """RTT·RSSI만 6개씩 생성(6회 측정 행마다 각 1개). 6개 평균 = 원본 장비 평균, 소수점 2자리."""
    rtt = 0.7 if avg_rtt is None else float(avg_rtt)
    rssi = -1.0 if avg_rssi is None else float(avg_rssi)
    # RTT: 원본 값에 맞는 구간 (0.7일 때 ±0.2, 16일 때 ±2 등). min(2.0, ...) 제거해 고RTT 반영
    rtt_margin = max(0.2, min(2.0, rtt * 0.15))
    rtt_lo = max(0.1, rtt - rtt_margin)
    rtt_hi = rtt + rtt_margin
    # RSSI: 원본 값에 맞는 구간. -52일 때 [-55,-49] 등, max(-50,...) 제거해 -50 이하 반영
    rssi_lo = max(-70, rssi - 3)
    rssi_hi = min(0, rssi + 3)
    rtts = _generate_six_in_range(6 * rtt, rtt_lo, rtt_hi, seed)
    rssis = _generate_six_in_range(6 * rssi, rssi_lo, rssi_hi, seed + 5000)
    return (rtts, rssis)


def _assign_times(device_list: list[dict]) -> list[tuple[dict, list[time]]]:
    """
    장비 순서대로 측정 시각 6개 배정.
    학교 간 30분, 장비 간 3~4분, 회차 내 DL-UL 15초, 회차 간 29초. 17:00 전 종료.
    반환: [(device, [t1..t6]), ...]
    """
    result = []
    current = datetime.combine(datetime.today(), START_TIME)
    end_dt = datetime.combine(datetime.today(), END_BEFORE)
    prev_school = None
    for dev in device_list:
        sc = dev.get("school_code") or ""
        if prev_school is not None and sc != prev_school:
            current += timedelta(seconds=SEC_SCHOOL_TRAVEL)
        if current.time() >= END_BEFORE:
            break
        # 이 장비 6회: 0, 15s, +29s, +15s, +29s, +15s
        times = []
        t = current
        for i in range(6):
            if t.time() >= END_BEFORE:
                break
            times.append(t.time())
            if i % 2 == 0:
                t += timedelta(seconds=SEC_DL_UL)
            else:
                t += timedelta(seconds=SEC_BETWEEN_SESSIONS)
        if len(times) < 6:
            times.extend([times[-1]] * (6 - len(times)))
        result.append((dev, times))
        # 다음 장비까지 3~4분
        current += timedelta(seconds=random.randint(SEC_MIN_DEVICE, SEC_MAX_DEVICE))
        prev_school = sc
    return result


def _build_output_rows(device_with_times: list[tuple[dict, list[time]]], phase_label: str) -> list[list[Any]]:
    """출력 시트용 행 리스트. 원본 측정 데이터는 절대 바꾸지 않고, 평균이 나오도록 생성만 함.
    - 6회 측정 = 다운로드 3번 + 업로드 3번 (DL 3개값, UL 3개값).
    - RTT, RSSI만 6개씩 생성(행마다 1개). 소수점 2자리."""
    out = []
    for idx, (dev, times) in enumerate(device_with_times):
        six = _generate_six_values(dev.get("dl"), dev.get("ul"))
        seed_rtt_rssi = idx * 1000 + int((dev.get("dl") or 0) * 10 + (dev.get("rtt") or 0) * 100) & 0x7FFFFFFF
        rtts, rssis = _generate_six_rtt_rssi(dev.get("rtt"), dev.get("rssi"), seed_rtt_rssi)
        ch_val = dev.get("ch")
        if ch_val is None:
            ch_val = 100
        ch_val = int(ch_val) if isinstance(ch_val, (int, float)) else 100
        base = {
            "학교": dev.get("school_name") or "",
            "학년": "",
            "반": "",
            "메모1": dev.get("mgmt") or dev.get("school_code") or "",
            "메모2": "AP#01",
            "측정유형": "속도",
            "측정서버": "ISP",
            "LOSS": LOSS_FIXED,
            "SSID": "wi_cne_class_S",
            "BSSID": "",
            "Standard": "802.11ac",
            "CH": ch_val,
            "Freq.": 5500 if ch_val > 36 else 2437,
            "BW": 80,
            "학교코드": dev.get("school_code") or "",
        }
        measure_date = datetime.today().date()
        for i, (dl, ul) in enumerate(six):
            seq = (i // 2) + 1  # 1,1,2,2,3,3
            row = [
                measure_date,
                times[i] if i < len(times) else None,
                base["학교"],
                base["학년"],
                base["반"],
                base["메모1"],
                base["메모2"],
                seq,
                base["측정유형"],
                base["측정서버"],
                dl,
                ul,
                rtts[i] if i < len(rtts) else round(0.7, 2),
                base["LOSS"],
                base["SSID"],
                base["BSSID"],
                base["Standard"],
                base["CH"],
                base["Freq."],
                rssis[i] if i < len(rssis) else round(-1.0, 2),
                base["BW"],
                base["학교코드"],
            ]
            out.append(row)
    return out


def _write_excel(rows: list[list[Any]], out_path: str, phase_label: str) -> None:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for c, h in enumerate(OUTPUT_HEADERS, 1):
        ws.cell(row=1, column=c, value=h)
    for r_idx, row in enumerate(rows, 2):
        for c_idx, val in enumerate(row, 1):
            if c_idx <= len(OUTPUT_HEADERS):
                ws.cell(row=r_idx, column=c_idx, value=val)
    wb.save(out_path)
    wb.close()
    _log(f"[저장] {out_path} ({len(rows)}행)")


def run_phase(wb, sheet_name: str, mapping: dict[str, int], phase_label: str, out_dir: str) -> str | None:
    """한 시트(1차 또는 2차)에 대해 원데이터 생성 후 저장. 반환: 저장 경로 또는 None."""
    ws = wb[sheet_name]
    header_row = HEADER_ROW
    data = _read_data_rows(ws, header_row, mapping)
    if not data:
        _log(f"[경고] '{sheet_name}'에서 유효한 데이터 행이 없습니다.")
        return None
    # 학교코드 순 → 장비 순 정렬
    data.sort(key=lambda x: (x.get("school_code") or "", x.get("mgmt") or ""))
    device_with_times = _assign_times(data)
    out_rows = _build_output_rows(device_with_times, phase_label)
    if not out_rows:
        return None
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = os.path.join(out_dir, f"ISP_원데이터_{phase_label}_{stamp}.xlsx")
    _write_excel(out_rows, out_path, phase_label)
    return out_path


def main() -> None:
    _log("=" * 60)
    _log("ISP 1차/2차 원데이터 생성")
    _log("=" * 60)

    try:
        from openpyxl import load_workbook
    except ImportError:
        _log("[오류] openpyxl이 필요합니다. pip install openpyxl")
        sys.exit(1)

    files = _collect_files()
    if not files:
        path = _input("엑셀 파일 경로를 입력하세요: ")
    else:
        path = _pick_file("기초 데이터 파일(1차/2차 평균 데이터)을 선택하세요.", files)
    if not path or not os.path.isfile(path):
        _log("[오류] 파일을 찾을 수 없습니다.")
        sys.exit(1)

    wb = load_workbook(path, data_only=True)
    _log(f"\n파일: {path}")

    # 1차 시트 선택 (제목행 3행 고정)
    sheet_1 = _pick_sheet(wb, "1차 데이터 시트를 선택하세요.")
    header_row = HEADER_ROW
    _show_header_and_sample(wb[sheet_1], header_row)
    mapping = _read_mapping(wb[sheet_1], header_row)
    if not mapping.get("학교코드") and not mapping.get("장비관리번호"):
        _log("[오류] 학교코드 또는 장비관리번호 열이 필요합니다.")
        wb.close()
        sys.exit(1)

    out_dir = os.path.dirname(path)
    out_dir_custom = _input(f"\n출력 폴더 (Enter: {out_dir}): ")
    if out_dir_custom.strip():
        out_dir = out_dir_custom.strip()

    path_1 = run_phase(wb, sheet_1, mapping, "1차", out_dir)
    if path_1:
        _log(f"1차 원데이터: {path_1}")

    # 2차 시트 선택 (같은 매핑 사용 가능)
    _log("\n2차 데이터도 생성합니다.")
    sheet_2 = _pick_sheet(wb, "2차 데이터 시트를 선택하세요 (같은 시트면 1차와 동일 번호).")
    use_same_mapping = _input("열 매핑을 1차와 동일하게 사용할까요? (Enter: 예): ")
    if use_same_mapping.strip().lower() in ("n", "no", "0"):
        _show_header_and_sample(wb[sheet_2], HEADER_ROW)
        mapping = _read_mapping(wb[sheet_2], HEADER_ROW)
    path_2 = run_phase(wb, sheet_2, mapping, "2차", out_dir)
    if path_2:
        _log(f"2차 원데이터: {path_2}")

    wb.close()
    _log("\n[완료]")


if __name__ == "__main__":
    main()
