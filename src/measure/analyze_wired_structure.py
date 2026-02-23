# -*- coding: utf-8 -*-
"""
유선망 측정 결과 파일 구조 분석

소스 폴더의 첫 번째 엑셀 파일 구조를 출력합니다.
실제 데이터 배치 후 실행하여 시트/열 구조를 확인하세요.
"""
import os
import sys
import glob

_MEASURE_DIR = os.path.dirname(os.path.abspath(__file__))
if _MEASURE_DIR not in sys.path:
    sys.path.insert(0, _MEASURE_DIR)

from wired_preprocess_config import WIRED_1ST_SOURCE

def main():
    for region, source_dir in WIRED_1ST_SOURCE.items():
        print(f"\n=== {region}: {source_dir} ===")
        if not os.path.isdir(source_dir):
            print("  폴더 없음")
            continue
        files = glob.glob(os.path.join(source_dir, "*.xlsx")) + glob.glob(os.path.join(source_dir, "*.xls"))
        if not files:
            print("  엑셀 파일 없음")
            continue
        path = files[0]
        print(f"  샘플: {os.path.basename(path)}")
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            for sn in wb.sheetnames:
                ws = wb[sn]
                print(f"  시트: {sn} | 행:{ws.max_row} 열:{ws.max_column}")
                for r in range(1, min(6, ws.max_row + 1)):
                    row_vals = [str(ws.cell(row=r, column=c).value)[:15] for c in range(1, min(12, ws.max_column + 1))]
                    print(f"    Row{r}: {row_vals}")
            wb.close()
        except Exception as e:
            print(f"  오류: {e}")

if __name__ == "__main__":
    main()
