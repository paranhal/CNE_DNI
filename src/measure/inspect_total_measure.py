# -*- coding: utf-8 -*-
"""TOTAL_MEASURE_LIST_V1.xlsx 시트별 구조 확인 - 학교코드 열 위치"""
import os
import sys
from openpyxl import load_workbook

BASE = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE = os.path.join(BASE, "CNE", "TOTAL_MEASURE_LIST_구조확인.txt")
CNE = os.path.join(BASE, "CNE")
path = os.path.join(CNE, "TOTAL_MEASURE_LIST_V1.xlsx")
if not os.path.isfile(path):
    path = os.path.join(CNE, "TOTAL_MEASURE_LIST_V1.XLSX")

lines = []
def log(s=""):
    lines.append(s)
    print(s)

log("=" * 70)
log("[TOTAL_MEASURE_LIST_V1.xlsx] 시트별 열 구조")
log("학교코드가 A열(1)에 있어야 함. 현재 상태:")
log("=" * 70)
wb = load_workbook(path, data_only=True)
for sn in wb.sheetnames:
    ws = wb[sn]
    headers = []
    for c in range(1, min(12, ws.max_column + 1)):
        v = ws.cell(1, c).value
        headers.append("%r" % (v,) if v else "")
    log("\n--- %s (%d행 x %d열) ---" % (sn, ws.max_row, ws.max_column))
    log("  헤더: " + ", ".join(["%d:%s" % (i, h) for i, h in enumerate(headers, 1)]))
    # A열(1)에 학교코드 있는지
    col1 = ws.cell(1, 1).value
    code_in_a = "학교코드" in str(col1 or "") if col1 else False
    log("  A열(1): %r  -> 학교코드? %s" % (col1, "예" if code_in_a else "아니오"))
    # 2행 샘플 (A열, B열 값)
    if ws.max_row >= 2:
        a2 = ws.cell(2, 1).value
        b2 = ws.cell(2, 2).value
        log("  2행 A=%r B=%r" % (a2, b2))
wb.close()

# 요약: 학교코드 열 위치
log("\n" + "=" * 70)
log("[요약] 학교코드 열 위치 (A열=1이어야 함)")
log("=" * 70)
wb2 = load_workbook(path, data_only=True)
for sn in wb2.sheetnames:
    ws = wb2[sn]
    code_col = None
    for c in range(1, min(6, ws.max_column + 1)):
        v = ws.cell(1, c).value
        if v and "학교코드" in str(v):
            code_col = c
            break
    status = "OK" if code_col == 1 else ">>> A열 아님 (현재 %d열)" % code_col
    log("  %s: 학교코드=%d열 %s" % (sn, code_col or 0, status))
wb2.close()

# 결과를 파일로 저장
with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
    f.write("\n".join(lines))
print("\n[저장] %s" % OUTPUT_FILE)
