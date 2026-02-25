# -*- coding: utf-8 -*-
"""
대전(DNI) 학교별 측정 리포트 생성 (템플릿 사용)

측정값_템플릿.xlsx 복사 → J열 측정값, L열 판정 입력 → 학교명_학교코드.xlsx 저장
"""
from __future__ import print_function
import sys
import io
import os
import csv
import re
import shutil
from datetime import datetime

if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

from openpyxl import load_workbook
from openpyxl.styles import Font
from tqdm import tqdm

_MEASURE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, _MEASURE_DIR)

from dni_school_report_config import (
    TEMPLATE_CANDIDATES,
    TOTAL_MEASURE_LIST,
    OUTPUT_DIR,
    J_OUTPUT_MAP,
    L_JUDGMENT_MAP,
    J_COL,
    L_COL,
    G_COL,
    FONT_BLACK_ROWS,
    JUDGE_BY_V2_ROWS,
    JUDGE_BOTH_ROWS,
    JUDGMENT_ROW_START,
    JUDGMENT_ROW_END,
    LOG_DIR,
    LOG_PREFIX,
    SCHOOL_LIST_FILES,
    ROUND_1_ROWS,
)

SCHOOL_LIST_SEARCH_DIRS = [_MEASURE_DIR, os.path.join(os.path.dirname(_MEASURE_DIR), "split")]

TEMPLATE_SHEET = "문제점분석"


def sanitize_filename(s):
    if not s:
        return ""
    s = str(s).strip()
    for c in r'\/:*?"<>|':
        s = s.replace(c, "_")
    return s[:50]


def find_template():
    for p in TEMPLATE_CANDIDATES:
        if os.path.isfile(p):
            return p
    for template_dir in [os.path.join(_MEASURE_DIR, "측정값_템플릿")]:
        if os.path.isdir(template_dir):
            for f in os.listdir(template_dir):
                if f.endswith((".xlsx", ".xls")):
                    return os.path.join(template_dir, f)
    return None


def load_full_school_list():
    for base_dir in SCHOOL_LIST_SEARCH_DIRS:
        for fname in SCHOOL_LIST_FILES:
            path = os.path.join(base_dir, fname)
            if not os.path.isfile(path):
                continue
            try:
                codes = []
                code_to_name = {}
                if path.endswith(".csv"):
                    with open(path, "r", encoding="utf-8-sig") as f:
                        reader = csv.reader(f)
                        header = next(reader, None)
                        rows = list(reader)
                    code_col = name_col = 0
                    for i, h in enumerate(header or []):
                        s = str(h or "").lower()
                        if "학교코드" in s or "code" in s:
                            code_col = i
                        if "학교명" in s or "name" in s:
                            name_col = i
                    for row in rows:
                        if len(row) > max(code_col, name_col):
                            code = str(row[code_col] or "").strip()
                            name = str(row[name_col] or "").strip()
                            if code:
                                codes.append(code)
                                code_to_name[code] = name
                else:
                    wb = load_workbook(path, read_only=True, data_only=True)
                    ws = wb.active
                    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
                    code_col = name_col = 0
                    for i, h in enumerate(header or []):
                        s = str(h or "").lower()
                        if "학교코드" in s or "code" in s:
                            code_col = i
                        if "학교명" in s or "name" in s:
                            name_col = i
                    for r in range(2, ws.max_row + 1):
                        code = str(ws.cell(r, code_col + 1).value or "").strip()
                        name = str(ws.cell(r, name_col + 1).value or "").strip()
                        if code:
                            codes.append(code)
                            code_to_name[code] = name
                    wb.close()
                if codes:
                    return codes, code_to_name
            except Exception:
                pass
    return [], {}


def load_stats_by_school(wb_stats):
    by_school = {}
    for sheet_name in wb_stats.sheetnames:
        ws = wb_stats[sheet_name]
        code_col = 1
        for c in range(1, min(50, ws.max_column + 1)):
            v = ws.cell(1, c).value
            if v and ("학교코드" in str(v) or "code" in str(v).lower()):
                code_col = c
                break
        for r in range(2, ws.max_row + 1):
            sc = str(ws.cell(r, code_col).value or "").strip()
            if not sc:
                continue
            if sc not in by_school:
                by_school[sc] = {}
            by_school[sc][sheet_name] = r
    return by_school


def get_school_values(wb_stats, school_code, school_data, row_def):
    sheet_name = row_def[1]
    if sheet_name == "fixed":
        return (row_def[2], None)
    if sheet_name == "h_only":
        return (None, None)
    if sheet_name not in wb_stats.sheetnames:
        return None, None
    if sheet_name not in school_data:
        return None, None
    data_row = school_data[sheet_name]
    ws = wb_stats[sheet_name]
    col1, col2 = row_def[2], row_def[3]
    if isinstance(col1, list):
        vals = [ws.cell(data_row, c).value for c in col1]
        return vals, None
    v1 = ws.cell(data_row, col1).value if col1 else None
    v2 = ws.cell(data_row, col2).value if col2 else None
    return v1, v2


def get_numeric(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "").replace("%", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def judge(val, op, threshold, val2=None):
    if op == "always":
        return str(threshold) if threshold else "정상"
    if op == "has_value":
        s = str(val or "").strip()
        return "개선필요" if s else "정상"
    if op == "zero_or_empty_ok":
        s = str(val or "").strip()
        if not s:
            return "정상"
        n = get_numeric(val)
        return "정상" if n is not None and n <= threshold else "개선필요"
    if op == "split_exact":
        s = str(val or "").strip()
        if s == "분리":
            return "정상"
        return "개선필요"
    if op == "both_ge":
        n1 = get_numeric(val)
        n2 = get_numeric(val2) if val2 is not None else None
        if n1 is not None and n1 <= threshold:
            return "개선필요"
        if n2 is not None and n2 <= threshold:
            return "개선필요"
        return "정상"
    if op == "split":
        s = str(val or "").strip()
        return "정상" if "분리" in s else "개선필요"
    if op == "ge_before_keyword":
        s = str(val or "").strip()
        m = re.search(r"(\d+)\s*계위", s)
        if not m:
            return ""
        n = get_numeric(m.group(1))
        if n is None:
            return ""
        return "개선필요" if n >= threshold else "정상"
    n = get_numeric(val)
    if n is None:
        return ""
    if op == "ge":
        return "정상" if n >= threshold else "개선필요"
    if op == "le":
        return "정상" if n <= threshold else "개선필요"
    return ""


def format_value(v):
    if v is None or (isinstance(v, str) and not str(v).strip()):
        return ""
    if isinstance(v, (int, float)):
        return v
    return str(v).strip()


def format_output_value(v1, v2, row_def):
    fmt = row_def[4] if len(row_def) >= 5 else None
    if fmt == "h_only":
        return ""
    if fmt == "fixed":
        return str(row_def[2]) if row_def[2] is not None else ""
    if fmt == "location5":
        g_val = format_value(v1) if v1 is not None else ""
        h_val = format_value(v2) if v2 is not None else ""
        parts = [
            f"위치 5 - {g_val}" if g_val else "위치 5 - ",
            f"위치 5 외 - {h_val}" if h_val else "위치 5 외 - ",
            "보관 및 확인불가 - 없음",
        ]
        return "\n".join(parts)
    if fmt == "n_ac_ax" and isinstance(v1, (list, tuple)):
        d_val = format_value(v1[0]) if len(v1) > 0 else ""
        e_val = format_value(v1[1]) if len(v1) > 1 else ""
        f_val = format_value(v1[2]) if len(v1) > 2 else ""
        parts = [
            f"N - {d_val}식",
            f"AC - {e_val}식",
            f"AX - {f_val}식",
        ]
        return "\n".join(parts)
    if fmt == "cable" and isinstance(v1, (list, tuple)):
        labels = ("SM", "MM", "CAT6", "CAT5e", "CAT5")
        parts = []
        for i, lbl in enumerate(labels):
            val = format_value(v1[i]) if i < len(v1) and v1[i] is not None else ""
            parts.append(f"{lbl}({val})" if val else f"{lbl}()")
        return "\n".join(parts)
    if fmt == "fullload":
        c_val = format_value(v1) if v1 is not None else ""
        d_val = format_value(v2) if v2 is not None else ""
        parts = []
        if c_val:
            parts.append(f"{c_val}(Mbps)")
        if d_val:
            parts.append(f"{d_val}(Mbps)")
        return "\n".join(parts) if parts else ""
    if row_def[3] is not None and not isinstance(row_def[2], list):
        return f"{format_value(v1)} / {format_value(v2)}" if (v1 or v2) else (format_value(v1) or format_value(v2))
    return format_value(v1)


def generate_school_report(template_path, wb_stats, school_code, school_data):
    wb = load_workbook(template_path)
    ws = wb[TEMPLATE_SHEET] if TEMPLATE_SHEET in wb.sheetnames else wb.active
    l_map = {row: (op, th) for row, op, th in L_JUDGMENT_MAP}
    for row_def in J_OUTPUT_MAP:
        row = row_def[0]
        v1, v2 = get_school_values(wb_stats, school_code, school_data, row_def)
        out_val = format_output_value(v1, v2, row_def)
        fmt = row_def[4] if len(row_def) >= 5 else None
        if fmt != "h_only":
            if row in ROUND_1_ROWS and isinstance(out_val, (int, float)):
                out_val = round(out_val, 1)
            cell = ws.cell(row=row, column=J_COL)
            cell.value = out_val
            if row in FONT_BLACK_ROWS:
                cell.font = Font(color="000000")
        if row in l_map:
            op, threshold = l_map[row]
            val_for_judge = v2 if row in JUDGE_BY_V2_ROWS else v1
            if isinstance(val_for_judge, (list, tuple)):
                val_for_judge = val_for_judge[0] if val_for_judge else None
            val2_for_judge = v2 if row in JUDGE_BOTH_ROWS and not isinstance(v1, (list, tuple)) else None
            result = judge(val_for_judge, op, threshold, val2_for_judge)
            if not result and v1 is not None and v2 is not None and not isinstance(v1, (list, tuple)):
                result = judge(v2, op, threshold)
            l_cell = ws.cell(row=row, column=L_COL, value=result)
            if result == "개선필요":
                l_cell.font = Font(color="FF0000")
    ws.cell(row=21, column=G_COL, value="375 Mhz 이상")
    count_정상 = 0
    count_개선필요 = 0
    for r in range(JUDGMENT_ROW_START, JUDGMENT_ROW_END + 1):
        val = ws.cell(row=r, column=L_COL).value
        s = str(val or "").strip()
        if s == "정상":
            count_정상 += 1
        elif s == "개선필요":
            count_개선필요 += 1
    ws.cell(row=36, column=G_COL, value="정상")
    ws.cell(row=36, column=L_COL, value=count_정상)
    ws.cell(row=37, column=G_COL, value="개선필요")
    ws.cell(row=37, column=L_COL, value=count_개선필요)
    return wb


def main():
    print("=" * 50)
    print("[대전(DNI) 학교별 측정 리포트] 생성 (템플릿 사용)")
    print("=" * 50)
    template_path = find_template()
    if not template_path:
        print(f"[오류] 템플릿 없음. 확인: {TEMPLATE_CANDIDATES}")
        sys.exit(1)
    print(f"템플릿: {template_path}")
    if not os.path.isfile(TOTAL_MEASURE_LIST):
        print(f"[오류] 통계 파일 없음: {TOTAL_MEASURE_LIST}")
        sys.exit(1)
    print(f"통계: {TOTAL_MEASURE_LIST}")
    all_schools, code_to_name = load_full_school_list()
    if not all_schools:
        print("[경고] 학교 리스트 없음. 통계에 있는 학교만 처리합니다.")
    wb_stats = load_workbook(TOTAL_MEASURE_LIST, data_only=True)
    by_school = load_stats_by_school(wb_stats)
    if not by_school:
        print("[오류] 학교별 데이터 없음")
        sys.exit(1)
    schools_with_data = set(by_school.keys())
    missing = [sc for sc in all_schools if sc not in schools_with_data]
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    os.makedirs(LOG_DIR, exist_ok=True)
    log_path = os.path.join(LOG_DIR, f"{LOG_PREFIX}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")
    log_lines = [
        f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 시작",
        f"대상 학교: {len(all_schools)}개",
        f"통계 데이터 있음: {len(schools_with_data)}개",
        f"통계 데이터 없음: {len(missing)}개",
    ]
    if missing:
        log_lines.append("")
        log_lines.append("[통계 데이터 없는 학교 (리포트 미생성)]")
        for sc in sorted(missing):
            log_lines.append(f"  {sc}  {code_to_name.get(sc, '')}")
    with open(log_path, "w", encoding="utf-8") as f:
        f.write("\n".join(log_lines))
    print(f"출력: {OUTPUT_DIR}")
    if missing:
        print(f"[로그] 통계 데이터 없는 학교 {len(missing)}개 → {log_path}")
    for school_code in tqdm(sorted(by_school.keys()), desc="학교별 생성", unit="교"):
        school_name = code_to_name.get(school_code, "")
        safe_name = sanitize_filename(school_name) or school_code
        school_data = by_school[school_code]
        wb = generate_school_report(template_path, wb_stats, school_code, school_data)
        out_name = f"{safe_name}_{school_code}.xlsx"
        out_path = os.path.join(OUTPUT_DIR, out_name)
        try:
            wb.save(out_path)
        except PermissionError:
            out_path = os.path.join(OUTPUT_DIR, f"{safe_name}_{school_code}_백업.xlsx")
            wb.save(out_path)
    wb_stats.close()
    print(f"[완료] {len(by_school)}개 학교 리포트 생성")

    # 지역별 폴더 복사
    copy_reports_by_region(wb_stats_path=TOTAL_MEASURE_LIST, output_dir=OUTPUT_DIR)


def load_region_map(wb_stats_path):
    """Sheet1에서 학교코드 → 지역 매핑 로드"""
    wb = load_workbook(wb_stats_path, data_only=True, read_only=True)
    code_to_region = {}
    if "Sheet1" in wb.sheetnames:
        ws = wb["Sheet1"]
        for r in range(2, ws.max_row + 1):
            code = str(ws.cell(r, 1).value or "").strip()
            region = str(ws.cell(r, 2).value or "").strip()
            if code and region:
                code_to_region[code] = region
    wb.close()
    return code_to_region


def copy_reports_by_region(wb_stats_path, output_dir):
    """생성된 리포트를 지역별 하위 폴더로 복사"""
    print("\n[지역별 폴더 복사] 시작")
    code_to_region = load_region_map(wb_stats_path)
    if not code_to_region:
        print("[경고] 지역 매핑 데이터 없음 (Sheet1). 복사 건너뜀.")
        return

    # 기존 지역 폴더 삭제 (재실행 시 중복 방지)
    for d in os.listdir(output_dir):
        dp = os.path.join(output_dir, d)
        if os.path.isdir(dp):
            shutil.rmtree(dp)

    files = [f for f in os.listdir(output_dir) if f.endswith(".xlsx")]
    region_files = {}
    no_region = 0
    for fname in files:
        parts = fname.rsplit("_", 1)
        if len(parts) < 2:
            continue
        school_code = parts[-1].replace(".xlsx", "").replace("_백업", "")
        region = code_to_region.get(school_code, "")
        if not region:
            no_region += 1
            continue
        region_files.setdefault(region, []).append(fname)

    copied = 0
    for region, fnames in sorted(region_files.items()):
        folder_name = f"{region}({len(fnames)})"
        region_dir = os.path.join(output_dir, folder_name)
        os.makedirs(region_dir, exist_ok=True)
        for fname in fnames:
            shutil.copy2(os.path.join(output_dir, fname), os.path.join(region_dir, fname))
            copied += 1

    print(f"[완료] {copied}개 파일 → {len(region_files)}개 지역 폴더 복사")
    for region in sorted(region_files):
        print(f"  {region}({len(region_files[region])})")
    if no_region:
        print(f"  (지역 매핑 없음: {no_region}개 파일 미복사)")


if __name__ == "__main__":
    main()
