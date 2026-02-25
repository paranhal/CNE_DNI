# -*- coding: utf-8 -*-
"""
DNI_FULLLOAD_MEASURE.xlsx 전부하_통합 시트 - 최종 측정값 선정

[열 구조]
- B(2): 학교명
- C(3)~H(8): 최종 측정값 (출력)
- I(9)~N(14): 1차 측정 (N=진단결과: 양호/미흡)
- O(15)~T(20): 2차 측정 (T=진단결과: 양호/미흡)

[선정 로직 - 학교별]
1. 1차/2차 측정 장비 수 중 큰 값 = 측정 장비수 (최대 10)
2. 1차_양호 수 == 측정 장비수 → 1차 전체 사용
3. 1차_양호 + 2차_양호 == 측정 장비수 → 1차_양호 + 2차_양호
4. 1차_양호 + 2차_양호 > 측정 장비수 → 1차_양호 + 2차_양호(잘라서 맞춤)
5. 1차_양호 + 2차_양호 < 측정 장비수 → 1차_양호 + 2차_양호 + 1차_미흡(부족분 채움)
"""

from __future__ import annotations

import os
from collections import defaultdict

from openpyxl import load_workbook


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DNI_DIR = os.path.join(BASE_DIR, "DNI")
INPUT_PATH = os.path.join(DNI_DIR, "DNI_FULLLOAD_MEASURE.xlsx")
MAX_DEVICES = 10


def _log(msg: str) -> None:
    print(msg, flush=True)


def main() -> None:
    if not os.path.isfile(INPUT_PATH):
        _log(f"[오류] 파일이 없습니다: {INPUT_PATH}")
        return

    _log("=" * 60)
    _log("[DNI 전부하] 최종 측정값 선정")
    _log(f"대상: {INPUT_PATH}")
    _log("=" * 60)

    wb = load_workbook(INPUT_PATH, data_only=False)
    ws = wb["전부하_통합"] if "전부하_통합" in wb.sheetnames else wb[wb.sheetnames[0]]
    max_row = ws.max_row
    _log(f"시트: {ws.title} | 행 수: {max_row}")

    COL_B = 2
    # 최종 출력 (C~H = 3~8)
    DST = (3, 4, 5, 6, 7, 8)
    # 1차 측정 (I~N = 9~14), N(14)=진단결과
    SRC_1ST = (9, 10, 11, 12, 13, 14)
    COL_N = 14
    # 2차 측정 (O~T = 15~20), T(20)=진단결과
    SRC_2ND = (15, 16, 17, 18, 19, 20)
    COL_T = 20

    def _s(v):
        return (str(v).strip() if v is not None else "")

    # ── 1단계: C~H 비우기 ──
    for r in range(2, max_row + 1):
        for c in DST:
            ws.cell(row=r, column=c).value = None
    _log("[1단계] C~H 비움")

    # ── 2단계: 학교별 행 수집 ──
    school_rows = defaultdict(list)
    for r in range(2, max_row + 1):
        name = _s(ws.cell(row=r, column=COL_B).value)
        if name:
            school_rows[name].append(r)
    _log(f"[2단계] {len(school_rows)}개 학교 수집")

    # ── 3단계: 학교별 최종 측정값 선정 ──
    total_written = 0
    over_limit_schools = []

    for school, rows in school_rows.items():
        rows_sorted = sorted(rows)

        # 1차/2차 데이터 수집
        data_1st_good = []   # 1차 양호: [(row, (i,j,k,l,m,n)), ...]
        data_1st_bad = []    # 1차 미흡
        data_2nd_good = []   # 2차 양호
        count_1st = 0
        count_2nd = 0

        for r in rows_sorted:
            # 1차 데이터 확인
            vals_1st = tuple(ws.cell(row=r, column=c).value for c in SRC_1ST)
            has_1st = any(v is not None for v in vals_1st)
            if has_1st:
                count_1st += 1
                judge_1st = _s(ws.cell(row=r, column=COL_N).value)
                if judge_1st == "양호":
                    data_1st_good.append((r, vals_1st))
                else:
                    data_1st_bad.append((r, vals_1st))

            # 2차 데이터 확인
            vals_2nd = tuple(ws.cell(row=r, column=c).value for c in SRC_2ND)
            has_2nd = any(v is not None for v in vals_2nd)
            if has_2nd:
                count_2nd += 1
                judge_2nd = _s(ws.cell(row=r, column=COL_T).value)
                if judge_2nd == "양호":
                    data_2nd_good.append((r, vals_2nd))

        # 측정 장비수 = max(1차, 2차), 최대 10
        device_count = max(count_1st, count_2nd)
        if device_count > MAX_DEVICES:
            over_limit_schools.append((school, device_count))
            device_count = MAX_DEVICES

        n_1g = len(data_1st_good)
        n_2g = len(data_2nd_good)
        n_1b = len(data_1st_bad)

        # 최종 데이터 선정
        final_data: list[tuple] = []

        if n_1g >= device_count:
            # Case 1: 1차 양호만으로 충분
            final_data = [v for _, v in data_1st_good[:device_count]]
        elif n_1g + n_2g == device_count:
            # Case 2: 1차 양호 + 2차 양호 = 딱 맞음
            final_data = [v for _, v in data_1st_good]
            final_data += [v for _, v in data_2nd_good]
        elif n_1g + n_2g > device_count:
            # Case 3: 1차 양호 + 2차 양호 > 장비수 → 2차 양호 잘라서 맞춤
            need_from_2nd = device_count - n_1g
            final_data = [v for _, v in data_1st_good]
            final_data += [v for _, v in data_2nd_good[:need_from_2nd]]
        else:
            # Case 4: 1차 양호 + 2차 양호 < 장비수 → 1차 미흡으로 채움
            final_data = [v for _, v in data_1st_good]
            final_data += [v for _, v in data_2nd_good]
            need_more = device_count - len(final_data)
            final_data += [v for _, v in data_1st_bad[:need_more]]

        # C~H에 기록
        for i, vals in enumerate(final_data):
            if i >= len(rows_sorted):
                break
            target_row = rows_sorted[i]
            for ci, dst_c in enumerate(DST):
                ws.cell(row=target_row, column=dst_c).value = vals[ci]
            total_written += 1

    _log(f"[3단계] 최종 측정값 선정 완료: {total_written}행")

    if over_limit_schools:
        _log(f"[경고] 측정 장비수 {MAX_DEVICES}개 초과 학교:")
        for name, cnt in over_limit_schools:
            _log(f"  {name}: {cnt}개 → {MAX_DEVICES}개로 제한")

    # 샘플 확인
    for school in list(school_rows.keys())[:5]:
        rows = school_rows[school]
        filled = sum(1 for r in rows if ws.cell(row=r, column=DST[0]).value is not None)
        _log(f"  {school}: 행={len(rows)}, C~H채움={filled}")

    # ── 4단계: 학교별 평균 → "전부하측정_학교별평균" 시트 생성 ──
    THRESHOLD = 375
    SHEET_AVG = "전부하측정_학교별평균"

    if SHEET_AVG in wb.sheetnames:
        del wb[SHEET_AVG]
    ws_avg = wb.create_sheet(SHEET_AVG)

    # 헤더
    avg_headers = ["학교코드", "학교명", "다운로드", "업로드", "*RSSI", "CH", "다운로드 진단", "업로드 진단"]
    for ci, h in enumerate(avg_headers, start=1):
        ws_avg.cell(row=1, column=ci).value = h

    COL_C = 3   # 장비관리번호 (학교코드 추출)
    COL_D = 4   # Down
    COL_E = 5   # Up
    COL_F = 6   # *RSSI
    COL_G = 7   # CH

    avg_row = 2
    for school, rows in school_rows.items():
        rows_sorted = sorted(rows)
        dl_vals = []
        ul_vals = []
        rssi_vals = []
        ch_vals = []
        school_code = ""

        for r in rows_sorted:
            mgmt = ws.cell(row=r, column=COL_C).value
            if mgmt is None:
                continue
            if not school_code:
                school_code = _s(mgmt)[:12]
            d = ws.cell(row=r, column=COL_D).value
            e = ws.cell(row=r, column=COL_E).value
            f = ws.cell(row=r, column=COL_F).value
            g = ws.cell(row=r, column=COL_G).value
            if isinstance(d, (int, float)):
                dl_vals.append(d)
            if isinstance(e, (int, float)):
                ul_vals.append(e)
            if isinstance(f, (int, float)):
                rssi_vals.append(f)
            if isinstance(g, (int, float)):
                ch_vals.append(g)

        if not dl_vals and not ul_vals:
            continue

        avg_dl = sum(dl_vals) / len(dl_vals) if dl_vals else 0
        avg_ul = sum(ul_vals) / len(ul_vals) if ul_vals else 0
        avg_rssi = sum(rssi_vals) / len(rssi_vals) if rssi_vals else 0
        avg_ch = sum(ch_vals) / len(ch_vals) if ch_vals else 0

        diag_dl = "양호" if avg_dl >= THRESHOLD else "미흡"
        diag_ul = "양호" if avg_ul >= THRESHOLD else "미흡"

        ws_avg.cell(row=avg_row, column=1).value = school_code
        ws_avg.cell(row=avg_row, column=2).value = school
        ws_avg.cell(row=avg_row, column=3).value = round(avg_dl, 2)
        ws_avg.cell(row=avg_row, column=4).value = round(avg_ul, 2)
        ws_avg.cell(row=avg_row, column=5).value = round(avg_rssi, 2)
        ws_avg.cell(row=avg_row, column=6).value = round(avg_ch, 2)
        ws_avg.cell(row=avg_row, column=7).value = diag_dl
        ws_avg.cell(row=avg_row, column=8).value = diag_ul
        avg_row += 1

    _log(f"[4단계] '{SHEET_AVG}' 시트 생성: {avg_row - 2}개 학교 평균")

    from datetime import datetime
    out_path = INPUT_PATH
    try:
        wb.save(out_path)
    except PermissionError:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = os.path.join(DNI_DIR, f"DNI_FULLLOAD_MEASURE_final_{stamp}.xlsx")
        _log(f"[경고] 원본 열림 → 대신 저장: {out_path}")
        wb.save(out_path)
    wb.close()
    _log(f"[완료] 저장: {out_path}")


if __name__ == "__main__":
    main()
