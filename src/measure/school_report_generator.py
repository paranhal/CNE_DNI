# -*- coding: utf-8 -*-
"""
학교별 측정 리포트 생성 (템플릿 사용)

측정값_템플릿.xlsx 복사 → J열 측정값, L열 판정 입력 → 학교명_학교코드.xlsx 저장
대상 719개 학교 중 통계 데이터 없는 학교는 로그에 기록
"""
from __future__ import print_function
import sys
import io
import os
import csv
import re
from datetime import datetime

if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

from openpyxl import load_workbook
from openpyxl.styles import Font
from tqdm import tqdm

_MEASURE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, _MEASURE_DIR)

from school_report_config import (
    TEMPLATE_CANDIDATES,
    TOTAL_MEASURE_LIST,
    OUTPUT_DIR,
    J_OUTPUT_MAP,
    L_JUDGMENT_MAP,
    J_COL,
    L_COL,
    LOG_DIR,
    LOG_PREFIX,
)

SCHOOL_LIST_SEARCH_DIRS = [_MEASURE_DIR, os.path.join(os.path.dirname(_MEASURE_DIR), "split")]

# 템플릿 시트명 (최종: 문제점분석)
TEMPLATE_SHEET = "문제점분석"


def sanitize_filename(s):
    """파일명에 사용 불가 문자 제거"""
    if not s:
        return ""
    s = str(s).strip()
    for c in r'\/:*?"<>|':
        s = s.replace(c, "_")
    return s[:50]


def find_template():
    for p in TEMPLATE_CANDIDATES:
        if os.path.isfile(p):
            return p
    for template_dir in [os.path.join(_MEASURE_DIR, "측정값_템플릿"), os.path.join(_MEASURE_DIR, "측정밗_템플릿")]:
        if os.path.isdir(template_dir):
            for f in os.listdir(template_dir):
                if f.endswith((".xlsx", ".xls")):
                    return os.path.join(template_dir, f)
    return None


def load_full_school_list():
    """719개 대상 학교 리스트 로드"""
    for base_dir in SCHOOL_LIST_SEARCH_DIRS:
        for fname in ["school_reg_list_CNE.xlsx", "school_reg_list_CNE.csv", "SCHOOL_REG_LIST_CNE.xlsx"]:
            path = os.path.join(base_dir, fname)
            if not os.path.isfile(path):
                continue
            try:
                codes = []
                code_to_name = {}
                if path.endswith(".csv"):
                    with open(path, "r", encoding="utf-8-sig") as f:
                        reader = csv.reader(f)
                        header = next(reader, None)
                        rows = list(reader)
                    code_col = name_col = 0
                    for i, h in enumerate(header or []):
                        s = str(h or "").lower()
                        if "학교코드" in s or "code" in s:
                            code_col = i
                        if "학교명" in s or "name" in s:
                            name_col = i
                    for row in rows:
                        if len(row) > max(code_col, name_col):
                            code = str(row[code_col] or "").strip()
                            name = str(row[name_col] or "").strip()
                            if code:
                                codes.append(code)
                                code_to_name[code] = name
                else:
                    wb = load_workbook(path, read_only=True, data_only=True)
                    ws = wb.active
                    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
                    code_col = name_col = 0
                    for i, h in enumerate(header or []):
                        s = str(h or "").lower()
                        if "학교코드" in s or "code" in s:
                            code_col = i
                        if "학교명" in s or "name" in s:
                            name_col = i
                    for r in range(2, ws.max_row + 1):
                        code = str(ws.cell(r, code_col + 1).value or "").strip()
                        name = str(ws.cell(r, name_col + 1).value or "").strip()
                        if code:
                            codes.append(code)
                            code_to_name[code] = name
                    wb.close()
                if codes:
                    return codes, code_to_name
            except Exception:
                pass
    return [], {}


def load_stats_by_school(wb_stats):
    """통계 워크북에서 학교코드별 데이터 행 로드"""
    by_school = {}
    for sheet_name in wb_stats.sheetnames:
        ws = wb_stats[sheet_name]
        code_col = 1
        for c in range(1, min(6, ws.max_column + 1)):
            v = ws.cell(1, c).value
            if v and ("학교코드" in str(v) or "code" in str(v).lower()):
                code_col = c
                break
        for r in range(2, ws.max_row + 1):
            sc = str(ws.cell(r, code_col).value or "").strip()
            if not sc:
                continue
            if sc not in by_school:
                by_school[sc] = {}
            by_school[sc][sheet_name] = r
    return by_school


def get_school_values(wb_stats, school_code, school_data, row_def):
    """한 학교의 특정 출력행에 대한 값 조회
    school_data = by_school[school_code] = {시트명: 행번호}
    """
    sheet_name = row_def[1]
    if sheet_name not in wb_stats.sheetnames:
        return None, None
    if sheet_name not in school_data:
        return None, None
    data_row = school_data[sheet_name]
    ws = wb_stats[sheet_name]
    col1, col2 = row_def[2], row_def[3]
    v1 = ws.cell(data_row, col1).value if col1 else None
    v2 = ws.cell(data_row, col2).value if col2 else None
    return v1, v2


def get_numeric(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "").replace("%", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def judge(val, op, threshold):
    n = get_numeric(val)
    if n is None:
        return ""
    if op == "ge":
        return "정상" if n >= threshold else "개선필요"
    if op == "le":
        return "정상" if n <= threshold else "개선필요"
    return ""


def format_value(v):
    if v is None or (isinstance(v, str) and not str(v).strip()):
        return ""
    if isinstance(v, (int, float)):
        return v
    return str(v).strip()


def generate_school_report(template_path, wb_stats, school_code, school_data):
    """템플릿 복사 후 F열(측정값), H열(평가결과)에 데이터 입력"""
    wb = load_workbook(template_path)
    ws = wb[TEMPLATE_SHEET] if TEMPLATE_SHEET in wb.sheetnames else wb.active
    l_map = {row: (op, th) for row, op, th in L_JUDGMENT_MAP}
    for row_def in J_OUTPUT_MAP:
        row = row_def[0]
        v1, v2 = get_school_values(wb_stats, school_code, school_data, row_def)
        if row_def[3] is not None:
            out_val = f"{format_value(v1)} / {format_value(v2)}" if (v1 or v2) else (format_value(v1) or format_value(v2))
        else:
            out_val = format_value(v1)
        cell = ws.cell(row=row, column=J_COL)
        cell.value = out_val
        # R28~R30(집선ISP): 측정값 폰트 검정색
        if row in (28, 29, 30):
            cell.font = Font(color="000000")
            cell.font = Font(color="000000")
        if row in l_map:
            op, threshold = l_map[row]
            val_for_judge = v1 if v1 is not None else v2
            result = judge(val_for_judge, op, threshold)
            if not result and v1 is not None and v2 is not None:
                result = judge(v2, op, threshold)
            l_cell = ws.cell(row=row, column=L_COL, value=result)
            if result == "개선필요":
                l_cell.font = Font(color="FF0000")
    return wb


def main():
    print("=" * 50)
    print("[학교별 측정 리포트] 생성 (템플릿 사용)")
    print("=" * 50)
    template_path = find_template()
    if not template_path:
        print(f"[오류] 템플릿 없음. 확인: {TEMPLATE_CANDIDATES}")
        sys.exit(1)
    print(f"템플릿: {template_path}")
    if not os.path.isfile(TOTAL_MEASURE_LIST):
        print(f"[오류] 통계 파일 없음: {TOTAL_MEASURE_LIST}")
        sys.exit(1)
    print(f"통계: {TOTAL_MEASURE_LIST}")
    all_schools, code_to_name = load_full_school_list()
    if not all_schools:
        print("[경고] 학교 리스트 없음. 통계에 있는 학교만 처리합니다.")
    wb_stats = load_workbook(TOTAL_MEASURE_LIST, data_only=True)
    by_school = load_stats_by_school(wb_stats)
    if not by_school:
        print("[오류] 학교별 데이터 없음")
        sys.exit(1)
    schools_with_data = set(by_school.keys())
    missing = [sc for sc in all_schools if sc not in schools_with_data]
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    os.makedirs(LOG_DIR, exist_ok=True)
    log_path = os.path.join(LOG_DIR, f"{LOG_PREFIX}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")
    log_lines = [
        f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 시작",
        f"대상 학교: {len(all_schools)}개",
        f"통계 데이터 있음: {len(schools_with_data)}개",
        f"통계 데이터 없음: {len(missing)}개",
    ]
    if missing:
        log_lines.append("")
        log_lines.append("[통계 데이터 없는 학교 (리포트 미생성)]")
        for sc in sorted(missing):
            log_lines.append(f"  {sc}  {code_to_name.get(sc, '')}")
    with open(log_path, "w", encoding="utf-8") as f:
        f.write("\n".join(log_lines))
    print(f"출력: {OUTPUT_DIR}")
    if missing:
        print(f"[로그] 통계 데이터 없는 학교 {len(missing)}개 → {log_path}")
    for school_code in tqdm(sorted(by_school.keys()), desc="학교별 생성", unit="교"):
        school_name = code_to_name.get(school_code, "")
        safe_name = sanitize_filename(school_name) or school_code
        school_data = by_school[school_code]
        wb = generate_school_report(template_path, wb_stats, school_code, school_data)
        out_name = f"{safe_name}_{school_code}.xlsx"
        out_path = os.path.join(OUTPUT_DIR, out_name)
        try:
            wb.save(out_path)
        except PermissionError:
            out_path = os.path.join(OUTPUT_DIR, f"{safe_name}_{school_code}_백업.xlsx")
            wb.save(out_path)
    wb_stats.close()
    print(f"[완료] {len(by_school)}개 학교 리포트 생성")


if __name__ == "__main__":
    main()
