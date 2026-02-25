# -*- coding: utf-8 -*-
"""
DNO_FULLLOAD_MEANSURE_수정.xlsx 전부하_통합 시트 2차 결과 복사 (관리번호 열 추가 버전)

[열 구조 - 관리번호 열 추가]
- B열(2): 1차 학교명
- V열(22): 관리번호 (2차 장비관리번호, 여기서 읽기)
- W열(23): 학교코드 (관리번호 앞 12자리 → 기록)
- X열(24): 학교명 (학교코드로 CSV 조회 → 기록)
- Y~AC(25~29): 2차 측정값 (Down, Up, RSSI, CH, 진단)
- N(14)=장비관리번호, O~R(15~18)=2차 측정값 (1차 행에 복사)

[동작]
- V열(22) 관리번호 앞 12자리 → 학교코드 추출 → W열(23) 기록.
- school_reg_list_DNI.csv로 학교코드→학교명 조회 → X열(24) 기록.
- B열(1차 학교명)과 X열(2차 학교명) 매칭 후, 같은 학교의 2차 데이터(Y~AC)를 1차 행의 N~R에 복사.
"""

from __future__ import annotations

import csv
import os
from collections import defaultdict

from openpyxl import load_workbook


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DNI_DIR = os.path.join(BASE_DIR, "DNI")
SPLIT_DIR = os.path.join(os.path.dirname(BASE_DIR), "split")
INPUT_PATH = os.path.join(DNI_DIR, "DNO_FULLLOAD_MEANSURE_수정.xlsx")
SCHOOL_LIST_CSV = os.path.join(SPLIT_DIR, "school_reg_list_DNI.csv")
SCHOOL_CODE_LEN = 12


def _log(msg: str) -> None:
    print(msg, flush=True)


def load_code_to_name(path: str) -> dict[str, str]:
    """school_reg_list_DNI.csv에서 학교코드 → 학교명 로드"""
    code_to_name: dict[str, str] = {}
    if not os.path.isfile(path):
        _log(f"[경고] 학교 리스트 없음: {path}")
        return code_to_name
    with open(path, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        header = next(reader, None)
        if not header:
            return code_to_name
        code_idx = name_idx = 0
        for i, h in enumerate(header):
            s = (h or "").strip()
            if "학교코드" in s or "code" in s:
                code_idx = i
            if "학교명" in s or "name" in s:
                name_idx = i
        for row in reader:
            if len(row) <= max(code_idx, name_idx):
                continue
            code = (row[code_idx] or "").strip()
            name = (row[name_idx] or "").strip()
            if code:
                code_to_name[code] = name
                code_to_name[code[:SCHOOL_CODE_LEN]] = name
    _log(f"[학교 리스트] {len(code_to_name)}개 로드: {path}")
    return code_to_name


def main() -> None:
    if not os.path.isfile(INPUT_PATH):
        _log(f"[오류] 파일이 없습니다: {INPUT_PATH}")
        return

    _log("=" * 60)
    _log("[DNI 전부하 v2] V열 관리번호 → W학교코드·X학교명, B-X 매칭 → N~R 복사")
    _log(f"대상 파일: {INPUT_PATH}")
    _log("=" * 60)

    wb = load_workbook(INPUT_PATH, data_only=False)
    sheetnames = wb.sheetnames
    ws = wb["전부하_통합"] if "전부하_통합" in sheetnames else wb[sheetnames[0]]

    max_row = ws.max_row
    _log(f"시트: {ws.title} | 행 수: {max_row}")

    # 관리번호 열 추가 기준
    COL_NAME_1ST = 2    # B열: 학교명(1차)
    COL_MGMT = 22       # V열(22): 관리번호 (2차, 여기서 읽기)
    COL_W = 23          # W열(23): 학교코드 (기록)
    COL_X = 24          # X열(24): 학교명 (기록, B열과 매칭)
    COL_DOWN, COL_UP, COL_RSSI, COL_CH, COL_DIAG = 25, 26, 27, 28, 29  # Y~AC: 2차 측정값
    COL_N, COL_O, COL_P, COL_Q, COL_R = 14, 15, 16, 17, 18

    def _norm(s):
        return (str(s).strip() if s is not None else "") or ""

    code_to_name = load_code_to_name(SCHOOL_LIST_CSV)
    if not code_to_name:
        _log("[오류] school_reg_list_DNI 로드 실패. 진행 중단.")
        wb.close()
        return

    # 1단계: N~R 비우기
    for r in range(2, max_row + 1):
        for c in (COL_N, COL_O, COL_P, COL_Q, COL_R):
            ws.cell(row=r, column=c, value=None)
    _log("[1단계] N~R 전부 비움")

    # 2단계: 1차 행(B열), 2차 행(V열 관리번호→W학교코드·X학교명) 수집
    rows_by_school_1st = defaultdict(list)
    rows_by_school_2nd = defaultdict(list)

    for r in range(2, max_row + 1):
        b_name = _norm(ws.cell(row=r, column=COL_NAME_1ST).value)
        if b_name:
            rows_by_school_1st[b_name].append(r)

    filled_wx = 0
    for r in range(2, max_row + 1):
        mgmt = ws.cell(row=r, column=COL_MGMT).value  # V열 관리번호
        mgmt_str = _norm(mgmt)
        if not mgmt_str or len(mgmt_str) < SCHOOL_CODE_LEN:
            continue
        school_code = mgmt_str[:SCHOOL_CODE_LEN]
        school_name = code_to_name.get(school_code) or code_to_name.get(school_code.strip())
        if not school_name:
            continue
        ws.cell(row=r, column=COL_W, value=school_code)
        ws.cell(row=r, column=COL_X, value=school_name)
        filled_wx += 1
        v_down = ws.cell(row=r, column=COL_DOWN).value
        v_up = ws.cell(row=r, column=COL_UP).value
        v_rssi = ws.cell(row=r, column=COL_RSSI).value
        v_ch = ws.cell(row=r, column=COL_CH).value
        v_diag = ws.cell(row=r, column=COL_DIAG).value
        if all(v is None for v in (v_down, v_up, v_rssi, v_ch, v_diag)):
            continue
        rows_by_school_2nd[school_name].append((r, (mgmt_str, v_down, v_up, v_rssi, v_ch, v_diag)))
    _log(f"[2단계] V열 관리번호 기준 W·X 채움: {filled_wx}행 (2차 측정값 있는 행: {sum(len(v) for v in rows_by_school_2nd.values())}행)")

    for name in rows_by_school_1st:
        rows_by_school_1st[name].sort()
    for name in rows_by_school_2nd:
        rows_by_school_2nd[name].sort(key=lambda x: x[0])

    set_1st = set(rows_by_school_1st.keys())
    set_2nd = set(rows_by_school_2nd.keys())
    common = set_1st & set_2nd
    _log(f"[매칭] 1차 학교 수: {len(set_1st)}, 2차 학교 수: {len(set_2nd)}, 동일명: {len(common)}")
    if len(common) == 0 and (set_1st or set_2nd):
        _log(f"  [참고] 1차 샘플: {list(set_1st)[:3]}")
        _log(f"  [참고] 2차 샘플: {list(set_2nd)[:3]}")

    updated = 0
    for school, first_rows in rows_by_school_1st.items():
        second_list = rows_by_school_2nd.get(school, [])
        for i, row_1st in enumerate(first_rows):
            if i >= len(second_list):
                break
            _, (mgmt_no, v_down, v_up, v_rssi, v_ch, v_diag) = second_list[i]
            ws.cell(row=row_1st, column=COL_N, value=mgmt_no)   # N열: 장비관리번호
            ws.cell(row=row_1st, column=COL_O, value=v_down)
            ws.cell(row=row_1st, column=COL_P, value=v_up)
            ws.cell(row=row_1st, column=COL_Q, value=v_rssi)
            ws.cell(row=row_1st, column=COL_R, value=v_diag)
            updated += 1

    from datetime import datetime
    out_path = INPUT_PATH
    try:
        wb.save(out_path)
    except PermissionError:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = os.path.join(DNI_DIR, f"DNO_FULLLOAD_MEANSURE_수정_merged_{stamp}.xlsx")
        _log(f"[경고] 원본이 열려 있어 대신 저장: {out_path}")
        wb.save(out_path)
    wb.close()

    _log(f"[완료] N=장비관리번호·O~R=측정값 복사: {updated}행")


if __name__ == "__main__":
    main()
