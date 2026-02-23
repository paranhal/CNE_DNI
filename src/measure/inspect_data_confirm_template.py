# -*- coding: utf-8 -*-
"""최종_데이터확인_템플릿.xlsx 구조 확인"""
import os
from openpyxl import load_workbook

BASE = os.path.dirname(os.path.abspath(__file__))
path = os.path.join(BASE, "최종_데이터확인_템플릿.xlsx")
out_path = os.path.join(BASE, "CNE", "data_confirm_template_structure.txt")
lines = []
if not os.path.isfile(path):
    lines.append("파일 없음: " + path)
else:
    wb = load_workbook(path, data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        lines.append("시트: %s %dx%d" % (sn, ws.max_row, ws.max_column))
        for r in range(1, min(5, ws.max_row + 1)):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                lines.append("  C%d: %s" % (c, str(v or "")[:40]))
    wb.close()
with open(out_path, "w", encoding="utf-8") as f:
    f.write("\n".join(lines))
print("Saved to", out_path)
