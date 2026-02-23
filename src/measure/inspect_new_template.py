# -*- coding: utf-8 -*-
"""최종_측정값_템플릿.xlsx 구조 확인 - 셀번호, 측정값/판정 위치"""
import os
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
path = os.path.join(BASE, "최종_측정값_템플릿.xlsx")
if not os.path.isfile(path):
    print("파일 없음:", path)
    exit(1)

wb = load_workbook(path, data_only=True)
print("시트:", wb.sheetnames)
# 셀참조 패턴: J30, L30, J32, L32 등 (알파벳+숫자)
cell_ref = re.compile(r"^([A-Z]+)(\d+)$", re.I)
for sn in wb.sheetnames:
    ws = wb[sn]
    print("\n=== %s (%d행 x %d열) ===" % (sn, ws.max_row, ws.max_column))
    # 전체 스캔: 셀번호 형태(J30, L32 등) 또는 분류코드(A20, A22 등) 찾기
    found = []
    for r in range(1, min(55, ws.max_row + 1)):
        for c in range(1, min(20, ws.max_column + 1)):
            v = ws.cell(r, c).value
            if v is not None:
                s = str(v).strip()
                # 셀참조 (J30, L32) 또는 분류코드 (A20, A22)
                if re.match(r"^[A-Z]+\d+$", s, re.I) or (len(s) <= 4 and s.startswith("A") and s[1:].isdigit()):
                    coord = get_column_letter(c) + str(r)
                    found.append((r, c, coord, s))
    for r, c, coord, s in found[:60]:
        print("  %s(행%d열%d)=%r" % (coord, r, c, s))
    # 헤더행 1~5
    print("\n헤더(1~3행):")
    for r in range(1, 4):
        row = [str(ws.cell(r, c).value or "")[:12] for c in range(1, 14)]
        print("  R%d: %s" % (r, row))
wb.close()
