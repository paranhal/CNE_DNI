# -*- coding: utf-8 -*-
"""
학교별 장비 현황 엑셀 분리 통합 스크립트 (V1)
- --AP / --switch / --security / --poe: 장비 선택
- --DNI / --CNE: 지역 선택
- 원본 경로·시트 규칙: split_config.py 참조
- 로그: split_log_{장비}_{지역}_{날짜}.csv
"""
import os
import csv
import re
import argparse
from datetime import datetime
from copy import copy
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side
from openpyxl import Workbook
from school_utils import (
    extract_school_code_from_mgmt_num, find_mgmt_col, find_school_code_col,
    get_output_cols, sort_schools_by_region, get_school_list_path,
)
from split_config import (
    BASE_DIR, get_source_path, get_sheet_candidates,
    OUTPUT_BASE_BY_REGION, OUTPUT_BASE_TEST,
)

MISSED_SCHOOLS_FILE = os.path.join(BASE_DIR, "missed_schools.csv")

# 장비별 처리/출력 설정 (원본 경로·시트는 split_config에서)
EQUIPMENT_CONFIG = {
    "AP": {
        "log_name": "AP",
        "device_label": "AP",
        "sheet_title": "AP 장비 현황",
        "code_from": "school_or_mgmt",
        "title_mode": "ap_b1",
        "use_border": True,
    },
    "switch": {
        "log_name": "switch",
        "device_label": "스위치",
        "sheet_title": "스위치 장비 현황",
        "code_from": "mgmt_only",
        "title_mode": "copy_a1",
        "use_border": False,
    },
    "security": {
        "log_name": "security",
        "device_label": "보안장비",
        "sheet_title": "보안장비 장비 현황",
        "code_from": "mgmt_only",
        "title_mode": "copy_a1",
        "use_border": False,
    },
    "poe": {
        "log_name": "poe",
        "device_label": "POE",
        "sheet_title": "POE 장비 현황",
        "code_from": "mgmt_only",
        "title_mode": "copy_a1",
        "use_border": False,
    },
}

THIN_BORDER = Border(
    left=Side(border_style='thin'), right=Side(border_style='thin'),
    top=Side(border_style='thin'), bottom=Side(border_style='thin')
)


def get_log_path(equipment, region_key, suffix=None):
    """split_log_{장비}_{지역}_{날짜}.csv"""
    cfg = EQUIPMENT_CONFIG.get(equipment, {})
    log_name = cfg.get("log_name", equipment)
    base = f"split_log_{log_name}_{region_key}_{datetime.now().strftime('%Y%m%d')}"
    if suffix:
        return os.path.join(BASE_DIR, f"{base}_{suffix}.csv")
    return os.path.join(BASE_DIR, f"{base}.csv")


def get_processed_school_codes(equipment, region_key):
    """지정 장비·지역의 split_log에서 처리된 학교코드 집합"""
    cfg = EQUIPMENT_CONFIG.get(equipment, {})
    log_name = cfg.get("log_name", equipment)
    prefix = f"split_log_{log_name}_{region_key}_"
    processed = set()
    for fname in os.listdir(BASE_DIR):
        if fname.startswith(prefix) and fname.endswith('.csv'):
            try:
                with open(os.path.join(BASE_DIR, fname), 'r', encoding='utf-8-sig') as f:
                    for row in csv.DictReader(f):
                        c = row.get('학교코드', '').strip()
                        if c:
                            processed.add(c)
            except Exception:
                pass
    return processed


def load_existing_log(path):
    entries = []
    if path and os.path.exists(path):
        for enc in ('utf-8-sig', 'utf-8', 'cp949'):
            try:
                with open(path, 'r', encoding=enc) as f:
                    for row in csv.DictReader(f):
                        entries.append(row)
                break
            except UnicodeDecodeError:
                continue
    return entries


def normalize_code(val):
    if val is None:
        return ''
    s = str(val).strip()
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        s = str(int(val)) if val == int(val) else str(val)
    return s


def sanitize_filename(name):
    if name is None:
        return "Unknown"
    return re.sub(r'[\\/:*?"<>|]', '_', str(name).strip())


def copy_cell_style(src_cell, tgt_cell, copy_fill=True):
    for attr in ('font', 'fill', 'border', 'alignment', 'number_format'):
        try:
            val = getattr(src_cell, attr, None)
            if val and (attr != 'fill' or (copy_fill and getattr(val, 'fill_type', None))):
                if attr == 'border' and val:
                    b = val
                    if (b.left and b.left.style) or (b.right and b.right.style) or (b.top and b.top.style) or (b.bottom and b.bottom.style):
                        tgt_cell.border = copy(val)
                else:
                    setattr(tgt_cell, attr, copy(val) if hasattr(val, '__copy__') else val)
        except Exception:
            pass


def copy_title_row_ap(ws_src, ws_out):
    """AP용: B1 값을 A1에, 왼쪽 맞춤"""
    src_cell_a1 = ws_src.cell(row=1, column=1)
    title_val = ws_src.cell(row=1, column=2).value
    ws_out['A1'] = title_val if title_val else "AP 자산 상세"
    copy_cell_style(src_cell_a1, ws_out['A1'])
    ws_out['A1'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    if 1 in ws_src.row_dimensions and ws_src.row_dimensions[1].height:
        ws_out.row_dimensions[1].height = ws_src.row_dimensions[1].height
    for merged in list(ws_src.merged_cells.ranges):
        if merged.min_row <= 1 <= merged.max_row:
            try:
                ws_out.merge_cells(str(merged))
            except Exception:
                pass


def copy_title_row_a1(ws_src, ws_out, src_row=1):
    """A1 셀 복사 + 병합"""
    src_cell = ws_src.cell(row=src_row, column=1)
    tgt_cell = ws_out.cell(row=src_row, column=1)
    tgt_cell.value = src_cell.value
    copy_cell_style(src_cell, tgt_cell)
    if src_row in ws_src.row_dimensions and ws_src.row_dimensions[src_row].height:
        ws_out.row_dimensions[src_row].height = ws_src.row_dimensions[src_row].height
    for merged in list(ws_src.merged_cells.ranges):
        if merged.min_row <= src_row <= merged.max_row:
            try:
                ws_out.merge_cells(str(merged))
            except Exception:
                pass


def main(equipment, missed_only=False, source_file=None, only_schools=None, schools_file=None,
         today_from_missed=False, from_log=None, new_log=False, test_output=False, region_key='DNI'):
    cfg = EQUIPMENT_CONFIG.get(equipment)
    if not cfg:
        print(f"오류: 지원하지 않는 장비: {equipment}")
        return

    source_path = source_file or get_source_path(region_key, equipment)
    output_base = OUTPUT_BASE_TEST if test_output else OUTPUT_BASE_BY_REGION.get(region_key, OUTPUT_BASE_BY_REGION["DNI"])
    device_label = cfg["device_label"]

    if test_output:
        print(f"[테스트] 출력 경로: {output_base}")
    if not source_path:
        print(f"오류: 원본 경로 규칙이 정의되지 않았습니다. (지역: {region_key}, 장비: {equipment}) → split_config.py 확인")
        return
    if not os.path.exists(source_path):
        print(f"오류: 원본 파일을 찾을 수 없습니다. {source_path}")
        return

    school_list_path = get_school_list_path(region_key, BASE_DIR)
    if not os.path.exists(school_list_path):
        print(f"오류: 학교 리스트 파일을 찾을 수 없습니다. (지역: {region_key})")
        return

    # 학교 리스트 로드 (xlsx 또는 csv)
    school_list = []
    if school_list_path.lower().endswith('.csv'):
        with open(school_list_path, 'r', encoding='utf-8-sig') as f:
            for row in csv.DictReader(f):
                code = row.get('학교코드', row.get('code', '')).strip()
                region = row.get('지역', row.get('region', '')).strip()
                name = row.get('학교명', row.get('name', '')).strip()
                if code or region or name:
                    school_list.append({'code': code, 'region': region, 'name': name})
    else:
        from openpyxl import load_workbook as load_wb
        wb_school = load_wb(school_list_path, read_only=True, data_only=True)
        for row in wb_school.active.iter_rows(min_row=1, values_only=True):
            row = list(row) if row else []
            while len(row) < 3:
                row.append(None)
            if row[0] is None and row[1] is None and row[2] is None:
                continue
            if row[0] or row[1] or row[2]:
                school_list.append({
                    'code': str(row[0]).strip() if row[0] else '',
                    'region': str(row[1]).strip() if row[1] else '',
                    'name': str(row[2]).strip() if row[2] else '',
                })
        wb_school.close()

    if school_list and (school_list[0].get('code') == '학교코드' or school_list[0].get('region') == '지역' or school_list[0].get('name') == '학교명'):
        school_list = school_list[1:]

    seen_codes = set()
    school_list_unique = []
    for s in school_list:
        if s.get('code') and s['code'] not in seen_codes:
            seen_codes.add(s['code'])
            school_list_unique.append(s)
    school_list = school_list_unique

    if missed_only:
        processed = get_processed_school_codes(equipment, region_key)
        school_list = [s for s in school_list if s['code'] and s['code'] not in processed]
        print(f"[빠진 학교만] 미처리: {len(school_list)}개")
        if not school_list:
            return

    only_codes = set()
    if from_log:
        log_path = from_log if os.path.isabs(from_log) else os.path.join(BASE_DIR, from_log)
        if os.path.exists(log_path):
            schools_file = log_path
    elif today_from_missed and not schools_file and os.path.exists(MISSED_SCHOOLS_FILE):
        schools_file = MISSED_SCHOOLS_FILE
    if only_schools:
        only_codes.update(c.strip() for c in only_schools.split(',') if c.strip())
    if schools_file and os.path.exists(schools_file):
        if schools_file.lower().endswith('.csv'):
            with open(schools_file, 'r', encoding='utf-8-sig') as f:
                for row in csv.DictReader(f):
                    c = row.get('학교코드', '').strip()
                    if c:
                        only_codes.add(c)
        else:
            with open(schools_file, 'r', encoding='utf-8-sig') as f:
                for line in f:
                    c = line.strip()
                    if c and not c.startswith('#'):
                        only_codes.add(c)
    if only_codes:
        school_list = [s for s in school_list if s['code'] and s['code'] in only_codes]
        if not school_list:
            return

    print(f"[{device_label}] 학교 리스트: {len(school_list)}개")

    wb_src = load_workbook(source_path, read_only=False, data_only=False)
    sheet_name = None
    for cand in get_sheet_candidates(equipment):
        if cand in wb_src.sheetnames:
            sheet_name = cand
            break
    if not sheet_name:
        sheet_name = wb_src.sheetnames[0] if wb_src.sheetnames else None
    if not sheet_name:
        print("오류: 시트를 찾을 수 없습니다.")
        wb_src.close()
        return

    ws_src = wb_src[sheet_name]
    max_row = ws_src.max_row
    HEADER_ROW = 2
    DATA_START_ROW = 3

    school_code_col = find_school_code_col(ws_src, HEADER_ROW) if cfg["code_from"] == "school_or_mgmt" else None
    mgmt_col = find_mgmt_col(ws_src, HEADER_ROW)
    if cfg["code_from"] == "mgmt_only" and not mgmt_col:
        print("경고: '관리번호' 열을 찾을 수 없습니다.")
        wb_src.close()
        return
    if cfg["code_from"] == "school_or_mgmt" and not school_code_col and not mgmt_col:
        print("경고: '학교코드' 또는 '관리번호' 열을 찾을 수 없습니다.")
        wb_src.close()
        return

    src_cols = get_output_cols(ws_src, HEADER_ROW, exclude_school_code=True)
    school_codes = {normalize_code(s['code']): s for s in school_list if s['code']}
    rows_by_school = {}

    for r in range(DATA_START_ROW, max_row + 1):
        if cfg["code_from"] == "school_or_mgmt" and school_code_col:
            extracted = normalize_code(ws_src.cell(row=r, column=school_code_col).value)
        else:
            if not mgmt_col:
                continue
            mgmt_val = ws_src.cell(row=r, column=mgmt_col).value
            if not mgmt_val:
                continue
            extracted = normalize_code(extract_school_code_from_mgmt_num(mgmt_val))
        if extracted and extracted in school_codes:
            rows_by_school.setdefault(extracted, []).append(r)

    schools_with_data = [(school_codes[code], sorted(rows)) for code, rows in rows_by_school.items()]
    schools_with_data = sort_schools_by_region(schools_with_data, region_key)

    if not schools_with_data:
        print("경고: 학교코드와 매칭되는 데이터가 없습니다.")
        wb_src.close()
        return

    log_entries = []
    failed_schools = []
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    use_border = cfg.get("use_border", False)

    for item in schools_with_data:
        school, data = item
        code, region, name = school['code'], school['region'], school['name']
        data_rows = sorted(data)
        start_row, end_row = min(data_rows), max(data_rows)
        row_count = len(data_rows)

        region_safe = sanitize_filename(region) if region else "기타"
        display_name = sanitize_filename(name) if name else f"학교_{code}"
        school_folder = f"{display_name}_{sanitize_filename(code)}"
        # 테스트: OUTPUT/DNI/시군구/학교폴더, OUTPUT/CNE/시군구/학교폴더
        if test_output:
            out_dir = os.path.join(output_base, region_key, region_safe, school_folder)
        else:
            out_dir = os.path.join(output_base, region_safe, school_folder)
        os.makedirs(out_dir, exist_ok=True)
        out_filename = f"{display_name}_{device_label} 장비 현황 상세.XLSX"
        out_path = os.path.join(out_dir, out_filename)

        try:
            wb_out = Workbook()
            ws_out = wb_out.active
            ws_out.title = cfg["sheet_title"]

            if cfg["title_mode"] == "ap_b1":
                copy_title_row_ap(ws_src, ws_out)
            else:
                copy_title_row_a1(ws_src, ws_out, 1)

            # 2행
            for c_idx, src_col in enumerate(src_cols, start=1):
                src_cell = ws_src.cell(row=HEADER_ROW, column=src_col)
                tgt_cell = ws_out.cell(row=2, column=c_idx)
                tgt_cell.value = src_cell.value
                copy_cell_style(src_cell, tgt_cell)
                tgt_cell.alignment = center_align
                if use_border:
                    tgt_cell.border = THIN_BORDER
            if HEADER_ROW in ws_src.row_dimensions and ws_src.row_dimensions[HEADER_ROW].height:
                ws_out.row_dimensions[2].height = ws_src.row_dimensions[HEADER_ROW].height

            # 3행~
            out_row = 3
            for src_row in data_rows:
                for c_idx, src_col in enumerate(src_cols, start=1):
                    src_cell = ws_src.cell(row=src_row, column=src_col)
                    tgt_cell = ws_out.cell(row=out_row, column=c_idx)
                    tgt_cell.value = src_cell.value
                    copy_cell_style(src_cell, tgt_cell, copy_fill=False)
                    tgt_cell.alignment = center_align
                    if use_border:
                        tgt_cell.border = THIN_BORDER
                if src_row in ws_src.row_dimensions and ws_src.row_dimensions[src_row].height:
                    ws_out.row_dimensions[out_row].height = ws_src.row_dimensions[src_row].height
                out_row += 1

            for c_idx, src_col in enumerate(src_cols, start=1):
                col_letter = get_column_letter(src_col)
                if col_letter in ws_src.column_dimensions and ws_src.column_dimensions[col_letter].width:
                    ws_out.column_dimensions[get_column_letter(c_idx)].width = ws_src.column_dimensions[col_letter].width

            wb_out.save(out_path)
            wb_out.close()
            log_entries.append({
                '학교명': name, '학교코드': code, '지역': region,
                '시작행': start_row, '끝행': end_row, '복사행수': row_count, '저장경로': out_path
            })
            print(f"  저장: {name} ({code}) - {row_count}행 -> {out_path}")
        except Exception as e:
            print(f"  [에러] {name} ({code}) {region}: {e}")
            failed_schools.append({'name': name, 'code': code, 'region': region, 'error': str(e)})

    if failed_schools:
        print(f"\n[실패 {len(failed_schools)}개]")
        for f in failed_schools:
            print(f"  - {f['name']} ({f['code']}) {f['region']}: {f['error']}")

    wb_src.close()

    log_path = get_log_path(equipment, region_key) if not new_log else get_log_path(equipment, region_key, datetime.now().strftime('%H%M%S'))
    existing = load_existing_log(log_path) if not new_log and os.path.exists(log_path) else []
    by_code = {e.get('학교코드', '').strip(): e for e in existing}
    for e in log_entries:
        by_code[e.get('학교코드', '').strip()] = e
    merged = list(by_code.values())
    with open(log_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=['학교명', '학교코드', '지역', '시작행', '끝행', '복사행수', '저장경로'])
        writer.writeheader()
        writer.writerows(merged)

    print(f"\n완료. 로그: {log_path} (총 {len(merged)}개, 이번 +{len(log_entries)}개)")


def _normalize_argv(argv):
    """옵션 대소문자 무관: --dni/--DNI, --poe/--POE 등 통일"""
    canonical = {'dni': '--DNI', 'cne': '--CNE', 'ap': '--AP', 'switch': '--switch',
                'security': '--security', 'poe': '--poe'}
    out = []
    for i, a in enumerate(argv):
        if a.startswith('--') and '=' not in a:
            key = a[2:].lower()
            if key in canonical:
                out.append(canonical[key])
                continue
        out.append(a)
    return out


if __name__ == "__main__":
    import sys
    sys.argv = [sys.argv[0]] + _normalize_argv(sys.argv[1:])

    parser = argparse.ArgumentParser(
        description='학교별 장비 현황 엑셀 분리 (통합)',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
예시:
  python split_school_all_v1.py --AP --DNI --test
  python split_school_all_v1.py --switch --CNE
  python split_school_all_v1.py -e poe --DNI --missed-only
        """,
    )
    def _device_type(s):
        m = {'ap': 'AP', 'switch': 'switch', 'security': 'security', 'poe': 'poe'}
        v = m.get(s.lower())
        if v is None:
            raise argparse.ArgumentTypeError(f"장비는 AP, switch, security, poe 중 하나: {s}")
        return v

    dev = parser.add_mutually_exclusive_group(required=True)
    dev.add_argument('--AP', '--ap', '--Ap', action='store_true', help='AP')
    dev.add_argument('--switch', '--Switch', '--SWITCH', action='store_true', help='스위치')
    dev.add_argument('--security', '--Security', '--SECURITY', action='store_true', help='보안(SEUTM)')
    dev.add_argument('--poe', '--Poe', '--POE', action='store_true', help='POE')
    parser.add_argument('-e', '--device', type=_device_type, help='장비 (AP/switch/security/poe, 대소문자 무관)')
    parser.add_argument('--DNI', '--dni', '--Dni', action='store_true', help='대전 (대소문자 무관)')
    parser.add_argument('--CNE', '--cne', '--Cne', action='store_true', help='충남 (대소문자 무관)')
    parser.add_argument('--missed-only', action='store_true', help='빠진 학교만 처리')
    parser.add_argument('--only-schools', type=str, default=None, help='학교코드 쉼표 구분')
    parser.add_argument('--schools-file', type=str, default=None, help='학교코드 목록 파일')
    parser.add_argument('--today-from-missed', action='store_true', help='missed_schools.csv 학교만')
    parser.add_argument('--from-log', type=str, default=None, help='로그 파일의 학교만')
    parser.add_argument('--new-log', action='store_true', help='별도 로그 파일 생성')
    parser.add_argument('--source', '-s', type=str, default=None, help='원본 파일 경로')
    parser.add_argument('--test', '-t', action='store_true', help='OUTPUT 폴더에 테스트 저장')
    args = parser.parse_args()

    region_key = 'CNE' if args.CNE else 'DNI'
    device = args.device or ('AP' if args.AP else 'switch' if args.switch else 'security' if args.security else 'poe')

    main(equipment=device, missed_only=args.missed_only, source_file=args.source,
         only_schools=args.only_schools, schools_file=args.schools_file,
         today_from_missed=args.today_from_missed, from_log=args.from_log, new_log=args.new_log,
         test_output=args.test, region_key=region_key)
