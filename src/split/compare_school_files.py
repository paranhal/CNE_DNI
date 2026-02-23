# -*- coding: utf-8 -*-
"""
어제/오늘 생성된 학교별 AP 파일 비교
- 데이터 행 수, 셀 값, 서식(폰트, 채우기, 맞춤 등) 비교
"""
import os
import sys
from openpyxl import load_workbook

# 충남: FILE_YESTERDAY = r"y:\CNE_DNI\_지역별 산출물 취합\_00.별첨자료_학교별\CNE\계룡\계룡고등학교_N108140237HS\계룡고등학교_AP 장비 현황 상세.XLSX"
# 충남: FILE_TODAY = r"y:\CNE_DNI\_지역별 산출물 취합\_00.별첨자료_학교별\CNE\보령\개화초등학교_N108171020ES\개화초등학교_AP 장비 현황 상세.XLSX"
FILE_YESTERDAY = r"y:\DJE_DNI\_지역별 산출물 취합\_00.별첨자료_학교별\DJE\동구\학교1_코드1\학교1_AP 장비 현황 상세.XLSX"  # 대전 예시
FILE_TODAY = r"y:\DJE_DNI\_지역별 산출물 취합\_00.별첨자료_학교별\DJE\중구\학교2_코드2\학교2_AP 장비 현황 상세.XLSX"  # 대전 예시
OUTPUT_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "compare_result.txt")


def cell_info(cell):
    """셀 서식 요약"""
    parts = []
    if cell.font:
        f = cell.font
        if f.bold: parts.append("굵게")
        if f.size: parts.append(f"폰트{f.size}")
        if f.name: parts.append(f.name[:8])
    if cell.fill and getattr(cell.fill, 'fill_type', None):
        parts.append("배경색O")
    if cell.alignment:
        a = cell.alignment
        if a.horizontal: parts.append(f"H:{a.horizontal}")
        if a.vertical: parts.append(f"V:{a.vertical}")
    return ",".join(parts) if parts else "-"


def analyze(wb, label):
    ws = wb.active
    info = {
        'label': label,
        'rows': ws.max_row,
        'cols': ws.max_column,
        'row_heights': {},
        'col_widths': {},
        'sample_data': [],
        'sample_format': []
    }
    for r in range(1, min(6, ws.max_row + 1)):
        row_vals = []
        row_fmt = []
        for c in range(1, min(13, ws.max_column + 1)):
            cell = ws.cell(row=r, column=c)
            v = cell.value
            row_vals.append(str(v)[:25] if v is not None else "")
            row_fmt.append(cell_info(cell))
        info['sample_data'].append(row_vals)
        info['sample_format'].append(row_fmt)
    for r in range(1, ws.max_row + 1):
        if r in ws.row_dimensions and ws.row_dimensions[r].height:
            info['row_heights'][r] = ws.row_dimensions[r].height
    for c in range(1, ws.max_column + 1):
        from openpyxl.utils import get_column_letter
        cl = get_column_letter(c)
        if cl in ws.column_dimensions and ws.column_dimensions[cl].width:
            info['col_widths'][c] = ws.column_dimensions[cl].width
    return info


def main():
    lines = []
    def log(s=""):
        lines.append(s)
        print(s)

    def save_result():
        """항상 결과 파일 저장"""
        try:
            with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
                f.write("\n".join(lines))
            log(f"\n결과 저장: {OUTPUT_FILE}")
        except Exception as e:
            lines.append(f"\n파일 저장 실패: {e}")
            lines.append(f"저장 경로: {OUTPUT_FILE}")

    log("=" * 70)
    log("어제 vs 오늘 생성 파일 비교")
    log("=" * 70)
    log(f"\n[어제] {FILE_YESTERDAY}")  # 충남: 계룡고등학교
    log(f"[오늘] 개화초등학교: {FILE_TODAY}")
    log()

    if not os.path.exists(FILE_YESTERDAY):
        log(f"오류: 어제 파일 없음 - {FILE_YESTERDAY}")
        save_result()
        return
    if not os.path.exists(FILE_TODAY):
        log(f"오류: 오늘 파일 없음 - {FILE_TODAY}")
        save_result()
        return

    try:
        wb1 = load_workbook(FILE_YESTERDAY, data_only=False)
        wb2 = load_workbook(FILE_TODAY, data_only=False)

        a1 = analyze(wb1, "어제")  # 충남: 계룡고
        a2 = analyze(wb2, "오늘(개화초)")

        log("--- 1. 행/열 수 비교 ---")
        log(f"  어제: {a1['rows']}행 x {a1['cols']}열")
        log(f"  오늘: {a2['rows']}행 x {a2['cols']}열")
        if a1['rows'] != a2['rows'] or a1['cols'] != a2['cols']:
            log("  ※ 행/열 수 다름!")

        log("\n--- 2. 1~5행 데이터 샘플 (어제) ---")
        for i, row in enumerate(a1['sample_data'], 1):
            log(f"  행{i}: {row}")

        log("\n--- 3. 1~5행 데이터 샘플 (오늘) ---")
        for i, row in enumerate(a2['sample_data'], 1):
            log(f"  행{i}: {row}")

        log("\n--- 4. 1~5행 서식 샘플 (어제) ---")
        for i, row in enumerate(a1['sample_format'], 1):
            log(f"  행{i}: {row}")

        log("\n--- 5. 1~5행 서식 샘플 (오늘) ---")
        for i, row in enumerate(a2['sample_format'], 1):
            log(f"  행{i}: {row}")

        log("\n--- 6. 행 높이 ---")
        log(f"  어제: {len(a1['row_heights'])}개 행에 높이 설정")
        log(f"  오늘: {len(a2['row_heights'])}개 행에 높이 설정")

        log("\n--- 7. 열 넓이 ---")
        log(f"  어제: {len(a1['col_widths'])}개 열에 넓이 설정")
        log(f"  오늘: {len(a2['col_widths'])}개 열에 넓이 설정")

        wb1.close()
        wb2.close()
    except Exception as e:
        log(f"\n오류 발생: {e}")
        import traceback
        log(traceback.format_exc())

    save_result()


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        # 스크립트가 아예 실패해도 파일은 남기기
        err_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "compare_result.txt")
        with open(err_path, 'w', encoding='utf-8') as f:
            f.write(f"실행 오류: {e}\n\n")
            import traceback
            f.write(traceback.format_exc())
        print(f"오류 발생. 로그 저장: {err_path}")
