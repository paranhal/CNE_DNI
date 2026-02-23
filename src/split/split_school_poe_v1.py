# -*- coding: utf-8 -*-
"""
학교별 POE 장비 현황 엑셀 분리 스크립트 (POE_V1)
- 스위치와 동일 구조·로직
- 원본: DJE_POE_LIST.xlsx (대전)
- 출력: {학교명}_POE_V1 장비 현황 상세.XLSX
- 로그: split_log_poe_{지역}_{날짜}.csv (예: split_log_poe_DNI_20260222.csv)
"""
import os
import csv
import re
import argparse
from datetime import datetime
from copy import copy
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from school_utils import extract_school_code_from_mgmt_num, find_mgmt_col, get_output_cols, sort_schools_by_region, get_school_list_path
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SOURCE_BY_REGION = {
    "DNI": os.path.join(BASE_DIR, "DJE_POE_LIST.xlsx"),
    "CNE": os.path.join(BASE_DIR, "CNE_POE_LIST.xlsx"),
}
OUTPUT_BASE_BY_REGION = {
    "DNI": r"Y:\DJE_DNI\_지역별 산출물 취합\_00.별첨자료_학교별\DJE",
    "CNE": r"Y:\CNE_DNI\_지역별 산출물 취합\_00.별첨자료_학교별\CNE",
}
MISSED_SCHOOLS_FILE = os.path.join(BASE_DIR, "missed_schools.csv")
OUTPUT_BASE_TEST = os.path.join(BASE_DIR, "OUTPUT")
DEVICE_NAME = "POE_V1"


def get_today_log_path(region='DNI', suffix=None):
    base = f"split_log_poe_{region}_{datetime.now().strftime('%Y%m%d')}"
    if suffix:
        return os.path.join(BASE_DIR, f"{base}_{suffix}.csv")
    return os.path.join(BASE_DIR, f"{base}.csv")


def get_processed_school_codes(region='DNI'):
    prefix = f"split_log_poe_{region}_"
    processed = set()
    for fname in os.listdir(BASE_DIR):
        if fname.startswith(prefix) and fname.endswith('.csv'):
            path = os.path.join(BASE_DIR, fname)
            try:
                with open(path, 'r', encoding='utf-8-sig') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        code = row.get('학교코드', '').strip()
                        if code:
                            processed.add(code)
            except Exception:
                pass
    return processed


def load_existing_log(path):
    entries = []
    if path and os.path.exists(path):
        for enc in ('utf-8-sig', 'utf-8', 'cp949'):
            try:
                with open(path, 'r', encoding=enc) as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        entries.append(row)
                break
            except UnicodeDecodeError:
                continue
    return entries


def normalize_code(val):
    if val is None:
        return ''
    s = str(val).strip()
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        s = str(int(val)) if val == int(val) else str(val)
    return s


def sanitize_filename(name):
    if name is None:
        return "Unknown"
    return re.sub(r'[\\/:*?"<>|]', '_', str(name).strip())


def copy_cell_style(src_cell, tgt_cell, copy_fill=True):
    try:
        if src_cell.font:
            tgt_cell.font = copy(src_cell.font)
    except Exception:
        pass
    try:
        if copy_fill and src_cell.fill and getattr(src_cell.fill, 'fill_type', None):
            tgt_cell.fill = copy(src_cell.fill)
    except Exception:
        pass
    try:
        if src_cell.border:
            b = src_cell.border
            if (b.left and b.left.style) or (b.right and b.right.style) or \
               (b.top and b.top.style) or (b.bottom and b.bottom.style):
                tgt_cell.border = copy(src_cell.border)
    except Exception:
        pass
    try:
        if src_cell.alignment:
            tgt_cell.alignment = copy(src_cell.alignment)
    except Exception:
        pass
    try:
        if src_cell.number_format:
            tgt_cell.number_format = src_cell.number_format
    except Exception:
        pass


def copy_title_row_with_merge(ws_src, ws_out, src_row=1):
    src_cell = ws_src.cell(row=src_row, column=1)
    tgt_cell = ws_out.cell(row=src_row, column=1)
    tgt_cell.value = src_cell.value
    copy_cell_style(src_cell, tgt_cell)
    if src_row in ws_src.row_dimensions and ws_src.row_dimensions[src_row].height:
        ws_out.row_dimensions[src_row].height = ws_src.row_dimensions[src_row].height
    for merged in list(ws_src.merged_cells.ranges):
        if merged.min_row <= src_row <= merged.max_row:
            try:
                ws_out.merge_cells(str(merged))
            except Exception:
                pass


def main(missed_only=False, source_file=None, only_schools=None, schools_file=None,
         today_from_missed=False, from_log=None, new_log=False, test_output=False, region_key='DNI'):
    source_path = source_file or SOURCE_BY_REGION.get(region_key, SOURCE_BY_REGION["DNI"])
    output_base = OUTPUT_BASE_TEST if test_output else OUTPUT_BASE_BY_REGION.get(region_key, OUTPUT_BASE_BY_REGION["DNI"])
    if test_output:
        print(f"[테스트] 출력 경로: {output_base}")
    if not os.path.exists(source_path):
        print(f"오류: 원본 파일을 찾을 수 없습니다. {source_path}")
        return
    school_list_path = get_school_list_path(region_key, BASE_DIR)
    if not os.path.exists(school_list_path):
        print(f"오류: 학교 리스트 파일을 찾을 수 없습니다. (지역: {region_key})")
        return

    from openpyxl import load_workbook as load_wb
    wb_school = load_wb(school_list_path, read_only=True, data_only=True)
    ws_school = wb_school.active
    school_list = []
    for row in ws_school.iter_rows(min_row=1, values_only=True):
        row = list(row) if row else []
        while len(row) < 3:
            row.append(None)
        if row[0] is None and row[1] is None and row[2] is None:
            continue
        code, region, name = row[0], row[1], row[2]
        if code or region or name:
            school_list.append({
                'code': str(code).strip() if code else '',
                'region': str(region).strip() if region else '',
                'name': str(name).strip() if name else ''
            })
    wb_school.close()

    if school_list and (school_list[0].get('code') == '학교코드' or
                        school_list[0].get('region') == '지역' or
                        school_list[0].get('name') == '학교명'):
        school_list = school_list[1:]

    seen_codes = set()
    school_list_unique = []
    for s in school_list:
        if s['code'] and s['code'] not in seen_codes:
            seen_codes.add(s['code'])
            school_list_unique.append(s)
    school_list = school_list_unique

    if missed_only:
        processed_codes = get_processed_school_codes(region_key)
        school_list = [s for s in school_list if s['code'] and s['code'] not in processed_codes]
        print(f"[빠진 학교만] 미처리: {len(school_list)}개")
        if not school_list:
            return

    only_codes = set()
    if from_log:
        log_path = from_log if os.path.isabs(from_log) else os.path.join(BASE_DIR, from_log)
        if os.path.exists(log_path):
            schools_file = log_path
    elif today_from_missed and not schools_file and os.path.exists(MISSED_SCHOOLS_FILE):
        schools_file = MISSED_SCHOOLS_FILE
    if only_schools:
        only_codes.update(c.strip() for c in only_schools.split(',') if c.strip())
    if schools_file and os.path.exists(schools_file):
        if schools_file.lower().endswith('.csv'):
            with open(schools_file, 'r', encoding='utf-8-sig') as f:
                for row in csv.DictReader(f):
                    code = row.get('학교코드', '').strip()
                    if code:
                        only_codes.add(code)
        else:
            with open(schools_file, 'r', encoding='utf-8-sig') as f:
                for line in f:
                    code = line.strip()
                    if code and not code.startswith('#'):
                        only_codes.add(code)
    if only_codes:
        school_list = [s for s in school_list if s['code'] and s['code'] in only_codes]
        if not school_list:
            return

    print(f"학교 리스트: {len(school_list)}개")

    wb_src = load_workbook(source_path, read_only=False, data_only=False)
    sheet_name = None
    for candidate in ["POE_대전전체", "POE", "Sheet1"]:  # "POE_충남전체" 주석
        if candidate in wb_src.sheetnames:
            sheet_name = candidate
            break
    if not sheet_name:
        sheet_name = wb_src.sheetnames[0] if wb_src.sheetnames else None
    if not sheet_name:
        print("오류: 시트를 찾을 수 없습니다.")
        wb_src.close()
        return

    ws_src = wb_src[sheet_name]
    max_row = ws_src.max_row
    max_col = ws_src.max_column
    HEADER_ROW = 2
    DATA_START_ROW = 3

    mgmt_col = find_mgmt_col(ws_src, HEADER_ROW)
    if not mgmt_col:
        print("경고: '관리번호' 열을 찾을 수 없습니다.")
        wb_src.close()
        return

    src_cols = get_output_cols(ws_src, HEADER_ROW, exclude_school_code=True)

    school_codes = {normalize_code(s['code']): s for s in school_list if s['code']}
    rows_by_school = {}
    for r in range(DATA_START_ROW, max_row + 1):
        mgmt_val = ws_src.cell(row=r, column=mgmt_col).value
        if not mgmt_val:
            continue
        extracted = normalize_code(extract_school_code_from_mgmt_num(mgmt_val))
        if extracted and extracted in school_codes:
            rows_by_school.setdefault(extracted, []).append(r)

    schools_with_data = [(school_codes[code], sorted(rows))
                         for code, rows in rows_by_school.items()]
    schools_with_data = sort_schools_by_region(schools_with_data, region_key)

    if not schools_with_data:
        print("경고: 학교코드와 매칭되는 데이터가 없습니다.")
        wb_src.close()
        return

    log_entries = []
    failed_schools = []
    from openpyxl import Workbook

    for item in schools_with_data:
        school, data = item
        code, region, name = school['code'], school['region'], school['name']
        data_rows = sorted(data)

        start_row = min(data_rows)
        end_row = max(data_rows)
        row_count = len(data_rows)

        region_safe = sanitize_filename(region) if region else "기타"
        display_name = sanitize_filename(name) if name else f"학교_{code}"
        school_folder = f"{display_name}_{sanitize_filename(code)}"
        out_dir = os.path.join(output_base, region_safe, school_folder)
        os.makedirs(out_dir, exist_ok=True)
        out_filename = f"{display_name}_{DEVICE_NAME} 장비 현황 상세.XLSX"
        out_path = os.path.join(out_dir, out_filename)

        try:
            wb_out = Workbook()
            ws_out = wb_out.active
            ws_out.title = f"{DEVICE_NAME} 장비 현황"

            copy_title_row_with_merge(ws_src, ws_out, src_row=1)

            for c_idx, src_col in enumerate(src_cols, start=1):
                src_cell = ws_src.cell(row=HEADER_ROW, column=src_col)
                tgt_cell = ws_out.cell(row=2, column=c_idx)
                tgt_cell.value = src_cell.value
                copy_cell_style(src_cell, tgt_cell)
                tgt_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            if HEADER_ROW in ws_src.row_dimensions and ws_src.row_dimensions[HEADER_ROW].height:
                ws_out.row_dimensions[2].height = ws_src.row_dimensions[HEADER_ROW].height

            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            out_row = 3
            for src_row in data_rows:
                for c_idx, src_col in enumerate(src_cols, start=1):
                    src_cell = ws_src.cell(row=src_row, column=src_col)
                    tgt_cell = ws_out.cell(row=out_row, column=c_idx)
                    tgt_cell.value = src_cell.value
                    copy_cell_style(src_cell, tgt_cell, copy_fill=False)
                    tgt_cell.alignment = center_align
                if src_row in ws_src.row_dimensions and ws_src.row_dimensions[src_row].height:
                    ws_out.row_dimensions[out_row].height = ws_src.row_dimensions[src_row].height
                out_row += 1

            for c_idx, src_col in enumerate(src_cols, start=1):
                col_letter = get_column_letter(src_col)
                if col_letter in ws_src.column_dimensions and ws_src.column_dimensions[col_letter].width:
                    ws_out.column_dimensions[get_column_letter(c_idx)].width = ws_src.column_dimensions[col_letter].width

            wb_out.save(out_path)
            wb_out.close()
            log_entries.append({
                '학교명': name, '학교코드': code, '지역': region,
                '시작행': start_row, '끝행': end_row, '복사행수': row_count, '저장경로': out_path
            })
            print(f"  저장: {name} ({code}) - {row_count}행 -> {out_path}")
        except Exception as e:
            print(f"  [에러] {name} ({code}) {region}: {e}")
            failed_schools.append({'name': name, 'code': code, 'region': region, 'error': str(e)})

    if failed_schools:
        print(f"\n[실패 {len(failed_schools)}개]")
        for f in failed_schools:
            print(f"  - {f['name']} ({f['code']}) {f['region']}: {f['error']}")

    wb_src.close()

    log_path = get_today_log_path(region_key) if not new_log else get_today_log_path(region_key, datetime.now().strftime('%H%M%S'))
    existing = load_existing_log(log_path) if not new_log and log_path and os.path.exists(log_path) else []
    by_code = {e.get('학교코드', '').strip(): e for e in existing}
    for e in log_entries:
        by_code[e.get('학교코드', '').strip()] = e
    merged = list(by_code.values())
    with open(log_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=['학교명', '학교코드', '지역', '시작행', '끝행', '복사행수', '저장경로'])
        writer.writeheader()
        writer.writerows(merged)

    print(f"\n완료. 로그: {log_path} (총 {len(merged)}개, 이번 +{len(log_entries)}개)")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='학교별 POE 장비 현황 엑셀 분리 (POE_V1)')
    parser.add_argument('--missed-only', action='store_true')
    parser.add_argument('--only-schools', type=str, default=None)
    parser.add_argument('--schools-file', type=str, default=None)
    parser.add_argument('--today-from-missed', action='store_true')
    parser.add_argument('--from-log', type=str, default=None)
    parser.add_argument('--new-log', action='store_true')
    parser.add_argument('--source', '-s', type=str, default=None)
    parser.add_argument('--test', '-t', action='store_true',
                        help='테스트: OUTPUT 폴더에 생성. 없으면 실제 목표 저장 폴더에 저장')
    parser.add_argument('--DNI', '--dni', action='store_true', help='대전 (로그: split_log_poe_DNI_YYYYMMDD.csv)')
    parser.add_argument('--CNE', '--cne', action='store_true', help='충남 (로그: split_log_poe_CNE_YYYYMMDD.csv)')
    args = parser.parse_args()
    region_key = 'CNE' if args.CNE else 'DNI'
    main(missed_only=args.missed_only, source_file=args.source,
         test_output=args.test,
         only_schools=args.only_schools, schools_file=args.schools_file,
         today_from_missed=args.today_from_missed, from_log=args.from_log, new_log=args.new_log,
         region_key=region_key)
