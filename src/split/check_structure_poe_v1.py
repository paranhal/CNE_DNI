# -*- coding: utf-8 -*-
"""POE 원본 구조 확인 (대전: DJE_POE_LIST.xlsx, POE_V1)"""
import openpyxl
from openpyxl.utils import get_column_letter
import os

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# 충남: SOURCE = os.path.join(BASE_DIR, "CNE_POE_LIST.xlsx")
SOURCE = os.path.join(BASE_DIR, "DJE_POE_LIST.xlsx")

def main():
    if not os.path.exists(SOURCE):
        print(f"파일 없음: {SOURCE}")
        return
    wb = openpyxl.load_workbook(SOURCE, read_only=False, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n시트: {sheet_name} | 행:{ws.max_row} 열:{ws.max_column}")
        for row in range(1, min(5, ws.max_row + 1)):
            vals = [ws.cell(row=row, column=c).value for c in range(1, min(13, ws.max_column + 1))]
            print(f"  Row{row}: {vals}")
    wb.close()

if __name__ == "__main__":
    main()
