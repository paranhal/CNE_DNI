# -*- coding: utf-8 -*-
"""AP 원본 구조 분석 → ap_structure_report.txt에 저장 (학교 리스트 + AP 시트)"""
import openpyxl
import os

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DNI_DIR = os.path.join(BASE_DIR, "DNI")
SCHOOL_LIST_FILE = os.path.join(BASE_DIR, "SCHOOL_REG_LIST.XLSX")
# 충남: SOURCE = os.path.join(BASE_DIR, "00.충남_AP_자산_첨부(충남전체)_.xlsx")
SOURCE = os.path.join(DNI_DIR, "DNI_AP_LIST.XLSX")  # 대전 1차 AP
OUTPUT = os.path.join(BASE_DIR, "ap_structure_report.txt")
TARGET_CODE = "N108140237HS"  # 검색용 (대전 학교코드로 변경 가능)

def main():
    lines = []

    # 1. 학교 리스트 (2번째 파일)
    lines.append("=" * 60)
    lines.append("1. SCHOOL_REG_LIST.XLSX 학교 리스트")
    lines.append("=" * 60)
    if os.path.exists(SCHOOL_LIST_FILE):
        wb_school = openpyxl.load_workbook(SCHOOL_LIST_FILE, read_only=True, data_only=True)
        ws_school = wb_school.active
        school_count = ws_school.max_row
        first_cell = str(ws_school.cell(1, 1).value or "").strip()
        if first_cell in ("학교코드", "지역", "학교명"):
            school_count = max(0, school_count - 1)  # 헤더 행 제외
        lines.append(f"총 {school_count}개 학교")
        wb_school.close()
    else:
        lines.append(f"파일 없음: {SCHOOL_LIST_FILE}")

    # 2. AP 원본
    lines.append("\n" + "=" * 60)
    lines.append("2. AP 원본 시트 구조")
    lines.append("=" * 60)
    if not os.path.exists(SOURCE):
        lines.append(f"파일 없음: {SOURCE}")
        with open(OUTPUT, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))
        return

    wb = openpyxl.load_workbook(SOURCE, read_only=False, data_only=True)
    sheet_name = None
    for candidate in ["AP자산", "AP_대전전체", "AP (2)", "AP(2)", "Sheet1"]:  # "AP_충남전체" 주석
        if candidate in wb.sheetnames:
            sheet_name = candidate
            break
    if not sheet_name:
        sheet_name = wb.sheetnames[0] if wb.sheetnames else None

    if sheet_name:
        ws = wb[sheet_name]
        lines.append(f"\n=== 시트: {sheet_name} | 행:{ws.max_row} 열:{ws.max_column} ===")
        for row in range(1, min(10, ws.max_row + 1)):
            vals = []
            for c in range(1, min(15, ws.max_column + 1)):
                v = ws.cell(row=row, column=c).value
                vals.append(str(v)[:20] if v is not None else "")
            lines.append(f"  Row{row}: {vals}")

        # N108140237HS 검색
        lines.append(f"\n  --- '{TARGET_CODE}' 검색 ---")
        found = []
        for r in range(1, min(ws.max_row + 1, 500)):
            for c in range(1, min(ws.max_column + 1, 20)):
                v = ws.cell(row=r, column=c).value
                if v and TARGET_CODE in str(v):
                    found.append((r, c, str(v)[:50]))
        lines.append(f"  발견: {found[:10]}")
    else:
        lines.append("AP 시트를 찾을 수 없습니다.")
    wb.close()

    with open(OUTPUT, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))
    print("저장:", OUTPUT)

if __name__ == "__main__":
    main()
