# -*- coding: utf-8 -*-
"""엑셀 파일 구조 확인 스크립트 - 학교 리스트 및 AP 시트 구조 파악"""
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import os

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DNI_DIR = os.path.join(BASE_DIR, "DNI")
# 충남: SOURCE_FILE = os.path.join(BASE_DIR, "00.충남_AP_자산_첨부(충남전체)_.xlsx")
SOURCE_FILE = os.path.join(DNI_DIR, "DNI_AP_LIST.XLSX")
SCHOOL_LIST_FILE = os.path.join(BASE_DIR, "SCHOOL_REG_LIST.XLSX")

def main():
    print("=" * 60)
    print("1. SCHOOL_REG_LIST.XLSX 학교 리스트 확인")
    print("=" * 60)
    
    if os.path.exists(SCHOOL_LIST_FILE):
        df_school = pd.read_excel(SCHOOL_LIST_FILE, header=None)
        df_school.columns = ['학교코드', '지역', '학교명']
        print(f"총 {len(df_school)}개 학교")
        print(df_school.to_string())
    else:
        print(f"파일 없음: {SCHOOL_LIST_FILE}")
    
    print("\n" + "=" * 60)
    print("2. AP 시트 구조 확인")
    print("=" * 60)
    
    if os.path.exists(SOURCE_FILE):
        wb = openpyxl.load_workbook(SOURCE_FILE, read_only=False, data_only=False)
        # 충남: sheet_name = "AP_충남전체" if "AP_충남전체" in wb.sheetnames else ...
        for cand in ["AP자산", "AP_대전전체", "AP (2)", "AP(2)", "Sheet1"]:
            if cand in wb.sheetnames:
                sheet_name = cand
                break
        else:
            sheet_name = wb.sheetnames[0] if wb.sheetnames else None
        if sheet_name:
            ws = wb[sheet_name]
            print(f"시트: {sheet_name}")
            print(f"최대 행: {ws.max_row}, 최대 열: {ws.max_column}")
            print("\n--- 1~5행, A~L열 데이터 ---")
            for row in range(1, min(6, ws.max_row + 1)):
                row_data = []
                for col in range(1, min(13, ws.max_column + 1)):
                    cell = ws.cell(row=row, column=col)
                    val = cell.value
                    col_letter = get_column_letter(col)
                    if val is not None:
                        row_data.append(f"{col_letter}{row}:{str(val)[:30]}")
                    else:
                        row_data.append(f"{col_letter}{row}:")
                print(f"  Row {row}: {row_data}")
            wb.close()
        else:
            print(f"AP 시트 없음. 시트 목록: {wb.sheetnames}")  # 충남: AP_충남전체
            wb.close()
    else:
        print(f"파일 없음: {SOURCE_FILE}")

if __name__ == "__main__":
    main()
