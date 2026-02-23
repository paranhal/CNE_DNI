# -*- coding: utf-8 -*-
"""AP 원본 구조 확인 (대전: 00.대전_AP_자산_첨부(대전전체)_.xlsx)"""
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import os

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DNI_DIR = os.path.join(BASE_DIR, "DNI")
# 충남: SOURCE = os.path.join(BASE_DIR, "00.충남_AP_자산_첨부(충남전체)_.xlsx")
SOURCE = os.path.join(DNI_DIR, "DNI_AP_LIST.XLSX")  # 대전 1차 AP
SCHOOL_LIST_FILE = os.path.join(BASE_DIR, "SCHOOL_REG_LIST.XLSX")

def main():
    print("=" * 60)
    print("1. SCHOOL_REG_LIST.XLSX 학교 리스트 확인")
    print("=" * 60)

    if os.path.exists(SCHOOL_LIST_FILE):
        df_school = pd.read_excel(SCHOOL_LIST_FILE, header=None)
        df_school.columns = ['학교코드', '지역', '학교명']
        print(f"총 {len(df_school)}개 학교")
        print(df_school.to_string())
    else:
        print(f"파일 없음: {SCHOOL_LIST_FILE}")

    print("\n" + "=" * 60)
    print("2. AP 시트 구조 확인")
    print("=" * 60)

    if not os.path.exists(SOURCE):
        print(f"파일 없음: {SOURCE}")
        return

    wb = openpyxl.load_workbook(SOURCE, read_only=False, data_only=True)
    sheet_name = None
    for candidate in ["AP자산", "AP_대전전체", "AP (2)", "AP(2)", "Sheet1"]:  # "AP_충남전체" 주석
        if candidate in wb.sheetnames:
            sheet_name = candidate
            break
    if sheet_name:
        ws = wb[sheet_name]
        print(f"\n시트: {sheet_name} | 행:{ws.max_row} 열:{ws.max_column}")
        for row in range(1, min(5, ws.max_row + 1)):
            vals = [ws.cell(row=row, column=c).value for c in range(1, min(13, ws.max_column + 1))]
            print(f"  Row{row}: {vals}")
    else:
        print(f"AP 시트 없음. 시트 목록: {wb.sheetnames}")
    wb.close()

if __name__ == "__main__":
    main()
