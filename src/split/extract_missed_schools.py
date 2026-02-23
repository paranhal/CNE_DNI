# -*- coding: utf-8 -*-
"""
빠진 학교 리스트 추출 스크립트
- SCHOOL_REG_LIST.XLSX (또는 SCHOOL_REG_LIST.csv)와 기존 로그(split_log.csv, split_log_*.csv)를 비교하여 미처리 학교 목록 생성
- 출력: missed_schools.csv (학교명, 학교코드, 지역 순)
"""
import os
import csv
from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SCHOOL_LIST_XLSX = os.path.join(BASE_DIR, "SCHOOL_REG_LIST.XLSX")
SCHOOL_LIST_CSV = os.path.join(BASE_DIR, "SCHOOL_REG_LIST.csv")
MISSED_OUTPUT = os.path.join(BASE_DIR, "missed_schools.csv")


def get_processed_school_codes():
    """기존 모든 로그 파일에서 처리된 학교코드 집합 (split_log.csv, split_log_*.csv)"""
    processed = set()
    for fname in os.listdir(BASE_DIR):
        if fname.startswith('split_log') and fname.endswith('.csv'):
            path = os.path.join(BASE_DIR, fname)
            try:
                with open(path, 'r', encoding='utf-8-sig') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        code = row.get('학교코드', '').strip()
                        if code:
                            processed.add(code)
            except Exception:
                pass
    return processed


def load_school_list_from_xlsx(path):
    """XLSX에서 학교 리스트 로드 (A=학교코드, B=지역, C=학교명)"""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    all_schools = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        row = list(row) if row else []
        while len(row) < 3:
            row.append(None)
        if row[0] is None and row[1] is None and row[2] is None:
            continue
        code = str(row[0]).strip() if row[0] else ''
        region = str(row[1]).strip() if row[1] else ''
        name = str(row[2]).strip() if row[2] else ''
        if code or region or name:
            all_schools.append({'code': code, 'region': region, 'name': name})
    wb.close()
    return all_schools


def load_school_list_from_csv(path):
    """CSV에서 학교 리스트 로드 (학교코드, 지역, 학교명 또는 A,B,C 컬럼 또는 위치 기반)"""
    all_schools = []
    with open(path, 'r', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        rows = list(reader)
    if not rows:
        return all_schools
    # 헤더 확인: 첫 행이 학교코드/지역/학교명이면 스킵
    first = rows[0]
    start = 0
    if len(first) >= 3 and (first[0] == '학교코드' or first[1] == '지역' or first[2] == '학교명'):
        start = 1
    for row in rows[start:]:
        row = row + [''] * (3 - len(row)) if len(row) < 3 else row[:3]
        code = str(row[0]).strip()
        region = str(row[1]).strip()
        name = str(row[2]).strip()
        if code or region or name:
            all_schools.append({'code': code, 'region': region, 'name': name})
    return all_schools


def main():
    school_list_path = SCHOOL_LIST_XLSX if os.path.exists(SCHOOL_LIST_XLSX) else SCHOOL_LIST_CSV
    if not os.path.exists(school_list_path):
        print(f"오류: 학교 리스트 파일을 찾을 수 없습니다.")
        print(f"  - {SCHOOL_LIST_XLSX}")
        print(f"  - 또는 {SCHOOL_LIST_CSV}")
        return

    # 처리된 학교코드 로드 (split_log.csv, split_log_*.csv)
    processed_codes = get_processed_school_codes()
    print(f"처리 완료된 학교: {len(processed_codes)}개 (기존 로그 파일 기준)")

    # 학교 리스트 로드
    if school_list_path.endswith('.csv'):
        all_schools = load_school_list_from_csv(school_list_path)
    else:
        all_schools = load_school_list_from_xlsx(school_list_path)

    # 헤더 제거
    if all_schools and (all_schools[0].get('code') == '학교코드' or
                        all_schools[0].get('region') == '지역' or
                        all_schools[0].get('name') == '학교명'):
        all_schools = all_schools[1:]

    # 학교코드 기준 중복 제거 (같은 이름 다른 학교 있음 - 코드가 유일 식별자)
    seen_codes = set()
    all_schools_unique = []
    for s in all_schools:
        if s['code'] and s['code'] not in seen_codes:
            seen_codes.add(s['code'])
            all_schools_unique.append(s)
    all_schools = all_schools_unique

    # 빠진 학교 필터 (학교코드 기준)
    missed = [s for s in all_schools if s['code'] not in processed_codes]

    # 저장 (학교명, 학교코드, 지역 순)
    with open(MISSED_OUTPUT, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=['학교명', '학교코드', '지역'])
        writer.writeheader()
        writer.writerows([{'학교명': s['name'], '학교코드': s['code'], '지역': s['region']} for s in missed])

    print(f"\n빠진 학교: {len(missed)}개")
    print(f"저장: {MISSED_OUTPUT}")
    for s in missed[:10]:
        print(f"  - {s['name']} ({s['code']}) / {s['region']}")
    if len(missed) > 10:
        print(f"  ... 외 {len(missed)-10}개")


if __name__ == "__main__":
    main()
