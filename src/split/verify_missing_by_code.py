# -*- coding: utf-8 -*-
"""
학교코드 기준 빠진 학교 검증 스크립트
- --DNI / --CNE: 지역 (대전/충남)
- --AP / --switch / --security / --poe: 장비 유형
- SCHOOL_REG_LIST와 split_log 비교 → 빠진 학교코드
- 원본 시트에 해당 코드 존재 여부 확인
"""
import os
import csv
import argparse
from openpyxl import load_workbook
from school_utils import extract_school_code_from_mgmt_num, find_mgmt_col, find_school_code_col

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DNI_DIR = os.path.join(BASE_DIR, "DNI")
SCHOOL_LIST_XLSX = os.path.join(BASE_DIR, "SCHOOL_REG_LIST.XLSX")
SCHOOL_LIST_CSV = os.path.join(BASE_DIR, "SCHOOL_REG_LIST.csv")

# 지역×장비별 설정 (로그: split_log_{장비}_{지역}_{날짜}.csv)
# school_list_candidates: 지역별 학교 리스트 (우선순위)
CONFIG = {
    "DNI": {  # 대전
        "school_list_candidates": [
            os.path.join(BASE_DIR, "SCHOOL_REG_LIST_DNI.xlsx"),
            os.path.join(BASE_DIR, "SCHOOL_REG_LIST_DNI.XLSX"),
            os.path.join(BASE_DIR, "school_reg_list_dni.xlsx"),
            os.path.join(BASE_DIR, "SCHOOL_REG_LIST.XLSX"),
            os.path.join(BASE_DIR, "school_reg_list.xlsx"),
            os.path.join(BASE_DIR, "SCHOOL_REG_LIST.csv"),
        ],
        "AP": {
            "log_prefix": "split_log_AP_DNI",
            "log_exclude": [],
            "source_candidates": [
                os.path.join(DNI_DIR, "DNI_AP_LIST.XLSX"),
                os.path.join(DNI_DIR, "DNI_AP_LIST.xlsx"),
                os.path.join(BASE_DIR, "DNI_AP_LIST.XLSX"),
                os.path.join(BASE_DIR, "00.대전_AP_자산_첨부(대전전체)_.xlsx"),
            ],
            "sheets": ["AP자산", "AP_대전전체", "AP (2)", "AP(2)", "Sheet1"],
            "code_from": "school_col",  # A열/학교코드
        },
        "switch": {
            "log_prefix": "split_log_switch_DNI",
            "log_exclude": [],
            "source_candidates": [
                os.path.join(BASE_DIR, "DJE_SWITCH_LIST.xlsx"),
                os.path.join(BASE_DIR, "DJE_SWITCH_LIST.XLSX"),
            ],
            "sheets": ["SW_대전전체", "스위치_대전전체", "스위치", "Switch", "Sheet1"],
            "code_from": "mgmt_col",
        },
        "security": {
            "log_prefix": "split_log_security_DNI",
            "log_exclude": [],
            "source_candidates": [
                os.path.join(BASE_DIR, "DJE_SEUTM_LIST.xlsx"),
                os.path.join(BASE_DIR, "DJE_SEUTM_LIST.XLSX"),
            ],
            "sheets": ["SEC_대전전체", "보안_대전전체", "보안장비", "Security", "Sheet1"],
            "code_from": "mgmt_col",
        },
        "poe": {
            "log_prefix": "split_log_poe_DNI",
            "log_exclude": [],
            "source_candidates": [
                os.path.join(BASE_DIR, "DJE_POE_LIST.xlsx"),
                os.path.join(BASE_DIR, "DJE_POE_LIST.XLSX"),
            ],
            "sheets": ["POE_대전전체", "POE", "Sheet1"],
            "code_from": "mgmt_col",
        },
    },
    "CNE": {  # 충남
        "school_list_candidates": [
            os.path.join(BASE_DIR, "SCHOOL_REG_LIST_CNE.xlsx"),
            os.path.join(BASE_DIR, "SCHOOL_REG_LIST_CNE.XLSX"),
            os.path.join(BASE_DIR, "school_reg_list_cne.xlsx"),
            os.path.join(BASE_DIR, "SCHOOL_REG_LIST.XLSX"),
            os.path.join(BASE_DIR, "school_reg_list.xlsx"),
            os.path.join(BASE_DIR, "SCHOOL_REG_LIST.csv"),
        ],
        "AP": {
            "log_prefix": "split_log_AP_CNE",
            "log_exclude": [],
            "source_candidates": [
                os.path.join(BASE_DIR, "00.충남_AP_자산_첨부(충남전체)_.xlsx"),
            ],
            "sheets": ["AP_충남전체", "AP자산", "AP (2)", "AP(2)", "Sheet1"],
            "code_from": "school_col",
        },
        "switch": {
            "log_prefix": "split_log_switch_CNE",
            "log_exclude": [],
            "source_candidates": [
                os.path.join(BASE_DIR, "CNE_SWITCH_LIST.xlsx"),
                os.path.join(BASE_DIR, "CNE_SWITCH_LIST.XLSX"),
            ],
            "sheets": ["SW_충남전체", "스위치_충남전체", "스위치", "Switch", "Sheet1"],
            "code_from": "mgmt_col",
        },
        "security": {
            "log_prefix": "split_log_security_CNE",
            "log_exclude": [],
            "source_candidates": [
                os.path.join(BASE_DIR, "CNE_SEUTM_LIST.xlsx"),
                os.path.join(BASE_DIR, "CNE_SEUTM_LIST.XLSX"),
            ],
            "sheets": ["SEC_충남전체", "보안_충남전체", "보안장비", "Security", "Sheet1"],
            "code_from": "mgmt_col",
        },
        "poe": {
            "log_prefix": "split_log_poe_CNE",
            "log_exclude": [],
            "source_candidates": [
                os.path.join(BASE_DIR, "CNE_POE_LIST.xlsx"),
                os.path.join(BASE_DIR, "CNE_POE_LIST.XLSX"),
            ],
            "sheets": ["POE_충남전체", "POE", "Sheet1"],
            "code_from": "mgmt_col",
        },
    },
}

DEVICE_LABELS = {"AP": "AP", "switch": "스위치", "security": "보안(SEUTM)", "poe": "POE"}


def get_processed_codes(region, device):
    """지정 지역·장비의 split_log에서 처리된 학교코드 집합"""
    cfg = CONFIG.get(region, {}).get(device)
    if not cfg:
        return set()
    prefix = cfg["log_prefix"]
    exclude = cfg.get("log_exclude", [])
    processed = set()
    for fname in os.listdir(BASE_DIR):
        if not (fname.startswith(prefix) and fname.endswith(".csv")):
            continue
        if any(fname.startswith(ex) for ex in exclude):
            continue
        path = os.path.join(BASE_DIR, fname)
        try:
            with open(path, "r", encoding="utf-8-sig") as f:
                for row in csv.DictReader(f):
                    c = row.get("학교코드", "").strip()
                    if c:
                        processed.add(c)
        except Exception:
            pass
    return processed


def load_all_schools(region):
    """지역(DNI/CNE)에 맞는 학교 리스트 로드"""
    region_cfg = CONFIG.get(region, {})
    candidates = region_cfg.get("school_list_candidates") or [
        SCHOOL_LIST_XLSX, os.path.join(BASE_DIR, "school_reg_list.xlsx"), SCHOOL_LIST_CSV
    ]
    path = None
    for cand in candidates:
        if os.path.exists(cand):
            path = cand
            break
    if not path:
        return []
    if region == "CNE" and "CNE" not in os.path.basename(path).upper():
        print(f"[참고] 충남용 SCHOOL_REG_LIST_CNE.xlsx 없음 → {os.path.basename(path)} 사용 (충남 학교 리스트인지 확인)")
    if path.endswith(".csv"):
        with open(path, "r", encoding="utf-8-sig") as f:
            rows = list(csv.reader(f))
        start = 1 if rows and rows[0][0] == "학교코드" else 0
        return [
            {"code": str(r[0]).strip(), "region": str(r[1]).strip(), "name": str(r[2]).strip()}
            for r in rows[start:]
            if len(r) >= 3 and (r[0] or r[1] or r[2])
        ]
    wb = load_workbook(path, read_only=True, data_only=True)
    schools = []
    for row in wb.active.iter_rows(min_row=1, values_only=True):
        row = list(row) if row else []
        while len(row) < 3:
            row.append(None)
        if row[0] or row[1] or row[2]:
            schools.append(
                {
                    "code": str(row[0]).strip() if row[0] else "",
                    "region": str(row[1]).strip() if row[1] else "",
                    "name": str(row[2]).strip() if row[2] else "",
                }
            )
    wb.close()
    return schools


def get_source_codes(region, device):
    """원본 시트에서 학교코드 집합 반환 (장비별 열 처리)"""
    cfg = CONFIG.get(region, {}).get(device)
    if not cfg:
        return set(), None, None

    path = None
    for cand in cfg["source_candidates"]:
        if os.path.exists(cand):
            path = cand
            break
    if not path:
        return set(), None, None

    wb = load_workbook(path, read_only=False, data_only=True)
    for sn in cfg["sheets"]:
        if sn in wb.sheetnames:
            ws = wb[sn]
            vals = set()
            if cfg["code_from"] == "school_col":
                school_col = find_school_code_col(ws, 2) or 1
                for row in ws.iter_rows(min_row=3, min_col=school_col, max_col=school_col, values_only=True):
                    v = row[0] if row else None
                    if v is not None:
                        s = str(v).strip()
                        if s:
                            vals.add(s)
            else:
                mgmt_col = find_mgmt_col(ws, 2)
                if mgmt_col:
                    for row in ws.iter_rows(min_row=3, min_col=mgmt_col, max_col=mgmt_col, values_only=True):
                        v = row[0] if row else None
                        if v is not None:
                            code = extract_school_code_from_mgmt_num(v)
                            if code:
                                vals.add(code)
            wb.close()
            return vals, sn, path
    wb.close()
    return set(), None, None


def main(region, device):
    cfg = CONFIG.get(region, {}).get(device)
    if not cfg:
        print(f"오류: 지원하지 않는 조합 (region={region}, device={device})")
        return

    device_label = DEVICE_LABELS.get(device, device)
    print(f"=== {region} / {device_label} 빠진 학교 검증 ===\n")

    processed = get_processed_codes(region, device)
    all_schools = load_all_schools(region)
    if not all_schools:
        print("SCHOOL_REG_LIST 없음")
        return
    if all_schools[0].get("code") == "학교코드":
        all_schools = all_schools[1:]
    seen = set()
    unique = []
    for s in all_schools:
        if s["code"] and s["code"] not in seen:
            seen.add(s["code"])
            unique.append(s)

    missed = [s for s in unique if s["code"] not in processed]
    print(f"학교코드 기준 빠진 학교: {len(missed)}개\n")

    region_cfg = CONFIG.get(region, {})
    candidates = region_cfg.get("school_list_candidates") or [SCHOOL_LIST_XLSX, os.path.join(BASE_DIR, "school_reg_list.xlsx"), SCHOOL_LIST_CSV]
    for cand in candidates:
        if os.path.exists(cand):
            print(f"[학교리스트] {cand}")
            break

    source_vals, sheet_name, source_path = get_source_codes(region, device)
    if source_path:
        print(f"[원본] {source_path}")
    if sheet_name:
        print(f"{device_label} 시트 '{sheet_name}' 고유 학교코드: {len(source_vals)}개\n")
    else:
        print(f"원본 파일/시트를 찾을 수 없습니다.\n")

    for s in missed:
        code = s["code"]
        in_src = "O" if code in source_vals else "X"
        print(f"  {s['name']} ({code}) / {s['region']}  → 원본: {in_src}")

    not_in_src = [s for s in missed if s["code"] not in source_vals]
    if not_in_src:
        print(f"\n※ 원본 시트에 코드가 없어 매칭 안 됨: {len(not_in_src)}개")
        for s in not_in_src:
            print(f"    - {s['code']} ({s['name']})")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="학교코드 기준 빠진 학교 검증 (지역·장비별)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
[지역] --DNI 또는 --CNE 중 하나 필수
[장비] --AP, --switch, --security, --poe 중 하나 필수 (디폴트 없음)

예시:
  python verify_missing_by_code.py --DNI --AP
  python verify_missing_by_code.py --CNE --switch
  python verify_missing_by_code.py --dni -e security
  python verify_missing_by_code.py --cne --poe
        """,
    )
    reg = parser.add_mutually_exclusive_group(required=True)
    reg.add_argument("--DNI", "--dni", action="store_true", help="대전 (DJE) 지역")
    reg.add_argument("--CNE", "--cne", action="store_true", help="충남 지역")

    dev = parser.add_mutually_exclusive_group(required=True)
    dev.add_argument("--AP", "--ap", action="store_true", help="AP 장비")
    dev.add_argument("--switch", "--Switch", "--SWITCH", action="store_true", help="스위치 장비")
    dev.add_argument("--security", "--Security", "--SECURITY", action="store_true", help="보안(SEUTM) 장비")
    dev.add_argument("--poe", "--Poe", "--POE", action="store_true", help="POE 장비")
    def _device_type(s):
        m = {"ap": "AP", "switch": "switch", "security": "security", "poe": "poe"}
        v = m.get((s or "").lower())
        if v is None:
            raise argparse.ArgumentTypeError(f"장비는 AP, switch, security, poe 중 하나: {s}")
        return v

    dev.add_argument("-e", "--device", type=_device_type,
                     help="장비 지정 (AP/switch/security/poe, 대소문자 무관)")

    args = parser.parse_args()

    region = "CNE" if args.CNE else "DNI"

    device = args.device if args.device else ("AP" if args.AP else "switch" if args.switch else "security" if args.security else "poe")

    main(region, device)
