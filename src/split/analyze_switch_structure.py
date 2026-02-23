# -*- coding: utf-8 -*-
"""스위치 원본 구조 분석 → structure_report.txt에 저장"""
import openpyxl
import os

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# 충남: SOURCE = os.path.join(BASE_DIR, "CNE_SWITCH_LIST.xlsx")
SOURCE = os.path.join(BASE_DIR, "DJE_SWITCH_LIST.xlsx")
OUTPUT = os.path.join(BASE_DIR, "structure_report.txt")
TARGET_CODE = "N108140237HS"  # 검색용 (충남: 계룡고등학교)

def main():
    lines = []
    if not os.path.exists(SOURCE):
        lines.append(f"파일 없음: {SOURCE}")
        with open(OUTPUT, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))
        return

    wb = openpyxl.load_workbook(SOURCE, read_only=False, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"\n=== 시트: {sheet_name} | 행:{ws.max_row} 열:{ws.max_column} ===")
        for row in range(1, min(10, ws.max_row + 1)):
            vals = []
            for c in range(1, min(15, ws.max_column + 1)):
                v = ws.cell(row=row, column=c).value
                vals.append(str(v)[:20] if v is not None else "")
            lines.append(f"  Row{row}: {vals}")

        # N108140237HS 검색
        lines.append(f"\n  --- '{TARGET_CODE}' 검색 ---")
        found = []
        for r in range(1, min(ws.max_row + 1, 500)):
            for c in range(1, min(ws.max_column + 1, 20)):
                v = ws.cell(row=r, column=c).value
                if v and TARGET_CODE in str(v):
                    found.append((r, c, str(v)[:50]))
        lines.append(f"  발견: {found[:10]}")
    wb.close()

    with open(OUTPUT, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))
    print("저장:", OUTPUT)

if __name__ == "__main__":
    main()
