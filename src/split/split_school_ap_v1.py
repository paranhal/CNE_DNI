


"""
학교별 AP 장비 현황 엑셀 분리 스크립트 (V1)
- 학교코드 기준 작업 (학교명은 지역별로 다를 수 있음)
- 원본 서식(행높이, 열넓이, 셀색상, 폰트, 맞춤) 유지
- OUTPUT/지역/학교명_학교코드/학교명_AP 장비 현황 상세.XLSX (A1: "AP 자산 상세")
- 로그: split_log_AP_{지역}_{날짜}.csv (예: split_log_AP_DNI_20260222.csv)
- --missed-only: 기존 로그에 없는 빠진 학교만 처리 (학교코드 기준)
- --only-schools: 지정한 학교코드만 처리 (쉼표 구분 또는 파일 경로, 어제 파일 건드리지 않음)
- --from-log: split_log_YYYYMMDD.csv 등 로그 파일의 학교만 처리 (오늘 작업분만 생성)
"""
import os
import csv
import re
import argparse
from datetime import datetime
from copy import copy
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side
from school_utils import extract_school_code_from_mgmt_num, find_mgmt_col, find_school_code_col, get_output_cols, sort_schools_by_region, get_school_list_path
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DNI_DIR = os.path.join(BASE_DIR, "DNI")
# 지역별 설정 (조건문으로 선택)
SOURCE_BY_REGION = {
    "DNI": os.path.join(DNI_DIR, "DNI_AP_LIST.XLSX"),
    "CNE": os.path.join(BASE_DIR, "00.충남_AP_자산_첨부(충남전체)_.xlsx"),
}
OUTPUT_BASE_BY_REGION = {
    "DNI": r"Y:\DJE_DNI\_지역별 산출물 취합\_00.별첨자료_학교별\DJE",
    "CNE": r"Y:\CNE_DNI\_지역별 산출물 취합\_00.별첨자료_학교별\CNE",
}
MISSED_SCHOOLS_FILE = os.path.join(BASE_DIR, "missed_schools.csv")
OUTPUT_BASE_TEST = os.path.join(BASE_DIR, "OUTPUT")


def get_today_log_path(region='DNI', suffix=None):
    """오늘 날짜가 붙은 로그 파일 경로 (split_log_AP_{지역}_{날짜}.csv)"""
    base = f"split_log_AP_{region}_{datetime.now().strftime('%Y%m%d')}"
    if suffix:
        return os.path.join(BASE_DIR, f"{base}_{suffix}.csv")
    return os.path.join(BASE_DIR, f"{base}.csv")


def get_processed_school_codes(region='DNI'):
    """지정 지역의 AP 로그에서 처리된 학교코드 집합 반환 (split_log_AP_{지역}_*.csv)"""
    prefix = f"split_log_AP_{region}_"
    processed = set()
    for fname in os.listdir(BASE_DIR):
        if fname.startswith(prefix) and fname.endswith('.csv'):
            path = os.path.join(BASE_DIR, fname)
            try:
                with open(path, 'r', encoding='utf-8-sig') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        code = row.get('학교코드', '').strip()
                        if code:
                            processed.add(code)
            except Exception:
                pass
    return processed


def load_existing_log(path):
    """지정 경로의 로그 파일 로드 (utf-8/cp949 자동 감지)"""
    entries = []
    if path and os.path.exists(path):
        for enc in ('utf-8-sig', 'utf-8', 'cp949'):
            try:
                with open(path, 'r', encoding=enc) as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        entries.append(row)
                break
            except UnicodeDecodeError:
                continue
    return entries

def normalize_code(val):
    """학교코드 정규화 (Excel 숫자/공백 등 처리)"""
    if val is None:
        return ''
    s = str(val).strip()
    # Excel이 숫자로 저장한 경우 (예: 1.08140237E+8)
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        s = str(int(val)) if val == int(val) else str(val)
    return s


# 테두리 (다른 장비 출력과 동일)
THIN_BORDER = Border(
    left=Side(border_style='thin'),
    right=Side(border_style='thin'),
    top=Side(border_style='thin'),
    bottom=Side(border_style='thin')
)


# 파일명에 사용할 수 없는 문자 제거
def sanitize_filename(name):
    if name is None:
        return "Unknown"
    return re.sub(r'[\\/:*?"<>|]', '_', str(name).strip())


def copy_cell_style(src_cell, tgt_cell, copy_fill=True):
    """셀 서식 복사 (폰트, 채우기, 테두리, 맞춤, 숫자형식). copy_fill=False면 실데이터용으로 배경색 제외"""
    try:
        if src_cell.font:
            tgt_cell.font = copy(src_cell.font)
    except Exception:
        pass
    try:
        if copy_fill and src_cell.fill and getattr(src_cell.fill, 'fill_type', None):
            tgt_cell.fill = copy(src_cell.fill)
    except Exception:
        pass
    try:
        if src_cell.border:
            b = src_cell.border
            if (b.left and b.left.style) or (b.right and b.right.style) or \
               (b.top and b.top.style) or (b.bottom and b.bottom.style):
                tgt_cell.border = copy(src_cell.border)
    except Exception:
        pass
    try:
        if src_cell.alignment:
            tgt_cell.alignment = copy(src_cell.alignment)
    except Exception:
        pass
    try:
        if src_cell.number_format:
            tgt_cell.number_format = src_cell.number_format
    except Exception:
        pass


def main(missed_only=False, source_file=None, only_schools=None, schools_file=None, today_from_missed=False, from_log=None, new_log=False, test_output=False, region_key='DNI'):
    source_path = source_file or SOURCE_BY_REGION.get(region_key, SOURCE_BY_REGION["DNI"])
    output_base = OUTPUT_BASE_TEST if test_output else OUTPUT_BASE_BY_REGION.get(region_key, OUTPUT_BASE_BY_REGION["DNI"])
    if test_output:
        print(f"[테스트] 출력 경로: {output_base}")
    if not os.path.exists(source_path):
        print(f"오류: 원본 파일을 찾을 수 없습니다. {source_path}")
        return
    school_list_path = get_school_list_path(region_key, BASE_DIR)
    if not os.path.exists(school_list_path):
        print(f"오류: 학교 리스트 파일을 찾을 수 없습니다. (지역: {region_key})")
        return

    # 학교 리스트 로드 (A:학교코드, B:지역, C:학교명)
    from openpyxl import load_workbook as load_wb
    wb_school = load_wb(school_list_path, read_only=True, data_only=True)
    ws_school = wb_school.active
    school_list = []
    for row in ws_school.iter_rows(min_row=1, values_only=True):
        row = list(row) if row else []
        while len(row) < 3:
            row.append(None)
        if row[0] is None and row[1] is None and row[2] is None:
            continue
        code = row[0]
        region = row[1]
        name = row[2]
        if code or region or name:
            school_list.append({
                'code': str(code).strip() if code else '',
                'region': str(region).strip() if region else '',
                'name': str(name).strip() if name else ''
            })
    wb_school.close()

    # 헤더 행 제거 (첫 행이 헤더일 수 있음)
    if school_list and (school_list[0].get('code') == '학교코드' or 
                        school_list[0].get('region') == '지역' or 
                        school_list[0].get('name') == '학교명'):
        school_list = school_list[1:]

    # 학교코드 기준 중복 제거 (같은 이름 다른 학교 있음 - 코드가 유일 식별자)
    seen_codes = set()
    school_list_unique = []
    for s in school_list:
        if s['code'] and s['code'] not in seen_codes:
            seen_codes.add(s['code'])
            school_list_unique.append(s)
    school_list = school_list_unique

    # --missed-only: 기존 로그에 없는 학교만 처리 (학교코드 기준, 해당 지역 로그만)
    if missed_only:
        processed_codes = get_processed_school_codes(region_key)
        school_list = [s for s in school_list if s['code'] and s['code'] not in processed_codes]
        print(f"[빠진 학교만 처리] 기존 처리: {len(processed_codes)}개, 미처리: {len(school_list)}개")
        if not school_list:
            print("처리할 빠진 학교가 없습니다.")
            return

    # --only-schools / --today-from-missed / --from-log: 지정한 학교만 처리 (오늘 작업한 파일만 생성, 어제 파일 건드리지 않음)
    only_codes = set()
    if from_log:
        log_path = from_log if os.path.isabs(from_log) else os.path.join(BASE_DIR, from_log)
        if os.path.exists(log_path):
            schools_file = log_path  # 로그 파일도 학교코드 컬럼 있음
    elif today_from_missed and not schools_file and os.path.exists(MISSED_SCHOOLS_FILE):
        schools_file = MISSED_SCHOOLS_FILE
    if only_schools:
        only_codes.update(c.strip() for c in only_schools.split(',') if c.strip())
    if schools_file and os.path.exists(schools_file):
        if schools_file.lower().endswith('.csv'):
            with open(schools_file, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    code = row.get('학교코드', '').strip()
                    if code:
                        only_codes.add(code)
        else:
            with open(schools_file, 'r', encoding='utf-8-sig') as f:
                for line in f:
                    code = line.strip()
                    if code and not code.startswith('#'):
                        only_codes.add(code)
    if only_codes:
        school_list = [s for s in school_list if s['code'] and s['code'] in only_codes]
        print(f"[오늘 작업만] 지정 학교 {len(only_codes)}개 중 원본에 데이터 있는 학교: {len(school_list)}개")
        if not school_list:
            print("처리할 학교가 없습니다. 원본에 해당 학교코드 데이터가 있는지 확인하세요.")
            return

    print(f"학교 리스트: {len(school_list)}개")
    for s in school_list[:5]:
        print(f"  - {s['name']} ({s['code']}) / {s['region']}")
    if len(school_list) > 5:
        print(f"  ... 외 {len(school_list)-5}개")

    # 원본 시트 로드 (서식 유지를 위해 data_only=False)
    # 충남: ["AP_충남전체", "AP (2)", "AP(2)"]
    wb_src = load_workbook(source_path, read_only=False, data_only=False)
    sheet_name = None
    for candidate in ["AP자산", "AP_대전전체", "AP (2)", "AP(2)", "Sheet1"]:  # "AP_충남전체" 주석
        if candidate in wb_src.sheetnames:
            sheet_name = candidate
            break
    if not sheet_name:
        print(f"오류: AP 시트를 찾을 수 없습니다. 시트 목록: {wb_src.sheetnames}")
        wb_src.close()
        return

    ws_src = wb_src[sheet_name]
    max_row = ws_src.max_row
    HEADER_ROW = 2
    DATA_START_ROW = 3

    # 학교코드 열 또는 관리번호 열로 매칭 (대전 DNI: A열 학교코드)  # 충남: 관리번호
    school_code_col = find_school_code_col(ws_src, HEADER_ROW)
    mgmt_col = find_mgmt_col(ws_src, HEADER_ROW)
    if not school_code_col and not mgmt_col:
        print("경고: '학교코드' 또는 '관리번호' 열을 찾을 수 없습니다.")
        wb_src.close()
        return

    # 출력할 열: 학교코드 제외
    src_cols = get_output_cols(ws_src, HEADER_ROW, exclude_school_code=True)

    # 행 그룹화: 학교코드 열 있으면 직접 사용, 없으면 관리번호 앞 12자리
    school_codes = {normalize_code(s['code']): s for s in school_list if s['code']}
    rows_by_school = {}
    for r in range(DATA_START_ROW, max_row + 1):
        if school_code_col:
            extracted = normalize_code(ws_src.cell(row=r, column=school_code_col).value)
        else:
            mgmt_val = ws_src.cell(row=r, column=mgmt_col).value
            if not mgmt_val:
                continue
            extracted = normalize_code(extract_school_code_from_mgmt_num(mgmt_val))
        if extracted and extracted in school_codes:
            rows_by_school.setdefault(extracted, []).append(r)

    schools_with_data = [(school_codes[code], sorted(rows))
                         for code, rows in rows_by_school.items()]
    schools_with_data = sort_schools_by_region(schools_with_data, region_key)

    if not schools_with_data:
        print("\n경고: 학교코드와 매칭되는 데이터가 없습니다.")
        print(f"  - 원본 {sheet_name} 시트의 A열에 학교코드가 있는지 확인하세요.")
        print("  - check_structure_ap.py를 실행하여 데이터 구조를 확인해보세요.")
        wb_src.close()
        return

    log_entries = []
    failed_schools = []

    for item in schools_with_data:
        school, data = item
        code = school['code']
        region = school['region']
        name = school['name']
        data_rows = sorted(data)
        start_row = min(data_rows)
        end_row = max(data_rows)
        row_count = len(data_rows)

        # 출력 경로: OUTPUT/지역/학교명_학교코드/학교명_AP 장비 현황 상세.XLSX
        # 학교코드 기준으로 폴더/파일 생성 (학교명은 지역 붙어서 다를 수 있음)
        region_safe = sanitize_filename(region) if region else "기타"
        display_name = sanitize_filename(name) if name else f"학교_{code}"
        school_folder = f"{display_name}_{sanitize_filename(code)}"  # 코드 포함해 동일명 구분
        out_dir = os.path.join(output_base, region_safe, school_folder)
        os.makedirs(out_dir, exist_ok=True)
        out_filename = f"{display_name}_AP 장비 현황 상세.XLSX"
        out_path = os.path.join(out_dir, out_filename)

        # 새 워크북 생성
        from openpyxl import Workbook
        try:
            wb_out = Workbook()
            ws_out = wb_out.active
            ws_out.title = "AP 장비 현황"

            # 1행: 제목 A1 셀 "AP 자산 상세" (원본 B1 값, 왼쪽 맞춤) + 병합 셀 복사
            src_cell_a1 = ws_src.cell(row=1, column=1)
            title_val = ws_src.cell(row=1, column=2).value  # B1에 "AP 자산 상세"
            ws_out['A1'] = title_val if title_val else "AP 자산 상세"
            copy_cell_style(src_cell_a1, ws_out['A1'])
            ws_out['A1'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            if 1 in ws_src.row_dimensions and ws_src.row_dimensions[1].height:
                ws_out.row_dimensions[1].height = ws_src.row_dimensions[1].height
            for merged in list(ws_src.merged_cells.ranges):
                if merged.min_row <= 1 <= merged.max_row:
                    try:
                        ws_out.merge_cells(str(merged))
                    except Exception:
                        pass

            # 2행: 제목행 (모두 중간 맞춤, 테두리)
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            for c_idx, src_col in enumerate(src_cols, start=1):
                src_cell = ws_src.cell(row=HEADER_ROW, column=src_col)
                tgt_cell = ws_out.cell(row=2, column=c_idx)
                tgt_cell.value = src_cell.value
                copy_cell_style(src_cell, tgt_cell)
                tgt_cell.alignment = center_align
                tgt_cell.border = THIN_BORDER

            # 행 높이 복사
            if HEADER_ROW in ws_src.row_dimensions and ws_src.row_dimensions[HEADER_ROW].height:
                ws_out.row_dimensions[2].height = ws_src.row_dimensions[HEADER_ROW].height

            # 3행~: 데이터 (모두 중간 맞춤, 테두리)
            out_row = 3
            for src_row in data_rows:
                for c_idx, src_col in enumerate(src_cols, start=1):
                    src_cell = ws_src.cell(row=src_row, column=src_col)
                    tgt_cell = ws_out.cell(row=out_row, column=c_idx)
                    tgt_cell.value = src_cell.value
                    copy_cell_style(src_cell, tgt_cell, copy_fill=False)
                    tgt_cell.alignment = center_align
                    tgt_cell.border = THIN_BORDER
                if src_row in ws_src.row_dimensions and ws_src.row_dimensions[src_row].height:
                    ws_out.row_dimensions[out_row].height = ws_src.row_dimensions[src_row].height
                out_row += 1

            # 열 넓이 복사
            for c_idx, src_col in enumerate(src_cols, start=1):
                col_letter = get_column_letter(src_col)
                if col_letter in ws_src.column_dimensions and ws_src.column_dimensions[col_letter].width:
                    out_col_letter = get_column_letter(c_idx)
                    ws_out.column_dimensions[out_col_letter].width = ws_src.column_dimensions[col_letter].width

            wb_out.save(out_path)
            wb_out.close()
            log_entries.append({
                '학교명': name, '학교코드': code, '지역': region,
                '시작행': start_row, '끝행': end_row, '복사행수': row_count, '저장경로': out_path
            })
            print(f"  저장: {name} ({code}) - {row_count}행 -> {out_path}")
        except Exception as e:
            print(f"  [에러] {name} ({code}) {region}: {e}")
            failed_schools.append({'name': name, 'code': code, 'region': region, 'error': str(e)})

    if failed_schools:
        print(f"\n[실패 {len(failed_schools)}개]")
        for f in failed_schools:
            print(f"  - {f['name']} ({f['code']}) {f['region']}: {f['error']}")

    wb_src.close()

    # 로그 CSV 저장: 기존 오늘 로그에 병합(학교코드 기준), --new-log이면 신규 파일 생성
    log_path = get_today_log_path(region_key) if not new_log else get_today_log_path(
        region_key, datetime.now().strftime('%H%M%S'))
    existing = load_existing_log(log_path) if not new_log and log_path and os.path.exists(log_path) else []
    by_code = {e.get('학교코드', '').strip(): e for e in existing}
    for e in log_entries:
        by_code[e.get('학교코드', '').strip()] = e
    merged = list(by_code.values())
    with open(log_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=['학교명', '학교코드', '지역', '시작행', '끝행', '복사행수', '저장경로'])
        writer.writeheader()
        writer.writerows(merged)

    print(f"\n완료. 로그: {log_path} (총 {len(merged)}개 학교, 이번 +{len(log_entries)}개)")
    print(f"이번 실행: {len(log_entries)}개 학교 파일 생성")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='학교별 AP 장비 현황 엑셀 분리 (V1)')
    parser.add_argument('--missed-only', action='store_true', help='빠진 학교만 처리')
    parser.add_argument('--only-schools', type=str, default=None,
                        help='오늘 작업한 학교만 처리 (학교코드 쉼표 구분, 예: 108140237,108140238)')
    parser.add_argument('--schools-file', type=str, default=None,
                        help='처리할 학교코드 목록 파일 (한 줄에 하나씩 또는 CSV)')
    parser.add_argument('--today-from-missed', action='store_true',
                        help='missed_schools.csv에 있는 학교만 처리 (오늘 데이터 넣은 학교만, 어제 파일 건드리지 않음)')
    parser.add_argument('--from-log', type=str, default=None,
                        help='지정한 로그 파일의 학교만 처리 (예: split_log_AP_DNI_20260222.csv)')
    parser.add_argument('--new-log', action='store_true',
                        help='이번 실행만 별도 로그 파일 생성 (split_log_YYYYMMDD_HHMMSS.csv, 기존 로그에 추가 안 함)')
    parser.add_argument('--source', '-s', type=str, default=None,
                        help='원본 엑셀 파일 경로 (기본: DNI/DNI_AP_LIST.XLSX)')
    parser.add_argument('--test', '-t', action='store_true',
                        help='테스트: OUTPUT 폴더에 생성. 없으면 실제 목표 저장 폴더(Y:\\...\\DJE)에 저장')
    parser.add_argument('--DNI', '--dni', action='store_true', help='대전 (로그: split_log_AP_DNI_YYYYMMDD.csv)')
    parser.add_argument('--CNE', '--cne', action='store_true', help='충남 (로그: split_log_AP_CNE_YYYYMMDD.csv)')
    args = parser.parse_args()
    region_key = 'CNE' if args.CNE else 'DNI'  # --CNE 우선, 기본 DNI
    main(missed_only=args.missed_only, source_file=args.source,
         only_schools=args.only_schools, schools_file=args.schools_file,
         today_from_missed=args.today_from_missed, from_log=args.from_log, new_log=args.new_log,
         test_output=args.test, region_key=region_key)
