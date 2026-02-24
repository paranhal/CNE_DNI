"""
사전 조사: 충남 전체 폴더·모든 엑셀·모든 시트 스캔.

- 빈 폴더 목록
- 모든 xlsx 파일에 대해 모든 시트 진입, 시트명·행 수 기록
- 719개 학교 자료 위치 파악용
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from tqdm import tqdm

# 충남 원본 루트 (사전조사 대상)
DEFAULT_CHUNNAM_ROOT = Path(
    "/Users/paranhal/Library/CloudStorage/GoogleDrive-paranhanl66@gmail.com"
    "/내 드라이브/20260215_D드라이브/Lee_20260202/충남"
)


def scan_folders(root: Path) -> list[dict[str, Any]]:
    """root 하위 모든 폴더 스캔. 빈 폴더·xlsx 개수 기록."""
    rows = []
    root = Path(root).resolve()
    for d in tqdm(sorted(root.rglob("*")), desc="폴더 스캔", unit="폴더"):
        if not d.is_dir():
            continue
        try:
            all_files = list(d.iterdir())
            xlsx = [f for f in all_files if f.suffix.lower() == ".xlsx"]
            is_empty = len(all_files) == 0
            rel = d.relative_to(root) if root != d else Path(".")
            rows.append({
                "folder_path": str(rel),
                "folder_abs": str(d),
                "is_empty": is_empty,
                "file_count": len(all_files),
                "xlsx_count": len(xlsx),
            })
        except Exception as e:
            rows.append({
                "folder_path": str(d.relative_to(root)) if d != root else ".",
                "folder_abs": str(d),
                "is_empty": None,
                "file_count": None,
                "xlsx_count": None,
                "error": str(e),
            })
    return rows


def scan_sheets_in_file(xlsx_path: Path) -> list[dict[str, Any]]:
    """엑셀 파일 하나 열어서 모든 시트의 이름·행 수·1행 샘플 반환."""
    import openpyxl
    out = []
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=False, data_only=True)
        for name in wb.sheetnames:
            try:
                ws = wb[name]
                max_row = ws.max_row
                max_col = ws.max_column
                first_row = []
                for c in range(1, min(max_col + 1, 16)):
                    v = ws.cell(row=1, column=c).value
                    first_row.append(str(v)[:25] if v is not None else "")
                out.append({
                    "sheet_name": name,
                    "max_row": max_row,
                    "max_column": max_col,
                    "first_row_sample": "|".join(first_row),
                })
            except Exception as e:
                out.append({"sheet_name": name, "max_row": None, "max_column": None, "first_row_sample": "", "error": str(e)})
        wb.close()
    except Exception as e:
        out.append({"sheet_name": None, "max_row": None, "max_column": None, "first_row_sample": "", "error": str(e)})
    return out


def run_survey(
    root: Path | str | None = None,
    output_dir: Path | str | None = None,
    max_files: int | None = None,
) -> dict[str, Any]:
    """
    사전 조사 실행: 폴더 스캔 → 빈 폴더·파일 목록 → 모든 xlsx의 모든 시트 조사.
    """
    root = Path(root) if root else DEFAULT_CHUNNAM_ROOT
    if not root.exists():
        return {"error": f"루트 없음: {root}"}

    if output_dir is None:
        output_dir = Path(__file__).resolve().parent.parent / "output"
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    # 1) 폴더 스캔
    folder_rows = scan_folders(root)
    df_folders = pd.DataFrame(folder_rows)
    empty_folders = df_folders[df_folders["is_empty"] == True] if "is_empty" in df_folders.columns else pd.DataFrame()
    folders_with_no_xlsx = df_folders[(df_folders["xlsx_count"] == 0) & (df_folders["is_empty"] == False)] if "xlsx_count" in df_folders.columns else pd.DataFrame()

    # 2) 모든 xlsx 목록
    xlsx_files = list(root.rglob("*.xlsx"))
    xlsx_files = [f for f in xlsx_files if f.is_file()]
    if max_files is not None and max_files > 0:
        xlsx_files = xlsx_files[: max_files]

    # 3) 파일별·시트별 조사
    file_sheet_rows = []
    for xpath in tqdm(xlsx_files, desc="엑셀 시트 조사", unit="파일"):
        try:
            rel = xpath.relative_to(root)
        except ValueError:
            rel = xpath
        for sh in scan_sheets_in_file(xpath):
            file_sheet_rows.append({
                "file_path": str(rel),
                "file_name": xpath.name,
                "sheet_name": sh.get("sheet_name"),
                "max_row": sh.get("max_row"),
                "max_column": sh.get("max_column"),
                "first_row_sample": sh.get("first_row_sample", ""),
                "error": sh.get("error"),
            })

    df_sheets = pd.DataFrame(file_sheet_rows)

    # 4) 저장
    df_folders.to_csv(output_dir / "사전조사_폴더목록.csv", index=False, encoding="utf-8-sig")
    df_sheets.to_csv(output_dir / "사전조사_파일별_모든시트.csv", index=False, encoding="utf-8-sig")
    if not empty_folders.empty:
        empty_folders.to_csv(output_dir / "사전조사_빈폴더목록.csv", index=False, encoding="utf-8-sig")
    if not folders_with_no_xlsx.empty:
        folders_with_no_xlsx.to_csv(output_dir / "사전조사_엑셀없는_폴더목록.csv", index=False, encoding="utf-8-sig")

    # 5) 요약 로그
    log_path = output_dir / "사전조사_요약.txt"
    with open(log_path, "w", encoding="utf-8") as f:
        f.write(f"사전 조사 루트: {root}\n")
        f.write(f"총 폴더 수: {len(df_folders)}\n")
        f.write(f"빈 폴더 수: {len(empty_folders)}\n")
        f.write(f"엑셀 없는 폴더 수(빈 폴더 제외): {len(folders_with_no_xlsx)}\n")
        f.write(f"엑셀 파일 수: {len(xlsx_files)}\n")
        f.write(f"시트 레코드 수(파일×시트): {len(df_sheets)}\n")
        if not df_sheets.empty and "max_row" in df_sheets.columns:
            with_data = df_sheets[df_sheets["max_row"].fillna(0).astype(int) > 0]
            f.write(f"데이터 있는 시트(행>0): {len(with_data)}\n")
        f.write("\n출력 파일:\n")
        f.write("  - 사전조사_폴더목록.csv\n")
        f.write("  - 사전조사_파일별_모든시트.csv\n")
        f.write("  - 사전조사_빈폴더목록.csv\n")
        f.write("  - 사전조사_엑셀없는_폴더목록.csv\n")

    return {
        "root": str(root),
        "output_dir": str(output_dir),
        "folder_count": len(df_folders),
        "empty_folder_count": len(empty_folders),
        "no_xlsx_folder_count": len(folders_with_no_xlsx),
        "xlsx_file_count": len(xlsx_files),
        "sheet_record_count": len(df_sheets),
        "log_path": str(log_path),
    }


def main() -> None:
    import sys
    try:
        from .config_loader import get_path, ensure_runtime_dirs
        ensure_runtime_dirs()
        out = get_path("output_root")
    except Exception:
        out = Path(__file__).resolve().parent.parent / "output"

    root = DEFAULT_CHUNNAM_ROOT
    max_files = None
    if len(sys.argv) > 1:
        root = Path(sys.argv[1])
    if len(sys.argv) > 2:
        out = Path(sys.argv[2])
    if len(sys.argv) > 3:
        try:
            max_files = int(sys.argv[3])
            print("제한: 엑셀 파일", max_files, "개만 조사")
        except ValueError:
            pass

    print("사전 조사 실행:", root)
    r = run_survey(root=root, output_dir=out, max_files=max_files)
    if "error" in r:
        print("오류:", r["error"])
        return
    print("폴더 수:", r["folder_count"])
    print("빈 폴더 수:", r["empty_folder_count"])
    print("엑셀 없는 폴더 수:", r["no_xlsx_folder_count"])
    print("엑셀 파일 수:", r["xlsx_file_count"])
    print("시트 레코드 수:", r["sheet_record_count"])
    print("요약 로그:", r["log_path"])


if __name__ == "__main__":
    main()
